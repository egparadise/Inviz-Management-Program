# -*- coding: utf-8 -*-
"""P8: 01_Dashboard에 KPI 수식·차트 삽입 + 검증 리포트 생성"""
import pandas as pd
from datetime import datetime
from openpyxl.chart import LineChart, BarChart, PieChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from common import load_master, save_master, MASTER

COLORS = {
    "primary": "1F4E79", "accent": "2E75B6",
    "dash_bg": "FCE4D6", "ok": "548235", "warning": "C00000",
}


def update_dashboard(wb):
    ws = wb["01_Dashboard"]

    # 기존 placeholder 제거 — 전체 클리어
    ws.delete_rows(1, ws.max_row)

    # === 타이틀 ===
    ws.column_dimensions["A"].width = 2
    for col in "BCDEFGHIJKLMNOP":
        ws.column_dimensions[col].width = 13

    ws["B2"] = "인비즈 경영 대시보드"
    ws["B2"].font = Font(name="맑은 고딕", size=20, bold=True, color=COLORS["primary"])
    ws["B3"] = f"갱신: {datetime.now().strftime('%Y-%m-%d %H:%M')} | FACT 시트 기준 자동 계산"
    ws["B3"].font = Font(name="맑은 고딕", size=10, italic=True, color="808080")

    # === 연간 KPI ===
    ws["B5"] = "■ 연간 실적 (단위: 원)"
    ws["B5"].font = Font(name="맑은 고딕", size=12, bold=True, color=COLORS["primary"])

    # 헤더
    yr_headers = ["구분", "2021", "2022", "2023", "2024", "2025", "2026"]
    for i, h in enumerate(yr_headers):
        c = ws.cell(row=7, column=2 + i, value=h)
        c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COLORS["primary"])
        c.alignment = Alignment(horizontal="center")

    metrics = [
        ("매출", '=SUMIFS(tbl_매출[공급가액],tbl_매출[년],{y})'),
        ("매입", '=SUMIFS(tbl_매입[공급가액],tbl_매입[년],{y})'),
        ("매출이익", "=B{r}-B{r2}"),  # 계산 식별 후 갱신
        ("급여", '=SUMIFS(tbl_급여[지급합계],tbl_급여[년],{y})'),
        ("비용", '=SUMIFS(tbl_비용[금액],tbl_비용[년],{y})'),
        ("영업이익", None),  # 후처리
        ("이익률(%)", None),
    ]
    years = [2021, 2022, 2023, 2024, 2025, 2026]
    for r_offset, (name, _) in enumerate(metrics):
        row = 8 + r_offset
        ws.cell(row=row, column=2, value=name).font = Font(name="맑은 고딕", size=10, bold=True)
        ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="D9E1F2")
        for ci, y in enumerate(years):
            col = 3 + ci
            cell = ws.cell(row=row, column=col)
            if name == "매출":
                cell.value = f'=SUMIFS(tbl_매출[공급가액],tbl_매출[년],{y})'
            elif name == "매입":
                cell.value = f'=SUMIFS(tbl_매입[공급가액],tbl_매입[년],{y})'
            elif name == "매출이익":
                cell.value = f'={get_column_letter(col)}8-{get_column_letter(col)}9'
            elif name == "급여":
                cell.value = f'=SUMIFS(tbl_급여[지급합계],tbl_급여[년],{y})'
            elif name == "비용":
                cell.value = f'=SUMIFS(tbl_비용[금액],tbl_비용[년],{y})'
            elif name == "영업이익":
                cell.value = f'={get_column_letter(col)}10-{get_column_letter(col)}11-{get_column_letter(col)}12'
            elif name == "이익률(%)":
                cell.value = f'=IFERROR({get_column_letter(col)}13/{get_column_letter(col)}8*100,0)'
            cell.number_format = "#,##0" if "이익률" not in name else "0.0"
            cell.alignment = Alignment(horizontal="right")

    # === 제품별 매출 비중 (당해, 2025) ===
    ws["B17"] = "■ 제품별 매출 (2025)"
    ws["B17"].font = Font(name="맑은 고딕", size=12, bold=True, color=COLORS["primary"])

    prod_headers = ["제품", "2025매출", "비중(%)"]
    for i, h in enumerate(prod_headers):
        c = ws.cell(row=19, column=2 + i, value=h)
        c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COLORS["primary"])
        c.alignment = Alignment(horizontal="center")

    # 제품 시드 10개 직접 조회 (DIM_제품에서)
    dim_p = wb["11_DIM_제품"]
    products = []
    for row in dim_p.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0]:
            products.append((row[0], row[1]))

    base_row = 20
    for i, (code, name) in enumerate(products):
        r = base_row + i
        ws.cell(row=r, column=2, value=name).font = Font(name="맑은 고딕", size=10)
        ws.cell(row=r, column=3, value=f'=SUMIFS(tbl_매출[공급가액],tbl_매출[제품코드],"{code}",tbl_매출[년],2025)').number_format = "#,##0"
        ws.cell(row=r, column=4, value=f'=IFERROR(C{r}/SUM(C{base_row}:C{base_row + len(products) - 1})*100,0)').number_format = "0.0"

    # === 미수금 잔액 TOP10 ===
    top_start_col = 7  # G열
    ws.cell(row=17, column=top_start_col, value="■ 미수금 잔액 TOP10 (24_FACT_미수금 누계)").font = Font(name="맑은 고딕", size=12, bold=True, color=COLORS["primary"])
    headers_top = ["거래처", "세금계산서누계", "입금누계", "잔액"]
    for i, h in enumerate(headers_top):
        c = ws.cell(row=19, column=top_start_col + i, value=h)
        c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COLORS["primary"])
        c.alignment = Alignment(horizontal="center")

    # 미수금 잔액은 거래처별 세금계산서-입금 차이
    # DIM_거래처에서 추출 — is_미수금 카테고리
    ws_dim = wb["10_DIM_거래처"]
    ar_parties = []
    for row in ws_dim.iter_rows(min_row=2, values_only=True):
        if row and row[1] and row[10] and "미수금" in str(row[10]):
            ar_parties.append((row[0], row[1]))
    ar_parties = ar_parties[:10]  # TOP 10

    for i, (code, name) in enumerate(ar_parties):
        r = 20 + i
        ws.cell(row=r, column=top_start_col, value=name).font = Font(name="맑은 고딕", size=10)
        ws.cell(row=r, column=top_start_col + 1, value=f'=SUMIFS(tbl_미수금[세금계산서금액(증)],tbl_미수금[거래처명],G{r})').number_format = "#,##0"
        ws.cell(row=r, column=top_start_col + 2, value=f'=SUMIFS(tbl_미수금[입금액(감)],tbl_미수금[거래처명],G{r})').number_format = "#,##0"
        ws.cell(row=r, column=top_start_col + 3, value=f'=H{r}-I{r}').number_format = "#,##0"

    # === 차입금 현황 ===
    ws.cell(row=35, column=2, value="■ 차입금 현황").font = Font(name="맑은 고딕", size=12, bold=True, color=COLORS["primary"])
    loan_headers = ["구분", "건수", "최초차입액 합계", "현재잔액 합계"]
    for i, h in enumerate(loan_headers):
        c = ws.cell(row=37, column=2 + i, value=h)
        c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COLORS["primary"])
        c.alignment = Alignment(horizontal="center")
    loan_kinds = ["은행", "개인", "개인(임원)", "사채"]
    for i, k in enumerate(loan_kinds):
        r = 38 + i
        ws.cell(row=r, column=2, value=k).font = Font(name="맑은 고딕", size=10, bold=True)
        ws.cell(row=r, column=3, value=f'=COUNTIFS(tbl_차입금마스터[구분(은행/개인/사채)],"{k}")').number_format = "#,##0"
        ws.cell(row=r, column=4, value=f'=SUMIFS(tbl_차입금마스터[최초차입액],tbl_차입금마스터[구분(은행/개인/사채)],"{k}")').number_format = "#,##0"
        ws.cell(row=r, column=5, value=f'=SUMIFS(tbl_차입금마스터[현재잔액],tbl_차입금마스터[구분(은행/개인/사채)],"{k}")').number_format = "#,##0"
    # 합계
    r = 38 + len(loan_kinds)
    ws.cell(row=r, column=2, value="합계").font = Font(name="맑은 고딕", size=10, bold=True)
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="D9E1F2")
    for col in [3, 4, 5]:
        cl = get_column_letter(col)
        ws.cell(row=r, column=col, value=f'=SUM({cl}38:{cl}{r - 1})').number_format = "#,##0"
        ws.cell(row=r, column=col).font = Font(bold=True)

    # === 계약 현황 ===
    ws.cell(row=35, column=top_start_col, value="■ 계약 현황").font = Font(name="맑은 고딕", size=12, bold=True, color=COLORS["primary"])
    contract_headers = ["상태", "건수", "계약금액 합계", "미수금 합계"]
    for i, h in enumerate(contract_headers):
        c = ws.cell(row=37, column=top_start_col + i, value=h)
        c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COLORS["primary"])
        c.alignment = Alignment(horizontal="center")
    statuses = ["진행", "만료", "해지"]
    for i, s in enumerate(statuses):
        r = 38 + i
        ws.cell(row=r, column=top_start_col, value=s).font = Font(name="맑은 고딕", size=10, bold=True)
        ws.cell(row=r, column=top_start_col + 1, value=f'=COUNTIFS(tbl_계약[활성상태(진행/만료/해지)],"{s}")').number_format = "#,##0"
        ws.cell(row=r, column=top_start_col + 2, value=f'=SUMIFS(tbl_계약[계약금액],tbl_계약[활성상태(진행/만료/해지)],"{s}")').number_format = "#,##0"
        ws.cell(row=r, column=top_start_col + 3, value=f'=SUMIFS(tbl_계약[미수금],tbl_계약[활성상태(진행/만료/해지)],"{s}")').number_format = "#,##0"

    # === 차트 ===
    # 연간 매출/매입/영업이익 추이 (B7:H13 사용)
    chart = BarChart()
    chart.type = "col"
    chart.title = "연간 매출·매입·영업이익"
    chart.style = 11
    chart.height = 8
    chart.width = 18

    # 데이터 — B8(매출), B9(매입), B13(영업이익) 3행
    data = Reference(ws, min_col=2, min_row=7, max_col=8, max_row=8)  # 매출
    chart.add_data(data, titles_from_data=True, from_rows=True)
    data = Reference(ws, min_col=2, min_row=9, max_col=8, max_row=9)  # 매입
    chart.add_data(data, titles_from_data=True, from_rows=True)
    data = Reference(ws, min_col=2, min_row=13, max_col=8, max_row=13)  # 영업이익
    chart.add_data(data, titles_from_data=True, from_rows=True)
    cats = Reference(ws, min_col=3, min_row=7, max_col=8, max_row=7)
    chart.set_categories(cats)
    ws.add_chart(chart, "L5")

    # 제품별 매출 파이
    pie = PieChart()
    pie.title = "2025 제품별 매출 비중"
    pie.height = 8
    pie.width = 10
    labels = Reference(ws, min_col=2, min_row=20, max_row=20 + len(products) - 1)
    data = Reference(ws, min_col=3, min_row=19, max_row=20 + len(products) - 1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList(showPercent=True)
    ws.add_chart(pie, "L17")


def validate(wb):
    """검증 — 각 FACT 시트의 행수·합계를 추출하여 검증 시트에 저장"""
    if "99_검증리포트" in wb.sheetnames:
        del wb["99_검증리포트"]
    ws = wb.create_sheet("99_검증리포트")
    ws.sheet_properties.tabColor = COLORS["warning"]
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 60

    ws["A1"] = "FACT 시트 검증 리포트"
    ws["A1"].font = Font(name="맑은 고딕", size=16, bold=True, color=COLORS["primary"])
    ws["A2"] = f"갱신: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="808080")

    rows = [
        ("시트", "행수", "주요 합계(공급가액)", "비고"),
        ("20_FACT_매출", "tbl_매출", '=SUMPRODUCT(tbl_매출[공급가액])', "2021~2026 매출 트랜잭션 union"),
        ("21_FACT_매입", "tbl_매입", '=SUMPRODUCT(tbl_매입[공급가액])', "2023~2026 매입 (2021~2022 보강 필요)"),
        ("22_FACT_급여", "tbl_급여", '=SUMPRODUCT(tbl_급여[지급합계])', "2023~2025 부서별 인건비 + 25.1~2월"),
        ("23_FACT_비용", "tbl_비용", '=SUMPRODUCT(tbl_비용[금액])', "2024년 직원별총합 (다른 연도 보강 필요)"),
        ("24_FACT_미수금", "tbl_미수금", '=SUMPRODUCT(tbl_미수금[세금계산서금액(증)])', "거래처별 ledger movements"),
        ("25_FACT_차입금", "tbl_차입금", '=SUMPRODUCT(tbl_차입금[차입(+)])', "임원 4명 단기차입 movements"),
        ("26_FACT_임대료", "tbl_임대료", '=SUMPRODUCT(tbl_임대료[금액])', "2024년 렌탈현황"),
        ("27_FACT_퇴직금", "tbl_퇴직금", '=SUMPRODUCT(tbl_퇴직금[기업납입금])', "직원별 월별 기업납입금 unpivot"),
        ("28_FACT_판독수수료", "tbl_판독수수료", '=SUMPRODUCT(tbl_판독수수료[매출액])', "[보강 필요] 시트 구조 추가 분석"),
        ("30_계약마스터", "tbl_계약", '=SUMPRODUCT(tbl_계약[계약금액])', "계약현황_2025 + 종료 계약"),
        ("31_차입금마스터", "tbl_차입금마스터", '=SUMPRODUCT(tbl_차입금마스터[현재잔액])', "장기차입금 + 임원 단기"),
    ]
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            c = ws.cell(row=4 + i, column=1 + j, value=val)
            if i == 0:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor=COLORS["primary"])
            elif j == 2:
                c.number_format = "#,##0"

    # 행수: 표 행수 함수 — ROWS 사용
    for i, row in enumerate(rows[1:], start=1):
        ws.cell(row=4 + i, column=2, value=f'=ROWS({row[1]})')


if __name__ == "__main__":
    print("=== P8: 대시보드 + 검증 ===")
    wb = load_master()
    update_dashboard(wb)
    print("  대시보드 갱신 완료")
    validate(wb)
    print("  검증 시트 추가 완료")
    save_master(wb)
    print(f"\n최종 저장: {MASTER}")
    print(f"\n총 시트 수: {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        print(f"  - {s}")
