# -*- coding: utf-8 -*-
"""
인비즈 경영관리 마스터 워크북 프레임워크 빌더 v1.0

생성 시트:
  00_README             — 사용법
  01_Dashboard          — KPI + 차트 (placeholder)
  10_DIM_거래처         — 거래처 마스터
  11_DIM_제품           — 제품/상품 마스터
  12_DIM_계정           — 계정과목 마스터
  13_DIM_직원           — 직원 마스터
  14_DIM_부서           — 부서 마스터
  20_FACT_매출          — 매출 트랜잭션 (long)
  21_FACT_매입          — 매입 트랜잭션 (long)
  22_FACT_급여          — 급여 월별 (long)
  23_FACT_비용          — 비용 트랜잭션
  24_FACT_미수금        — 미수금 ledger
  25_FACT_차입금        — 차입금 movements
  26_FACT_임대료        — 임대료 수입/지출
  27_FACT_퇴직금        — 퇴직연금 월별
  28_FACT_판독수수료    — 판독수수료 매출/매입
  30_계약마스터         — 계약 list
  31_차입금마스터       — 차입금/대출 list
  32_제품매핑           — 품명→제품 코드 매핑 룰
  40_월별집계           — 월별 매출/매입/비용 pivot
  41_분기반기연집계     — 분기·반기·연간
  42_제품별집계         — 제품별 매출/이익률
  43_거래처별집계       — 거래처별 매출/매입 순위
  44_미수금현황         — 거래처별 미수금 잔액
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from datetime import datetime

OUT_PATH = r"C:\Users\scpar\OneDrive - Inviz\5.Inviz_Corporation\14.경영정보\00.경영관리마스터\인비즈_경영관리마스터_v1.xlsx"

# ===== 스타일 정의 =====
COLORS = {
    "primary": "1F4E79",      # 진한 파랑 (헤더)
    "accent": "2E75B6",       # 파랑 (sub)
    "warning": "C00000",      # 빨강 (경고)
    "ok": "548235",           # 초록 (정상)
    "neutral_bg": "D9E1F2",   # 연한 파랑 (DIM/입력)
    "fact_bg": "FFF2CC",      # 연한 노랑 (FACT/거래)
    "agg_bg": "E2EFDA",       # 연한 초록 (집계)
    "dash_bg": "FCE4D6",      # 연한 주황 (대시보드)
}

THIN = Side(border_style="thin", color="808080")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def header_font():
    return Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")


def title_font():
    return Font(name="맑은 고딕", size=16, bold=True, color=COLORS["primary"])


def section_font():
    return Font(name="맑은 고딕", size=11, bold=True, color=COLORS["primary"])


def header_fill(color_key="primary"):
    return PatternFill(start_color=COLORS[color_key], end_color=COLORS[color_key], fill_type="solid")


def cat_fill(category):
    """카테고리별 배경색"""
    return PatternFill(start_color=COLORS[category], end_color=COLORS[category], fill_type="solid")


def write_headers(ws, headers, start_row=1, fill_key="primary"):
    fill = header_fill(fill_key)
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=col_idx, value=h)
        c.font = header_font()
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[start_row].height = 30


def autosize(ws, headers, default=14):
    for col_idx, h in enumerate(headers, 1):
        w = max(default, min(len(str(h)) * 2 + 2, 40))
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def add_table(ws, headers, name, start_row=1, n_rows=100):
    """엑셀 표(Table)로 등록 — 필터·서식 자동"""
    last_col = get_column_letter(len(headers))
    ref = f"A{start_row}:{last_col}{start_row + n_rows}"
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)


# ===== 워크북 생성 =====
wb = Workbook()
wb.remove(wb.active)


# ----- 00_README -----
ws = wb.create_sheet("00_README")
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 80

ws["B2"] = "인비즈 경영관리 마스터 워크북"
ws["B2"].font = Font(name="맑은 고딕", size=20, bold=True, color=COLORS["primary"])
ws["B3"] = f"v1.0  |  생성일: {datetime.now().strftime('%Y-%m-%d')}"
ws["B3"].font = Font(name="맑은 고딕", size=10, italic=True, color="595959")

ws["B5"] = "1. 목적"
ws["B5"].font = section_font()
ws["C5"] = "흩어진 매출·매입·자금·인사·계약·차입금 자료를 하나의 마스터로 통합하여, 월·분기·반기·연간 경영실적과 제품별 수익성을 즉시 조회한다."

ws["B7"] = "2. 시트 컬러 규약"
ws["B7"].font = section_font()
rules = [
    ("DIM (기준정보)", "neutral_bg", "거래처·제품·계정·직원 마스터. 수동 유지보수 대상."),
    ("FACT (트랜잭션)", "fact_bg", "일자별 거래 데이터. ETL 스크립트로 적재."),
    ("집계 (분석)", "agg_bg", "FACT 기반 자동 집계. 수식만 — 수동 입력 금지."),
    ("Dashboard", "dash_bg", "KPI 카드 + 핵심 차트. 경영진 뷰."),
]
for i, (label, color_key, desc) in enumerate(rules, start=8):
    c1 = ws.cell(row=i, column=2, value=label)
    c1.fill = cat_fill(color_key)
    c1.font = Font(name="맑은 고딕", size=10, bold=True)
    c1.alignment = Alignment(horizontal="center")
    c1.border = BORDER
    ws.cell(row=i, column=3, value=desc)

ws["B13"] = "3. 시트 인덱스"
ws["B13"].font = section_font()
sheet_index = [
    ("00_README", "현재 시트 (사용법)"),
    ("01_Dashboard", "KPI 카드 + 월·분기·연 매출 추이 + 제품별 비중"),
    ("10_DIM_거래처", "거래처 마스터 (코드·이름·카테고리·활성여부)"),
    ("11_DIM_제품", "제품 마스터 (Cloud Care Life, PACS, Vision Maker, Ai Echo Care, 기타)"),
    ("12_DIM_계정", "계정과목 마스터 (매출·매입·비용·자산·부채)"),
    ("13_DIM_직원", "직원 마스터 (사번·이름·부서·입사일)"),
    ("14_DIM_부서", "부서 마스터"),
    ("20_FACT_매출", "매출 트랜잭션 (장-format)"),
    ("21_FACT_매입", "매입 트랜잭션"),
    ("22_FACT_급여", "급여 월별 (지급·공제 항목)"),
    ("23_FACT_비용", "판관비 트랜잭션"),
    ("24_FACT_미수금", "미수금 ledger movements"),
    ("25_FACT_차입금", "차입금 입출금 movements"),
    ("26_FACT_임대료", "임대료 수입/지출"),
    ("27_FACT_퇴직금", "퇴직연금 월별 적립"),
    ("28_FACT_판독수수료", "원격판독 수수료 매출·매입"),
    ("30_계약마스터", "계약 list (시작·만료·금액)"),
    ("31_차입금마스터", "차입금/대출 list (장기·단기·은행·개인)"),
    ("32_제품매핑", "품명 → 제품코드 자동 매핑 룰"),
    ("40_월별집계", "월별 매출·매입·비용·이익 자동 집계"),
    ("41_분기반기연집계", "분기·반기·연간 집계"),
    ("42_제품별집계", "제품별 매출·매입·이익률"),
    ("43_거래처별집계", "거래처별 매출/매입 순위"),
    ("44_미수금현황", "거래처별 미수금 잔액"),
]
for i, (sheet, desc) in enumerate(sheet_index, start=14):
    ws.cell(row=i, column=2, value=sheet).font = Font(name="맑은 고딕", size=10, bold=True, color=COLORS["accent"])
    ws.cell(row=i, column=3, value=desc)

last_row = 14 + len(sheet_index)
ws.cell(row=last_row + 2, column=2, value="4. 운영 원칙").font = section_font()
principles = [
    "FACT 시트의 데이터는 ETL 스크립트로만 갱신한다 (수동 입력 금지).",
    "DIM 시트는 신규 거래처/제품 추가 시 수동으로 행 추가한다.",
    "32_제품매핑에 신규 품명 패턴이 들어오면 즉시 매핑 룰을 추가한다.",
    "집계 시트(40~44)는 수식·피벗으로 자동 갱신된다.",
    "백업: 매월 말 사본을 '_yyyymm' 접미로 보관.",
]
for i, p in enumerate(principles, start=last_row + 3):
    ws.cell(row=i, column=2, value=f"  {i - last_row - 2}.")
    ws.cell(row=i, column=3, value=p)


# ----- 01_Dashboard (placeholder layout) -----
ws = wb.create_sheet("01_Dashboard")
ws.sheet_properties.tabColor = COLORS["primary"]
ws.column_dimensions["A"].width = 2
for col in "BCDEFGHIJKLMN":
    ws.column_dimensions[col].width = 14

ws["B2"] = "인비즈 경영 대시보드"
ws["B2"].font = title_font()
ws["B3"] = "데이터 적재 후 자동 계산됩니다. (FACT 시트 채워진 뒤)"
ws["B3"].font = Font(name="맑은 고딕", size=10, italic=True, color="808080")

# KPI 카드 영역 (placeholder)
ws["B5"] = "■ 당월 KPI"
ws["B5"].font = section_font()
kpi_labels = ["당월 매출", "당월 매입", "당월 매출이익", "당월 비용", "당월 영업이익", "미수금 잔액"]
for i, label in enumerate(kpi_labels):
    col = chr(ord("B") + i * 2)
    next_col = chr(ord("B") + i * 2 + 1)
    ws[f"{col}7"] = label
    ws[f"{col}7"].font = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")
    ws[f"{col}7"].fill = header_fill("accent")
    ws[f"{col}7"].alignment = Alignment(horizontal="center")
    ws.merge_cells(f"{col}7:{next_col}7")
    ws[f"{col}8"] = "(ETL 후)"
    ws[f"{col}8"].font = Font(name="맑은 고딕", size=14, bold=True)
    ws[f"{col}8"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"{col}8:{next_col}9")
    for r in [7, 8, 9]:
        for c in [col, next_col]:
            ws[f"{c}{r}"].border = BORDER

ws["B12"] = "■ 월별 매출·매입 추이"
ws["B12"].font = section_font()
ws["B13"] = "(40_월별집계 시트 적재 후 차트 삽입)"
ws["B13"].font = Font(name="맑은 고딕", size=10, italic=True, color="808080")

ws["B25"] = "■ 제품별 매출 비중"
ws["B25"].font = section_font()
ws["B26"] = "(42_제품별집계 시트 적재 후 차트 삽입)"
ws["B26"].font = Font(name="맑은 고딕", size=10, italic=True, color="808080")

ws["I25"] = "■ 거래처별 매출 TOP10"
ws["I25"].font = section_font()
ws["I26"] = "(43_거래처별집계 시트 적재 후 차트 삽입)"
ws["I26"].font = Font(name="맑은 고딕", size=10, italic=True, color="808080")


# ----- DIM Sheets -----

# 10_DIM_거래처
ws = wb.create_sheet("10_DIM_거래처")
ws.sheet_properties.tabColor = COLORS["neutral_bg"]
headers = ["거래처코드", "거래처명", "사업자번호", "구분(병원/대리점/공급사/기타)", "주거래제품", "활성여부", "최초거래일", "최종거래일", "주담당자", "연락처", "비고"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=16)
add_table(ws, headers, "tbl_거래처", n_rows=500)
ws.freeze_panes = "A2"

# 11_DIM_제품
ws = wb.create_sheet("11_DIM_제품")
ws.sheet_properties.tabColor = COLORS["neutral_bg"]
headers = ["제품코드", "제품명", "카테고리(상품/제품/용역)", "그룹", "단가기준", "비고"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=18)
# 시드 데이터
products_seed = [
    ("P001", "Cloud Care Life", "용역", "원격판독", "건별", "원격판독 서비스"),
    ("P002", "Saintview PACS", "제품", "PACS", "장비", "PACS 시스템"),
    ("P003", "Vision Maker", "제품", "AI 영상", "라이선스", "영상분석"),
    ("P004", "Ai Echo Care", "제품", "AI 영상", "라이선스", "심초음파 AI"),
    ("P005", "AI CXR", "용역", "AI 영상", "건별", "흉부 AI 판독"),
    ("P006", "AI MMG", "용역", "AI 영상", "건별", "유방촬영 AI 판독"),
    ("P007", "CR 장비", "상품", "장비", "대당", "Computed Radiography"),
    ("P008", "유지보수", "용역", "서비스", "월/연", "장비·SW 유지보수"),
    ("P009", "출장서비스", "용역", "서비스", "건별", "현장 출장 서비스"),
    ("P999", "기타", "기타", "기타", "—", "분류 불가/소액"),
]
for i, row in enumerate(products_seed, start=2):
    for j, val in enumerate(row, 1):
        ws.cell(row=i, column=j, value=val)
add_table(ws, headers, "tbl_제품", n_rows=max(50, len(products_seed) + 20))
ws.freeze_panes = "A2"

# 12_DIM_계정
ws = wb.create_sheet("12_DIM_계정")
ws.sheet_properties.tabColor = COLORS["neutral_bg"]
headers = ["계정코드", "계정과목", "대분류(B/S, P/L)", "중분류", "소분류", "차/대"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=16)
account_seed = [
    ("4101", "제품매출", "P/L", "매출", "제품", "대"),
    ("4102", "상품매출", "P/L", "매출", "상품", "대"),
    ("4103", "용역매출", "P/L", "매출", "용역", "대"),
    ("4104", "임대료수입", "P/L", "매출", "임대", "대"),
    ("4501", "수출매출", "P/L", "매출", "수출", "대"),
    ("5101", "원재료비", "P/L", "매출원가", "재료", "차"),
    ("5102", "상품매입", "P/L", "매출원가", "상품", "차"),
    ("5103", "용역매입", "P/L", "매출원가", "용역", "차"),
    ("6101", "급여", "P/L", "판관비", "인건비", "차"),
    ("6102", "퇴직급여", "P/L", "판관비", "인건비", "차"),
    ("6103", "복리후생비", "P/L", "판관비", "인건비", "차"),
    ("6201", "지급임차료", "P/L", "판관비", "임차", "차"),
    ("6301", "지급수수료", "P/L", "판관비", "수수료", "차"),
    ("6401", "이자비용", "P/L", "영업외", "금융", "차"),
    ("1101", "현금및예금", "B/S", "유동자산", "현금", "차"),
    ("1102", "외상매출금", "B/S", "유동자산", "매출채권", "차"),
    ("1103", "미수금", "B/S", "유동자산", "기타채권", "차"),
    ("2101", "외상매입금", "B/S", "유동부채", "매입채무", "대"),
    ("2102", "단기차입금", "B/S", "유동부채", "차입금", "대"),
    ("2201", "장기차입금", "B/S", "비유동부채", "차입금", "대"),
    ("2202", "퇴직급여충당부채", "B/S", "비유동부채", "퇴직", "대"),
]
for i, row in enumerate(account_seed, start=2):
    for j, val in enumerate(row, 1):
        ws.cell(row=i, column=j, value=val)
add_table(ws, headers, "tbl_계정", n_rows=max(80, len(account_seed) + 20))
ws.freeze_panes = "A2"

# 13_DIM_직원
ws = wb.create_sheet("13_DIM_직원")
ws.sheet_properties.tabColor = COLORS["neutral_bg"]
headers = ["사번", "성명", "부서", "직급", "고용형태", "입사일", "퇴사일", "재직여부", "주민등록번호(뒤4자리)", "기준임금", "퇴직연금가입여부", "비고"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=14)
add_table(ws, headers, "tbl_직원", n_rows=100)
ws.freeze_panes = "A2"

# 14_DIM_부서
ws = wb.create_sheet("14_DIM_부서")
ws.sheet_properties.tabColor = COLORS["neutral_bg"]
headers = ["부서코드", "부서명", "상위부서", "주요기능", "활성여부"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=16)
depts_seed = [
    ("D01", "연구개발", "본사", "R&D", "Y"),
    ("D02", "영업", "본사", "영업·마케팅", "Y"),
    ("D03", "서비스", "본사", "원격판독·기술지원", "Y"),
    ("D04", "관리", "본사", "경영지원·재무·인사", "Y"),
    ("D05", "어플리케이션", "본사", "SW 개발·운영", "Y"),
]
for i, row in enumerate(depts_seed, start=2):
    for j, val in enumerate(row, 1):
        ws.cell(row=i, column=j, value=val)
add_table(ws, headers, "tbl_부서", n_rows=20)
ws.freeze_panes = "A2"


# ----- FACT Sheets -----

# 20_FACT_매출
ws = wb.create_sheet("20_FACT_매출")
ws.sheet_properties.tabColor = COLORS["fact_bg"]
headers = ["거래ID", "전표일자", "년", "월", "분기", "반기", "거래처코드", "거래처명", "제품코드", "제품명", "품명(원본)", "계정코드", "계정과목", "매출유형(정기/신규/일회성/기타)", "공급가액", "부가세", "합계", "결제수단", "비고", "원본파일", "원본시트", "원본행"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=12)
add_table(ws, headers, "tbl_매출", n_rows=5000)
ws.freeze_panes = "C2"

# 21_FACT_매입
ws = wb.create_sheet("21_FACT_매입")
ws.sheet_properties.tabColor = COLORS["fact_bg"]
headers = ["거래ID", "전표일자", "년", "월", "분기", "반기", "거래처코드", "거래처명", "제품코드", "제품명", "품명(원본)", "계정코드", "계정과목", "매입유형(정기/일회성/기타)", "공급가액", "부가세", "합계", "결제수단", "비고", "원본파일", "원본시트", "원본행"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=12)
add_table(ws, headers, "tbl_매입", n_rows=5000)
ws.freeze_panes = "C2"

# 22_FACT_급여
ws = wb.create_sheet("22_FACT_급여")
ws.sheet_properties.tabColor = COLORS["fact_bg"]
headers = ["귀속년월", "년", "월", "사번", "성명", "부서", "기본급", "식대", "차량유지비", "연구수당", "기타수당", "연차수당", "연장근로수당", "야간근로수당", "성과급", "지급합계", "국민연금", "건강보험", "장기요양", "고용보험", "소득세", "지방소득세", "기타공제", "공제합계", "실지급액", "4대보험(기업부담)", "원본파일", "원본행"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=12)
add_table(ws, headers, "tbl_급여", n_rows=2000)
ws.freeze_panes = "C2"

# 23_FACT_비용
ws = wb.create_sheet("23_FACT_비용")
ws.sheet_properties.tabColor = COLORS["fact_bg"]
headers = ["거래ID", "사용일", "년", "월", "분기", "사용자(직원)", "부서", "거래처/사용처", "금액", "계정코드", "계정과목", "구분(대)", "상세구분(소)", "결제수단", "비고", "원본파일", "원본행"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=12)
add_table(ws, headers, "tbl_비용", n_rows=3000)
ws.freeze_panes = "C2"

# 24_FACT_미수금
ws = wb.create_sheet("24_FACT_미수금")
ws.sheet_properties.tabColor = COLORS["fact_bg"]
headers = ["거래ID", "일자", "년", "월", "거래처코드", "거래처명", "적요", "세금계산서금액(증)", "입금액(감)", "잔액", "전표번호", "비고", "원본파일", "원본시트"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=12)
add_table(ws, headers, "tbl_미수금", n_rows=3000)
ws.freeze_panes = "C2"

# 25_FACT_차입금
ws = wb.create_sheet("25_FACT_차입금")
ws.sheet_properties.tabColor = COLORS["fact_bg"]
headers = ["거래ID", "일자", "년", "월", "차입처구분(은행/개인)", "차입처명", "차입ID", "적요", "차입(+)", "상환(-)", "잔액", "이자", "비고", "원본파일", "원본시트"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=12)
add_table(ws, headers, "tbl_차입금", n_rows=1500)
ws.freeze_panes = "C2"

# 26_FACT_임대료
ws = wb.create_sheet("26_FACT_임대료")
ws.sheet_properties.tabColor = COLORS["fact_bg"]
headers = ["거래ID", "일자", "년", "월", "구분(수입/지출)", "거래처", "물건명(사무실/장비)", "항목(임차료/관리비/공과금/렌탈료)", "금액", "결제수단", "비고", "원본파일"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=14)
add_table(ws, headers, "tbl_임대료", n_rows=1500)
ws.freeze_panes = "C2"

# 27_FACT_퇴직금
ws = wb.create_sheet("27_FACT_퇴직금")
ws.sheet_properties.tabColor = COLORS["fact_bg"]
headers = ["귀속년월", "년", "월", "사번", "성명", "기준급여", "기업납입금", "개인납입금", "납입일자", "구분(적립/지급/중도인출)", "비고", "원본파일"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=12)
add_table(ws, headers, "tbl_퇴직금", n_rows=3000)
ws.freeze_panes = "C2"

# 28_FACT_판독수수료
ws = wb.create_sheet("28_FACT_판독수수료")
ws.sheet_properties.tabColor = COLORS["fact_bg"]
headers = ["거래ID", "귀속년월", "년", "월", "매출/매입", "병원/공급처", "대리점", "제품코드", "제품명", "판독건수", "단가", "판독료", "인비즈수익", "대리점수수료", "매출액", "원가", "순이익", "이익률", "비고", "원본파일"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=12)
add_table(ws, headers, "tbl_판독수수료", n_rows=3000)
ws.freeze_panes = "C2"


# ----- Master Sheets -----

# 30_계약마스터
ws = wb.create_sheet("30_계약마스터")
ws.sheet_properties.tabColor = COLORS["neutral_bg"]
headers = ["계약ID", "계약명", "구분(유지보수/장비/AI/판독/임대/기타)", "거래처코드", "공급받는자(거래처명)", "제품코드", "품명", "계약체결일", "계약시작일", "계약만료일", "계약기간(개월)", "자동연장(Y/N)", "잔여일수", "계약금액", "발행금액", "미수금", "대금지불(월/분기/연/일시)", "결제일", "설치일", "하자보수만료일", "계약서유무", "담당자", "연락처", "활성상태(진행/만료/해지)", "비고"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=13)
add_table(ws, headers, "tbl_계약", n_rows=500)
ws.freeze_panes = "C2"

# 31_차입금마스터
ws = wb.create_sheet("31_차입금마스터")
ws.sheet_properties.tabColor = COLORS["neutral_bg"]
headers = ["차입ID", "구분(은행/개인/사채)", "장단기", "금융기관/차주", "계좌번호/식별", "약정한도", "최초차입액", "현재잔액", "차입종류", "이자율(%)", "상환방법", "차입일", "만기일", "담보", "담보설정액", "대표이사지급보증", "활성상태", "비고"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=14)
add_table(ws, headers, "tbl_차입금마스터", n_rows=100)
ws.freeze_panes = "B2"

# 32_제품매핑
ws = wb.create_sheet("32_제품매핑")
ws.sheet_properties.tabColor = COLORS["neutral_bg"]
headers = ["우선순위", "매칭패턴(품명에 포함)", "제품코드", "제품명", "매출유형 default", "비고"]
write_headers(ws, headers, fill_key="primary")
autosize(ws, headers, default=20)
mapping_seed = [
    (10, "Cloud Care", "P001", "Cloud Care Life", "정기", ""),
    (10, "원격판독", "P001", "Cloud Care Life", "정기", "원격판독 수수료 매출"),
    (10, "Saintview", "P002", "Saintview PACS", "일회성", ""),
    (10, "PACS", "P002", "Saintview PACS", "정기", ""),
    (10, "Vision Maker", "P003", "Vision Maker", "정기", ""),
    (10, "Echo Care", "P004", "Ai Echo Care", "정기", ""),
    (10, "AI Echo", "P004", "Ai Echo Care", "정기", ""),
    (10, "CXR", "P005", "AI CXR", "정기", "흉부 AI 판독"),
    (10, "MMG", "P006", "AI MMG", "정기", "유방촬영 AI"),
    (20, "CR ", "P007", "CR 장비", "일회성", ""),
    (30, "유지보수", "P008", "유지보수", "정기", ""),
    (30, "출장", "P009", "출장서비스", "일회성", ""),
    (99, "*", "P999", "기타", "기타", "매칭 실패 시"),
]
for i, row in enumerate(mapping_seed, start=2):
    for j, val in enumerate(row, 1):
        ws.cell(row=i, column=j, value=val)
add_table(ws, headers, "tbl_제품매핑", n_rows=max(50, len(mapping_seed) + 20))
ws.freeze_panes = "A2"


# ----- 집계 Sheets (수식 골격) -----

# 40_월별집계
ws = wb.create_sheet("40_월별집계")
ws.sheet_properties.tabColor = COLORS["agg_bg"]
ws["A1"] = "월별 매출·매입·이익 집계"
ws["A1"].font = title_font()
ws["A2"] = "FACT 시트 적재 후 SUMIFS 수식으로 자동 계산. 컬러스케일·데이터바 적용."
ws["A2"].font = Font(name="맑은 고딕", size=10, italic=True, color="808080")

headers = ["년월", "년", "월", "분기", "매출", "매입", "매출이익", "이익률(%)", "급여", "비용", "영업이익", "비고"]
write_headers(ws, headers, start_row=4, fill_key="primary")
autosize(ws, headers, default=14)

# 2022-01 ~ 2026-12 (60개월) 골격
import datetime as dt
months = []
y, m = 2022, 1
while (y, m) <= (2026, 12):
    months.append((y, m))
    m += 1
    if m > 12:
        y += 1; m = 1

for i, (yr, mo) in enumerate(months, start=5):
    qtr = (mo - 1) // 3 + 1
    ws.cell(row=i, column=1, value=f"{yr}-{mo:02d}")
    ws.cell(row=i, column=2, value=yr)
    ws.cell(row=i, column=3, value=mo)
    ws.cell(row=i, column=4, value=f"Q{qtr}")
    # 수식 — 매출/매입/급여/비용
    r = i
    ws.cell(row=r, column=5, value=f'=SUMIFS(tbl_매출[공급가액],tbl_매출[년],B{r},tbl_매출[월],C{r})')
    ws.cell(row=r, column=6, value=f'=SUMIFS(tbl_매입[공급가액],tbl_매입[년],B{r},tbl_매입[월],C{r})')
    ws.cell(row=r, column=7, value=f"=E{r}-F{r}")
    ws.cell(row=r, column=8, value=f'=IFERROR(G{r}/E{r}*100,0)')
    ws.cell(row=r, column=9, value=f'=SUMIFS(tbl_급여[지급합계],tbl_급여[년],B{r},tbl_급여[월],C{r})')
    ws.cell(row=r, column=10, value=f'=SUMIFS(tbl_비용[금액],tbl_비용[년],B{r},tbl_비용[월],C{r})')
    ws.cell(row=r, column=11, value=f"=G{r}-I{r}-J{r}")
    # 숫자 포맷
    for col in [5, 6, 7, 9, 10, 11]:
        ws.cell(row=r, column=col).number_format = "#,##0"
    ws.cell(row=r, column=8).number_format = "0.0"

ws.freeze_panes = "A5"

# 컬러스케일 (영업이익 컬럼)
last_row = 4 + len(months)
ws.conditional_formatting.add(
    f"K5:K{last_row}",
    ColorScaleRule(start_type="min", start_color="F8696B",
                   mid_type="num", mid_value=0, mid_color="FFEB84",
                   end_type="max", end_color="63BE7B"),
)

# 41_분기반기연집계
ws = wb.create_sheet("41_분기반기연집계")
ws.sheet_properties.tabColor = COLORS["agg_bg"]
ws["A1"] = "분기·반기·연간 집계"
ws["A1"].font = title_font()

# 분기 블록
ws["A3"] = "■ 분기별 집계"
ws["A3"].font = section_font()
qheaders = ["연도", "분기", "매출", "매입", "매출이익", "이익률(%)", "비용", "영업이익"]
write_headers(ws, qheaders, start_row=4, fill_key="primary")
for col, h in enumerate(qheaders, 1):
    ws.column_dimensions[get_column_letter(col)].width = 14

row = 5
for yr in range(2022, 2027):
    for q in range(1, 5):
        ws.cell(row=row, column=1, value=yr)
        ws.cell(row=row, column=2, value=f"Q{q}")
        m1, m2, m3 = (q - 1) * 3 + 1, (q - 1) * 3 + 2, (q - 1) * 3 + 3
        cond = f'tbl_매출[년],{yr},tbl_매출[월],"<="&{m3},tbl_매출[월],">="&{m1}'
        ws.cell(row=row, column=3, value=f'=SUMIFS(tbl_매출[공급가액],tbl_매출[년],{yr},tbl_매출[분기],"Q{q}")')
        ws.cell(row=row, column=4, value=f'=SUMIFS(tbl_매입[공급가액],tbl_매입[년],{yr},tbl_매입[분기],"Q{q}")')
        ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
        ws.cell(row=row, column=6, value=f'=IFERROR(E{row}/C{row}*100,0)')
        ws.cell(row=row, column=7, value=f'=SUMIFS(tbl_비용[금액],tbl_비용[년],{yr},tbl_비용[분기],"Q{q}")')
        ws.cell(row=row, column=8, value=f"=E{row}-G{row}")
        for col in [3, 4, 5, 7, 8]:
            ws.cell(row=row, column=col).number_format = "#,##0"
        ws.cell(row=row, column=6).number_format = "0.0"
        row += 1

# 반기 블록
row += 2
ws.cell(row=row, column=1, value="■ 반기별 집계").font = section_font()
row += 1
write_headers(ws, ["연도", "반기", "매출", "매입", "매출이익", "이익률(%)"], start_row=row, fill_key="primary")
row += 1
for yr in range(2022, 2027):
    for h in [1, 2]:
        q1, q2 = (h - 1) * 2 + 1, (h - 1) * 2 + 2
        ws.cell(row=row, column=1, value=yr)
        ws.cell(row=row, column=2, value=f"H{h}")
        ws.cell(row=row, column=3, value=f'=SUMPRODUCT((tbl_매출[년]={yr})*((tbl_매출[분기]="Q{q1}")+(tbl_매출[분기]="Q{q2}"))*tbl_매출[공급가액])')
        ws.cell(row=row, column=4, value=f'=SUMPRODUCT((tbl_매입[년]={yr})*((tbl_매입[분기]="Q{q1}")+(tbl_매입[분기]="Q{q2}"))*tbl_매입[공급가액])')
        ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
        ws.cell(row=row, column=6, value=f'=IFERROR(E{row}/C{row}*100,0)')
        for col in [3, 4, 5]:
            ws.cell(row=row, column=col).number_format = "#,##0"
        ws.cell(row=row, column=6).number_format = "0.0"
        row += 1

# 연간 블록
row += 2
ws.cell(row=row, column=1, value="■ 연간 집계").font = section_font()
row += 1
write_headers(ws, ["연도", "매출", "매입", "매출이익", "이익률(%)", "급여", "비용", "영업이익"], start_row=row, fill_key="primary")
row += 1
for yr in range(2022, 2027):
    ws.cell(row=row, column=1, value=yr)
    ws.cell(row=row, column=2, value=f'=SUMIFS(tbl_매출[공급가액],tbl_매출[년],{yr})')
    ws.cell(row=row, column=3, value=f'=SUMIFS(tbl_매입[공급가액],tbl_매입[년],{yr})')
    ws.cell(row=row, column=4, value=f"=B{row}-C{row}")
    ws.cell(row=row, column=5, value=f'=IFERROR(D{row}/B{row}*100,0)')
    ws.cell(row=row, column=6, value=f'=SUMIFS(tbl_급여[지급합계],tbl_급여[년],{yr})')
    ws.cell(row=row, column=7, value=f'=SUMIFS(tbl_비용[금액],tbl_비용[년],{yr})')
    ws.cell(row=row, column=8, value=f"=D{row}-F{row}-G{row}")
    for col in [2, 3, 4, 6, 7, 8]:
        ws.cell(row=row, column=col).number_format = "#,##0"
    ws.cell(row=row, column=5).number_format = "0.0"
    row += 1

# 42_제품별집계
ws = wb.create_sheet("42_제품별집계")
ws.sheet_properties.tabColor = COLORS["agg_bg"]
ws["A1"] = "제품별 매출·매입·이익률 (연도별)"
ws["A1"].font = title_font()
ws["A2"] = "DIM_제품 × 연도 매트릭스. 신규 제품은 11_DIM_제품에 추가 후 본 시트에 행 추가."
ws["A2"].font = Font(name="맑은 고딕", size=10, italic=True, color="808080")

headers = ["제품코드", "제품명", "구분", "2022매출", "2023매출", "2024매출", "2025매출", "2026매출", "2022매입", "2023매입", "2024매입", "2025매입", "2026매입", "누적매출", "누적매입", "누적이익", "누적이익률(%)"]
write_headers(ws, headers, start_row=4, fill_key="primary")
for col, h in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(col)].width = 13

# 제품 시드 데이터 기반
for i, (code, name, cat, *_) in enumerate(products_seed, start=5):
    ws.cell(row=i, column=1, value=code)
    ws.cell(row=i, column=2, value=name)
    ws.cell(row=i, column=3, value=cat)
    for yi, yr in enumerate([2022, 2023, 2024, 2025, 2026]):
        ws.cell(row=i, column=4 + yi, value=f'=SUMIFS(tbl_매출[공급가액],tbl_매출[제품코드],A{i},tbl_매출[년],{yr})')
        ws.cell(row=i, column=9 + yi, value=f'=SUMIFS(tbl_매입[공급가액],tbl_매입[제품코드],A{i},tbl_매입[년],{yr})')
        ws.cell(row=i, column=4 + yi).number_format = "#,##0"
        ws.cell(row=i, column=9 + yi).number_format = "#,##0"
    ws.cell(row=i, column=14, value=f"=SUM(D{i}:H{i})")
    ws.cell(row=i, column=15, value=f"=SUM(I{i}:M{i})")
    ws.cell(row=i, column=16, value=f"=N{i}-O{i}")
    ws.cell(row=i, column=17, value=f'=IFERROR(P{i}/N{i}*100,0)')
    for col in [14, 15, 16]:
        ws.cell(row=i, column=col).number_format = "#,##0"
    ws.cell(row=i, column=17).number_format = "0.0"

# 합계 행
total_row = 5 + len(products_seed)
ws.cell(row=total_row, column=1, value="합계")
ws.cell(row=total_row, column=1).font = Font(bold=True)
for col in range(4, 17):
    cl = get_column_letter(col)
    ws.cell(row=total_row, column=col, value=f"=SUM({cl}5:{cl}{total_row - 1})")
    ws.cell(row=total_row, column=col).number_format = "#,##0" if col != 17 else "0.0"
    ws.cell(row=total_row, column=col).font = Font(bold=True)

# 43_거래처별집계
ws = wb.create_sheet("43_거래처별집계")
ws.sheet_properties.tabColor = COLORS["agg_bg"]
ws["A1"] = "거래처별 매출·매입 순위"
ws["A1"].font = title_font()
ws["A2"] = "10_DIM_거래처 적재 후 SUMIFS로 자동 계산. 매출 기준 내림차순 정렬 권장."
ws["A2"].font = Font(name="맑은 고딕", size=10, italic=True, color="808080")

headers = ["거래처코드", "거래처명", "구분", "당해매출", "당해매입", "당해이익", "전년매출", "YoY증감액", "YoY증감률(%)", "최근거래일", "활성여부"]
write_headers(ws, headers, start_row=4, fill_key="primary")
for col, h in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(col)].width = 14
ws["A5"] = "(거래처 DIM 적재 후 행 채워짐)"
ws["A5"].font = Font(italic=True, color="808080")

# 44_미수금현황
ws = wb.create_sheet("44_미수금현황")
ws.sheet_properties.tabColor = COLORS["agg_bg"]
ws["A1"] = "거래처별 미수금 잔액"
ws["A1"].font = title_font()
ws["A2"] = "FACT_미수금 ledger 누계로 잔액 계산."
ws["A2"].font = Font(name="맑은 고딕", size=10, italic=True, color="808080")

headers = ["거래처코드", "거래처명", "세금계산서 누계", "입금 누계", "현재 잔액", "최종 거래일", "경과일수", "약정회수기일", "비고"]
write_headers(ws, headers, start_row=4, fill_key="primary")
for col, h in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(col)].width = 16
ws["A5"] = "(거래처 DIM 적재 후 행 채워짐)"
ws["A5"].font = Font(italic=True, color="808080")


# ===== 저장 =====
wb.save(OUT_PATH)
print(f"저장 완료: {OUT_PATH}")
print(f"총 시트 수: {len(wb.sheetnames)}")
print("시트 목록:")
for s in wb.sheetnames:
    print(f"  - {s}")
