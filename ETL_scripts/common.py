# -*- coding: utf-8 -*-
"""공통 헬퍼 — 경로, 워크북 로드/저장, 시트 데이터 갱신"""
import sys, io, os, re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Windows console 한글 출력 고정
if sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

ROOT = Path(r"C:\Users\scpar\OneDrive - Inviz\5.Inviz_Corporation\14.경영정보")
MASTER = ROOT / "00.경영관리마스터" / "인비즈_경영관리마스터_v1.xlsx"

SRC = {
    "외상매출금": ROOT / "26.결산자료/26.결산자료/외상매출금 (20260422).xlsx",
    "외상매입금": ROOT / "26.결산자료/26.결산자료/외상매입금 (20260422).xlsx",
    "단기차입금": ROOT / "26.결산자료/26.결산자료/단기차입금 및 임원 급여 미지급비용 (20260415).xlsx",
    "퇴직연금":   ROOT / "26.결산자료/26.결산자료/(주)인비즈_퇴직연금_월별입금액표_251231.xlsx",
    "주요계정":   ROOT / "26.결산자료/26.결산자료/주요계정명세서_(20251231).xlsx",
    "계약관리":   ROOT / "10.계약관리/10.계약관리/(주)인비즈_계약관리 (20260415).xlsx",
    "미수금":     ROOT / "10.계약관리/10.계약관리/미수금 현황 (20250813).xlsx",
    "급여대장":   ROOT / "04. 인사/04. 인사/2.급여/급여대장2025 3.xlsx",
    "부서인건비": ROOT / "04. 인사/04. 인사/2.급여/부서별 인건비_23-24.v1.xlsx",
    "월별비용":   ROOT / "14.경영정보" / "08. 보고서류/08. 보고서류/2024년 월별 비용정리.v2.xlsx",
    "매출분류":   ROOT / "08. 보고서류/08. 보고서류/매출분류 인비즈 (보고용).xlsx",
    "거래처세계": ROOT / "08. 보고서류/08. 보고서류/03. 거래처별매입매출세금계산서.xlsx",
    "판독수수료": ROOT / "08. 보고서류/08. 보고서류/5) 판독수수료 매출매입 정산_250620_SY.xlsx",
    "대리점매출": ROOT / "08. 보고서류/08. 보고서류/원격판독,AI 관련 인비즈 대리점 매출 (2024.03.15).xlsx",
    "관리비렌탈": ROOT / "08. 보고서류/08. 보고서류/관리비 및 렌탈현황 (20240715).xlsx",
}
SRC["월별비용"] = ROOT / "08. 보고서류/08. 보고서류/2024년 월별 비용정리.v2.xlsx"


def load_master():
    return load_workbook(MASTER)


def save_master(wb):
    wb.save(MASTER)
    print(f"[저장] {MASTER.name}")


def write_dim_or_fact(wb, sheet_name, df, table_name, start_row=2):
    """시트의 기존 데이터(헤더 제외)를 지우고 df를 적재. 표(Table) 범위 재설정.

    df의 컬럼 순서는 시트 헤더와 정확히 일치해야 함. 컬럼 수만 맞으면 됨(이름 무시).
    """
    ws = wb[sheet_name]
    # 헤더 행 확인 (헤더가 1행에 있는 시트만 처리)
    header_row = 1
    n_cols = sum(1 for c in ws[header_row] if c.value is not None and str(c.value).strip())
    if df.shape[1] != n_cols:
        raise ValueError(f"[{sheet_name}] 컬럼 수 불일치: df={df.shape[1]} vs sheet={n_cols}")

    # 기존 데이터 삭제 (max_row > 1인 경우)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # 데이터 적재
    for i, row in enumerate(df.itertuples(index=False), start=start_row):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val if val is not None else None)

    # 표 범위 재설정
    n_data = len(df)
    last_col = get_column_letter(n_cols)
    new_ref = f"A1:{last_col}{max(start_row + n_data - 1, start_row)}"

    # 기존 표 제거
    to_remove = [t for t in ws.tables.values() if t.displayName == table_name]
    for t in to_remove:
        del ws.tables[t.displayName]

    tbl = Table(displayName=table_name, ref=new_ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)
    print(f"  [{sheet_name}] {n_data}행 적재, 표 범위 {new_ref}")


def normalize_name(name):
    if name is None:
        return None
    s = str(name).strip()
    if not s or s.lower() in {"nan", "none"}:
        return None
    # 공백 정리
    s = re.sub(r"\s+", " ", s)
    # 후행 특수문자
    s = s.rstrip(".,;:")
    return s


def year_quarter_half(year, month):
    q = (int(month) - 1) // 3 + 1
    h = 1 if int(month) <= 6 else 2
    return int(year), int(month), f"Q{q}", f"H{h}"


def classify_party(name):
    """거래처명으로 구분 추정"""
    if not name:
        return "기타"
    s = name.lower()
    if any(k in name for k in ["의원", "병원", "한방", "내과", "이비인후과", "정형외과", "안과", "치과", "성형", "재활", "한의원", "재활의학", "요양"]):
        return "병원"
    if any(k in name for k in ["영상", "센터", "판독"]):
        return "영상센터"
    if any(k in name for k in ["메디칼", "메디컬", "Medical", "메디"]):
        return "공급사"
    if any(k in name for k in ["대학교", "연구원", "학교", "교육원"]):
        return "교육/연구"
    if any(k in name for k in ["대리점", "딜러"]):
        return "대리점"
    if any(k in name for k in ["은행", "Bank", "캐피탈", "카드", "보험"]):
        return "금융"
    if any(k in name for k in ["주식회사", "(주)", "㈜", "Corp", "Inc", "Ltd", "Co"]):
        return "법인기타"
    return "기타"


def apply_product_mapping(item_name, mapping_rules):
    """품명 → 제품코드. mapping_rules는 [(priority, pattern, code, name), ...]"""
    if not item_name:
        return ("P999", "기타")
    s = str(item_name)
    for priority, pattern, code, name in mapping_rules:
        if pattern == "*":
            return (code, name)
        if pattern.lower() in s.lower():
            return (code, name)
    return ("P999", "기타")


def get_mapping_rules(wb):
    """32_제품매핑 시트에서 룰 읽기"""
    ws = wb["32_제품매핑"]
    rules = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        priority, pattern, code, name = row[0], row[1], row[2], row[3]
        if pattern and code:
            rules.append((int(priority) if priority else 99, str(pattern), str(code), str(name)))
    rules.sort(key=lambda x: x[0])
    return rules


if __name__ == "__main__":
    print(f"MASTER: {MASTER}")
    print(f"EXISTS: {MASTER.exists()}")
    print("SOURCE files:")
    for k, p in SRC.items():
        print(f"  {k}: {'OK' if p.exists() else 'MISSING'}  {p.name}")
