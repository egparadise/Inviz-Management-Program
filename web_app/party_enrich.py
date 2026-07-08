# -*- coding: utf-8 -*-
"""거래처 정보 자동 보강 (Party Enrichment)

매출/매입 Excel 원본 파일에서 거래처별 추가 정보를 추출하여
dim_party 테이블의 누락된 칸을 채운다.

수집 대상:
- 사업자번호 (biz_no)
- 담당자 (contact_person)
- 직급/직책
- 전화 / 휴대폰 / 팩스
- 이메일
- 주소
- 대표자(ceo)

스캔 파일 우선순위(우상위가 더 신뢰):
1. `(영업계약건)/대리점 연락처.xlsx` — 가장 풍부한 연락정보
2. `(영업계약건)/수수료 - 원격판독 대리점 및 PACS 수수료 정리 ...xlsx`
3. `08. 보고서류/매출분류 인비즈 (보고용).xlsx` — biz_no + 거래처 매칭 (수천 건)
4. `더존자료/04. 전체매입매출목록2023년~.xlsx`
5. `더존자료/(주)인비즈_사이트 정리.xlsx`
6. `08. 보고서류/03. 거래처별매입매출세금계산서.xlsx` — 담당지역
7. 사용자가 지정한 추가 파일(generic scanner)

매칭 키:
- 이름 normalize: 공백 제거, (주)/주식회사 제거, 의원/병원/약국 유지
- biz_no 정규화: 숫자만, 10자리

작동 모드:
- dry_run=True (기본): 매칭 결과만 반환 — 실제 DB 변경 없음
- dry_run=False: 변경 적용 + enrich_source 기록 + 빈 칸만 채움 (보수적)
"""
import re
from pathlib import Path
from typing import Optional
from sqlalchemy import select
from sqlalchemy.orm import Session

from models import Party

# 14.경영정보/ 루트
DATA_ROOT = Path(__file__).parent.parent.parent.parent

# 후보 파일들 (없으면 skip)
SOURCE_FILES = [
    ("01. 기본서류/(영업계약건)/대리점 연락처.xlsx", "dealer_contact"),
    ("01. 기본서류/(영업계약건)/수수료 - 원격판독 대리점 및 PACS 수수료 정리 (24.03.05).xlsx", "fee_master"),
    ("08. 보고서류/매출분류 인비즈 (보고용).xlsx", "sale_classify"),
    ("08. 보고서류/03. 거래처별매입매출세금계산서.xlsx", "tax_invoice"),
    ("01. 기본서류/더존자료/04. 전체매입매출목록2023년~.xlsx", "duzon_invoice"),
    ("01. 기본서류/더존자료/(주)인비즈_사이트 정리.xlsx", "sites"),
]


# ========== 정규화 ==========
_PAREN_RE = re.compile(r"[()（）]")
_NOISE_RE = re.compile(r"(주식회사|\(주\)|\(유\)|주식\s*회사|（주）|（유）|（재）|\(재\)|\(사\)|\(학\))")
_SPACE_RE = re.compile(r"\s+")

def normalize_name(name: str) -> str:
    """매칭용 이름 정규화 — 공백·괄호·법인격 제거 후 소문자"""
    if not name:
        return ""
    s = str(name).strip()
    s = _NOISE_RE.sub("", s)
    s = _PAREN_RE.sub("", s)
    s = _SPACE_RE.sub("", s)
    return s.lower()


def normalize_bizno(s) -> Optional[str]:
    if s is None:
        return None
    digits = re.sub(r"\D", "", str(s))
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:5]}-{digits[5:]}"
    return None


_PHONE_RE = re.compile(r"^[\d\-\s()+]{7,20}$")
_EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")
_MOBILE_PREFIX = re.compile(r"^01[016789]")


def normalize_phone(s) -> Optional[str]:
    if s is None:
        return None
    s = re.sub(r"\s+", "", str(s).strip())
    if not s or s.lower() in ("nan", "none", "-"):
        return None
    if not _PHONE_RE.match(s):
        return None
    # 숫자만 추출 후 형식화
    digits = re.sub(r"\D", "", s)
    if len(digits) < 9 or len(digits) > 11:
        return s  # 그대로 (해외 등)
    if _MOBILE_PREFIX.match(digits):
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}" if len(digits) == 11 else \
               f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    if digits.startswith("02"):
        if len(digits) == 10: return f"02-{digits[2:6]}-{digits[6:]}"
        if len(digits) == 9: return f"02-{digits[2:5]}-{digits[5:]}"
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    return s


def is_mobile(phone: str) -> bool:
    if not phone:
        return False
    digits = re.sub(r"\D", "", phone)
    return bool(_MOBILE_PREFIX.match(digits))


def normalize_email(s) -> Optional[str]:
    if s is None:
        return None
    s = str(s).strip()
    return s if _EMAIL_RE.match(s) else None


# ========== 컬럼 헤더 자동 매핑 ==========
HEADER_HINTS = {
    "name": ["거래처", "회사명", "거래처명", "업체명", "상호", "고객명"],
    "biz_no": ["사업자", "사업자번호", "사업자(주민)번호", "사업자등록번호"],
    "ceo": ["대표자", "대표", "사장", "ceo"],
    "contact_person": ["담당자", "담당", "성명", "이름", "directly", "directly"],
    "title": ["직급", "직책"],
    "phone": ["연락처", "유선", "전화", "전화번호", "tel"],
    "mobile": ["핸드폰", "휴대폰", "휴대전화", "모바일", "mobile", "hp"],
    "fax": ["팩스", "fax"],
    "email": ["이메일", "메일", "e-mail", "email", "mail"],
    "address": ["주소", "사업장주소", "address"],
    "region": ["담당지역", "지역"],
}


def map_headers(header_row: list) -> dict:
    """헤더 행에서 우리 필드명 → 컬럼 인덱스 매핑"""
    mapping = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        c = str(cell).strip().lower().replace(" ", "")
        for fld, hints in HEADER_HINTS.items():
            for h in hints:
                if h.lower().replace(" ", "") in c:
                    if fld not in mapping:
                        mapping[fld] = idx
                    break
    return mapping


# ========== 파일별 추출기 ==========
def _open(path: Path):
    """openpyxl read_only — 큰 파일에서도 빠르게"""
    from openpyxl import load_workbook
    return load_workbook(path, read_only=True, data_only=True)


def scan_sheet_generic(ws, source_label: str) -> list[dict]:
    """헤더 자동 인식 — 첫 1~10행 중에서 '거래처/회사명' 포함된 행을 헤더로 본다"""
    out = []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return out
    # 헤더 행 후보
    header_idx = None; mapping = {}
    for i, row in enumerate(rows[:12]):
        m = map_headers(list(row))
        if "name" in m and (len(m) >= 2):  # 이름 + 1개 이상의 다른 필드
            header_idx = i; mapping = m
            break
    if header_idx is None:
        return out

    for row in rows[header_idx + 1:]:
        if not row:
            continue
        name_raw = row[mapping["name"]] if mapping["name"] < len(row) else None
        if not name_raw or str(name_raw).strip() in ("", "-", "합계", "소계"):
            continue
        rec = {"name": str(name_raw).strip(), "_source": source_label}
        for fld, idx in mapping.items():
            if fld in ("name",) or idx >= len(row):
                continue
            val = row[idx]
            if val is None:
                continue
            sval = str(val).strip()
            if not sval or sval.lower() in ("nan","none","-"):
                continue
            if fld == "biz_no":
                v = normalize_bizno(sval)
                if v: rec["biz_no"] = v
            elif fld in ("phone","mobile","fax"):
                v = normalize_phone(sval)
                if v:
                    # 휴대폰 자동 분기
                    if fld == "phone" and is_mobile(v):
                        rec["mobile"] = v
                    else:
                        rec[fld] = v
            elif fld == "email":
                v = normalize_email(sval)
                if v: rec["email"] = v
            elif fld == "title":
                rec["_title"] = sval
            else:
                rec[fld] = sval[:300]
        if len(rec) > 2:  # name·_source 외 1개 이상 정보
            out.append(rec)
    return out


def scan_sale_classify(path: Path) -> list[dict]:
    """매출분류 인비즈 (보고용).xlsx — 시트 '2021','2022','2023': 거래처+사업자번호+품명"""
    out = []
    try:
        wb = _open(path)
        for sn in wb.sheetnames:
            if not re.match(r"^\d{4}$", sn):  # 연도 시트만
                continue
            ws = wb[sn]
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            try:
                name_idx = header.index("거래처")
                biz_idx = next(i for i,h in enumerate(header) if h and "사업자" in str(h))
            except (ValueError, StopIteration):
                continue
            seen = set()
            for row in rows:
                if not row:
                    continue
                name = row[name_idx] if name_idx < len(row) else None
                biz = row[biz_idx] if biz_idx < len(row) else None
                if not name or not biz:
                    continue
                key = (str(name).strip(), normalize_bizno(biz))
                if not key[1] or key in seen:
                    continue
                seen.add(key)
                out.append({"name": key[0], "biz_no": key[1],
                            "_source": f"매출분류 인비즈 ({sn})"})
        wb.close()
    except Exception as e:
        print(f"[enrich] sale_classify 실패: {e}")
    return out


# ========== 매칭 + 적용 ==========
def collect_all(roots: list[str] = None) -> list[dict]:
    """모든 소스 파일을 스캔하여 enrichment record 리스트 반환"""
    records = []
    base = DATA_ROOT
    if roots:
        roots = [base / r for r in roots]
    else:
        roots = [base / rel for rel, _ in SOURCE_FILES]

    for rel, label in SOURCE_FILES:
        p = base / rel
        if not p.exists():
            continue
        try:
            if "매출분류" in p.name:
                recs = scan_sale_classify(p)
            else:
                # generic
                wb = _open(p)
                recs = []
                for sn in wb.sheetnames:
                    try:
                        recs.extend(scan_sheet_generic(wb[sn], f"{p.name} / {sn}"))
                    except Exception as e:
                        print(f"[enrich] {p.name}::{sn} 시트 실패: {e}")
                wb.close()
            print(f"[enrich] {p.name}: {len(recs)}건 추출")
            records.extend(recs)
        except Exception as e:
            print(f"[enrich] {p.name} 실패: {e}")
    return records


def merge_records(records: list[dict]) -> dict[str, dict]:
    """정규화된 이름 기준으로 record를 병합 — 같은 거래처의 정보가 여러 파일에 흩어져 있을 때 합침"""
    merged: dict[str, dict] = {}
    for r in records:
        key = normalize_name(r.get("name") or "")
        if not key:
            continue
        if key not in merged:
            merged[key] = {"_names": set(), "_sources": []}
        m = merged[key]
        m["_names"].add(r["name"])
        m["_sources"].append(r.get("_source", "?"))
        # 빈 칸만 채움 (먼저 들어온 값 우선)
        for k in ("biz_no","ceo","contact_person","phone","mobile","fax","email","address"):
            v = r.get(k)
            if v and not m.get(k):
                m[k] = v
    return merged


def apply_to_db(db: Session, merged: dict, *, dry_run: bool = True,
                fields: tuple = ("biz_no","ceo","contact_person","phone","mobile","fax","email","address")) -> dict:
    """dim_party에 적용. 빈 칸만 채움 (기존 값 우선).

    반환: {matched, updated, unmatched_count, unmatched_names, fill_counts}
    """
    parties = db.execute(select(Party)).scalars().all()
    name_idx = {}
    for p in parties:
        key = normalize_name(p.name)
        if key:
            name_idx.setdefault(key, []).append(p)

    matched_codes = set()
    updated_codes = set()
    fill_counts = {f: 0 for f in fields}
    unmatched = []

    for key, info in merged.items():
        plist = name_idx.get(key)
        if not plist:
            unmatched.append({"key": key, "names": list(info["_names"])[:3],
                              "info_fields": [f for f in fields if info.get(f)]})
            continue
        # 첫 매칭만 적용
        p = plist[0]
        matched_codes.add(p.code)
        changed = False
        for f in fields:
            v = info.get(f)
            if v and not getattr(p, f, None):
                if not dry_run:
                    setattr(p, f, v)
                fill_counts[f] += 1
                changed = True
        if changed:
            updated_codes.add(p.code)
            if not dry_run:
                src = list(dict.fromkeys(info["_sources"]))[:3]
                p.enrich_source = " · ".join(src)[:200]
    if not dry_run:
        db.commit()

    return {
        "matched": len(matched_codes),
        "updated": len(updated_codes),
        "fill_counts": fill_counts,
        "unmatched_count": len(unmatched),
        "unmatched_samples": unmatched[:30],
        "total_party": len(parties),
        "total_records": sum(len(v["_sources"]) for v in merged.values()),
        "merged_keys": len(merged),
    }


def run_enrichment(db: Session, *, dry_run: bool = True) -> dict:
    records = collect_all()
    merged = merge_records(records)
    result = apply_to_db(db, merged, dry_run=dry_run)
    result["dry_run"] = dry_run
    return result
