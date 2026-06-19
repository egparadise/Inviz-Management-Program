# -*- coding: utf-8 -*-
"""Excel/PDF Export 유틸 — 인비즈 브랜드 적용

- build_xlsx: openpyxl로 .xlsx 바이트 생성 (인비즈 보라 헤더, 합계 행)
- build_pdf: reportlab으로 PDF 바이트 생성 (맑은 고딕, 인비즈 로고)
"""
import io
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 인비즈 컬러
INVIZ_PURPLE = "6B2C91"
INVIZ_PURPLE_DARK = "4F1D6B"
INVIZ_ORANGE = "F47521"
INVIZ_PURPLE_LIGHT = "F5EDFA"

# 한글 폰트 경로 (Windows)
MALGUN_TTF = Path(r"C:\Windows\Fonts\malgun.ttf")
MALGUN_BOLD_TTF = Path(r"C:\Windows\Fonts\malgunbd.ttf")

STATIC_DIR = Path(__file__).parent / "static" / "img"
LOGO_PNG = STATIC_DIR / "inviz_logo.png"
LOGO_SVG = STATIC_DIR / "inviz_logo.svg"


# ====================== Excel ======================
def build_xlsx(title: str, filter_desc: str, headers: list[str], rows: list[list],
               sums: dict[str, float] = None, money_cols: list[int] = None) -> bytes:
    """엑셀 파일 생성 → bytes 반환

    headers: 컬럼명 list
    rows: 각 행은 list (헤더와 같은 길이)
    sums: {컬럼명: 합계값} — 합계 행에 표시
    money_cols: 콤마 포맷 적용할 컬럼 인덱스 (0-based)
    """
    money_cols = money_cols or []
    wb = Workbook()
    ws = wb.active
    ws.title = "데이터"

    # 헤더 위 — 제목 + 메타
    ws["A1"] = f"인비즈  {title}"
    ws["A1"].font = Font(name="맑은 고딕", size=18, bold=True, color=INVIZ_PURPLE)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 32

    ws["A2"] = f"필터: {filter_desc or '-'}"
    ws["A2"].font = Font(name="맑은 고딕", size=10, italic=True, color="595959")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    ws["A3"] = f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  총 {len(rows)}건"
    ws["A3"].font = Font(name="맑은 고딕", size=10, color="808080")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))

    # 헤더 행
    HDR_ROW = 5
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    fill = PatternFill("solid", fgColor=INVIZ_PURPLE)
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=HDR_ROW, column=col_idx, value=h)
        c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[HDR_ROW].height = 24

    # 데이터 행
    for ri, row in enumerate(rows, start=HDR_ROW + 1):
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="맑은 고딕", size=10)
            c.border = border
            if (ci - 1) in money_cols:
                c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="right")

    # 합계 행
    if sums:
        sum_row = HDR_ROW + 1 + len(rows)
        ws.cell(row=sum_row, column=1, value="합계").font = Font(name="맑은 고딕", size=11, bold=True, color=INVIZ_PURPLE)
        ws.cell(row=sum_row, column=1).fill = PatternFill("solid", fgColor="FEF3E2")
        ws.cell(row=sum_row, column=1).border = border
        # 헤더 중에서 sums에 해당하는 컬럼만 채움
        for ci, h in enumerate(headers, start=1):
            if h in sums:
                c = ws.cell(row=sum_row, column=ci, value=sums[h])
                c.font = Font(name="맑은 고딕", size=11, bold=True, color=INVIZ_PURPLE)
                c.fill = PatternFill("solid", fgColor="FEF3E2")
                c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="right")
                c.border = border
            else:
                if ci != 1:
                    c = ws.cell(row=sum_row, column=ci)
                    c.fill = PatternFill("solid", fgColor="FEF3E2")
                    c.border = border

    # 열 너비
    for col_idx, h in enumerate(headers, 1):
        w = 14 if (col_idx - 1) in money_cols else max(12, min(len(str(h)) * 2 + 4, 28))
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = f"A{HDR_ROW + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ====================== PDF ======================
_pdf_font_registered = False


def _register_pdf_fonts():
    """맑은 고딕을 reportlab에 등록 (한 번만)"""
    global _pdf_font_registered
    if _pdf_font_registered:
        return
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    try:
        if MALGUN_TTF.exists():
            pdfmetrics.registerFont(TTFont("Malgun", str(MALGUN_TTF)))
        if MALGUN_BOLD_TTF.exists():
            pdfmetrics.registerFont(TTFont("MalgunBold", str(MALGUN_BOLD_TTF)))
        from reportlab.pdfbase.pdfmetrics import registerFontFamily
        registerFontFamily("Malgun", normal="Malgun",
                           bold="MalgunBold" if MALGUN_BOLD_TTF.exists() else "Malgun",
                           italic="Malgun", boldItalic="MalgunBold" if MALGUN_BOLD_TTF.exists() else "Malgun")
        _pdf_font_registered = True
    except Exception as e:
        print(f"[export_util] 폰트 등록 실패: {e}")


def build_pdf(title: str, filter_desc: str, headers: list[str], rows: list[list],
              sums: dict[str, float] = None, money_cols: list[int] = None) -> bytes:
    """PDF 생성 → bytes 반환 (가로 A4)"""
    _register_pdf_fonts()
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.platypus.flowables import HRFlowable

    money_cols = money_cols or []
    base_font = "Malgun" if MALGUN_TTF.exists() else "Helvetica"
    bold_font = "MalgunBold" if MALGUN_BOLD_TTF.exists() else base_font

    purple = colors.HexColor("#" + INVIZ_PURPLE)
    orange = colors.HexColor("#" + INVIZ_ORANGE)
    light = colors.HexColor("#" + INVIZ_PURPLE_LIGHT)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title=title, author="Inviz",
    )
    story = []

    # 헤더 — 로고 + 제목
    title_style = ParagraphStyle(
        "title", fontName=bold_font, fontSize=18, textColor=purple, leading=22,
    )
    meta_style = ParagraphStyle(
        "meta", fontName=base_font, fontSize=9, textColor=colors.HexColor("#666666"), leading=12,
    )

    # 로고 배치
    logo_flow = None
    if LOGO_PNG.exists():
        try:
            logo_flow = Image(str(LOGO_PNG), width=42 * mm, height=14 * mm, kind="proportional")
        except Exception:
            logo_flow = None

    if logo_flow:
        header_tbl = Table([[logo_flow, Paragraph(f"<b>{title}</b>", title_style)]],
                          colWidths=[50 * mm, None])
        header_tbl.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(header_tbl)
    else:
        story.append(Paragraph(f"<b>인비즈 — {title}</b>", title_style))

    story.append(HRFlowable(width="100%", thickness=1.5, color=orange, spaceBefore=4, spaceAfter=6))

    story.append(Paragraph(
        f"필터: {filter_desc or '-'}  ·  생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}  ·  총 {len(rows):,}건",
        meta_style,
    ))
    story.append(Spacer(1, 4))

    # 테이블 데이터
    def fmt(v, is_money):
        if v is None:
            return ""
        if is_money:
            try:
                return f"{float(v):,.0f}"
            except Exception:
                return str(v)
        return str(v)

    table_data = [headers]
    for r in rows:
        table_data.append([fmt(v, (i in money_cols)) for i, v in enumerate(r)])

    # 합계 행
    if sums:
        sum_row = []
        for ci, h in enumerate(headers):
            if ci == 0:
                sum_row.append("합계")
            elif h in sums:
                sum_row.append(f"{sums[h]:,.0f}")
            else:
                sum_row.append("")
        table_data.append(sum_row)

    # 열 너비 자동 계산 (간단 휴리스틱)
    page_width = landscape(A4)[0] - 30 * mm
    n = len(headers)
    # 머니 컬럼은 좀더 좁게, 텍스트는 길게
    base_widths = []
    for i, h in enumerate(headers):
        if i in money_cols:
            base_widths.append(20)  # 단위 가중치
        else:
            base_widths.append(25)
    total = sum(base_widths)
    col_widths = [page_width * w / total for w in base_widths]

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    style = [
        ("FONTNAME", (0, 0), (-1, -1), base_font),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        # 헤더
        ("BACKGROUND", (0, 0), (-1, 0), purple),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), bold_font),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        # 그리드
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        # 짝수 행 배경
        ("ROWBACKGROUNDS", (0, 1), (-1, -2 if sums else -1),
            [colors.white, colors.HexColor("#FAFAFA")]),
    ]
    # 머니 컬럼 우정렬
    for ci in money_cols:
        style.append(("ALIGN", (ci, 1), (ci, -1), "RIGHT"))
    # 합계 행
    if sums:
        style.extend([
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FEF3E2")),
            ("FONTNAME", (0, -1), (-1, -1), bold_font),
            ("TEXTCOLOR", (0, -1), (-1, -1), purple),
            ("LINEABOVE", (0, -1), (-1, -1), 1.5, orange),
        ])
    tbl.setStyle(TableStyle(style))
    story.append(tbl)

    # 푸터
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        '<font color="#999999" size="8">© Inviz Corporation — 경영관리 시스템</font>',
        meta_style,
    ))

    doc.build(story)
    return buf.getvalue()
