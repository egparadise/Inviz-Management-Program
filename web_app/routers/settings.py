# -*- coding: utf-8 -*-
"""통합 설정 라우터 — 외관·AI·출력·검색·계정 설정

설정값은 settings_store(app_setting 테이블)에 저장되며,
Jinja 전역 setting()으로 모든 페이지에서 적용된다.
"""
import io
import re
from datetime import datetime, timedelta
from pathlib import Path
from fastapi import APIRouter, Request, Form, Depends, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from sqlalchemy import select, func, desc, and_
from sqlalchemy.orm import Session

from fastapi import HTTPException
from database import get_db
from helpers import templates
from models import ActivityLog, Product, BankAccount, Card
import settings_store as ss

router = APIRouter()

STATIC_IMG = Path(__file__).parent.parent / "static" / "img"
STATIC_IMG.mkdir(parents=True, exist_ok=True)

# 설정 좌측 메뉴 — 각 항목이 /settings/<key> 페이지
SETTINGS_PAGES = [
    ("general", "일반 / 화면", "🎨"),
    ("network", "네트워크·사내접속", "🌐"),
    ("ai", "AI 모델·공급자", "🤖"),
    ("integrations", "연동·알림", "🔔"),
    ("selfdev", "자가발전 / 학습", "🛡"),
    ("schedule", "자동 실행", "⏰"),
    ("products", "제품 관리", "📦"),
    ("payroll", "급여 요율", "🧮"),
    ("banking", "은행·카드", "💳"),
    ("txn-mgmt", "매출·매입 관리", "🗑"),
    ("db", "DB 탐색기", "🗄"),
    ("menu", "메뉴 순서", "🧭"),
    ("account", "계정·보안·접속", "🔐"),
    ("logs", "활동 로그", "🧾"),
]
SETTINGS_PAGE_KEYS = {p[0] for p in SETTINGS_PAGES}
# 저장 섹션(_section) → 돌아갈 페이지
SECTION_TO_PAGE = {
    "base": "general", "appearance": "general", "output": "general", "search": "general", "logo": "general",
    "ai": "ai", "provider": "ai", "integrations": "integrations", "selfdev": "selfdev",
    "openbank": "banking", "payroll": "payroll", "network": "network",
}


def _net_info():
    """서버 LAN 주소(들) — 같은 네트워크의 다른 PC 접속용"""
    import socket
    ips = set()
    try:
        host = socket.gethostname()
        for info in socket.getaddrinfo(host, None):
            ip = info[4][0]
            if ":" not in ip and not ip.startswith("127."):
                ips.add(ip)
    except Exception:
        pass
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80)); ips.add(s.getsockname()[0]); s.close()
    except Exception:
        pass
    return sorted(ips)


def _render_settings(request, page="general", log_category="", log_limit=100, log_from="", log_to=""):
    if page not in SETTINGS_PAGE_KEYS:
        page = "general"
    cfg = ss.all()
    users = ss.get_users()
    pw_is_custom = bool(ss.get("auth_password_hash"))
    key_set = {
        "openai": bool(ss.get("ai_openai_key")),
        "anthropic": bool(ss.get("ai_anthropic_key")),
        "gemini": bool(ss.get("ai_gemini_key")),
        "popbill": bool(ss.get("tax_popbill_secret")),
        "barobill": bool(ss.get("tax_barobill_certkey")),
        "kakao": bool(ss.get("kakao_access_token")),
        "smtp": bool(ss.get("mail_smtp_pass")),
        "imap": bool(ss.get("mail_imap_pass")),
        "telegram": bool(ss.get("telegram_bot_token")),
    }
    try:
        import integrations as ig
        integ = {"mail_send": ig.mail_send_ready(), "mail_recv": ig.mail_recv_ready(),
                 "kakao": ig.kakao_ready(), "asp": ig.asp_ready(), "telegram": ig.telegram_ready()}
    except Exception:
        integ = {"mail_send": False, "mail_recv": False, "kakao": False, "asp": False, "telegram": False}
    try:
        import llm_provider
        provider_label = llm_provider.active_label()
        provider_ok, provider_msg = llm_provider.provider_ready()
    except Exception as e:
        provider_label, provider_ok, provider_msg = "?", False, str(e)
    lf = (cfg.get("learning_folder") or "").strip()
    learning_folder_exists = bool(lf) and Path(lf).exists()
    bf = (cfg.get("base_data_folder") or "").strip()
    base_folder_exists = bool(bf) and Path(bf).exists()
    selfdev_status = [{"path": p, "exists": Path(p).exists()} for p in ss.selfdev_path_list()]

    # 제품 목록(제품 페이지) + 활동 로그(로그 페이지만 — 항목·날짜 필터)
    from database import SessionLocal
    db = SessionLocal()
    logs, log_total, cats = [], 0, []
    bank_accounts, cards, acc_name_map = [], [], {}
    pay_rates, pay_rate_defaults = {}, {}
    if page == "payroll":
        try:
            import routers.payroll as _pr
            pay_rates = _pr._rates()
            pay_rate_defaults = dict(_pr._RATE_DEFAULTS)
        except Exception:
            pay_rates, pay_rate_defaults = {}, {}
    try:
        prod_rows = db.execute(select(Product).order_by(Product.code)).scalars().all()
        if page == "banking":
            bank_accounts = db.execute(select(BankAccount).order_by(
                BankAccount.sort_order, BankAccount.bank_name, BankAccount.id)).scalars().all()
            cards = db.execute(select(Card).order_by(Card.sort_order, Card.id)).scalars().all()
            acc_name_map = {a.id: f"{a.bank_name}{(' · ' + a.account_alias) if a.account_alias else ''}"
                            for a in bank_accounts}
        if page == "logs":
            cats = [r[0] for r in db.execute(
                select(ActivityLog.category, func.count()).group_by(ActivityLog.category)
                .order_by(func.count().desc())).all() if r[0]]
            stmt = select(ActivityLog)
            if log_category:
                stmt = stmt.where(ActivityLog.category == log_category)
            try:
                if log_from:
                    stmt = stmt.where(ActivityLog.ts >= datetime.fromisoformat(log_from))
                if log_to:
                    stmt = stmt.where(ActivityLog.ts <= datetime.fromisoformat(log_to + "T23:59:59"))
            except Exception:
                pass
            log_limit = max(10, min(int(log_limit or 100), 1000))
            log_total = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
            logs = db.execute(stmt.order_by(desc(ActivityLog.id)).limit(log_limit)).scalars().all()
    finally:
        db.close()

    return templates.TemplateResponse("settings/view.html", {
        "request": request, "cfg": cfg, "users": users,
        "pw_is_custom": pw_is_custom, "key_set": key_set,
        "provider_label": provider_label, "provider_ok": provider_ok, "provider_msg": provider_msg,
        "learning_folder_exists": learning_folder_exists, "selfdev_status": selfdev_status,
        "base_folder_exists": base_folder_exists,
        "logs": logs, "log_total": log_total, "log_categories": cats,
        "log_category": log_category, "log_limit": log_limit,
        "log_from": log_from, "log_to": log_to,
        "net_ips": _net_info(), "prod_rows": prod_rows, "integ": integ,
        "bank_accounts": bank_accounts, "cards": cards, "acc_name_map": acc_name_map,
        "banks": ["신한은행", "광주은행", "하나은행", "기업은행", "국민은행", "우리은행", "농협은행", "기타"],
        "card_issuers": ["우리카드", "광주카드", "하나카드", "신한카드", "삼성카드", "현대카드", "기타"],
        "pay_rates": pay_rates, "pay_rate_defaults": pay_rate_defaults,
        "page": page, "pages": SETTINGS_PAGES,
        "db_inventory": _db_inventory() if page == "db" else None,
    })


@router.get("/logs/download.{fmt}")
def settings_logs_download(fmt: str, log_category: str = "", log_from: str = "", log_to: str = "",
                           db: Session = Depends(get_db)):
    """활동 로그 다운로드 (txt 또는 xlsx) — 항목·날짜 필터 적용"""
    stmt = select(ActivityLog)
    if log_category:
        stmt = stmt.where(ActivityLog.category == log_category)
    try:
        if log_from:
            stmt = stmt.where(ActivityLog.ts >= datetime.fromisoformat(log_from))
        if log_to:
            stmt = stmt.where(ActivityLog.ts <= datetime.fromisoformat(log_to + "T23:59:59"))
    except Exception:
        pass
    rows = db.execute(stmt.order_by(desc(ActivityLog.id)).limit(50000)).scalars().all()
    today = datetime.now().strftime("%Y%m%d_%H%M")

    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook(); wsx = wb.active; wsx.title = "활동로그"
        headers = ["일시", "항목", "동작", "메서드", "경로", "상태", "IP", "소요(ms)"]
        wsx.append(headers)
        for c in range(1, len(headers) + 1):
            cell = wsx.cell(row=1, column=c)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill("solid", fgColor="6B2C91")
        for r in rows:
            wsx.append([
                r.ts.strftime("%Y-%m-%d %H:%M:%S") if r.ts else "",
                r.category or "", r.action or "", r.method or "",
                r.path or "", r.status_code or "", r.client_ip or "", r.duration_ms or "",
            ])
        for col, w in zip("ABCDEFGH", (20, 12, 30, 8, 40, 8, 16, 10)):
            wsx.column_dimensions[col].width = w
        buf = io.BytesIO(); wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="activity_log_{today}.xlsx"'},
        )
    # txt
    lines = ["일시\t항목\t동작\t상태\tIP\t소요(ms)"]
    for r in rows:
        lines.append("\t".join([
            r.ts.strftime("%Y-%m-%d %H:%M:%S") if r.ts else "",
            r.category or "", r.action or "", str(r.status_code or ""),
            r.client_ip or "", str(r.duration_ms or ""),
        ]))
    data = "\n".join(lines).encode("utf-8-sig")
    return Response(content=data, media_type="text/plain; charset=utf-8",
                    headers={"Content-Disposition": f'attachment; filename="activity_log_{today}.txt"'})


@router.get("/browse")
def settings_browse(path: str = ""):
    """서버 측 폴더 탐색기 — 지정 경로의 하위 폴더·파일 목록 반환 (로컬 PC 전용)."""
    import os
    info = {"cwd": "", "parent": None, "dirs": [], "files": [], "drives": [], "error": None}
    # 보안: 인터넷 공개 배포(INVIZ_PUBLIC=1)에서는 서버 파일시스템 탐색 비활성화
    if os.environ.get("INVIZ_PUBLIC") == "1":
        info["error"] = "공개 배포 모드에서는 보안상 서버 폴더 탐색이 비활성화되어 있습니다. (로컬 설치본에서만 사용)"
        return info
    try:
        if not path:
            # 루트: OneDrive · 빠른 위치 · 회사 기본 폴더 · 홈 · 드라이브
            try:
                drives = [str(d) for d in os.listdrives()]
            except Exception:
                drives = [f"{c}:\\" for c in "CDEFGHIJ" if Path(f"{c}:\\").exists()]
            info["drives"] = drives
            home = Path.home()
            shortcuts = []
            seen = set()

            def _add(label, p):
                try:
                    pp = Path(p)
                except Exception:
                    return
                key = str(pp).lower()
                if key in seen or not pp.is_dir():
                    return
                seen.add(key)
                shortcuts.append({"name": label, "path": str(pp)})

            # 1) OneDrive 루트(회사/개인) — 탐색기의 '성철 - Inviz' = OneDrive - Inviz
            od_roots = []
            for env in ("OneDriveCommercial", "OneDrive", "OneDriveConsumer"):
                v = os.environ.get(env)
                if v:
                    od_roots.append(v)
            for d in sorted(home.glob("OneDrive*")):
                od_roots.append(str(d))
            for od in od_roots:
                _add(f"☁️ {Path(od).name}", od)

            # 2) 빠른 위치 (바탕화면/다운로드/문서/사진/음악/동영상) — OneDrive 리디렉션·한글 폴더명 고려
            od_main = (os.environ.get("OneDriveCommercial") or os.environ.get("OneDrive")
                       or str(home / "OneDrive - Inviz"))
            quick = [
                ("🖥 바탕 화면", ["Desktop", "바탕 화면"]),
                ("⬇ 다운로드", ["Downloads", "다운로드"]),
                ("📄 문서", ["Documents", "문서"]),
                ("🖼 사진", ["Pictures", "그림"]),
                ("🎵 음악", ["Music", "음악"]),
                ("🎬 동영상", ["Videos", "동영상"]),
            ]
            for label, names in quick:
                for nm in names:
                    hit = None
                    for bdir in (home, Path(od_main)):
                        cp = bdir / nm
                        if cp.is_dir():
                            hit = cp
                            break
                    if hit:
                        _add(label, hit)
                        break

            # 3) 회사 기본 폴더
            _add("🏢 14.경영정보", home / "OneDrive - Inviz" / "5.Inviz_Corporation" / "14.경영정보")
            # 4) 홈 / 5) 드라이브
            _add("🏠 내 폴더", home)
            for d in drives:
                _add(f"💽 {d}", d)

            info["dirs"] = shortcuts
            # 처음 열 때 자동 진입할 위치: OneDrive 루트 → 14.경영정보 → 홈
            start = None
            for cand in (od_main, str(home / "OneDrive - Inviz"),
                         str(home / "OneDrive - Inviz" / "5.Inviz_Corporation" / "14.경영정보"),
                         str(home)):
                if cand and Path(cand).is_dir():
                    start = str(Path(cand))
                    break
            info["start_path"] = start
            return info
        p = Path(path)
        if not p.exists() or not p.is_dir():
            info["error"] = "폴더를 찾을 수 없습니다."
            return info
        info["cwd"] = str(p)
        parent = p.parent
        info["parent"] = str(parent) if str(parent) != str(p) else ""
        for child in sorted(p.iterdir(), key=lambda x: x.name.lower()):
            try:
                if child.name.startswith("."):
                    continue
                if child.is_dir():
                    info["dirs"].append({"name": child.name, "path": str(child)})
                elif child.suffix.lower() in (".xlsx", ".xls", ".xlsm", ".csv", ".pdf"):
                    info["files"].append({"name": child.name, "path": str(child)})
            except (PermissionError, OSError):
                continue
        return info
    except PermissionError:
        info["error"] = "접근 권한이 없습니다."
        return info
    except Exception as e:
        info["error"] = str(e)
        return info


@router.post("/save")
async def settings_save(request: Request):
    """설정 저장 — 섹션별(_section) 또는 전체. 전송된 필드만 갱신."""
    form = await request.form()
    section = form.get("_section", "")
    keys = [
        "ui_font_scale", "ui_logo_show", "ui_input_size", "app_title",
        "ai_default_model", "ai_default_mode",
        "ai_provider", "ai_openai_model", "ai_anthropic_model", "ai_gemini_model",
        "export_default", "pdf_orientation",
        "search_default_period", "search_per_page",
        "learning_model", "learning_folder", "selfdev_paths",
        "base_data_folder",
        # 국세청/홈택스(ASP)
        "tax_corp_no", "tax_corp_name", "tax_ceo", "tax_issuer_email",
        "tax_asp", "tax_popbill_linkid", "tax_popbill_test",
        # 카카오 / 이메일 (비밀값 제외)
        "mail_smtp_host", "mail_smtp_port", "mail_smtp_user", "mail_from",
        "mail_imap_host", "mail_imap_port", "mail_imap_user", "mail_notify_to", "mail_imap_days",
        "telegram_chat_id",
        # 오픈뱅킹(기관 API) 연동 — 준비
        "openbank_org_code", "openbank_client_id", "openbank_api_base",
        # 급여 4대보험 요율·상한
        "pay_rate_pension", "pay_rate_health", "pay_rate_longterm", "pay_rate_employment",
        "pay_rate_local_tax", "pay_rate_emp_employment", "pay_rate_accident",
        "pay_rate_pension_cap", "pay_rate_pension_floor", "pay_rate_meal_taxfree",
        # 네트워크 / 사내접속
        "net_bind_host", "net_port", "net_domain", "net_https",
        "net_cert_path", "net_key_path", "net_intranet_enabled",
    ]
    updates = {}
    for k in keys:
        v = form.get(k)
        if v is not None:
            updates[k] = v.strip() if isinstance(v, str) else v
    # 비밀값(API키·비밀번호·토큰) — 입력값이 있을 때만 갱신(빈 값으로 덮어쓰지 않음)
    for kk in ("ai_openai_key", "ai_anthropic_key", "ai_gemini_key",
               "tax_popbill_secret", "tax_barobill_certkey",
               "kakao_rest_key", "kakao_access_token", "kakao_refresh_token",
               "mail_smtp_pass", "mail_imap_pass", "telegram_bot_token",
               "openbank_client_secret"):
        v = form.get(kk)
        if isinstance(v, str) and v.strip():
            updates[kk] = v.strip()
    # 사용 모델 체크박스 (공급자 섹션) → ai_provider 도출
    if section in ("provider", ""):
        if form.get("prov_openai"):
            updates["ai_provider"] = "openai"
        elif form.get("prov_anthropic"):
            updates["ai_provider"] = "anthropic"
        elif form.get("prov_gemini"):
            updates["ai_provider"] = "gemini"
        else:
            updates["ai_provider"] = "ollama"
    # 체크박스 기본값 — 해당 섹션 저장 시에만
    if section in ("appearance", ""):
        updates["ui_logo_show"] = "1" if form.get("ui_logo_show") else "0"
    if section in ("selfdev", ""):
        updates["learning_include_subfolders"] = "1" if form.get("learning_include_subfolders") else "0"
    # 네트워크 — 체크박스 기본값 + intranet_enabled 자동 처리
    if section in ("network", ""):
        intranet = "1" if form.get("net_intranet_enabled") else "0"
        updates["net_intranet_enabled"] = intranet
        updates["net_https"] = "1" if form.get("net_https") else "0"
        # 사내망 허용이 켜졌고 사용자가 호스트를 127.0.0.1으로 둔 경우 자동 0.0.0.0으로
        if intranet == "1":
            current_host = (form.get("net_bind_host") or "").strip()
            if not current_host or current_host == "127.0.0.1":
                updates["net_bind_host"] = "0.0.0.0"
    if section in ("integrations", ""):
        updates["kakao_enabled"] = "1" if form.get("kakao_enabled") else "0"
        updates["mail_smtp_tls"] = "1" if form.get("mail_smtp_tls") else "0"
        updates["telegram_enabled"] = "1" if form.get("telegram_enabled") else "0"
    if section in ("openbank", ""):
        updates["openbank_enabled"] = "1" if form.get("openbank_enabled") else "0"
    ss.save(updates)
    return RedirectResponse(f"/settings/{SECTION_TO_PAGE.get(section, 'general')}?saved=1", status_code=303)


@router.post("/base-folder/apply")
def settings_base_apply():
    """기본 데이터 폴더 전체를 AI가 분석하여 시스템에 적용 (백그라운드 안전 동기화) — 폼 폴백"""
    import scheduler
    scheduler.run_task_async("selfdev", manual=True)
    return RedirectResponse("/settings/general?base_applying=1", status_code=303)


# ====== 백그라운드 AI 분석 — 진행 추적 + 보고서 ======
import threading as _thr
import uuid as _uuid
import time as _time
_BASE_JOBS: dict = {}


def _capture_kpis_snapshot() -> dict:
    """현재 DB의 주요 테이블 행수·합계 스냅샷 (분석 전후 비교용)."""
    from database import SessionLocal
    from sqlalchemy import select, func
    from models import (Sale, Purchase, Payroll, Expense, Contract,
                        LoanMaster, Party, Employee, Product, Document)
    db = SessionLocal()
    try:
        snap = {
            "fact_sale": db.execute(select(func.count()).select_from(Sale)).scalar() or 0,
            "fact_sale_sum": float(db.execute(select(func.coalesce(func.sum(Sale.supply), 0))).scalar() or 0),
            "fact_purchase": db.execute(select(func.count()).select_from(Purchase)).scalar() or 0,
            "fact_purchase_sum": float(db.execute(select(func.coalesce(func.sum(Purchase.supply), 0))).scalar() or 0),
            "fact_payroll": db.execute(select(func.count()).select_from(Payroll)).scalar() or 0,
            "fact_expense": db.execute(select(func.count()).select_from(Expense)).scalar() or 0,
            "master_contract": db.execute(select(func.count()).select_from(Contract)).scalar() or 0,
            "master_loan": db.execute(select(func.count()).select_from(LoanMaster)).scalar() or 0,
            "dim_party": db.execute(select(func.count()).select_from(Party)).scalar() or 0,
            "dim_employee": db.execute(select(func.count()).select_from(Employee)).scalar() or 0,
            "dim_product": db.execute(select(func.count()).select_from(Product)).scalar() or 0,
            "document": db.execute(select(func.count()).select_from(Document)).scalar() or 0,
        }
        try:
            from models import KnowledgeChunk
            snap["knowledge_chunk"] = db.execute(select(func.count()).select_from(KnowledgeChunk)).scalar() or 0
        except Exception:
            snap["knowledge_chunk"] = 0
        try:
            from models import FileRegistry
            snap["file_registry"] = db.execute(select(func.count()).select_from(FileRegistry)).scalar() or 0
        except Exception:
            snap["file_registry"] = 0
        return snap
    finally:
        db.close()


def _run_base_analysis(run_id: str, folder_path: str):
    job = _BASE_JOBS[run_id]
    import traceback
    try:
        # 1) 폴더 등록
        job.update(stage="📂 폴더 등록", progress=3)
        if folder_path:
            ss.save({"base_data_folder": folder_path})
            ss.invalidate()

        # 2) 사전 KPI
        job.update(stage="📊 분석 전 상태 측정", progress=8)
        job["kpis_before"] = _capture_kpis_snapshot()

        # 3) sync_core (폴더 스캔 + ETL)
        job.update(stage="🔍 폴더 스캔·ETL 적재", progress=15)
        import sync_core
        result = sync_core.run_sync(triggered_by="base-apply", force=True, verbose=False)
        job["sync_run_id"] = result.id
        job["files_processed"] = result.files_processed
        job["files_errored"] = result.files_errored
        job["rows_added"] = result.rows_added
        job.update(progress=50)

        # 4) LLM 분류 (미분류 파일들 도메인 추정)
        job.update(stage="🤖 LLM 자동 도메인 분류")
        import self_dev
        from database import SessionLocal
        total_llm = {"classified": 0, "auto_assigned": 0, "queued": 0}
        for i in range(5):
            db = SessionLocal()
            try:
                r = self_dev.review_unmapped_files(db, max_files=100, model="llama3.1:latest")
                for k, v in r.items():
                    total_llm[k] += v
                job.update(progress=50 + 6 * (i + 1), llm_classify=total_llm)
                if r.get("classified", 0) == 0:
                    break
            finally:
                db.close()
        job["llm_classify"] = total_llm

        # 5) RAG 인덱싱
        job.update(stage="📚 RAG 벡터 인덱싱", progress=85)
        try:
            import rag_ingest
            rag_result = rag_ingest.run_full_ingest()
            job["rag"] = rag_result
        except Exception as e:
            job["rag_error"] = str(e)[:300]

        # 6) 사후 KPI + 변동 분석
        job.update(stage="📈 변동 분석·보고서 생성", progress=95)
        job["kpis_after"] = _capture_kpis_snapshot()

        job.update(stage="✅ 완료", progress=100, ok=True, finished_at=_time.time())
    except Exception as e:
        job["error"] = f"{type(e).__name__}: {e}"
        job["traceback"] = traceback.format_exc()[:2000]
        job.update(stage="❌ 오류", ok=False, finished_at=_time.time())


@router.post("/base-folder/start")
async def settings_base_start(request: Request):
    """비동기 AI 분석 시작. run_id 반환."""
    form = dict(await request.form())
    folder = (form.get("base_data_folder") or ss.get("base_data_folder", "") or "").strip()
    if not folder:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "폴더 경로가 비어있습니다"}, status_code=400)
    if not Path(folder).exists():
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": f"폴더가 존재하지 않습니다: {folder}"}, status_code=400)

    run_id = _uuid.uuid4().hex[:12]
    _BASE_JOBS[run_id] = {
        "run_id": run_id, "started_at": _time.time(), "folder": folder,
        "stage": "🚀 초기화", "progress": 0, "ok": None, "finished_at": None,
    }
    t = _thr.Thread(target=_run_base_analysis, args=(run_id, folder), daemon=True)
    t.start()
    from fastapi.responses import JSONResponse
    return JSONResponse({"run_id": run_id})


@router.get("/base-folder/status")
def settings_base_status(run_id: str):
    from fastapi.responses import JSONResponse
    job = _BASE_JOBS.get(run_id)
    if not job:
        return JSONResponse({"error": "unknown_run_id"}, status_code=404)
    return JSONResponse({k: v for k, v in job.items() if k != "traceback"})


@router.get("/base-folder/report")
def settings_base_report(run_id: str):
    from fastapi.responses import JSONResponse
    job = _BASE_JOBS.get(run_id)
    if not job:
        return JSONResponse({"error": "unknown_run_id"}, status_code=404)
    if not job.get("finished_at"):
        return JSONResponse({"error": "not_finished"}, status_code=400)

    before = job.get("kpis_before") or {}
    after = job.get("kpis_after") or {}
    LABELS = {
        "fact_sale": "매출 건수", "fact_sale_sum": "매출 공급가 합계(원)",
        "fact_purchase": "매입 건수", "fact_purchase_sum": "매입 공급가 합계(원)",
        "fact_payroll": "급여 건수", "fact_expense": "지출 건수",
        "master_contract": "계약 건수", "master_loan": "차입금 건수",
        "dim_party": "거래처 마스터", "dim_employee": "직원 마스터",
        "dim_product": "제품 마스터", "document": "서류·인증",
        "knowledge_chunk": "RAG 지식 청크", "file_registry": "추적 파일 메타",
    }
    changes = []
    for k in [k for k in LABELS if k in before or k in after]:
        b = before.get(k, 0); a = after.get(k, 0)
        diff = a - b
        if diff != 0:
            changes.append({
                "key": k, "label": LABELS[k],
                "before": b, "after": a, "diff": diff,
                "pct": (diff / b * 100) if b else None,
            })

    # 개선된 기능 요약
    improvements = []
    if (job.get("rows_added") or 0) > 0:
        improvements.append(f"📊 {job['rows_added']:,}개 데이터 행이 신규 적재되어 매출·매입·계약 등 통계가 최신화됨")
    if (job.get("llm_classify", {}) or {}).get("auto_assigned", 0) > 0:
        improvements.append(f"🤖 LLM이 {job['llm_classify']['auto_assigned']}개 파일을 자동 도메인 분류 (다음 sync에서 ETL 처리 예정)")
    rag = job.get("rag") or {}
    if rag.get("embedded", 0) > 0:
        improvements.append(f"📚 RAG에 {rag['embedded']}개 청크가 임베딩되어 AI 챗 답변 품질·근거 자료 강화")
    if (job.get("files_processed") or 0) > 0:
        improvements.append(f"📂 {job['files_processed']}개 Excel/CSV 파일이 검증·처리됨 (오류 {job.get('files_errored', 0)}건)")
    if not improvements:
        improvements.append("ℹ️ 신규 변경 사항 없음 — 모든 데이터가 이미 최신 상태입니다.")

    return JSONResponse({
        "run_id": run_id, "ok": job.get("ok"),
        "folder": job.get("folder"),
        "elapsed_s": int((job["finished_at"] - job["started_at"])),
        "files_processed": job.get("files_processed", 0),
        "files_errored": job.get("files_errored", 0),
        "rows_added": job.get("rows_added", 0),
        "llm_classify": job.get("llm_classify") or {},
        "rag": rag, "rag_error": job.get("rag_error"),
        "error": job.get("error"),
        "kpis_before": before, "kpis_after": after,
        "changes": changes, "improvements": improvements,
    })


@router.post("/logo")
async def settings_logo(file: UploadFile = File(...)):
    """로고 이미지 업로드 → static/img/custom_logo.* 저장"""
    name = (file.filename or "").lower()
    ext = ".png"
    for e in (".png", ".jpg", ".jpeg", ".svg", ".webp", ".gif"):
        if name.endswith(e):
            ext = e
            break
    content = await file.read()
    fname = f"custom_logo{ext}"
    (STATIC_IMG / fname).write_bytes(content)
    ss.save({"ui_logo_custom": fname, "ui_logo_show": "1"})
    return RedirectResponse("/settings/general?saved=1", status_code=303)


@router.post("/logo/reset")
def settings_logo_reset():
    ss.save({"ui_logo_custom": ""})
    return RedirectResponse("/settings/general?saved=1", status_code=303)


@router.post("/schedule")
async def settings_schedule(request: Request):
    """자동 실행 스케줄 저장 (동기화 / AI 학습 / 자가발전)"""
    form = await request.form()
    updates = {}
    for t in ("sync", "learning", "selfdev"):
        updates[f"sched_{t}_enabled"] = "1" if form.get(f"sched_{t}_enabled") else "0"
        for fld in ("freq", "time", "dow", "dom"):
            v = form.get(f"sched_{t}_{fld}")
            if v is not None:
                updates[f"sched_{t}_{fld}"] = v.strip() if isinstance(v, str) else v
    ss.save(updates)
    return RedirectResponse("/settings/schedule?saved=1", status_code=303)


@router.post("/schedule/run/{task}")
def settings_schedule_run(task: str):
    """지금 즉시 실행 (수동 트리거)"""
    import scheduler
    if task in scheduler.TASKS:
        scheduler.run_task_async(task, manual=True)
        return RedirectResponse(f"/settings/schedule?ran={task}", status_code=303)
    return RedirectResponse("/settings/schedule", status_code=303)


@router.post("/product/add")
def settings_product_add(db: Session = Depends(get_db),
                         code: str = Form(...), name: str = Form(...),
                         category: str = Form(""), group: str = Form("")):
    code = code.strip(); name = name.strip()
    if not code or not name:
        return RedirectResponse("/settings/products?prod_err=코드·제품명+필수", status_code=303)
    if db.get(Product, code):
        return RedirectResponse("/settings/products?prod_err=코드+중복", status_code=303)
    db.add(Product(code=code, name=name, category=category.strip() or None, group=group.strip() or None))
    db.commit()
    return RedirectResponse("/settings/products?prod_saved=1", status_code=303)


@router.post("/product/{code}")
def settings_product_update(code: str, db: Session = Depends(get_db),
                            name: str = Form(...), category: str = Form(""), group: str = Form("")):
    row = db.get(Product, code)
    if not row:
        raise HTTPException(404)
    row.name = name.strip(); row.category = category.strip() or None; row.group = group.strip() or None
    db.commit()
    return RedirectResponse("/settings/products?prod_saved=1", status_code=303)


@router.post("/product/{code}/delete")
def settings_product_delete(code: str, db: Session = Depends(get_db)):
    row = db.get(Product, code)
    if row:
        db.delete(row); db.commit()
    return RedirectResponse("/settings/products?prod_saved=1", status_code=303)


@router.post("/payroll/reset")
def settings_payroll_reset():
    """급여 요율을 기본값으로 초기화 (오버라이드 제거 → 코드 기본값 사용)."""
    ss.save({k: "" for k in (
        "pay_rate_pension", "pay_rate_health", "pay_rate_longterm", "pay_rate_employment",
        "pay_rate_local_tax", "pay_rate_emp_employment", "pay_rate_accident",
        "pay_rate_pension_cap", "pay_rate_pension_floor", "pay_rate_meal_taxfree")})
    return RedirectResponse("/settings/payroll?saved=1", status_code=303)


@router.post("/nav-order")
def settings_nav_order(order: str = Form(...)):
    """메뉴 순서 저장 — order는 키 배열 JSON"""
    import json
    from helpers import NAV_DEFAULT_ORDER
    try:
        keys = json.loads(order)
        if not isinstance(keys, list):
            keys = []
    except Exception:
        keys = []
    # 유효 키만 + 빈 입력이면 기본 순서
    valid = set(NAV_DEFAULT_ORDER)
    keys = [k for k in keys if k in valid]
    if not keys:
        keys = list(NAV_DEFAULT_ORDER)
    ss.save({"nav_order": json.dumps(keys, ensure_ascii=False)})
    return RedirectResponse("/settings/menu?saved=1", status_code=303)


@router.post("/nav-order/reset")
def settings_nav_order_reset():
    from helpers import NAV_DEFAULT_ORDER
    import json
    ss.save({"nav_order": json.dumps(list(NAV_DEFAULT_ORDER), ensure_ascii=False)})
    return RedirectResponse("/settings/menu?saved=1", status_code=303)


@router.post("/password")
def settings_password(current: str = Form(""), new1: str = Form(...), new2: str = Form(...)):
    """공통 비밀번호 변경"""
    if new1 != new2:
        return RedirectResponse("/settings/account?pw_error=비밀번호가+일치하지+않습니다", status_code=303)
    if len(new1) < 4:
        return RedirectResponse("/settings/account?pw_error=4자+이상+입력하세요", status_code=303)
    # 현재 비밀번호 확인 (설정된 경우)
    import auth
    res = ss.check_password(current)
    if res is None:
        # 환경변수 비번과 비교
        if current and current != auth.SHARED_PASSWORD:
            return RedirectResponse("/settings/account?pw_error=현재+비밀번호가+틀립니다", status_code=303)
    elif res is False:
        return RedirectResponse("/settings/account?pw_error=현재+비밀번호가+틀립니다", status_code=303)
    ss.set_password(new1)
    return RedirectResponse("/settings/account?pw_saved=1", status_code=303)


@router.post("/user/add")
def settings_user_add(username: str = Form(...), password: str = Form(...)):
    ok = ss.add_user(username, password)
    if not ok:
        return RedirectResponse("/settings/account?user_error=입력값+확인", status_code=303)
    return RedirectResponse("/settings/account?user_saved=1", status_code=303)


@router.post("/user/remove")
def settings_user_remove(username: str = Form(...)):
    ss.remove_user(username)
    return RedirectResponse("/settings/account?user_saved=1", status_code=303)


# ===== 설정 진입/페이지 라우트 (반드시 파일 끝 — /browse·/logs/download 보다 늦게 등록) =====
@router.get("", response_class=HTMLResponse)
@router.get("/", response_class=HTMLResponse)
def settings_root():
    return RedirectResponse("/settings/general", status_code=307)


# ====== DB 탐색기 (읽기 전용) ======
_DB_SECRET_PATTERNS = ("_key", "_pass", "_password", "_token", "_secret",
                       "_hash", "_certkey", "linkid")


def _is_secret_key(key: str) -> bool:
    k = (key or "").lower()
    return any(p in k for p in _DB_SECRET_PATTERNS)


def _mask_secret(value):
    if value is None:
        return ""
    s = str(value)
    if not s:
        return ""
    if len(s) <= 4:
        return "*" * len(s)
    return s[:2] + "*" * (len(s) - 4) + s[-2:]


def _db_inventory():
    """app.db의 모든 테이블 목록 + 행수 + 카테고리."""
    from database import DB_PATH
    import sqlite3 as _sqlite
    conn = _sqlite.connect(DB_PATH)
    try:
        rows = conn.execute(
            "SELECT name, type, sql FROM sqlite_master "
            "WHERE type IN ('table','view') AND name NOT LIKE 'sqlite_%' "
            "ORDER BY name"
        ).fetchall()
        out = []
        for name, ttype, sql in rows:
            try:
                n = conn.execute(f'SELECT COUNT(*) FROM "{name}"').fetchone()[0]
            except Exception:
                n = -1
            # 카테고리 분류
            if name.startswith("dim_"):       cat, icon, color = "DIM", "🧩", "#7C3AED"
            elif name.startswith("fact_"):    cat, icon, color = "FACT", "📊", "#10B981"
            elif name.startswith("master_"):  cat, icon, color = "마스터", "📋", "#F59E0B"
            elif name in ("file_registry", "sync_run", "sync_run_detail",
                          "integrity_check", "unmapped_file_review", "import_batch"):
                cat, icon, color = "동기화", "🔄", "#3B82F6"
            elif name in ("knowledge_chunk", "chat_history", "ai_classify_run"):
                cat, icon, color = "AI", "🤖", "#EF4444"
            elif name in ("app_setting", "user_account", "activity_log",
                          "bank_account", "card", "document", "contract"):
                cat, icon, color = "운영", "⚙", "#6B7280"
            else:
                cat, icon, color = "기타", "📦", "#94A3B8"
            out.append({"name": name, "type": ttype, "count": n,
                        "category": cat, "icon": icon, "color": color})
        return sorted(out, key=lambda x: (x["category"], x["name"]))
    finally:
        conn.close()


def _table_schema(name: str):
    """PRAGMA로 컬럼·인덱스·외래키 조회."""
    from database import DB_PATH
    import sqlite3 as _sqlite
    conn = _sqlite.connect(DB_PATH)
    try:
        cols = [{"cid": r[0], "name": r[1], "type": r[2], "notnull": bool(r[3]),
                 "default": r[4], "pk": r[5]}
                for r in conn.execute(f'PRAGMA table_info("{name}")').fetchall()]
        idx_rows = conn.execute(f'PRAGMA index_list("{name}")').fetchall()
        indexes = []
        for ir in idx_rows:
            iname = ir[1]
            icols = [r[2] for r in conn.execute(f'PRAGMA index_info("{iname}")').fetchall()]
            indexes.append({"name": iname, "unique": bool(ir[2]), "columns": icols})
        fks = [{"id": r[0], "from": r[3], "to_table": r[2], "to_col": r[4]}
               for r in conn.execute(f'PRAGMA foreign_key_list("{name}")').fetchall()]
        return {"columns": cols, "indexes": indexes, "foreign_keys": fks}
    finally:
        conn.close()


def _table_rows(name: str, limit: int = 50, offset: int = 0,
                sort_col: str = "", sort_dir: str = "desc", q: str = ""):
    """테이블 데이터 조회 (LIMIT + 옵션 필터). 비밀값 마스킹."""
    from database import DB_PATH
    import sqlite3 as _sqlite
    conn = _sqlite.connect(DB_PATH)
    try:
        cols = [r[1] for r in conn.execute(f'PRAGMA table_info("{name}")').fetchall()]
        if not cols:
            return {"columns": [], "rows": [], "total": 0}

        where, params = "", []
        if q:
            terms = []
            for c in cols:
                terms.append(f'CAST("{c}" AS TEXT) LIKE ?')
                params.append(f"%{q}%")
            where = " WHERE " + " OR ".join(terms)

        order = ""
        if sort_col and sort_col in cols:
            d = "ASC" if (sort_dir or "").lower() == "asc" else "DESC"
            order = f' ORDER BY "{sort_col}" {d}'
        elif "id" in cols:
            order = ' ORDER BY "id" DESC'

        total = conn.execute(f'SELECT COUNT(*) FROM "{name}"' + where, params).fetchone()[0]
        sql = (f'SELECT * FROM "{name}"' + where + order +
               f' LIMIT {int(limit)} OFFSET {int(offset)}')
        raw = conn.execute(sql, params).fetchall()

        # app_setting 테이블은 key 컬럼 기준으로 마스킹
        is_setting = (name == "app_setting")
        rows = []
        for r in raw:
            d = {}
            for i, c in enumerate(cols):
                v = r[i]
                if is_setting and c == "value":
                    key_val = r[cols.index("key")] if "key" in cols else ""
                    if _is_secret_key(str(key_val)):
                        v = _mask_secret(v)
                elif _is_secret_key(c):
                    v = _mask_secret(v)
                d[c] = v
            rows.append(d)
        return {"columns": cols, "rows": rows, "total": total}
    finally:
        conn.close()


# sqlite-web 외부 GUI 프로세스 추적 (단일 인스턴스)
_SQLITEWEB_PROC = None
_SQLITEWEB_PORT = 18001


@router.post("/db/launch-external")
def db_launch_external():
    """sqlite-web 외부 DB GUI를 백그라운드로 띄우고 접속 URL 반환."""
    global _SQLITEWEB_PROC
    from fastapi.responses import JSONResponse
    import subprocess as _sp
    import socket as _sock
    from database import DB_PATH

    # 이미 가동 중이면 그대로 사용
    if _SQLITEWEB_PROC and _SQLITEWEB_PROC.poll() is None:
        return JSONResponse({
            "ok": True,
            "url": f"http://127.0.0.1:{_SQLITEWEB_PORT}/",
            "msg": "이미 가동 중인 sqlite-web 인스턴스에 접속합니다.",
        })

    # 포트 가용성 체크
    try:
        s = _sock.socket(_sock.AF_INET, _sock.SOCK_STREAM)
        s.bind(("127.0.0.1", _SQLITEWEB_PORT))
        s.close()
    except OSError:
        return JSONResponse({
            "ok": False,
            "error": f"포트 {_SQLITEWEB_PORT}가 이미 사용 중입니다. "
                     f"기존 sqlite-web 또는 다른 프로세스가 점유 중일 수 있습니다.",
        }, status_code=409)

    try:
        # readonly 모드로 띄움 — 데이터 수정 불가, 조회만
        _SQLITEWEB_PROC = _sp.Popen(
            [
                sys.executable if (sys := __import__("sys")) else "python",
                "-m", "sqlite_web",
                "--host", "127.0.0.1",
                "--port", str(_SQLITEWEB_PORT),
                "--no-browser",
                "--read-only",
                str(DB_PATH),
            ],
            stdout=_sp.DEVNULL, stderr=_sp.DEVNULL,
            creationflags=getattr(_sp, "CREATE_NEW_PROCESS_GROUP", 0),
        )
        # 부팅 대기 (최대 5초)
        import time as _t
        for _ in range(50):
            _t.sleep(0.1)
            try:
                s = _sock.socket(_sock.AF_INET, _sock.SOCK_STREAM)
                s.settimeout(0.2)
                r = s.connect_ex(("127.0.0.1", _SQLITEWEB_PORT))
                s.close()
                if r == 0:
                    break
            except Exception:
                pass
        return JSONResponse({
            "ok": True,
            "url": f"http://127.0.0.1:{_SQLITEWEB_PORT}/",
            "msg": "sqlite-web GUI가 새 포트에 띄워졌습니다.",
            "readonly": True,
        })
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


@router.post("/db/stop-external")
def db_stop_external():
    global _SQLITEWEB_PROC
    from fastapi.responses import JSONResponse
    if _SQLITEWEB_PROC and _SQLITEWEB_PROC.poll() is None:
        try:
            _SQLITEWEB_PROC.terminate()
            _SQLITEWEB_PROC.wait(timeout=5)
        except Exception:
            try: _SQLITEWEB_PROC.kill()
            except Exception: pass
        _SQLITEWEB_PROC = None
        return JSONResponse({"ok": True})
    return JSONResponse({"ok": True, "msg": "이미 중지됨"})


@router.get("/db/api/inventory")
def db_api_inventory():
    from fastapi.responses import JSONResponse
    return JSONResponse({"tables": _db_inventory()})


@router.get("/db/api/schema")
def db_api_schema(name: str):
    from fastapi.responses import JSONResponse
    return JSONResponse(_table_schema(name))


@router.get("/db/api/rows")
def db_api_rows(name: str, limit: int = 50, offset: int = 0,
                sort_col: str = "", sort_dir: str = "desc", q: str = ""):
    from fastapi.responses import JSONResponse
    if limit > 500:
        limit = 500
    return JSONResponse(_table_rows(name, limit, offset, sort_col, sort_dir, q))


# ====== 매출·매입 삭제 관리 ======
def _txn_model(kind: str):
    from models import Sale, Purchase
    return Sale if kind == "sale" else Purchase


def _txn_filter_conds(model, year=None, quarter=None, month=None, day=None):
    """범위 조건 빌더 — 연/분기/월/일 조합으로 WHERE 절."""
    conds = []
    if year:
        conds.append(model.year == int(year))
    if quarter:
        q_to_months = {"Q1": (1, 3), "Q2": (4, 6), "Q3": (7, 9), "Q4": (10, 12)}
        if quarter in q_to_months:
            mfrom, mto = q_to_months[quarter]
            conds.append(model.month >= mfrom)
            conds.append(model.month <= mto)
    if month:
        conds.append(model.month == int(month))
    if day:
        from datetime import date as _date
        if year and month:
            try:
                d = _date(int(year), int(month), int(day))
                conds.append(model.txn_date == d)
            except (ValueError, TypeError):
                pass
    return conds


@router.get("/txn-mgmt", response_class=HTMLResponse)
def settings_txn_mgmt(request: Request, log_category: str = "", log_limit: int = 100,
                     log_from: str = "", log_to: str = ""):
    """매출·매입 관리 페이지 (POST 삭제 후 redirect 대상)."""
    return _render_settings(request, "txn-mgmt", log_category, log_limit, log_from, log_to)


@router.get("/txn-mgmt/preview")
def txn_mgmt_preview(kind: str, year: str = "", quarter: str = "",
                     month: str = "", day: str = "",
                     db: Session = Depends(get_db)):
    """삭제 대상 건수·합계 미리보기 (JSON)."""
    from fastapi.responses import JSONResponse
    if kind not in ("sale", "purchase"):
        return JSONResponse({"error": "kind must be sale|purchase"}, status_code=400)
    model = _txn_model(kind)
    conds = _txn_filter_conds(model, year or None, quarter or None,
                              month or None, day or None)
    stmt = select(func.count(), func.coalesce(func.sum(model.supply), 0),
                  func.coalesce(func.sum(model.total), 0))
    if conds:
        stmt = stmt.where(and_(*conds) if len(conds) > 1 else conds[0])
    row = db.execute(stmt).one()
    return JSONResponse({
        "kind": kind, "year": year, "quarter": quarter, "month": month, "day": day,
        "count": int(row[0] or 0),
        "sum_supply": float(row[1] or 0),
        "sum_total": float(row[2] or 0),
    })


@router.post("/txn-mgmt/delete")
async def txn_mgmt_delete(request: Request, db: Session = Depends(get_db)):
    """삭제 실행 — DB 자동 백업 후 진행."""
    from fastapi.responses import JSONResponse
    form = dict(await request.form())
    kind = form.get("kind", "")
    if kind not in ("sale", "purchase"):
        return JSONResponse({"error": "kind must be sale|purchase"}, status_code=400)
    confirm = (form.get("confirm") or "").strip().upper()
    if confirm != "DELETE":
        return RedirectResponse(
            "/settings/txn-mgmt?_msg=" + f"❌ 확인 입력이 'DELETE'와 일치하지 않습니다",
            status_code=303)

    model = _txn_model(kind)
    year = form.get("year") or None
    quarter = form.get("quarter") or None
    month = form.get("month") or None
    day = form.get("day") or None
    conds = _txn_filter_conds(model, year, quarter, month, day)

    # 1) DB 자동 백업
    import shutil
    from database import DB_PATH
    backup_dir = DB_PATH.parent / "db_backup"
    backup_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"pre_txn_delete_{kind}_{ts}.db"
    try:
        shutil.copy2(DB_PATH, backup_path)
    except Exception as e:
        return JSONResponse({"error": f"백업 실패 — 삭제 중단: {e}"}, status_code=500)

    # 2) 대상 건수 측정
    stmt = select(func.count()).select_from(model)
    if conds:
        stmt = stmt.where(and_(*conds) if len(conds) > 1 else conds[0])
    n_before = db.execute(stmt).scalar() or 0

    if n_before == 0:
        return RedirectResponse(
            f"/settings/txn-mgmt?_msg=" + f"{kind} 삭제 대상이 없습니다 (조건과 일치하는 행 0건)",
            status_code=303)

    # 3) 삭제 실행
    from sqlalchemy import delete as _delete
    del_stmt = _delete(model)
    if conds:
        del_stmt = del_stmt.where(and_(*conds) if len(conds) > 1 else conds[0])
    db.execute(del_stmt)
    db.commit()

    label = "매출" if kind == "sale" else "매입"
    scope = []
    if year: scope.append(f"{year}년")
    if quarter: scope.append(quarter)
    if month: scope.append(f"{month}월")
    if day: scope.append(f"{day}일")
    if not scope: scope.append("전체")
    msg = f"✅ {label} {n_before:,}건 삭제 완료 ({' '.join(scope)}). 백업: {backup_path.name}"
    return RedirectResponse(f"/settings/txn-mgmt?_msg=" + msg, status_code=303)


@router.get("/ai/ping")
def ai_ping():
    """4종 AI 공급자(Ollama·OpenAI·Anthropic·Gemini) 실제 응답 헬스체크."""
    from fastapi.responses import JSONResponse
    try:
        import llm_provider
        return JSONResponse(llm_provider.ping_all())
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@router.get("/{page}", response_class=HTMLResponse)
def settings_page(page: str, request: Request, log_category: str = "", log_limit: int = 100,
                  log_from: str = "", log_to: str = ""):
    """설정 — 좌측 메뉴 항목별 페이지"""
    return _render_settings(request, page, log_category, log_limit, log_from, log_to)
