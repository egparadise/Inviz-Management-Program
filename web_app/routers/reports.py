# -*- coding: utf-8 -*-
"""보고서 라우터 — 양식(.xlsx) 업로드 → AI 분석 (placeholder 치환) → 화면 미리보기 → Excel/PDF 다운로드

플로우:
  1) GET /reports — 저장된 양식 목록 + 신규 업로드 폼
  2) POST /reports/upload — 양식 업로드, placeholder 자동 스캔
  3) GET /reports/{id} — 양식 상세 + AI 분석 버튼
  4) POST /reports/{id}/analyze — placeholder를 DB 값으로 치환 → 미리보기 HTML
  5) GET /reports/{id}/export.xlsx — 치환된 Excel 다운로드
  6) GET /reports/{id}/export.pdf — 치환된 PDF 다운로드
  7) GET /reports/template — 표준 양식(가이드 포함) 다운로드

Placeholder 문법:
  {{매출.올해}}, {{매출.지난달}}, {{매출.이번달}}, {{매출.YYYY-MM}}, {{매출.YYYY}}, {{매출.YYYY-Qn}}
  {{매입.올해}}, {{매입.이번달}}, ...
  {{매출총이익.올해}}, {{영업이익.올해}}, {{순이익.올해}}
  {{매출원가.올해}}, {{판관비.올해}}, {{급여.올해}}, {{비용.올해}}, {{임차료.올해}}
  {{거래처.TOP3}}, {{거래처.TOP5}}, {{거래처.TOP10}}
  {{회사명}}, {{사업자번호}}, {{대표자}}
  {{현재일자}}, {{현재월}}, {{현재연도}}, {{현재분기}}, {{기준일}}
"""
import io
import json
import re
import shutil
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional
from fastapi import APIRouter, Request, Form, Depends, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, Response, JSONResponse
from sqlalchemy import select, func
from sqlalchemy.orm import Session

from database import get_db
from helpers import templates
from models import (Sale, Purchase, Payroll, Expense, Rental, Party,
                    ReportTemplate, ReportSnapshot)

router = APIRouter()

# 양식 저장 디렉토리
TEMPLATE_DIR = Path(__file__).parent.parent / "report_templates"
TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
# 저장본(스냅샷) 디렉토리 — DB 폴더(db_backup 계열)와 함께 web_app 하위에 보관
SNAPSHOT_DIR = Path(__file__).parent.parent / "report_snapshots"
SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)

# 회사 정보 기본값 (DB 미설정 시 폴백)
COMPANY_INFO = {
    "회사명": "(주)인비즈",
    "사업자번호": "000-00-00000",
    "대표자": "박성철",
}


def _company_vars(db: Session) -> dict:
    """회사 기본정보(CompanyInfo)를 placeholder 변수 dict로 변환. DB 우선, 없으면 폴백."""
    from models import CompanyInfo
    vals = dict(COMPANY_INFO)
    try:
        ci = db.get(CompanyInfo, 1)
        if ci:
            if ci.name: vals["회사명"] = ci.name
            if ci.name_en: vals["영문회사명"] = ci.name_en
            if ci.biz_no: vals["사업자번호"] = ci.biz_no
            if ci.corp_no: vals["법인등록번호"] = ci.corp_no
            if ci.ceo: vals["대표자"] = ci.ceo
            if ci.established: vals["설립일"] = ci.established
            if ci.address: vals["주소"] = ci.address
            if ci.phone: vals["전화"] = ci.phone
            if ci.fax: vals["팩스"] = ci.fax
            if ci.email: vals["이메일"] = ci.email
            if ci.website: vals["홈페이지"] = ci.website
            if ci.industry: vals["업종"] = ci.industry
            if ci.capital is not None: vals["자본금"] = f"{int(round(float(ci.capital))):,}"
            if ci.employee_count: vals["임직원수"] = f"{ci.employee_count:,}"
        # 임직원수가 미입력(0/None)이면 직원 마스터 재직 인원으로 자동 집계
        if "임직원수" not in vals:
            from models import Employee
            n = db.scalar(select(func.count()).select_from(Employee)
                          .where(Employee.active.in_(["재직", "Y"]))) or 0
            if n:
                vals["임직원수"] = f"{n:,}"
    except Exception as e:
        print(f"[reports] 회사정보 로드 실패: {e}")
    return vals


# ====== Placeholder 파싱·치환 ======
PLACEHOLDER_RE = re.compile(r"\{\{\s*([^{}]+?)\s*\}\}")


def scan_placeholders_xlsx(file_path: str):
    """양식 xlsx에서 모든 {{...}} placeholder 추출 → 중복 제거된 리스트"""
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=False)
    found = set()
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                s = str(cell)
                for m in PLACEHOLDER_RE.finditer(s):
                    found.add(m.group(1).strip())
    return sorted(found)


def resolve_placeholder(var: str, db: Session) -> tuple[str, bool]:
    """단일 placeholder를 값으로 치환. 반환: (값_문자열, 매칭됨)"""
    var = var.strip()
    today = date.today()

    # 회사 기본정보 (DB 우선)
    cvars = _company_vars(db)
    if var in cvars:
        return cvars[var], True
    if var == "현재일자":
        return today.strftime("%Y-%m-%d"), True
    if var == "기준일":
        return today.strftime("%Y-%m-%d"), True
    if var == "현재월":
        return f"{today.month}월", True
    if var == "현재연도":
        return f"{today.year}년", True
    if var == "현재분기":
        return f"{(today.month - 1) // 3 + 1}분기", True

    # 거래처 TOP — {{거래처.TOPn}}
    m = re.match(r"^거래처\.TOP(\d+)$", var)
    if m:
        n = int(m.group(1))
        rows = db.execute(
            select(Sale.party_name, func.sum(Sale.supply))
            .where(Sale.year == today.year, Sale.party_name.is_not(None))
            .group_by(Sale.party_name)
            .order_by(func.sum(Sale.supply).desc()).limit(n)
        ).all()
        if not rows:
            return "(데이터 없음)", True
        return "\n".join([f"{i+1}. {r[0]} — {int(r[1] or 0):,}원"
                          for i, r in enumerate(rows)]), True

    # 도메인.기간 형식 — {{매출.올해}}, {{매출.YYYY-MM}}, {{매출.YYYY}}, {{매출.YYYY-Q1}}
    parts = var.split(".")
    if len(parts) != 2:
        return "", False
    domain, period = parts[0].strip(), parts[1].strip()

    # 기간 → (year, month, quarter, from_date, to_date)
    y, m_, q, fd, td = _parse_period_token(period, today)

    # 도메인별 집계
    val = _aggregate(db, domain, y=y, m=m_, q=q, fd=fd, td=td)
    if val is None:
        return "", False
    return f"{int(round(val)):,}", True


def _parse_period_token(period: str, today: date):
    """기간 토큰 파싱 → (year, month, quarter, from_date, to_date)"""
    y = m = q = None
    fd = td = None
    p = period.strip()
    if p in ("올해", "당해", "금년", "YTD", "ytd"):
        y = today.year
    elif p in ("작년", "전년"):
        y = today.year - 1
    elif p in ("이번달", "이달", "당월"):
        y = today.year; m = today.month
    elif p in ("지난달", "전월"):
        if today.month == 1:
            y = today.year - 1; m = 12
        else:
            y = today.year; m = today.month - 1
    elif p == "이번분기":
        y = today.year; q = (today.month - 1) // 3 + 1
    elif re.match(r"^\d{4}$", p):
        y = int(p)
    elif re.match(r"^\d{4}-\d{1,2}$", p):
        parts = p.split("-")
        y = int(parts[0]); m = int(parts[1])
    elif re.match(r"^\d{4}-Q[1-4]$", p, re.IGNORECASE):
        parts = p.split("-")
        y = int(parts[0]); q = int(parts[1][1:])
    return y, m, q, fd, td


def _aggregate(db, domain, *, y=None, m=None, q=None, fd=None, td=None):
    """도메인+기간 집계 — 매출, 매입, 매출총이익, 영업이익, 순이익, 매출원가, 판관비, 급여, 비용, 임차료"""
    sale_conds = []
    if y: sale_conds.append(Sale.year == y)
    if m: sale_conds.append(Sale.month == m)
    if q: sale_conds.append(Sale.quarter == f"Q{q}")
    if fd: sale_conds.append(Sale.txn_date >= fd)
    if td: sale_conds.append(Sale.txn_date <= td)

    pur_conds = []
    if y: pur_conds.append(Purchase.year == y)
    if m: pur_conds.append(Purchase.month == m)
    if q: pur_conds.append(Purchase.quarter == f"Q{q}")
    if fd: pur_conds.append(Purchase.txn_date >= fd)
    if td: pur_conds.append(Purchase.txn_date <= td)

    pay_conds = []
    if y: pay_conds.append(Payroll.year == y)
    if m: pay_conds.append(Payroll.month == m)

    exp_conds = []
    if y: exp_conds.append(Expense.year == y)
    if m: exp_conds.append(Expense.month == m)
    if q: exp_conds.append(Expense.quarter == f"Q{q}")
    if fd: exp_conds.append(Expense.use_date >= fd)
    if td: exp_conds.append(Expense.use_date <= td)

    ren_conds = [Rental.direction == "지출"]
    if y: ren_conds.append(Rental.year == y)
    if m: ren_conds.append(Rental.month == m)

    def s_sum():
        return float(db.scalar(select(func.coalesce(func.sum(Sale.supply), 0)).where(*sale_conds)) or 0)
    def p_sum():
        return float(db.scalar(select(func.coalesce(func.sum(Purchase.supply), 0)).where(*pur_conds)) or 0)
    def pay_sum():
        return float(db.scalar(select(func.coalesce(func.sum(Payroll.gross_pay), 0)).where(*pay_conds)) or 0)
    def emp_ins_sum():
        return float(db.scalar(select(func.coalesce(func.sum(Payroll.employer_insurance), 0)).where(*pay_conds)) or 0)
    def exp_sum():
        return float(db.scalar(select(func.coalesce(func.sum(Expense.amount), 0)).where(*exp_conds)) or 0)
    def ren_sum():
        try:
            return float(db.scalar(select(func.coalesce(func.sum(Rental.amount), 0)).where(*ren_conds)) or 0)
        except Exception:
            return 0.0

    if domain == "매출":
        return s_sum()
    if domain == "매입" or domain == "매출원가":
        return p_sum()
    if domain == "매출총이익":
        return s_sum() - p_sum()
    if domain == "급여":
        return pay_sum()
    if domain == "비용":
        return exp_sum()
    if domain == "임차료":
        return ren_sum()
    if domain == "판관비" or domain == "판매비와관리비":
        return pay_sum() + emp_ins_sum() + exp_sum() + ren_sum()
    if domain == "영업이익":
        return (s_sum() - p_sum()) - (pay_sum() + emp_ins_sum() + exp_sum() + ren_sum())
    if domain == "세전순이익":
        op = (s_sum() - p_sum()) - (pay_sum() + emp_ins_sum() + exp_sum() + ren_sum())
        return op  # 영업외 미적용
    if domain == "법인세":
        op = (s_sum() - p_sum()) - (pay_sum() + emp_ins_sum() + exp_sum() + ren_sum())
        return op * 0.10 if op > 0 else 0
    if domain == "순이익" or domain == "당기순이익":
        op = (s_sum() - p_sum()) - (pay_sum() + emp_ins_sum() + exp_sum() + ren_sum())
        tax = op * 0.10 if op > 0 else 0
        return op - tax
    return None


# ===== 보고서 카테고리(결산보고/재무제표) + PDF 양식 AI 분석 =====
REPORT_CATS = {
    "closing": {"label": "결산 보고", "icon": "🧾",
                "desc": "결산 보고 양식(Excel 또는 PDF)을 올리면 AI가 회사 결산 데이터로 채워 Excel·PDF로 출력합니다."},
    "financial": {"label": "재무제표", "icon": "📊",
                  "desc": "재무제표 양식(Excel 또는 PDF)을 올리면 AI가 회사 재무 데이터로 채워 Excel·PDF로 출력합니다."},
    "investor": {"label": "투자사 보고", "icon": "📈",
                 "desc": "투자사 보고 양식(Excel 또는 PDF)을 올리면 AI가 회사 실적·재무 데이터로 채워 Excel·PDF로 출력합니다."},
}


def _is_pdf(tpl) -> bool:
    return (getattr(tpl, "file_kind", "") or "").lower() == "pdf" or str(tpl.file_path).lower().endswith(".pdf")


def _extract_pdf_text(path: str, max_chars: int = 9000) -> str:
    try:
        import pdfplumber
        out = []
        with pdfplumber.open(path) as pdf:
            for pg in pdf.pages:
                out.append(pg.extract_text() or "")
        return "\n".join(out).strip()[:max_chars]
    except Exception as e:
        return f"(PDF 텍스트 추출 실패: {e})"


def _financial_context(db, year: int) -> str:
    def f(dom):
        v = _aggregate(db, dom, y=year)
        return f"{int(v):,}" if v is not None else "0"
    return "\n".join([
        f"[회사 재무 요약 — {year}년 (단위: 원, VAT 별도)]",
        f"매출액: {f('매출')}",
        f"매출원가(매입): {f('매입')}",
        f"매출총이익: {f('매출총이익')}",
        f"판매비와관리비: {f('판관비')}  (급여 {f('급여')} / 경비 {f('비용')} / 임차료 {f('임차료')})",
        f"영업이익: {f('영업이익')}",
        f"당기순이익(법인세 10% 가정): {f('순이익')}",
    ])


def _parse_json_rows(raw: str):
    """LLM 응답에서 JSON 배열을 견고하게 추출 → [{항목,금액,비고}]"""
    if not raw:
        return []
    m = re.search(r"\[.*\]", raw, re.DOTALL)
    if not m:
        return []
    try:
        data = json.loads(m.group(0))
    except Exception:
        return []
    out = []
    for d in (data if isinstance(data, list) else []):
        if not isinstance(d, dict):
            continue
        item = str(d.get("항목") or d.get("item") or d.get("과목") or "").strip()
        if not item:
            continue
        amt = d.get("금액", d.get("amount", d.get("금 액", 0)))
        try:
            amt = float(str(amt).replace(",", "").replace("원", "").strip() or 0)
        except Exception:
            amt = 0.0
        out.append({"항목": item, "금액": amt, "비고": str(d.get("비고") or d.get("note") or "")})
    return out


def _fallback_rows(db, year: int):
    items = [("매출액", "매출"), ("매출원가", "매입"), ("매출총이익", "매출총이익"),
             ("판매비와관리비", "판관비"), ("영업이익", "영업이익"), ("당기순이익", "순이익")]
    out = []
    for label, dom in items:
        v = _aggregate(db, dom, y=year)
        out.append({"항목": label, "금액": float(v or 0), "비고": f"{year}년 집계"})
    return out


def _latest_pdf_rows(db, tpl_id: int, snapshot: str = ""):
    snap = None
    if (snapshot or "").isdigit():
        snap = db.get(ReportSnapshot, int(snapshot))
        if snap and snap.template_id != tpl_id:
            snap = None
    if not snap:
        snap = db.execute(select(ReportSnapshot).where(ReportSnapshot.template_id == tpl_id)
                          .order_by(ReportSnapshot.id.desc())).scalars().first()
    if not snap:
        return []
    try:
        return json.loads(snap.cells_json or "[]")
    except Exception:
        return []


def _build_pdf_xlsx(tpl, rows) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "보고서"
    ws["A1"] = tpl.name
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"생성일: {date.today().isoformat()}"
    ws["A2"].font = Font(size=9, color="888888")
    hdr = ["항목", "금액(원)", "비고"]
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for j, h in enumerate(hdr, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="6B2C91")
        c.alignment = Alignment(horizontal="center")
        c.border = border
    r = 5
    for row in rows:
        ws.cell(row=r, column=1, value=row.get("항목", "")).border = border
        amt_cell = ws.cell(row=r, column=2, value=row.get("금액", 0))
        amt_cell.number_format = "#,##0"
        amt_cell.alignment = Alignment(horizontal="right")
        amt_cell.border = border
        ws.cell(row=r, column=3, value=row.get("비고", "")).border = border
        r += 1
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 24
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def replace_in_xlsx(src_path: str, db: Session) -> tuple[bytes, dict]:
    """양식 xlsx의 모든 placeholder를 DB 값으로 치환 → 새 xlsx bytes 반환
    반환: (xlsx bytes, 치환 결과 dict {placeholder: 값})
    """
    from openpyxl import load_workbook
    wb = load_workbook(src_path, data_only=False)
    resolved = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                s = str(cell.value)
                if "{{" not in s:
                    continue
                def repl(m):
                    var = m.group(1).strip()
                    if var not in resolved:
                        val, ok = resolve_placeholder(var, db)
                        resolved[var] = val if ok else f"{{{{{var}}}}}"
                    return str(resolved[var])
                new_val = PLACEHOLDER_RE.sub(repl, s)
                # 숫자만 있으면 number로 변환 (Excel 정렬·서식)
                clean = new_val.replace(",", "").strip()
                if clean and re.match(r"^-?\d+(\.\d+)?$", clean):
                    try:
                        cell.value = float(clean) if "." in clean else int(clean)
                        cell.number_format = "#,##0"
                    except Exception:
                        cell.value = new_val
                else:
                    cell.value = new_val
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), resolved


def build_preview_html(src_path: str, db: Session, override_values: dict | None = None) -> tuple[str, dict]:
    """양식 xlsx → HTML 표 미리보기 (placeholder 치환된 상태). 정적 출력(인쇄/PDF용).
    override_values(좌표→값) 있으면 해당 좌표는 저장본 값을 표시."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    wb = load_workbook(src_path, data_only=False)
    override_values = override_values or {}
    resolved = {}
    html_parts = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        html_parts.append(f'<div class="mb-6"><h3 class="font-semibold text-sm mb-2" style="color: var(--inviz-purple);">📄 시트: {sn}</h3>')
        html_parts.append('<table class="border-collapse w-full" style="font-size:13px;">')
        # merged cells map
        merged_map = {}
        for r in ws.merged_cells.ranges:
            min_r, min_c = r.min_row, r.min_col
            for rr in range(r.min_row, r.max_row + 1):
                for cc in range(r.min_col, r.max_col + 1):
                    if (rr, cc) != (min_r, min_c):
                        merged_map[(rr, cc)] = "hidden"
                    else:
                        merged_map[(rr, cc)] = (r.max_row - r.min_row + 1, r.max_col - r.min_col + 1)
        max_col = ws.max_column or 1
        max_row = ws.max_row or 1
        for ri in range(1, max_row + 1):
            html_parts.append("<tr>")
            for ci in range(1, max_col + 1):
                meta = merged_map.get((ri, ci))
                if meta == "hidden":
                    continue
                rs, cs = (1, 1) if not meta or meta == "hidden" else meta
                cell = ws.cell(row=ri, column=ci)
                coord = f"{sn}!{ri}!{ci}"
                v = cell.value
                if coord in override_values:
                    text = str(override_values[coord])
                elif v is None:
                    text = ""
                else:
                    s = str(v)
                    if "{{" in s:
                        def repl(m):
                            var = m.group(1).strip()
                            if var not in resolved:
                                val, ok = resolve_placeholder(var, db)
                                resolved[var] = val if ok else f"<span class='text-red-600'>{{{{{var}}}}}</span>"
                            return str(resolved[var])
                        s = PLACEHOLDER_RE.sub(repl, s)
                    text = s
                # 스타일
                style = "border:1px solid #CBD5E1; padding:6px 10px; vertical-align:middle;"
                try:
                    if cell.font and cell.font.bold:
                        style += "font-weight:bold;"
                except Exception:
                    pass
                # 배경색 — solid 패턴 + ARGB 문자열인 경우만 안전하게 적용
                try:
                    fill = cell.fill
                    if fill is not None and getattr(fill, "patternType", None) == "solid":
                        fg = fill.fgColor
                        rgb = getattr(fg, "rgb", None)
                        # rgb는 문자열(예: 'FF6B2C91')일 때만 사용. RGB 객체/None/테마색은 제외
                        if isinstance(rgb, str) and len(rgb) >= 6 and rgb not in ("00000000",):
                            style += f"background:#{rgb[-6:]};"
                except Exception:
                    pass
                try:
                    if cell.alignment and cell.alignment.horizontal:
                        style += f"text-align:{cell.alignment.horizontal};"
                except Exception:
                    pass
                span_attr = ""
                if rs > 1: span_attr += f' rowspan="{rs}"'
                if cs > 1: span_attr += f' colspan="{cs}"'
                # 텍스트가 숫자면 우측 정렬
                if text and re.match(r"^-?[\d,]+(\.\d+)?$", text.replace(" ", "")):
                    if "text-align" not in style:
                        style += "text-align:right; font-variant-numeric:tabular-nums;"
                html_parts.append(f'<td style="{style}"{span_attr}>{text.replace(chr(10), "<br>")}</td>')
            html_parts.append("</tr>")
        html_parts.append("</table></div>")
    return "\n".join(html_parts), resolved


def _esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;"))


def build_editable_grid(src_path: str, db: Session, override_values: dict | None = None):
    """양식 xlsx → 편집 가능한 그리드 HTML.
    - 각 셀: data-cell="시트!행!열", 값은 contenteditable
    - placeholder 셀: data-template(원본 식) + 🔄 갱신 버튼
    - override_values(저장본) 있으면 해당 좌표 값 우선 사용
    반환: (html, resolved_dict, missing_list)
    """
    from openpyxl import load_workbook
    wb = load_workbook(src_path, data_only=False)
    override_values = override_values or {}
    resolved = {}
    missing = []
    html_parts = []

    for sn in wb.sheetnames:
        ws = wb[sn]
        html_parts.append(
            f'<div class="mb-6"><h3 class="font-semibold text-sm mb-2" style="color: var(--inviz-purple);">📄 시트: {_esc(sn)}</h3>')
        html_parts.append('<table class="border-collapse" style="font-size:13px;">')
        merged_map = {}
        for r in ws.merged_cells.ranges:
            min_r, min_c = r.min_row, r.min_col
            for rr in range(r.min_row, r.max_row + 1):
                for cc in range(r.min_col, r.max_col + 1):
                    if (rr, cc) != (min_r, min_c):
                        merged_map[(rr, cc)] = "hidden"
                    else:
                        merged_map[(rr, cc)] = (r.max_row - r.min_row + 1, r.max_col - r.min_col + 1)
        max_col = ws.max_column or 1
        max_row = ws.max_row or 1
        for ri in range(1, max_row + 1):
            html_parts.append("<tr>")
            for ci in range(1, max_col + 1):
                meta = merged_map.get((ri, ci))
                if meta == "hidden":
                    continue
                rs, cs = (1, 1) if not meta or meta == "hidden" else meta
                cell = ws.cell(row=ri, column=ci)
                coord = f"{sn}!{ri}!{ci}"
                raw = "" if cell.value is None else str(cell.value)
                has_ph = "{{" in raw

                # 표시 값 결정: override(저장본) > placeholder 치환 > 원본
                if coord in override_values:
                    display = str(override_values[coord])
                elif has_ph:
                    def repl(m):
                        var = m.group(1).strip()
                        # 캐시 — 같은 placeholder 재해석 방지(속도 개선)
                        if var in resolved:
                            return str(resolved[var])
                        if var in missing:
                            return f"{{{{{var}}}}}"
                        val, ok = resolve_placeholder(var, db)
                        if ok:
                            resolved[var] = val
                            return str(val)
                        else:
                            missing.append(var)
                            return f"{{{{{var}}}}}"
                    display = PLACEHOLDER_RE.sub(repl, raw)
                else:
                    display = raw

                # 스타일
                style = "border:1px solid #CBD5E1; padding:4px 8px; vertical-align:middle; min-width:60px;"
                try:
                    if cell.font and cell.font.bold:
                        style += "font-weight:bold;"
                except Exception:
                    pass
                try:
                    fill = cell.fill
                    if fill is not None and getattr(fill, "patternType", None) == "solid":
                        rgb = getattr(fill.fgColor, "rgb", None)
                        if isinstance(rgb, str) and len(rgb) >= 6 and rgb not in ("00000000",):
                            style += f"background:#{rgb[-6:]};"
                except Exception:
                    pass
                align = ""
                try:
                    if cell.alignment and cell.alignment.horizontal:
                        align = cell.alignment.horizontal
                except Exception:
                    pass
                is_num = bool(display) and re.match(r"^-?[\d,]+(\.\d+)?$", display.replace(" ", ""))
                if not align and is_num:
                    align = "right"
                if align:
                    style += f"text-align:{align};"

                span_attr = ""
                if rs > 1: span_attr += f' rowspan="{rs}"'
                if cs > 1: span_attr += f' colspan="{cs}"'

                cls = "rpt-cell"
                ph_attr = ""
                refresh_btn = ""
                why_btn = ""
                if has_ph:
                    cls += " rpt-ph"
                    ph_attr = f' data-template="{_esc(raw)}"'
                    refresh_btn = ('<button type="button" class="rpt-refresh" title="이 항목을 최신 데이터로 갱신" '
                                   'onclick="refreshCell(this)">🔄</button>')
                    why_btn = ('<button type="button" class="rpt-why" title="이 값의 산출 근거 보기" '
                               'onclick="showBreakdown(this)">🔍</button>')

                val_html = _esc(display).replace("\n", "<br>")
                html_parts.append(
                    f'<td class="{cls}" data-cell="{coord}"{ph_attr}{span_attr} style="{style}">'
                    f'<span class="rpt-val" contenteditable="true" oninput="markEdited(this)">{val_html}</span>'
                    f'{refresh_btn}{why_btn}</td>'
                )
            html_parts.append("</tr>")
        html_parts.append("</table></div>")
    return "\n".join(html_parts), resolved, missing


def render_filled_xlsx(src_path: str, db: Session, override_values: dict | None = None) -> bytes:
    """양식 xlsx를 값으로 채워 bytes 반환.
    override_values(좌표→값) 있으면 그 값을, 없으면 placeholder를 DB로 치환.
    """
    from openpyxl import load_workbook
    wb = load_workbook(src_path, data_only=False)
    override_values = override_values or {}
    cache = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                coord = f"{sn}!{cell.row}!{cell.column}"
                if coord in override_values:
                    new_val = str(override_values[coord])
                elif cell.value is not None and "{{" in str(cell.value):
                    def repl(m):
                        var = m.group(1).strip()
                        if var not in cache:
                            val, ok = resolve_placeholder(var, db)
                            cache[var] = val if ok else f"{{{{{var}}}}}"
                        return str(cache[var])
                    new_val = PLACEHOLDER_RE.sub(repl, str(cell.value))
                else:
                    continue
                clean = new_val.replace(",", "").strip()
                if clean and re.match(r"^-?\d+(\.\d+)?$", clean):
                    try:
                        cell.value = float(clean) if "." in clean else int(clean)
                        cell.number_format = "#,##0"
                    except Exception:
                        cell.value = new_val
                else:
                    cell.value = new_val
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ====== 사업현황 입력·수정 (투자사 보고서 연동) ======
BIZ_CATEGORIES = ["영업", "연구개발", "재무 성장"]


def _load_biz_status(db, y: int, q: int) -> dict:
    """{(category, period_col): content}"""
    from models import BizStatus
    out = {}
    for r in db.execute(select(BizStatus).where(BizStatus.year == y, BizStatus.quarter == q)).scalars():
        out[(r.category, r.period_col)] = r.content or ""
    return out


@router.get("/biz-status", response_class=HTMLResponse)
def biz_status_page(request: Request, db: Session = Depends(get_db),
                    year: int = 0, quarter: int = 0):
    """사업현황 입력·수정 — 구분(영업/연구개발/재무 성장) × 당분기실적/차분기계획."""
    today = date.today()
    y = year or today.year
    q = quarter or ((today.month - 1) // 3 + 1)
    data = _load_biz_status(db, y, q)
    ny, nq = (y + 1, 1) if q == 4 else (y, q + 1)
    return templates.TemplateResponse("reports/biz_status.html", {
        "request": request, "year": y, "quarter": q,
        "next_label": f"{ny}년 {nq}분기",
        "categories": BIZ_CATEGORIES, "data": data,
    })


@router.post("/biz-status/save")
async def biz_status_save(request: Request, db: Session = Depends(get_db)):
    from models import BizStatus
    form = await request.form()
    y = int(form.get("year") or 0)
    q = int(form.get("quarter") or 0)
    if not y or not q:
        raise HTTPException(400, "year/quarter 필요")
    saved = 0
    for cat in BIZ_CATEGORIES:
        for col in ("current", "next"):
            key = f"{cat}__{col}"
            content = (form.get(key) or "").strip()
            row = db.execute(select(BizStatus).where(
                BizStatus.year == y, BizStatus.quarter == q,
                BizStatus.category == cat, BizStatus.period_col == col)).scalar_one_or_none()
            if row:
                row.content = content
            else:
                db.add(BizStatus(year=y, quarter=q, category=cat, period_col=col, content=content))
            saved += 1
    db.commit()
    from urllib.parse import quote
    return RedirectResponse(f"/reports/biz-status?year={y}&quarter={q}&_msg=" +
                            quote("✅ 사업현황이 저장되었습니다 — 투자사 보고서 생성 시 자동 반영"), status_code=303)


# ====== 투자사 표준 보고서 (5시트: 회사개요·임직원현황·매출현황·재무현황·사업현황) ======
@router.get("/investor-report.xlsx")
def investor_report(db: Session = Depends(get_db), year: int = 0, quarter: int = 0):
    """투자사 보고서 자동 생성 — 회사·주주·임직원·분기매출·차입금을 시스템 데이터로 채움."""
    import json as _json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from models import CompanyInfo, Employee, LoanMaster, BankAccount

    today = date.today()
    y = year or today.year
    q = quarter or ((today.month - 1) // 3 + 1)

    # 스타일
    NAVY = PatternFill("solid", fgColor="16365C")
    GRAY = PatternFill("solid", fgColor="D9D9D9")
    LGRAY = PatternFill("solid", fgColor="F2F2F2")
    W_BOLD = Font(color="FFFFFF", bold=True, size=10)
    BOLD = Font(bold=True, size=10)
    N10 = Font(size=10)
    CENTER = Alignment(horizontal="center", vertical="center")
    RIGHT = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="808080")
    BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

    def th(ws, cell, text, fill=GRAY, font=BOLD):
        c = ws[cell]; c.value = text; c.fill = fill; c.font = font
        c.alignment = CENTER; c.border = BOX

    def td(ws, cell, val, *, num=False, align=None, fill=None, font=N10):
        c = ws[cell]; c.value = val; c.border = BOX; c.font = font
        if num:
            c.number_format = "#,##0"; c.alignment = RIGHT
        if align: c.alignment = align
        if fill: c.fill = fill

    def title_row(ws, row, label):
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True, size=11)

    def period_head(ws):
        th(ws, "B1", "대상기간:", GRAY)
        td(ws, "C1", f"{y}년 {q}분기", align=CENTER, font=BOLD)
        ws.column_dimensions["A"].width = 3

    wb = Workbook()

    # ===== 1) 회사개요 =====
    ws = wb.active; ws.title = "회사개요"
    period_head(ws)
    ci = db.get(CompanyInfo, 1)
    title_row(ws, 3, "1. 회사 개요")
    th(ws, "B4", "구분"); th(ws, "C4", "내용")
    emp_active = db.scalar(select(func.count()).select_from(Employee)
                           .where(Employee.active.in_(["재직", "Y"]))) or 0
    rows_ci = [("회사명", ci.name if ci else "(주)인비즈"),
               ("주소", (ci.address if ci else "") or ""),
               ("사업자등록번호", (ci.biz_no if ci else "") or ""),
               ("대표이사", (ci.ceo if ci else "") or ""),
               ("자본금", int(float(ci.capital or 0)) if ci else 0),
               ("임직원수", emp_active)]
    for i, (k, v) in enumerate(rows_ci, start=5):
        th(ws, f"B{i}", k, LGRAY)
        td(ws, f"C{i}", v, num=isinstance(v, int))
    ws["B12"] = "(*) 변동 사항은 붉은색으로 표시"; ws["B12"].font = Font(size=9, color="808080")

    title_row(ws, 14, "2. 주주현황")
    for col, h in zip("BCDEF", ["주주명", "보통주", "우선주", "합계", "지분율"]):
        th(ws, f"{col}15", h)
    try:
        shs = _json.loads(ci.shareholders_json or "[]") if ci else []
    except Exception:
        shs = []
    r = 16; tot_c = tot_p = 0
    for s in shs:
        cnt = int(str(s.get("shares", "0")).replace(",", "") or 0)
        is_pref = "우선" in (s.get("note") or "")
        td(ws, f"B{r}", s.get("name", ""))
        td(ws, f"C{r}", "" if is_pref else cnt, num=not is_pref)
        td(ws, f"D{r}", cnt if is_pref else "", num=is_pref)
        td(ws, f"E{r}", cnt, num=True, font=BOLD)
        ratio = str(s.get("ratio", "")).replace("%", "")
        c = ws[f"F{r}"]; c.border = BOX; c.alignment = RIGHT
        try:
            c.value = float(ratio) / 100; c.number_format = "0.0%"
        except Exception:
            c.value = ratio
        tot_p += cnt if is_pref else 0
        tot_c += 0 if is_pref else cnt
        r += 1
    td(ws, f"B{r}", "합계", align=CENTER, font=BOLD, fill=GRAY)
    td(ws, f"C{r}", tot_c, num=True, font=BOLD, fill=GRAY)
    td(ws, f"D{r}", tot_p, num=True, font=BOLD, fill=GRAY)
    td(ws, f"E{r}", tot_c + tot_p, num=True, font=BOLD, fill=GRAY)
    c = ws[f"F{r}"]; c.value = 1; c.number_format = "0.0%"; c.border = BOX
    c.font = BOLD; c.fill = GRAY; c.alignment = RIGHT

    er = r + 3
    title_row(ws, er - 1, "3. 등기이사 현황")
    for col, h in zip("BCD", ["구분", "성명", "직급 및 역할"]):
        th(ws, f"{col}{er}", h)
    try:
        exs = _json.loads(ci.executives_json or "[]") if ci else []
    except Exception:
        exs = []
    for i, e in enumerate(exs, start=er + 1):
        td(ws, f"B{i}", e.get("title", ""), align=CENTER)
        td(ws, f"C{i}", e.get("name", ""), align=CENTER)
        td(ws, f"D{i}", e.get("note", ""), align=CENTER)
    for col, wd in zip("BCDEF", [18, 26, 16, 14, 10]):
        ws.column_dimensions[col].width = wd
    ws.column_dimensions["C"].width = 40

    # ===== 2) 임직원현황 =====
    ws = wb.create_sheet("임직원현황")
    period_head(ws)
    title_row(ws, 3, "1. 임직원 현황")
    th(ws, "B4", "구분"); th(ws, "C4", ""); th(ws, "D4", "인원수")
    n_exec = len(exs)
    emps = db.execute(select(Employee).where(Employee.active.in_(["재직", "Y"]))).scalars().all()
    exec_names = {e.get("name") for e in exs}
    staff = [e for e in emps if e.name not in exec_names]
    by_dept = {}
    for e in staff:
        by_dept[e.department or "기타"] = by_dept.get(e.department or "기타", 0) + 1
    r = 5
    td(ws, f"B{r}", "임원", align=CENTER); td(ws, f"C{r}", "등기", align=CENTER); td(ws, f"D{r}", n_exec, num=True)
    r += 1
    td(ws, f"B{r}", "", align=CENTER); td(ws, f"C{r}", "미등기", align=CENTER); td(ws, f"D{r}", "", num=False)
    r += 1
    td(ws, f"B{r}", "", align=CENTER); td(ws, f"C{r}", "임원계", align=CENTER, font=BOLD)
    td(ws, f"D{r}", n_exec, num=True, font=BOLD, fill=LGRAY)
    r += 1
    first_staff = r
    for dept, n in sorted(by_dept.items(), key=lambda x: -x[1]):
        td(ws, f"B{r}", "직원" if r == first_staff else "", align=CENTER)
        td(ws, f"C{r}", dept, align=CENTER); td(ws, f"D{r}", n, num=True)
        r += 1
    td(ws, f"B{r}", "", align=CENTER); td(ws, f"C{r}", "직원계", align=CENTER, font=BOLD)
    td(ws, f"D{r}", len(staff), num=True, font=BOLD, fill=LGRAY)
    r += 1
    td(ws, f"B{r}", "합계", align=CENTER, font=BOLD, fill=GRAY); td(ws, f"C{r}", "", fill=GRAY)
    td(ws, f"D{r}", n_exec + len(staff), num=True, font=BOLD, fill=GRAY)

    # 분기 경계
    q_start = date(y, (q - 1) * 3 + 1, 1)
    q_end = date(y + (1 if q == 4 else 0), 1 if q == 4 else q * 3 + 1, 1)

    r += 3
    title_row(ws, r - 1, "2. 신규 채용")
    for col, h in zip("BCD", ["성명", "직급", "역할"]):
        th(ws, f"{col}{r}", h)
    hires = db.execute(select(Employee).where(
        Employee.hire_date.is_not(None), Employee.hire_date >= q_start, Employee.hire_date < q_end)).scalars().all()
    rr = r + 1
    for e in hires or [None, None]:
        td(ws, f"B{rr}", e.name if e else "", align=CENTER)
        td(ws, f"C{rr}", (e.rank or "") if e else "", align=CENTER)
        td(ws, f"D{rr}", (e.department or "") if e else "", align=CENTER)
        rr += 1

    rr += 2
    title_row(ws, rr - 1, "3. 퇴사")
    for col, h in zip("BCDE", ["성명", "직급", "역할", "퇴사사유"]):
        th(ws, f"{col}{rr}", h)
    resigns = db.execute(select(Employee).where(
        Employee.resign_date.is_not(None), Employee.resign_date >= q_start, Employee.resign_date < q_end)).scalars().all()
    rr2 = rr + 1
    for e in resigns or [None, None]:
        td(ws, f"B{rr2}", e.name if e else "", align=CENTER)
        td(ws, f"C{rr2}", (e.rank or "") if e else "", align=CENTER)
        td(ws, f"D{rr2}", (e.department or "") if e else "", align=CENTER)
        td(ws, f"E{rr2}", (e.note or "") if e else "", align=CENTER)
        rr2 += 1

    rr2 += 2
    title_row(ws, rr2 - 1, "4. 차분기 신규 채용 계획")
    for i, lbl in enumerate(["영업사원", "사무직 사원", "개발", "서비스"], start=rr2):
        td(ws, f"B{i}", f"{i - rr2 + 1}"); td(ws, f"C{i}", lbl); td(ws, f"D{i}", "0명", align=CENTER)
    for col, wd in zip("BCDE", [10, 16, 16, 20]):
        ws.column_dimensions[col].width = wd

    # ===== 3) 매출현황 (최근 4개 분기, 천원) =====
    ws = wb.create_sheet("매출현황")
    period_head(ws)
    title_row(ws, 3, "1. 분기 매출 현황")
    ws["J3"] = "단위: 천원"; ws["J3"].font = Font(size=9, color="808080"); ws["J3"].alignment = RIGHT
    # 분기열: 선택 분기 포함 최근 4개 (최신 → 과거)
    quarters = []
    yy, qq = y, q
    for _ in range(4):
        quarters.append((yy, qq))
        qq -= 1
        if qq == 0: yy, qq = yy - 1, 4
    quarters = quarters[::-1]  # 과거 → 최신? 이미지의 열 순서는 최신 2개 먼저 — 여기선 시간순
    hdr = ["제품(품목)", "매출유형"] + [f"{qy % 100}Q{qn}" if False else f"{qn}Q{qy}" for qy, qn in quarters] + ["합계"]
    for i, h in enumerate(hdr):
        c = ws.cell(row=4, column=2 + i)
        c.value = h; c.fill = NAVY; c.font = W_BOLD; c.alignment = CENTER; c.border = BOX
    # 제품별 분기 합계
    data = {}
    for (qy, qn) in quarters:
        rows_q = db.execute(
            select(Sale.product_name, func.coalesce(func.sum(Sale.supply), 0))
            .where(Sale.year == qy, Sale.quarter == f"Q{qn}")
            .group_by(Sale.product_name)).all()
        for pn, s in rows_q:
            data.setdefault(pn or "기타", {})[(qy, qn)] = float(s or 0)
    r = 5
    col_tot = {qk: 0 for qk in quarters}
    for pn in sorted(data.keys(), key=lambda k: -sum(data[k].values())):
        td(ws, f"B{r}", pn)
        td(ws, f"C{r}", "제조/서비스", align=CENTER)
        row_sum = 0
        for i, qk in enumerate(quarters):
            v = round(data[pn].get(qk, 0) / 1000)
            td(ws, ws.cell(row=r, column=4 + i).coordinate, v or "", num=bool(v))
            row_sum += v; col_tot[qk] += v
        td(ws, ws.cell(row=r, column=8).coordinate, row_sum, num=True, font=BOLD)
        r += 1
    # 매출액계
    td(ws, f"B{r}", "매출액계", align=CENTER, font=W_BOLD, fill=NAVY)
    td(ws, f"C{r}", "", fill=NAVY)
    gtot = 0
    for i, qk in enumerate(quarters):
        c = ws.cell(row=r, column=4 + i)
        c.value = col_tot[qk]; c.number_format = "#,##0"; c.alignment = RIGHT
        c.fill = NAVY; c.font = W_BOLD; c.border = BOX
        gtot += col_tot[qk]
    c = ws.cell(row=r, column=8); c.value = gtot; c.number_format = "#,##0"
    c.fill = NAVY; c.font = W_BOLD; c.border = BOX; c.alignment = RIGHT

    r += 3
    title_row(ws, r - 1, "2. R&D 과제수주")
    for i, h in enumerate(["과제명", "기관", "기간", "정부출연금", "비고"]):
        c = ws.cell(row=r, column=2 + i)
        c.value = h; c.fill = NAVY; c.font = W_BOLD; c.alignment = CENTER; c.border = BOX
    for blank in range(r + 1, r + 4):
        for i in range(5):
            ws.cell(row=blank, column=2 + i).border = BOX
    ws.column_dimensions["B"].width = 24
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 13

    # ===== 4) 재무현황 =====
    ws = wb.create_sheet("재무현황")
    period_head(ws)
    title_row(ws, 3, "1. 가용 현금 및 예금")
    th(ws, "B4", "구분"); th(ws, "C4", "금액")
    bal = float(db.scalar(select(func.coalesce(func.sum(BankAccount.balance), 0))
                          .where(BankAccount.active == "Y")) or 0)
    rows_c = [("현금및현금성자산", round(bal / 1000)), ("단기금융상품", ""), ("장기금융상품", "")]
    r = 5
    for k, v in rows_c:
        td(ws, f"B{r}", k); td(ws, f"C{r}", v, num=isinstance(v, (int, float)))
        r += 1
    td(ws, f"B{r}", "합계", align=CENTER, font=BOLD, fill=GRAY)
    td(ws, f"C{r}", round(bal / 1000), num=True, font=BOLD, fill=GRAY)
    ws[f"D{r-3}"] = "(단위: 천원 · 은행계좌 잔액 합계)"; ws[f"D{r-3}"].font = Font(size=9, color="808080")

    r += 3
    title_row(ws, r - 1, "2. 차입금 현황")
    for i, h in enumerate(["은행명", "대출금액", "대출잔액", "차입일", "만기", "상환방법"]):
        c = ws.cell(row=r, column=2 + i)
        c.value = h; c.fill = NAVY; c.font = W_BOLD; c.alignment = CENTER; c.border = BOX
    loans = db.execute(select(LoanMaster).where(
        (LoanMaster.status.is_(None)) | (~LoanMaster.status.in_(["상환완료", "종료"])))
        .order_by(LoanMaster.start_date)).scalars().all()
    rr = r + 1
    t_init = t_bal = 0
    for l in loans:
        td(ws, f"B{rr}", l.institution or l.kind or "")
        td(ws, f"C{rr}", int(float(l.initial_amount or 0)), num=True)
        td(ws, f"D{rr}", int(float(l.current_balance or 0)), num=True)
        td(ws, f"E{rr}", str(l.start_date or ""), align=CENTER)
        td(ws, f"F{rr}", str(l.end_date or ""), align=CENTER)
        td(ws, f"G{rr}", l.repayment_method or "", align=CENTER)
        t_init += float(l.initial_amount or 0); t_bal += float(l.current_balance or 0)
        rr += 1
    td(ws, f"B{rr}", "합계", align=CENTER, font=W_BOLD, fill=NAVY)
    td(ws, f"C{rr}", int(t_init), num=True, font=W_BOLD, fill=NAVY)
    td(ws, f"D{rr}", int(t_bal), num=True, font=W_BOLD, fill=NAVY)
    for col in "EFG":
        c = ws[f"{col}{rr}"]; c.fill = NAVY; c.border = BOX

    rr += 3
    title_row(ws, rr - 1, "3. 투자금 사용 내역")
    for i, h in enumerate(["사용내역", "출금", "비용인정", "입금"]):
        c = ws.cell(row=rr, column=2 + i)
        c.value = h; c.fill = GRAY; c.font = BOLD; c.alignment = CENTER; c.border = BOX
    for blank in range(rr + 1, rr + 6):
        for i in range(4):
            ws.cell(row=blank, column=2 + i).border = BOX
    for col, wd in zip("BCDEFG", [22, 15, 15, 12, 12, 12]):
        ws.column_dimensions[col].width = wd

    # ===== 5) 사업현황 (구분 × 당분기실적/차분기계획 — /reports/biz-status에서 입력) =====
    ws = wb.create_sheet("사업현황")
    ny, nq = (y + 1, 1) if q == 4 else (y, q + 1)
    # 대상기간 헤더 (당분기 | 차분기)
    th(ws, "A1", "대상기간:", GRAY)
    c = ws["B1"]; c.value = f"{y}년 {q}분기"; c.fill = NAVY; c.font = W_BOLD; c.alignment = CENTER; c.border = BOX
    c = ws["C1"]; c.value = f"{ny}년 {nq}분기"; c.font = BOLD; c.alignment = CENTER; c.border = BOX
    # 컬럼 헤더
    th(ws, "A2", "구분", NAVY, W_BOLD)
    th(ws, "B2", "당분기실적", NAVY, W_BOLD)
    th(ws, "C2", "차분기계획", NAVY, W_BOLD)
    biz = _load_biz_status(db, y, q)
    r = 3
    for cat in BIZ_CATEGORIES:
        cur_lines = [l for l in (biz.get((cat, "current"), "") or "").split("\n") if l.strip()] or [""]
        nxt_lines = [l for l in (biz.get((cat, "next"), "") or "").split("\n") if l.strip()] or [""]
        n_rows = max(len(cur_lines), len(nxt_lines))
        start = r
        for i in range(n_rows):
            td(ws, f"A{r}", cat if i == 0 else "", align=CENTER, font=BOLD)
            cc = ws[f"B{r}"]; cc.value = cur_lines[i] if i < len(cur_lines) else ""
            cc.border = BOX; cc.alignment = Alignment(vertical="top", wrap_text=True); cc.font = N10
            cn = ws[f"C{r}"]; cn.value = nxt_lines[i] if i < len(nxt_lines) else ""
            cn.border = BOX; cn.alignment = Alignment(vertical="top", wrap_text=True); cn.font = N10
            ws.row_dimensions[r].height = 34
            r += 1
        if n_rows > 1:
            ws.merge_cells(f"A{start}:A{r-1}")
        ws[f"A{start}"].alignment = CENTER
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 58

    buf = io.BytesIO()
    wb.save(buf)
    safe = f"인비즈_투자사보고서_{y}년{q}분기"
    from urllib.parse import quote
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(safe)}.xlsx"})


# ====== 표준 양식 다운로드 ======
@router.get("/template")
def standard_template():
    """B5 표준 양식 — placeholder 사용 가이드 포함 xlsx"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "월간 보고서"

    purple = PatternFill("solid", fgColor="6B2C91")
    orange = PatternFill("solid", fgColor="F47521")
    gray = PatternFill("solid", fgColor="F1F5F9")
    white = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    # 제목
    ws.merge_cells("A1:D1")
    ws["A1"] = "{{회사명}} 월간 경영보고서"
    ws["A1"].font = Font(size=18, bold=True, color="6B2C91")
    ws["A1"].alignment = center

    ws.merge_cells("A2:D2")
    ws["A2"] = "기준일: {{현재일자}} (작성)"
    ws["A2"].alignment = center
    ws["A2"].font = Font(color="64748B", italic=True)

    # 본문 헤더
    ws["A4"] = "구분"; ws["B4"] = "이번달"; ws["C4"] = "지난달"; ws["D4"] = "올해 누계"
    for c in ["A4", "B4", "C4", "D4"]:
        ws[c].fill = purple; ws[c].font = white; ws[c].alignment = center; ws[c].border = thin

    rows_data = [
        ("Ⅰ. 매출액",          "{{매출.이번달}}",      "{{매출.지난달}}",      "{{매출.올해}}"),
        ("Ⅱ. 매출원가",        "{{매입.이번달}}",      "{{매입.지난달}}",      "{{매입.올해}}"),
        ("Ⅲ. 매출총이익",      "{{매출총이익.이번달}}", "{{매출총이익.지난달}}", "{{매출총이익.올해}}"),
        ("Ⅳ. 판매비와관리비",  "{{판관비.이번달}}",    "{{판관비.지난달}}",    "{{판관비.올해}}"),
        ("  급여",             "{{급여.이번달}}",      "{{급여.지난달}}",      "{{급여.올해}}"),
        ("  비용",             "{{비용.이번달}}",      "{{비용.지난달}}",      "{{비용.올해}}"),
        ("  임차료",           "{{임차료.이번달}}",    "{{임차료.지난달}}",    "{{임차료.올해}}"),
        ("Ⅴ. 영업이익",        "{{영업이익.이번달}}",  "{{영업이익.지난달}}",  "{{영업이익.올해}}"),
        ("Ⅵ. 당기순이익",      "{{순이익.이번달}}",    "{{순이익.지난달}}",    "{{순이익.올해}}"),
    ]
    for i, (a, b, c_, d) in enumerate(rows_data, start=5):
        ws.cell(row=i, column=1, value=a).border = thin
        ws.cell(row=i, column=2, value=b).border = thin
        ws.cell(row=i, column=3, value=c_).border = thin
        ws.cell(row=i, column=4, value=d).border = thin
        if a.startswith("Ⅲ") or a.startswith("Ⅴ") or a.startswith("Ⅵ"):
            for col in range(1, 5):
                ws.cell(row=i, column=col).font = bold
                ws.cell(row=i, column=col).fill = orange

    # 거래처 TOP 5
    ws["A15"] = "주요 거래처 TOP 5 (올해 매출 기준)"
    ws["A15"].font = bold; ws["A15"].fill = gray
    ws.merge_cells("A15:D15")
    ws["A16"] = "{{거래처.TOP5}}"
    ws.merge_cells("A16:D24")
    ws["A16"].alignment = Alignment(vertical="top", wrap_text=True)

    # 사용 가이드 시트
    ws2 = wb.create_sheet("📘 사용 방법")
    ws2["A1"] = "Placeholder 사용 가이드"
    ws2["A1"].font = Font(size=14, bold=True, color="6B2C91")
    guide_rows = [
        ["", ""],
        ["사용법", "셀에 {{변수.기간}} 형식으로 입력하면 AI 분석 시 자동 치환됩니다."],
        ["", ""],
        ["▶ 도메인", "매출, 매입, 매출원가, 매출총이익, 판관비, 영업이익, 순이익, 급여, 비용, 임차료"],
        ["▶ 기간", "올해, 작년, 이번달, 지난달, 이번분기, YYYY (예: 2025), YYYY-MM (예: 2025-06), YYYY-Q1"],
        ["", ""],
        ["▶ 예시", ""],
        ["  {{매출.올해}}", "올해 1월 1일~오늘까지 매출 합계"],
        ["  {{매출.이번달}}", "이번 달 매출 합계"],
        ["  {{매출.2025-06}}", "2025년 6월 매출 합계"],
        ["  {{영업이익.올해}}", "올해 영업이익 (매출 - 매출원가 - 판관비)"],
        ["  {{순이익.올해}}", "올해 당기순이익 (영업이익 - 추정 법인세 10%)"],
        ["", ""],
        ["▶ 정보 변수", ""],
        ["  {{회사명}}", COMPANY_INFO["회사명"]],
        ["  {{사업자번호}}", COMPANY_INFO["사업자번호"]],
        ["  {{대표자}}", COMPANY_INFO["대표자"]],
        ["  {{현재일자}}", "오늘 날짜 (YYYY-MM-DD)"],
        ["  {{기준일}}", "오늘 날짜"],
        ["  {{현재월}}", "1월, 2월 등"],
        ["  {{현재연도}}", "2026년 등"],
        ["  {{현재분기}}", "1분기, 2분기 등"],
        ["", ""],
        ["▶ 거래처 TOP", ""],
        ["  {{거래처.TOP3}}", "올해 매출 TOP3 거래처 리스트"],
        ["  {{거래처.TOP5}}", "올해 매출 TOP5 거래처 리스트"],
        ["  {{거래처.TOP10}}", "올해 매출 TOP10 거래처 리스트"],
    ]
    for i, (a, b) in enumerate(guide_rows, start=3):
        ws2.cell(row=i, column=1, value=a)
        ws2.cell(row=i, column=2, value=b)
        if a.startswith("▶"):
            ws2.cell(row=i, column=1).font = Font(bold=True, color="F47521")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 60

    # 컬럼 너비
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="report_template.xlsx"'},
    )


# ====== 라우트 ======
def _list_response(request, db, category=""):
    q = select(ReportTemplate)
    if category:
        q = q.where(ReportTemplate.category == category)
    rows = db.execute(q.order_by(ReportTemplate.id.desc())).scalars().all()
    cat = REPORT_CATS.get(category)
    return templates.TemplateResponse("reports/list.html", {
        "request": request, "templates_list": rows, "preview_html": None,
        "category": category,
        "cat_label": cat["label"] if cat else "보고서",
        "cat_icon": cat["icon"] if cat else "📑",
        "cat_desc": cat["desc"] if cat else None,
    })


@router.get("", response_class=HTMLResponse)
@router.get("/", response_class=HTMLResponse)
def reports_list(request: Request, db: Session = Depends(get_db)):
    """전체 보고서 양식 목록 + 신규 업로드 폼"""
    return _list_response(request, db, "")


@router.get("/closing", response_class=HTMLResponse)
def reports_closing(request: Request, db: Session = Depends(get_db)):
    """결산 보고 — 양식 업로드 + AI 분석 + Excel/PDF 출력"""
    return _list_response(request, db, "closing")


@router.get("/financial", response_class=HTMLResponse)
def reports_financial(request: Request, db: Session = Depends(get_db)):
    """재무제표 — 양식 업로드 + AI 분석 + Excel/PDF 출력"""
    return _list_response(request, db, "financial")


@router.get("/investor", response_class=HTMLResponse)
def reports_investor(request: Request, db: Session = Depends(get_db)):
    """투자사 보고 — 양식 업로드 + AI 분석 + Excel/PDF 출력"""
    return _list_response(request, db, "investor")


def _tax_data(db, yi, view):
    """세금 집계 — (periods, annual, corp_tax)."""
    pay = {m: {"wht_nat": 0.0, "wht_local": 0.0, "ins_emp": 0.0, "ins_corp": 0.0} for m in range(1, 13)}
    for m, it, lt, ie, ic in db.execute(select(
            Payroll.month,
            func.coalesce(func.sum(Payroll.income_tax), 0),
            func.coalesce(func.sum(Payroll.local_tax), 0),
            func.coalesce(func.sum(Payroll.pension + Payroll.health + Payroll.longterm + Payroll.employment), 0),
            func.coalesce(func.sum(Payroll.employer_insurance), 0),
    ).where(Payroll.year == yi).group_by(Payroll.month)).all():
        if m in pay:
            pay[m] = {"wht_nat": float(it), "wht_local": float(lt), "ins_emp": float(ie), "ins_corp": float(ic)}

    svat = {m: 0.0 for m in range(1, 13)}
    pvat = {m: 0.0 for m in range(1, 13)}
    for m, v in db.execute(select(Sale.month, func.coalesce(func.sum(Sale.vat), 0)).where(Sale.year == yi).group_by(Sale.month)).all():
        if m in svat:
            svat[m] = float(v)
    for m, v in db.execute(select(Purchase.month, func.coalesce(func.sum(Purchase.vat), 0)).where(Purchase.year == yi).group_by(Purchase.month)).all():
        if m in pvat:
            pvat[m] = float(v)

    def row_for(months, label):
        wn = sum(pay[m]["wht_nat"] for m in months)
        wl = sum(pay[m]["wht_local"] for m in months)
        ie = sum(pay[m]["ins_emp"] for m in months)
        ic = sum(pay[m]["ins_corp"] for m in months)
        sv = sum(svat[m] for m in months)
        pv = sum(pvat[m] for m in months)
        return {"label": label,
                "wht_nat": wn, "wht_local": wl, "wht": wn + wl,
                "ins_emp": ie, "ins_corp": ic, "ins": ie + ic,
                "svat": sv, "pvat": pv, "vat": sv - pv,
                "tax_total": wn + wl + (sv - pv)}

    periods = []
    if view == "month":
        for m in range(1, 13):
            periods.append(row_for([m], f"{m}월"))
    else:
        for qi in range(1, 5):
            periods.append(row_for([(qi - 1) * 3 + 1, (qi - 1) * 3 + 2, (qi - 1) * 3 + 3], f"{qi}분기"))
    annual = row_for(list(range(1, 13)), f"{yi} 연간")
    corp_tax = float(_aggregate(db, "법인세", y=yi) or 0)
    return periods, annual, corp_tax


@router.get("/tax", response_class=HTMLResponse)
def reports_tax(request: Request, db: Session = Depends(get_db), year: str = "", view: str = "month"):
    """세금 — 국세(원천세·법인세·부가세), 지방세, 4대보험을 급여·매출 자료로 연/월/분기 자동 집계."""
    yi = int(year) if (year or "").isdigit() else date.today().year
    if view not in ("month", "quarter"):
        view = "month"
    periods, annual, corp_tax = _tax_data(db, yi, view)
    return templates.TemplateResponse("reports/tax.html", {
        "request": request, "year": yi, "view": view, "periods": periods, "annual": annual,
        "corp_tax": corp_tax,
        "sale_supply": float(_aggregate(db, "매출", y=yi) or 0),
        "pur_supply": float(_aggregate(db, "매입", y=yi) or 0),
        "years": list(range(2021, date.today().year + 1)),
    })


@router.get("/tax/export.xlsx")
def reports_tax_export(db: Session = Depends(get_db), year: str = "", view: str = "month"):
    """세금 집계 Excel 내보내기."""
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    yi = int(year) if (year or "").isdigit() else date.today().year
    if view not in ("month", "quarter"):
        view = "month"
    periods, annual, corp_tax = _tax_data(db, yi, view)
    wb = Workbook(); ws = wb.active; ws.title = f"{yi}년 세금"
    headers = ["기간", "원천세(국세)", "지방소득세", "원천세 계",
               "매출세액", "매입세액", "부가세 납부",
               "4대보험(근로자)", "4대보험(회사)", "세금 합계"]
    keys = ["label", "wht_nat", "wht_local", "wht", "svat", "pvat", "vat", "ins_emp", "ins_corp", "tax_total"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="6B2C91")
        cell.alignment = Alignment(horizontal="center")
    for p in periods:
        ws.append([p[k] for k in keys])
    ws.append([annual[k] for k in keys])
    last = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(row=last, column=c).font = Font(bold=True)
        ws.cell(row=last, column=c).fill = PatternFill("solid", fgColor="EDE9FE")
    ws.append([])
    ws.append(["법인세(연간 추정)", corp_tax])
    for col, w in zip("ABCDEFGHIJ", (12, 14, 13, 13, 14, 14, 13, 15, 14, 14)):
        ws.column_dimensions[col].width = w
    for r in range(2, ws.max_row + 1):
        for c in range(2, len(headers) + 1):
            ws.cell(row=r, column=c).number_format = "#,##0"
    buf = _io.BytesIO(); wb.save(buf)
    fname = f"tax_{yi}_{view}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.post("/tax/calendar")
def reports_tax_calendar(db: Session = Depends(get_db), year: str = Form("")):
    """해당 연도 세금 신고·납부 기한을 캘린더에 자동 등록(중복 방지)."""
    from models import CalendarEvent
    from calendar import monthrange
    yi = int(year) if (year or "").isdigit() else date.today().year
    events = []
    # 원천세·4대보험 — 매월 10일 (전월 귀속분)
    for m in range(1, 13):
        events.append((date(yi, m, 10), "원천세·4대보험 신고·납부 (전월분)", f"tax:wht:{yi}-{m:02d}"))
    # 부가가치세 — 법인 분기
    events += [
        (date(yi, 1, 25), "부가가치세 2기 확정 신고·납부 (전년 10~12월)", f"tax:vat:{yi}-01"),
        (date(yi, 4, 25), "부가가치세 1기 예정 신고·납부 (1~3월)", f"tax:vat:{yi}-04"),
        (date(yi, 7, 25), "부가가치세 1기 확정 신고·납부 (4~6월)", f"tax:vat:{yi}-07"),
        (date(yi, 10, 25), "부가가치세 2기 예정 신고·납부 (7~9월)", f"tax:vat:{yi}-10"),
    ]
    # 법인세 — 12월 결산 법인, 3/31
    events.append((date(yi, 3, 31), "법인세 신고·납부 (전년도 결산, 12월 결산 법인)", f"tax:corp:{yi}"))
    added = 0
    for ev_date, title, tag in events:
        exists = db.scalar(select(func.count()).select_from(CalendarEvent).where(
            CalendarEvent.category == "tax", CalendarEvent.note == tag))
        if exists:
            continue
        db.add(CalendarEvent(event_date=ev_date, title=f"🧾 {title}", category="tax",
                             repeat="none", done="N", note=tag))
        added += 1
    db.commit()
    return RedirectResponse(f"/reports/tax?year={yi}&cal_added={added}", status_code=303)


@router.post("/upload")
async def reports_upload(
    request: Request,
    db: Session = Depends(get_db),
    file: UploadFile = File(...),
    name: str = Form(""),
    description: str = Form(""),
    category: str = Form(""),
):
    """양식 업로드 (Excel 또는 PDF) — xlsx는 placeholder 스캔, pdf는 AI 분석 대상으로 저장"""
    back = f"/reports/{category}" if category in REPORT_CATS else "/reports"
    fn = (file.filename or "").lower()
    is_xlsx = fn.endswith((".xlsx", ".xlsm", ".xls"))
    is_pdf = fn.endswith(".pdf")
    if not (is_xlsx or is_pdf):
        return RedirectResponse(f"{back}?error=Excel(.xlsx)+또는+PDF만+허용됩니다", status_code=303)
    content = await file.read()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = re.sub(r"[^\w가-힣.-]", "_", file.filename)
    save_path = TEMPLATE_DIR / f"{ts}_{base}"
    save_path.write_bytes(content)
    placeholders = []
    if is_xlsx:
        try:
            placeholders = scan_placeholders_xlsx(str(save_path))
        except Exception as e:
            save_path.unlink(missing_ok=True)
            return RedirectResponse(f"{back}?error=양식+분석+실패:+{e}", status_code=303)
    rec = ReportTemplate(
        name=name.strip() or file.filename,
        description=description.strip() or None,
        file_path=str(save_path),
        file_name=file.filename,
        file_size=len(content),
        placeholders_json=json.dumps(placeholders, ensure_ascii=False),
        category=category if category in REPORT_CATS else "",
        file_kind="pdf" if is_pdf else "xlsx",
    )
    db.add(rec); db.commit(); db.refresh(rec)
    return RedirectResponse(f"/reports/{rec.id}", status_code=303)


def _load_snapshot_values(db, tpl_id, snapshot_id):
    """스냅샷 id로 저장된 좌표→값 dict 로드. 없으면 (None, None)"""
    snap = db.get(ReportSnapshot, snapshot_id)
    if not snap or snap.template_id != tpl_id:
        return None, None
    try:
        return json.loads(snap.cells_json or "{}"), snap
    except Exception:
        return {}, snap


@router.get("/{tpl_id}", response_class=HTMLResponse)
def report_detail(tpl_id: int, request: Request, db: Session = Depends(get_db),
                  snapshot: str = ""):
    tpl = db.get(ReportTemplate, tpl_id)
    if not tpl:
        raise HTTPException(404, "양식 없음")
    # PDF 양식: 텍스트 추출 + AI 분석 화면
    if _is_pdf(tpl):
        pdf_text = _extract_pdf_text(tpl.file_path)
        last = db.execute(select(ReportSnapshot).where(ReportSnapshot.template_id == tpl_id)
                          .order_by(ReportSnapshot.id.desc())).scalars().first()
        rows = []
        if last:
            try:
                rows = json.loads(last.cells_json or "[]")
            except Exception:
                rows = []
        return templates.TemplateResponse("reports/detail_pdf.html", {
            "request": request, "tpl": tpl, "pdf_text": pdf_text, "rows": rows,
            "snap_id": last.id if last else None,
            "cat_label": REPORT_CATS.get(tpl.category or "", {}).get("label", "보고서"),
        })
    placeholders = json.loads(tpl.placeholders_json or "[]")
    grid_html, resolved, missing, loaded_snap = None, None, [], None
    # 저장본 불러오기 (?snapshot=ID)
    snap_id = None
    try:
        snap_id = int(snapshot) if snapshot else None
    except ValueError:
        snap_id = None
    if snap_id and Path(tpl.file_path).exists():
        override, loaded_snap = _load_snapshot_values(db, tpl_id, snap_id)
        if override is not None:
            grid_html, resolved, missing = build_editable_grid(tpl.file_path, db, override)
    return templates.TemplateResponse("reports/detail.html", {
        "request": request, "tpl": tpl, "placeholders": placeholders,
        "grid_html": grid_html, "resolved": resolved, "missing": missing,
        "loaded_snap": loaded_snap,
    })


@router.post("/{tpl_id}/analyze", response_class=HTMLResponse)
def report_analyze(tpl_id: int, request: Request, db: Session = Depends(get_db)):
    """🤖 AI 분석 — 모든 placeholder를 DB 값으로 치환 → 편집 가능한 그리드 반환"""
    tpl = db.get(ReportTemplate, tpl_id)
    if not tpl:
        raise HTTPException(404, "양식 없음")
    if not Path(tpl.file_path).exists():
        raise HTTPException(404, "양식 파일이 디스크에서 사라졌습니다.")
    grid_html, resolved, missing = build_editable_grid(tpl.file_path, db)
    placeholders = json.loads(tpl.placeholders_json or "[]")
    tpl.last_used_at = datetime.utcnow()
    tpl.use_count = (tpl.use_count or 0) + 1
    db.commit()
    return templates.TemplateResponse("reports/detail.html", {
        "request": request, "tpl": tpl, "placeholders": placeholders,
        "grid_html": grid_html, "resolved": resolved, "missing": missing,
        "loaded_snap": None,
    })


@router.post("/{tpl_id}/analyze-pdf", response_class=HTMLResponse)
def report_analyze_pdf(tpl_id: int, request: Request, db: Session = Depends(get_db),
                       year: str = Form("")):
    """PDF 양식 AI 분석 — 양식 텍스트 + 회사 재무데이터 → LLM이 항목별 금액 채움 → 편집표 반환"""
    tpl = db.get(ReportTemplate, tpl_id)
    if not tpl or not _is_pdf(tpl):
        raise HTTPException(404, "PDF 양식 없음")
    try:
        yr = int(year) if str(year).isdigit() else date.today().year
    except Exception:
        yr = date.today().year
    text = _extract_pdf_text(tpl.file_path)
    ctx = _financial_context(db, yr)
    cat_label = REPORT_CATS.get(tpl.category or "", {}).get("label", "보고서")
    system = (
        f"당신은 한국 회계 비서입니다. 사용자가 올린 '{cat_label}' 양식(PDF에서 추출한 텍스트)의 "
        "각 항목을 아래 [회사 재무 요약] 수치로 채워주세요. "
        "반드시 JSON 배열만 출력하세요. 형식: "
        '[{"항목":"매출액","금액":123456,"비고":""}, ...]. '
        "금액은 콤마 없는 숫자(원)만. 양식의 항목 순서를 최대한 유지하고, "
        "자료로 알 수 없는 항목은 금액 0, 비고에 '자료없음'으로 표기하세요. JSON 외 다른 말은 쓰지 마세요."
    )
    user = f"[회사 재무 요약]\n{ctx}\n\n[업로드한 {cat_label} 양식 텍스트]\n{text[:6000]}"
    rows, err = [], None
    try:
        import llm_provider
        raw = llm_provider.chat_complete(
            [{"role": "system", "content": system}, {"role": "user", "content": user}],
            temperature=0.1, max_tokens=1500)
        rows = _parse_json_rows(raw)
    except Exception as e:
        err = str(e)
    if not rows:
        rows = _fallback_rows(db, yr)  # AI 실패/파싱불가 시 기본 손익 표
    snap = ReportSnapshot(
        template_id=tpl_id, template_name=tpl.name,
        title=f"{tpl.name} AI분석 {datetime.now().strftime('%m-%d %H:%M')}",
        note=("AI 분석" + (f" (LLM오류: {err}, 기본표 사용)" if err else "")),
        cells_json=json.dumps(rows, ensure_ascii=False))
    db.add(snap); db.commit(); db.refresh(snap)
    return templates.TemplateResponse("reports/_pdf_table.html", {
        "request": request, "tpl": tpl, "rows": rows, "snap_id": snap.id, "year": yr,
    })


@router.post("/{tpl_id}/pdf-save")
async def report_pdf_save(tpl_id: int, request: Request, db: Session = Depends(get_db)):
    """PDF 분석 결과(편집본) 저장 — 새 스냅샷 생성, snap_id 반환"""
    tpl = db.get(ReportTemplate, tpl_id)
    if not tpl:
        raise HTTPException(404, "양식 없음")
    body = await request.json()
    rows = body.get("rows") or []
    title = (body.get("title") or f"{tpl.name} 저장본 {datetime.now().strftime('%m-%d %H:%M')}").strip()
    snap = ReportSnapshot(template_id=tpl_id, template_name=tpl.name, title=title,
                          note="PDF 편집 저장", cells_json=json.dumps(rows, ensure_ascii=False))
    db.add(snap); db.commit(); db.refresh(snap)
    return JSONResponse({"ok": True, "snap_id": snap.id})


@router.post("/{tpl_id}/resolve-cell")
def report_resolve_cell(tpl_id: int, db: Session = Depends(get_db),
                        template: str = Form(...)):
    """단일 셀 최신값 갱신 — 원본 셀 텍스트(placeholder 포함)를 받아 현재 DB값으로 치환"""
    matched_all = True
    def repl(m):
        nonlocal matched_all
        var = m.group(1).strip()
        val, ok = resolve_placeholder(var, db)
        if not ok:
            matched_all = False
            return f"{{{{{var}}}}}"
        return str(val)
    value = PLACEHOLDER_RE.sub(repl, template)
    return {"value": value, "ok": matched_all}


# ============================================================
# 항목별 산출 근거 (Breakdown) — 재무제표 셀 클릭 시 호출
# ============================================================
def _period_label(y, m_, q, fd, td) -> str:
    parts = []
    if y: parts.append(f"{y}년")
    if m_: parts.append(f"{m_}월")
    if q: parts.append(f"Q{q}")
    if fd: parts.append(f"{fd} ~")
    if td: parts.append(f"~ {td}")
    return " ".join(parts) if parts else "전체"


def _sample_rows(db, model, conds, amount_col, *, fields, label_fn, limit=30):
    """집계 대상 source 행을 amount 내림차순으로 sample. 합계도 같이 반환."""
    from sqlalchemy import desc as _desc
    q = select(model).where(*conds).order_by(_desc(amount_col)).limit(limit)
    rows = db.execute(q).scalars().all()
    total_q = select(func.count(), func.coalesce(func.sum(amount_col), 0)).where(*conds)
    cnt, tot = db.execute(total_q).first() or (0, 0)
    out = []
    for r in rows:
        d = {}
        for k in fields:
            v = getattr(r, k, None)
            if hasattr(v, "isoformat"):
                v = v.isoformat()
            elif v is not None and not isinstance(v, (str, int, float, bool)):
                v = str(v)
            d[k] = v
        d["__label"] = label_fn(r)
        out.append(d)
    return out, int(cnt or 0), float(tot or 0)


def _domain_breakdown(db, domain: str, *, y=None, m_=None, q=None, fd=None, td=None) -> dict:
    """단일 도메인의 산출 근거 — 원자 도메인은 source rows, 복합은 components."""
    sale_conds, pur_conds, pay_conds, exp_conds = [], [], [], []
    ren_conds = [Rental.direction == "지출"]
    if y: sale_conds.append(Sale.year == y); pur_conds.append(Purchase.year == y); pay_conds.append(Payroll.year == y); exp_conds.append(Expense.year == y); ren_conds.append(Rental.year == y)
    if m_: sale_conds.append(Sale.month == m_); pur_conds.append(Purchase.month == m_); pay_conds.append(Payroll.month == m_); exp_conds.append(Expense.month == m_); ren_conds.append(Rental.month == m_)
    if q: sale_conds.append(Sale.quarter == f"Q{q}"); pur_conds.append(Purchase.quarter == f"Q{q}"); exp_conds.append(Expense.quarter == f"Q{q}")
    if fd: sale_conds.append(Sale.txn_date >= fd); pur_conds.append(Purchase.txn_date >= fd); exp_conds.append(Expense.use_date >= fd)
    if td: sale_conds.append(Sale.txn_date <= td); pur_conds.append(Purchase.txn_date <= td); exp_conds.append(Expense.use_date <= td)

    label = _period_label(y, m_, q, fd, td)

    # 원자 도메인
    if domain == "매출":
        rows, cnt, tot = _sample_rows(db, Sale, sale_conds, Sale.supply,
            fields=["txn_date","party_name","product_name","item_raw","supply","vat","total","source_file"],
            label_fn=lambda r: f"{r.txn_date} · {r.party_name or '-'} · {r.product_name or r.item_raw or '-'}")
        return {"kind":"atomic","domain":"매출","label":f"매출 ({label})","total":tot,
                "formula":"Σ fact_sale.supply (공급가액)",
                "source_table":"fact_sale", "row_count":cnt, "rows":rows,
                "fields":[("txn_date","일자"),("party_name","거래처"),("product_name","제품"),("item_raw","품목"),("supply","공급가액"),("vat","부가세"),("total","합계"),("source_file","출처")],
                "amount_field":"supply"}
    if domain in ("매입","매출원가"):
        rows, cnt, tot = _sample_rows(db, Purchase, pur_conds, Purchase.supply,
            fields=["txn_date","party_name","product_name","item_raw","supply","vat","total","source_file"],
            label_fn=lambda r: f"{r.txn_date} · {r.party_name or '-'} · {r.product_name or r.item_raw or '-'}")
        return {"kind":"atomic","domain":domain,"label":f"{domain} ({label})","total":tot,
                "formula":"Σ fact_purchase.supply (공급가액)",
                "source_table":"fact_purchase", "row_count":cnt, "rows":rows,
                "fields":[("txn_date","일자"),("party_name","거래처"),("product_name","제품"),("item_raw","품목"),("supply","공급가액"),("vat","부가세"),("total","합계"),("source_file","출처")],
                "amount_field":"supply"}
    if domain == "급여":
        rows, cnt, tot = _sample_rows(db, Payroll, pay_conds, Payroll.gross_pay,
            fields=["period","employee_name","department","gross_pay","net_pay","total_deduction","employer_insurance"],
            label_fn=lambda r: f"{r.period} · {r.employee_name or '-'} ({r.department or '-'})")
        return {"kind":"atomic","domain":"급여","label":f"급여 ({label})","total":tot,
                "formula":"Σ fact_payroll.gross_pay (지급총액)",
                "source_table":"fact_payroll", "row_count":cnt, "rows":rows,
                "fields":[("period","월"),("employee_name","성명"),("department","부서"),("gross_pay","지급총액"),("net_pay","실수령"),("total_deduction","공제계"),("employer_insurance","회사부담분")],
                "amount_field":"gross_pay"}
    if domain == "비용":
        rows, cnt, tot = _sample_rows(db, Expense, exp_conds, Expense.amount,
            fields=["use_date","employee_name","department","party_or_place","category_main","category_sub","amount","payment_method"],
            label_fn=lambda r: f"{r.use_date} · {r.party_or_place or '-'} · {r.category_main or '-'}")
        return {"kind":"atomic","domain":"비용","label":f"비용/경비 ({label})","total":tot,
                "formula":"Σ fact_expense.amount",
                "source_table":"fact_expense", "row_count":cnt, "rows":rows,
                "fields":[("use_date","일자"),("employee_name","사용자"),("department","부서"),("party_or_place","사용처"),("category_main","분류"),("category_sub","세부"),("amount","금액"),("payment_method","결제")],
                "amount_field":"amount"}
    if domain == "임차료":
        rows, cnt, tot = _sample_rows(db, Rental, ren_conds, Rental.amount,
            fields=["txn_date","party","asset_name","item","amount","payment_method"],
            label_fn=lambda r: f"{r.txn_date} · {r.party or '-'} · {r.item or '-'}")
        return {"kind":"atomic","domain":"임차료","label":f"임차료 ({label})","total":tot,
                "formula":"Σ fact_rental.amount where direction='지출'",
                "source_table":"fact_rental", "row_count":cnt, "rows":rows,
                "fields":[("txn_date","일자"),("party","임대인"),("asset_name","자산"),("item","항목"),("amount","금액"),("payment_method","결제")],
                "amount_field":"amount"}

    # 복합 도메인 — 구성요소 분해
    def sub(d):
        return _domain_breakdown(db, d, y=y, m_=m_, q=q, fd=fd, td=td)
    def total_only(d):
        v = _aggregate(db, d, y=y, m=m_, q=q, fd=fd, td=td)
        return float(v or 0)
    if domain == "매출총이익":
        s = total_only("매출"); p = total_only("매입")
        return {"kind":"composite","domain":"매출총이익","label":f"매출총이익 ({label})",
                "total": s - p, "formula":"매출 − 매출원가(매입)",
                "components":[
                    {"label":"매출","value":s,"sign":"+","domain":"매출"},
                    {"label":"매출원가(매입)","value":p,"sign":"−","domain":"매입"},
                ]}
    if domain in ("판관비","판매비와관리비"):
        pay = total_only("급여"); exp = total_only("비용"); ren = total_only("임차료")
        # 사용자분 4대보험 분리 집계
        emp_ins = float(db.scalar(select(func.coalesce(func.sum(Payroll.employer_insurance), 0)).where(*pay_conds)) or 0)
        return {"kind":"composite","domain":"판관비","label":f"판매비와관리비 ({label})",
                "total": pay + emp_ins + exp + ren,
                "formula":"급여 + 회사부담 4대보험 + 경비 + 임차료",
                "components":[
                    {"label":"급여","value":pay,"sign":"+","domain":"급여"},
                    {"label":"회사부담 4대보험","value":emp_ins,"sign":"+","domain":"급여"},
                    {"label":"경비","value":exp,"sign":"+","domain":"비용"},
                    {"label":"임차료","value":ren,"sign":"+","domain":"임차료"},
                ]}
    if domain == "영업이익":
        gp = total_only("매출총이익")
        sga = total_only("판관비")
        return {"kind":"composite","domain":"영업이익","label":f"영업이익 ({label})",
                "total": gp - sga, "formula":"매출총이익 − 판매비와관리비",
                "components":[
                    {"label":"매출총이익","value":gp,"sign":"+","domain":"매출총이익"},
                    {"label":"판매비와관리비","value":sga,"sign":"−","domain":"판관비"},
                ]}
    if domain in ("세전순이익","법인세","순이익","당기순이익"):
        op = total_only("영업이익")
        tax = op * 0.10 if op > 0 else 0
        if domain == "세전순이익":
            return {"kind":"composite","domain":"세전순이익","label":f"세전순이익 ({label})",
                    "total": op, "formula":"영업이익 (영업외손익 미적용)",
                    "components":[{"label":"영업이익","value":op,"sign":"+","domain":"영업이익"}]}
        if domain == "법인세":
            return {"kind":"composite","domain":"법인세","label":f"법인세 가정 ({label})",
                    "total": tax, "formula":"영업이익 × 10% (영업이익 > 0 일 때)",
                    "components":[{"label":"영업이익","value":op,"sign":"+","domain":"영업이익"}]}
        return {"kind":"composite","domain":"당기순이익","label":f"당기순이익 ({label})",
                "total": op - tax, "formula":"영업이익 − 법인세(10% 가정)",
                "components":[
                    {"label":"영업이익","value":op,"sign":"+","domain":"영업이익"},
                    {"label":"법인세(10%)","value":tax,"sign":"−","domain":"법인세"},
                ]}
    return {"kind":"unknown","domain":domain,"label":f"{domain} ({label})","total":0,
            "formula":"(해당 도메인의 산출식이 정의되지 않았습니다)",
            "components":[]}


@router.get("/{tpl_id}/breakdown")
def report_breakdown(tpl_id: int, db: Session = Depends(get_db),
                     template: str = "", variable: str = ""):
    """셀(또는 placeholder)의 산출 근거 — 도메인·기간·공식·구성요소·source rows.

    parameter:
      - template: 원본 셀 텍스트 (예: "{{매출.올해}}" 또는 "매출 {{매출.올해}} / 매입 {{매입.올해}}")
      - variable: 단일 변수 (예: "매출.올해"). template과 둘 다 비면 400.
    """
    today = date.today()

    # variable 단독 → {{variable}}로 처리
    vars_to_process = []
    if variable.strip():
        vars_to_process = [variable.strip()]
    else:
        for m in PLACEHOLDER_RE.finditer(template or ""):
            v = m.group(1).strip()
            if v not in vars_to_process:
                vars_to_process.append(v)
    if not vars_to_process:
        return JSONResponse({"ok": False, "message": "산출 근거를 표시할 placeholder가 없습니다."}, status_code=400)

    items = []
    cvars = _company_vars(db)
    for var in vars_to_process:
        # 회사정보 / 현재일자 등 — 단순 텍스트
        if var in cvars:
            items.append({"variable": var, "kind":"company","label":var,
                          "total": cvars[var], "formula":"회사 기본정보 (DB)",
                          "components":[], "rows":[]})
            continue
        if var in ("현재일자","기준일","현재월","현재연도","현재분기"):
            val, _ = resolve_placeholder(var, db)
            items.append({"variable": var, "kind":"datetime","label":var,
                          "total": val, "formula":"오늘 날짜 기준 자동 계산",
                          "components":[], "rows":[]})
            continue
        # 거래처.TOPn
        m = re.match(r"^거래처\.TOP(\d+)$", var)
        if m:
            n = int(m.group(1)); y = today.year
            rows_q = db.execute(
                select(Sale.party_name, func.sum(Sale.supply), func.count())
                .where(Sale.year == y, Sale.party_name.is_not(None))
                .group_by(Sale.party_name)
                .order_by(func.sum(Sale.supply).desc()).limit(n)
            ).all()
            rows = [{"party_name": r[0], "supply": float(r[1] or 0), "txn_count": int(r[2] or 0),
                     "__label": f"{i+1}위 {r[0]}"} for i, r in enumerate(rows_q)]
            tot = sum(r["supply"] for r in rows)
            items.append({"variable": var, "kind":"ranking",
                          "label": f"거래처 TOP {n} ({y}년)",
                          "total": tot,
                          "formula": f"fact_sale.supply 합산 → 거래처별 그룹 → 상위 {n}",
                          "source_table":"fact_sale",
                          "fields":[("party_name","거래처"),("supply","연 매출 공급가액"),("txn_count","거래 건수")],
                          "amount_field":"supply",
                          "row_count": len(rows), "rows": rows, "components":[]})
            continue
        # 도메인.기간
        parts = var.split(".")
        if len(parts) != 2:
            items.append({"variable": var, "kind":"unknown","label":var,
                          "total":0, "formula":"인식되지 않은 placeholder 형식",
                          "components":[], "rows":[]})
            continue
        domain, period = parts[0].strip(), parts[1].strip()
        y, m_, q, fd, td = _parse_period_token(period, today)
        info = _domain_breakdown(db, domain, y=y, m_=m_, q=q, fd=fd, td=td)
        info["variable"] = var
        info["period"] = {"raw": period, "year": y, "month": m_, "quarter": q,
                          "from": fd.isoformat() if fd else None,
                          "to": td.isoformat() if td else None,
                          "label": _period_label(y, m_, q, fd, td)}
        items.append(info)

    return {"ok": True, "template": template, "items": items}


@router.post("/{tpl_id}/ai-edit")
async def report_ai_edit(tpl_id: int, request: Request, db: Session = Depends(get_db)):
    """AI 수정 지시 — 사용자 지시문 + 현재 셀들을 LLM에 보내 변경할 셀 값을 추론.
    반환: {ok, changes: {coord: new_value}, message}
    LLM이 돌려준 값에 placeholder가 있으면 서버에서 실제 DB값으로 재치환한다.
    """
    import llm_provider
    tpl = db.get(ReportTemplate, tpl_id)
    if not tpl:
        return JSONResponse({"ok": False, "message": "양식 없음"}, status_code=404)

    form = await request.form()
    instruction = (form.get("instruction") or "").strip()
    cells_raw = form.get("cells") or "[]"
    if not instruction:
        return JSONResponse({"ok": False, "message": "수정 지시 내용을 입력하세요."}, status_code=400)
    try:
        cells = json.loads(cells_raw)  # [{coord, value, template}]
    except Exception:
        return JSONResponse({"ok": False, "message": "cells 파싱 실패"}, status_code=400)

    ready, msg = llm_provider.provider_ready()
    if not ready:
        return JSONResponse({
            "ok": False,
            "message": f"AI 공급자가 준비되지 않았습니다: {msg} (설정 → AI 공급자)",
        }, status_code=503)

    # 셀을 LLM에 보낼 간단 표현으로 (최대 120개)
    items = []
    for c in cells[:120]:
        coord = c.get("coord")
        val = c.get("value", "")
        tmpl = c.get("template", "")
        if not coord:
            continue
        line = f'"{coord}": {json.dumps(val, ensure_ascii=False)}'
        if tmpl and tmpl != val:
            line += f'   (원본식: {tmpl})'
        items.append(line)
    cells_text = "\n".join(items)

    system = (
        "당신은 한국 의료 IT 회사 인비즈(Inviz)의 보고서 편집 비서입니다. "
        "사용자의 수정 지시에 따라 변경이 필요한 셀만 골라 JSON으로 반환하세요. "
        "변경할 값에는 {{매출.2025}}, {{영업이익.이번달}}, {{회사명}} 같은 placeholder를 쓸 수 있으며 "
        "시스템이 실제 데이터로 자동 치환합니다. 변경이 필요 없는 셀은 포함하지 마세요."
    )
    user = f"""[현재 셀 목록] (좌표: 현재값)
{cells_text}

[placeholder 사용 가능 예]
- 도메인: 매출, 매입, 매출원가, 매출총이익, 판관비, 영업이익, 순이익, 급여, 비용, 임차료
- 기간: 올해, 작년, 이번달, 지난달, 이번분기, 2025, 2025-06, 2025-Q1
- 회사정보: 회사명, 대표자, 주소, 사업자번호, 자본금, 임직원수, 현재일자, 현재월

[사용자 수정 지시]
{instruction}

[출력] 변경할 셀만 JSON 객체로. 형식: {{"좌표": "새 값", ...}}
JSON만 출력. 설명·마크다운 금지."""

    try:
        resp = llm_provider.chat_complete(
            [{"role": "system", "content": system},
             {"role": "user", "content": user}],
            temperature=0.1, json_mode=True, max_tokens=800,
        )
    except Exception as e:
        return JSONResponse({"ok": False, "message": f"LLM 호출 실패: {e}"}, status_code=502)

    # JSON 추출
    m = re.search(r"\{[\s\S]*\}", resp)
    if not m:
        return JSONResponse({"ok": False, "message": "AI 응답을 해석하지 못했습니다.", "raw": resp[:200]}, status_code=200)
    try:
        raw_changes = json.loads(m.group(0))
    except Exception:
        return JSONResponse({"ok": False, "message": "AI 응답 JSON 파싱 실패", "raw": resp[:200]}, status_code=200)

    # 유효 좌표만 + placeholder 재치환
    valid_coords = {c.get("coord") for c in cells}
    changes = {}
    for coord, new_val in raw_changes.items():
        if coord not in valid_coords:
            continue
        s = str(new_val)
        if "{{" in s:
            def repl(mm):
                var = mm.group(1).strip()
                val, ok = resolve_placeholder(var, db)
                return str(val) if ok else f"{{{{{var}}}}}"
            s = PLACEHOLDER_RE.sub(repl, s)
        changes[coord] = s

    msg = f"{len(changes)}개 항목을 수정했습니다." if changes else "수정할 항목을 찾지 못했습니다. 지시를 더 구체적으로 입력해 보세요."
    return JSONResponse({"ok": True, "changes": changes, "message": msg})


@router.post("/{tpl_id}/save")
async def report_save(tpl_id: int, request: Request, db: Session = Depends(get_db)):
    """현재 그리드 상태(갱신·수기수정 포함)를 스냅샷으로 저장 — DB + xlsx 파일"""
    tpl = db.get(ReportTemplate, tpl_id)
    if not tpl:
        raise HTTPException(404, "양식 없음")
    form = await request.form()
    title = (form.get("title") or "").strip() or f"{tpl.name} {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    note = (form.get("note") or "").strip() or None
    cells_raw = form.get("cells") or "{}"
    try:
        cells = json.loads(cells_raw)
    except Exception:
        return JSONResponse({"error": "cells 파싱 실패"}, status_code=400)

    # xlsx 파일 생성·저장
    file_path = None
    try:
        if Path(tpl.file_path).exists():
            xlsx_bytes = render_filled_xlsx(tpl.file_path, db, override_values=cells)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe = re.sub(r"[^\w가-힣.-]", "_", tpl.name)
            sp = SNAPSHOT_DIR / f"{tpl_id}_{ts}_{safe}.xlsx"
            sp.write_bytes(xlsx_bytes)
            file_path = str(sp)
    except Exception as e:
        print(f"[reports] 스냅샷 xlsx 저장 실패: {e}")

    snap = ReportSnapshot(
        template_id=tpl_id, template_name=tpl.name,
        title=title, note=note,
        cells_json=json.dumps(cells, ensure_ascii=False),
        file_path=file_path,
    )
    db.add(snap); db.commit(); db.refresh(snap)
    return JSONResponse({"ok": True, "snapshot_id": snap.id, "title": snap.title})


@router.get("/{tpl_id}/snapshots")
def report_snapshots(tpl_id: int, db: Session = Depends(get_db)):
    """이 양식의 과거 저장본 목록 (JSON)"""
    rows = db.execute(
        select(ReportSnapshot).where(ReportSnapshot.template_id == tpl_id)
        .order_by(ReportSnapshot.id.desc())
    ).scalars().all()
    return [{
        "id": s.id, "title": s.title, "note": s.note,
        "created_at": s.created_at.strftime("%Y-%m-%d %H:%M") if s.created_at else "",
        "has_file": bool(s.file_path),
    } for s in rows]


@router.post("/{tpl_id}/snapshot/{snap_id}/delete")
def report_snapshot_delete(tpl_id: int, snap_id: int, db: Session = Depends(get_db)):
    snap = db.get(ReportSnapshot, snap_id)
    if snap and snap.template_id == tpl_id:
        try:
            if snap.file_path:
                Path(snap.file_path).unlink(missing_ok=True)
        except Exception:
            pass
        db.delete(snap); db.commit()
    return RedirectResponse(f"/reports/{tpl_id}", status_code=303)


@router.get("/{tpl_id}/export.xlsx")
def report_export_xlsx(tpl_id: int, db: Session = Depends(get_db), snapshot: str = ""):
    tpl = db.get(ReportTemplate, tpl_id)
    if not tpl:
        raise HTTPException(404, "양식 없음")
    if _is_pdf(tpl):
        rows = _latest_pdf_rows(db, tpl_id, snapshot)
        xlsx_bytes = _build_pdf_xlsx(tpl, rows)
        safe_name = re.sub(r"[^\w가-힣.-]", "_", tpl.name)
        fname = f"{safe_name}_{date.today().strftime('%Y%m%d')}.xlsx"
        try:
            fname.encode("ascii"); cd = f'attachment; filename="{fname}"'
        except Exception:
            from urllib.parse import quote
            cd = f"attachment; filename*=UTF-8''{quote(fname)}"
        return Response(content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": cd})
    if not Path(tpl.file_path).exists():
        raise HTTPException(404, "양식 파일 없음")
    override = None
    snap_id = int(snapshot) if (snapshot or "").isdigit() else None
    if snap_id:
        override, _ = _load_snapshot_values(db, tpl_id, snap_id)
    xlsx_bytes = render_filled_xlsx(tpl.file_path, db, override_values=override)
    safe_name = re.sub(r"[^\w가-힣.-]", "_", tpl.name)
    fname = f"{safe_name}_{date.today().strftime('%Y%m%d')}.xlsx"
    try:
        fname.encode("ascii"); cd = f'attachment; filename="{fname}"'
    except Exception:
        from urllib.parse import quote
        cd = f"attachment; filename*=UTF-8''{quote(fname)}"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": cd},
    )


@router.get("/{tpl_id}/export.pdf")
def report_export_pdf(tpl_id: int, db: Session = Depends(get_db), snapshot: str = ""):
    """PDF 다운로드 — 미리보기 HTML을 weasyprint/xhtml2pdf로 PDF화. 없으면 인쇄용 HTML."""
    tpl = db.get(ReportTemplate, tpl_id)
    if not tpl:
        raise HTTPException(404, "양식 없음")
    override = None
    snap_id = int(snapshot) if (snapshot or "").isdigit() else None
    if snap_id:
        override, _ = _load_snapshot_values(db, tpl_id, snap_id)
    if _is_pdf(tpl):
        rows = _latest_pdf_rows(db, tpl_id, snapshot)
        trs = "".join(
            f"<tr><td>{_esc(r.get('항목',''))}</td>"
            f"<td style='text-align:right'>{int(round(float(r.get('금액',0) or 0))):,}</td>"
            f"<td>{_esc(r.get('비고',''))}</td></tr>" for r in rows)
        preview_html = ("<table border='1' cellspacing='0' cellpadding='6'>"
                        "<thead><tr><th>항목</th><th>금액(원)</th><th>비고</th></tr></thead>"
                        f"<tbody>{trs}</tbody></table>")
    else:
        preview_html, _ = build_preview_html(tpl.file_path, db, override)
    # 전체 HTML 래핑
    full_html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
body {{ font-family: 'Malgun Gothic', 'Apple SD Gothic Neo', sans-serif; padding: 20px; }}
table {{ border-collapse: collapse; margin-bottom: 20px; }}
h3 {{ color: #6B2C91; }}
</style></head><body>
<h2>{tpl.name}</h2>
<p style="color:#64748B; font-size:12px;">생성일: {date.today().isoformat()}</p>
{preview_html}
</body></html>"""

    # 우선 weasyprint → 없으면 xhtml2pdf → 없으면 HTML 그대로
    pdf_bytes = None
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=full_html).write_pdf()
    except ImportError:
        try:
            from xhtml2pdf import pisa
            buf = io.BytesIO()
            pisa.CreatePDF(io.StringIO(full_html), dest=buf, encoding="utf-8")
            pdf_bytes = buf.getvalue()
        except ImportError:
            pass

    safe_name = re.sub(r"[^\w가-힣.-]", "_", tpl.name)
    if pdf_bytes:
        fname = f"{safe_name}_{date.today().strftime('%Y%m%d')}.pdf"
        try:
            fname.encode("ascii"); cd = f'attachment; filename="{fname}"'
        except Exception:
            from urllib.parse import quote
            cd = f"attachment; filename*=UTF-8''{quote(fname)}"
        return Response(content=pdf_bytes, media_type="application/pdf",
                        headers={"Content-Disposition": cd})
    # PDF 라이브러리 없으면 HTML로 (브라우저에서 인쇄 → PDF 저장)
    return HTMLResponse(content=full_html + """
<script>setTimeout(() => window.print(), 500);</script>
""")


@router.post("/{tpl_id}/delete")
def report_delete(tpl_id: int, db: Session = Depends(get_db)):
    tpl = db.get(ReportTemplate, tpl_id)
    if not tpl:
        raise HTTPException(404, "양식 없음")
    try:
        Path(tpl.file_path).unlink(missing_ok=True)
    except Exception:
        pass
    db.delete(tpl); db.commit()
    return RedirectResponse("/reports", status_code=303)
