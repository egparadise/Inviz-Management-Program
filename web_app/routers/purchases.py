# -*- coding: utf-8 -*-
"""매입 라우터 — 기간 필터, 합계 카드, Excel/PDF 다운로드, CSV 업로드"""
import csv, io, json
from datetime import date, datetime
from pathlib import Path
from fastapi import APIRouter, Request, Form, Depends, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from sqlalchemy import select, func, or_
from sqlalchemy.orm import Session

from database import get_db
from helpers import templates
from models import Purchase, Party, Product
from export_util import build_xlsx, build_pdf

router = APIRouter()


def parse_d(s):
    if not s: return None
    try: return datetime.strptime(s, "%Y-%m-%d").date()
    except: return None


def parse_i(s):
    """빈 문자열·None·잘못된 입력을 None으로 안전 변환"""
    if s is None or s == "":
        return None
    try:
        return int(s)
    except (ValueError, TypeError):
        return None


def build_purchase_query(year, month, fd, td, party, product, q, purchase_type=""):
    conds = []
    if fd: conds.append(Purchase.txn_date >= fd)
    if td: conds.append(Purchase.txn_date <= td)
    if year: conds.append(Purchase.year == year)
    if month: conds.append(Purchase.month == month)
    if party: conds.append(Purchase.party_name.contains(party))
    if product: conds.append(Purchase.product_code == product)
    if purchase_type: conds.append(Purchase.purchase_type == purchase_type)
    if q: conds.append(or_(Purchase.party_name.contains(q), Purchase.item_raw.contains(q)))
    return conds


def _qs(year, month, from_date, to_date, party, product, q, purchase_type="", sort="", dir=""):
    parts = []
    if year: parts.append(f"year={year}")
    if month: parts.append(f"month={month}")
    if from_date: parts.append(f"from_date={from_date}")
    if to_date: parts.append(f"to_date={to_date}")
    if party: parts.append(f"party={party}")
    if product: parts.append(f"product={product}")
    if purchase_type: parts.append(f"purchase_type={purchase_type}")
    if q: parts.append(f"q={q}")
    if sort: parts.append(f"sort={sort}")
    if dir: parts.append(f"dir={dir}")
    return "&".join(parts)


def _filter_desc(year, month, fd, td, party, product, q):
    parts = []
    if fd or td:
        parts.append(f"기간 {fd or '시작'} ~ {td or '종료'}")
    if year and not (fd or td):
        parts.append(f"{year}년" + (f" {month}월" if month else ""))
    if party: parts.append(f"거래처:{party}")
    if product: parts.append(f"제품:{product}")
    if q: parts.append(f"검색:{q}")
    return " · ".join(parts) if parts else "전체 기간"


@router.get("", response_class=HTMLResponse)
def list_purchases(
    request: Request, db: Session = Depends(get_db),
    year: str = "", month: str = "",
    from_date: str = "", to_date: str = "",
    party: str = "", product: str = "", q: str = "",
    purchase_type: str = "",
    page: int = 1, per_page: int = 50,
    sort: str = "", dir: str = "desc",
):
    year_i = parse_i(year); month_i = parse_i(month)
    fd = parse_d(from_date); td = parse_d(to_date)
    import settings_store as _ss
    if not request.query_params:
        _y, _fd, _td = _ss.default_search_filter()
        if _y: year_i = _y
        if _fd: fd = _fd
        if _td: td = _td
    if "per_page" not in request.query_params:
        per_page = _ss.get_int("search_per_page", per_page)
    conds = build_purchase_query(year_i, month_i, fd, td, party, product, q, purchase_type)
    base = select(Purchase).where(*conds) if conds else select(Purchase)

    total_count = db.scalar(select(func.count()).select_from(base.subquery())) or 0
    sum_q = select(
        func.coalesce(func.sum(Purchase.supply), 0),
        func.coalesce(func.sum(Purchase.vat), 0),
        func.coalesce(func.sum(Purchase.total), 0),
    ).where(*conds) if conds else select(
        func.coalesce(func.sum(Purchase.supply), 0),
        func.coalesce(func.sum(Purchase.vat), 0),
        func.coalesce(func.sum(Purchase.total), 0),
    )
    s_row = db.execute(sum_q).one()
    sum_supply, sum_vat, sum_total = float(s_row[0] or 0), float(s_row[1] or 0), float(s_row[2] or 0)

    party_stmt = select(
        Purchase.party_name, func.count().label("cnt"), func.sum(Purchase.supply).label("supply")
    ).where(*conds, Purchase.party_name.is_not(None)) if conds else select(
        Purchase.party_name, func.count().label("cnt"), func.sum(Purchase.supply).label("supply")
    ).where(Purchase.party_name.is_not(None))
    party_top = db.execute(
        party_stmt.group_by(Purchase.party_name).order_by(func.sum(Purchase.supply).desc()).limit(10)
    ).all()

    prod_stmt = select(
        Purchase.product_code, Purchase.product_name, func.count().label("cnt"), func.sum(Purchase.supply).label("supply")
    ).where(*conds) if conds else select(
        Purchase.product_code, Purchase.product_name, func.count().label("cnt"), func.sum(Purchase.supply).label("supply")
    )
    prod_sum = db.execute(
        prod_stmt.group_by(Purchase.product_code, Purchase.product_name).order_by(func.sum(Purchase.supply).desc())
    ).all()

    SORT_MAP = {
        "txn_date": Purchase.txn_date, "party_name": Purchase.party_name,
        "product_name": Purchase.product_name, "item_raw": Purchase.item_raw,
        "purchase_type": Purchase.purchase_type, "supply": Purchase.supply,
        "vat": Purchase.vat, "total": Purchase.total, "note": Purchase.note,
    }
    col = SORT_MAP.get(sort)
    direction = "asc" if dir == "asc" else "desc"
    if col is not None:
        order_col = col.asc() if direction == "asc" else col.desc()
        order_stmt = base.order_by(order_col, Purchase.id.desc())
    else:
        order_stmt = base.order_by(Purchase.txn_date.desc(), Purchase.id.desc())

    rows = db.execute(
        order_stmt.offset((page - 1) * per_page).limit(per_page)
    ).scalars().all()

    products = db.execute(select(Product).order_by(Product.code)).scalars().all()
    years = list(range(2021, datetime.now().year + 1))

    # 최근 업로드 적용 내역 (되돌리기용, 안 되돌린 것만)
    from models import ImportBatch
    recent_batches = db.execute(
        select(ImportBatch).where(ImportBatch.domain == "purchase", ImportBatch.undone == "N")
        .order_by(ImportBatch.id.desc()).limit(5)
    ).scalars().all()

    return templates.TemplateResponse("purchases/list.html", {
        "request": request, "rows": rows,
        "total_count": total_count,
        "sum_supply": sum_supply, "sum_vat": sum_vat, "sum_total": sum_total,
        "party_top": [(r[0], r[1], float(r[2] or 0)) for r in party_top],
        "prod_sum": [(r[0], r[1], r[2], float(r[3] or 0)) for r in prod_sum],
        "products": products, "years": years,
        "filter": {"year": year_i, "month": month_i, "from_date": from_date, "to_date": to_date,
                   "party": party, "product": product, "q": q, "purchase_type": purchase_type,
                   "sort": sort, "dir": direction},
        "purchase_types": [r[0] for r in db.execute(
            select(Purchase.purchase_type).where(Purchase.purchase_type.is_not(None))
            .group_by(Purchase.purchase_type).order_by(Purchase.purchase_type)
        ).all() if r[0]],
        "page": page, "per_page": per_page,
        "total_pages": (total_count + per_page - 1) // per_page,
        "qs": _qs(year_i, month_i, from_date, to_date, party, product, q, purchase_type, sort, direction),
        "recent_batches": recent_batches,
    })


@router.get("/export.xlsx")
def export_xlsx(
    db: Session = Depends(get_db),
    year: str = "", month: str = "",
    from_date: str = "", to_date: str = "",
    party: str = "", product: str = "", q: str = "",
    purchase_type: str = "",
):
    year_i = parse_i(year); month_i = parse_i(month)
    fd = parse_d(from_date); td = parse_d(to_date)
    conds = build_purchase_query(year_i, month_i, fd, td, party, product, q, purchase_type)
    base = select(Purchase).where(*conds) if conds else select(Purchase)
    rows = db.execute(base.order_by(Purchase.txn_date, Purchase.id)).scalars().all()

    headers = ["일자", "거래처", "제품코드", "제품명", "품명(원본)",
               "매입유형", "공급가액", "부가세", "합계", "비고"]
    data_rows = [[r.txn_date, r.party_name, r.product_code, r.product_name,
                  r.item_raw, r.purchase_type,
                  float(r.supply or 0), float(r.vat or 0), float(r.total or 0),
                  r.note or ""] for r in rows]
    sums = {"공급가액": sum(r[6] for r in data_rows),
            "부가세": sum(r[7] for r in data_rows),
            "합계": sum(r[8] for r in data_rows)}
    xlsx_bytes = build_xlsx("매입 명세", _filter_desc(year_i, month_i, fd, td, party, product, q),
                            headers, data_rows, sums, money_cols=[6, 7, 8])
    fn = f"inviz_purchases_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(content=xlsx_bytes,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{fn}"'})


@router.get("/export.pdf")
def export_pdf(
    db: Session = Depends(get_db),
    year: str = "", month: str = "",
    from_date: str = "", to_date: str = "",
    party: str = "", product: str = "", q: str = "",
):
    year_i = parse_i(year); month_i = parse_i(month)
    fd = parse_d(from_date); td = parse_d(to_date)
    conds = build_purchase_query(year_i, month_i, fd, td, party, product, q, purchase_type)
    base = select(Purchase).where(*conds) if conds else select(Purchase)
    rows = db.execute(base.order_by(Purchase.txn_date, Purchase.id)).scalars().all()

    headers = ["일자", "거래처", "제품", "품명", "공급가액", "부가세", "합계"]
    data_rows = [[
        r.txn_date.strftime("%Y-%m-%d") if r.txn_date else "",
        (r.party_name or "")[:18], (r.product_name or "")[:12], (r.item_raw or "")[:20],
        float(r.supply or 0), float(r.vat or 0), float(r.total or 0),
    ] for r in rows]
    sums = {"공급가액": sum(r[4] for r in data_rows),
            "부가세": sum(r[5] for r in data_rows),
            "합계": sum(r[6] for r in data_rows)}
    pdf_bytes = build_pdf("매입 명세서", _filter_desc(year_i, month_i, fd, td, party, product, q),
                          headers, data_rows, sums, money_cols=[4, 5, 6])
    fn = f"inviz_purchases_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{fn}"'})


# ====================== CSV 업로드 ======================
CSV_HEADERS_P = ["일자", "거래처명", "거래처코드", "제품코드", "품명",
                 "매입유형", "공급가액", "부가세", "결제수단", "비고"]


# ====================== Excel 업로드 ======================
XLSX_HEADERS_P = ["일자", "거래처명", "거래처코드", "제품코드", "품명",
                  "매입유형", "공급가액", "부가세", "결제수단", "비고"]


@router.get("/import-xlsx", response_class=HTMLResponse)
def import_xlsx_form(request: Request):
    return templates.TemplateResponse("purchases/import_xlsx.html", {
        "request": request, "headers": XLSX_HEADERS_P,
        "preview": None, "errors": None, "sheets": None,
    })


@router.get("/import-xlsx/template")
def xlsx_template_p():
    """인비즈 매입 표준 Excel 양식 다운로드"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "매입"

    purple = "6B2C91"
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws["A1"] = "인비즈 매입 일괄 등록 양식"
    ws["A1"].font = Font(name="맑은 고딕", size=14, bold=True, color=purple)
    ws.merge_cells("A1:J1")
    ws["A2"] = "2행은 안내, 3행은 헤더, 4행부터 데이터 입력. 헤더 행은 수정 금지."
    ws["A2"].font = Font(name="맑은 고딕", size=9, italic=True, color="808080")
    ws.merge_cells("A2:J2")

    fill = PatternFill("solid", fgColor=purple)
    for col_idx, h in enumerate(XLSX_HEADERS_P, 1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    example_fill = PatternFill("solid", fgColor="FFF2CC")
    examples = [
        ["2026-06-01", "예시공급사", "", "P002", "PACS 부품", "정기", 800000, 80000, "이체", ""],
        ["2026-06-02", "예시메디칼", "", "P007", "CR 장비 부속", "일회성", 1200000, 120000, "카드", "예시"],
    ]
    for ri, row in enumerate(examples, start=4):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="맑은 고딕", size=10, italic=True, color="999999")
            c.fill = example_fill
            c.border = border
            if ci in (7, 8):
                c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="right")

    widths = [12, 18, 12, 10, 30, 10, 14, 12, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A4"

    dv = DataValidation(type="list", formula1='"정기,일회성,기타"', allow_blank=True)
    dv.add("F4:F1000")
    ws.add_data_validation(dv)

    import io as _io
    buf = _io.BytesIO()
    wb.save(buf)
    fn = f"inviz_purchases_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


@router.post("/import-xlsx/preview", response_class=HTMLResponse)
async def import_xlsx_preview_p(
    request: Request,
    file: UploadFile = File(...),
    sheet_name: str = Form(""),
    header_row: int = Form(3),
):
    from helpers import load_workbook_any
    raw = await file.read()
    try:
        wb = load_workbook_any(raw)
    except Exception as e:
        return templates.TemplateResponse("purchases/import_xlsx.html", {
            "request": request, "headers": XLSX_HEADERS_P,
            "preview": None, "sheets": None,
            "errors": [f"Excel 파일을 읽을 수 없습니다: {e}"],
        })

    sheets = wb.sheetnames
    sel = sheet_name if sheet_name in sheets else sheets[0]
    ws = wb[sel]

    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = max(0, int(header_row) - 1) if header_row else 0

    def is_header(r):
        if not r: return False
        cells = [str(c).strip() if c is not None else "" for c in r]
        return any(c in ("일자", "거래처명", "공급가액") for c in cells)

    if not is_header(rows[hdr_idx] if hdr_idx < len(rows) else None):
        for i in range(min(8, len(rows))):
            if is_header(rows[i]):
                hdr_idx = i; break

    if hdr_idx >= len(rows):
        return templates.TemplateResponse("purchases/import_xlsx.html", {
            "request": request, "headers": XLSX_HEADERS_P,
            "preview": None, "sheets": sheets, "selected_sheet": sel,
            "errors": [f"헤더를 찾을 수 없습니다 (시트 '{sel}', 행 {hdr_idx + 1})"],
        })

    from helpers import normalize_header
    raw_hdr = rows[hdr_idx]
    hdr = normalize_header([str(c).strip() if c is not None else "" for c in raw_hdr])
    HOMETAX_REQUIRED = {"작성일자", "공급가액", "세액", "상호"}
    is_standard = hdr[:len(XLSX_HEADERS_P)] == XLSX_HEADERS_P
    is_hometax = HOMETAX_REQUIRED.issubset(set(hdr))
    if not (is_standard or is_hometax):
        return templates.TemplateResponse("purchases/import_xlsx.html", {
            "request": request, "headers": XLSX_HEADERS_P,
            "preview": None, "sheets": sheets, "selected_sheet": sel,
            "errors": [f"헤더 불일치 (시트 '{sel}', 행 {hdr_idx + 1})",
                       f"받은 헤더: {hdr[:10]}",
                       f"기대 헤더(표준): {XLSX_HEADERS_P}",
                       f"또는 홈택스 양식 필수 컬럼: {sorted(HOMETAX_REQUIRED)}"],
        })

    errs, preview = [], []
    if is_hometax:
        ix = {h: i for i, h in enumerate(hdr)}
        def _v(row, key):
            i = ix.get(key)
            return row[i] if (i is not None and i < len(row)) else None
        for i, row in enumerate(rows[hdr_idx + 1:], start=hdr_idx + 2):
            if not row or not any(c not in (None, "") for c in row):
                continue
            d_raw = _v(row, "작성일자")
            if isinstance(d_raw, datetime):
                dt = d_raw.date()
            else:
                dt = parse_d(str(d_raw).strip() if d_raw else "")
            if not dt:
                errs.append(f"행 {i}: 작성일자 형식 오류 ({d_raw})"); continue
            party_name = str(_v(row, "상호") or "").strip()
            if not party_name:
                errs.append(f"행 {i}: 공급자 상호 누락"); continue
            try:
                supply = float(_v(row, "공급가액") or 0)
                vat = float(_v(row, "세액") or 0)
            except Exception:
                errs.append(f"행 {i}: 공급가액/세액 숫자 오류"); continue
            preview.append({
                "row_no": i, "txn_date": dt, "party_name": party_name,
                "party_code": str(_v(row, "공급자사업자등록번호") or "").strip(),
                "product_code": "",
                "item_raw": str(_v(row, "품목명") or "").strip(),
                "purchase_type": "기타",
                "supply": supply, "vat": vat, "total": supply + vat,
                "payment_method": str(_v(row, "영수/청구 구분") or "").strip(),
                "note": str(_v(row, "비고") or "").strip(),
            })
    else:
        for i, row in enumerate(rows[hdr_idx + 1:], start=hdr_idx + 2):
            if not row or not any(c not in (None, "") for c in row):
                continue
            vals = list(row) + [None] * (len(XLSX_HEADERS_P) - len(row))
            d_raw = vals[0]
            if isinstance(d_raw, datetime):
                dt = d_raw.date()
            else:
                dt = parse_d(str(d_raw).strip() if d_raw else "")
            if not dt:
                errs.append(f"행 {i}: 일자 형식 오류 ({d_raw})"); continue
            party_name = str(vals[1]).strip() if vals[1] else ""
            if not party_name:
                errs.append(f"행 {i}: 거래처명 누락"); continue
            try:
                supply = float(vals[6]) if vals[6] not in (None, "") else 0
                vat = float(vals[7]) if vals[7] not in (None, "") else 0
            except Exception:
                errs.append(f"행 {i}: 숫자 형식 오류"); continue
            preview.append({
                "row_no": i, "txn_date": dt, "party_name": party_name,
                "party_code": str(vals[2]).strip() if vals[2] else "",
                "product_code": str(vals[3]).strip() if vals[3] else "",
                "item_raw": str(vals[4]).strip() if vals[4] else "",
                "purchase_type": str(vals[5]).strip() if vals[5] else "기타",
                "supply": supply, "vat": vat, "total": supply + vat,
                "payment_method": str(vals[8]).strip() if vals[8] else "",
                "note": str(vals[9]).strip() if vals[9] else "",
            })

    encoded = json.dumps(preview, default=str, ensure_ascii=False)
    return templates.TemplateResponse("purchases/import_xlsx.html", {
        "request": request, "headers": XLSX_HEADERS_P,
        "preview": preview, "errors": errs, "encoded": encoded,
        "sheets": sheets, "selected_sheet": sel, "header_row": hdr_idx + 1,
    })


def _record_purchase_batch(db, kind: str, ids: list, note: str = ""):
    """매입 업로드 적용 배치 기록 (되돌리기용). 배치 id 반환."""
    from models import ImportBatch
    if not ids:
        return ""
    b = ImportBatch(domain="purchase", kind=kind, count=len(ids),
                    row_ids=json.dumps(ids), note=note)
    db.add(b); db.commit(); db.refresh(b)
    return b.id


@router.post("/import-xlsx/commit")
def import_xlsx_commit_p(db: Session = Depends(get_db), payload: str = Form(...)):
    rows = json.loads(payload)
    n = 0
    ids = []
    for r in rows:
        dt = datetime.strptime(r["txn_date"], "%Y-%m-%d").date() if isinstance(r["txn_date"], str) else r["txn_date"]
        product_code = r.get("product_code") or "P999"
        prod = db.get(Product, product_code)
        product_name = prod.name if prod else "기타"
        q = (dt.month - 1) // 3 + 1
        p = Purchase(
            txn_date=dt, year=dt.year, month=dt.month,
            quarter=f"Q{q}", half="H1" if dt.month <= 6 else "H2",
            party_code=r.get("party_code") or None, party_name=r["party_name"],
            product_code=product_code, product_name=product_name,
            item_raw=r.get("item_raw") or None,
            purchase_type=r.get("purchase_type") or "기타",
            supply=r["supply"], vat=r["vat"], total=r["supply"] + r["vat"],
            payment_method=r.get("payment_method") or None,
            note=r.get("note") or None,
            source_file="web_app", source_sheet="xlsx_import",
        )
        db.add(p); db.flush()
        p.txn_id = f"P-XLSX-{p.id:06d}"
        ids.append(p.id); n += 1
    db.commit()
    bid = _record_purchase_batch(db, "xlsx", ids, note=f"Excel 업로드 {n}건")
    return RedirectResponse(f"/purchases?imported_xlsx={n}&batch={bid}", status_code=303)


@router.get("/import-csv", response_class=HTMLResponse)
def import_csv_form(request: Request):
    return templates.TemplateResponse("purchases/import_csv.html", {
        "request": request, "headers": CSV_HEADERS_P, "preview": None, "errors": None,
    })


@router.get("/import-csv/template")
def csv_template():
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(CSV_HEADERS_P)
    w.writerow(["2026-05-27", "예시공급사", "", "P002", "PACS 부품 매입", "정기", "800000", "80000", "이체", ""])
    return Response(
        content=buf.getvalue().encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="inviz_purchases_template.csv"'},
    )


@router.post("/import-csv/preview", response_class=HTMLResponse)
async def import_csv_preview(request: Request, file: UploadFile = File(...)):
    raw = await file.read()
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            text = raw.decode(enc); break
        except Exception:
            continue
    if text is None:
        return templates.TemplateResponse("purchases/import_csv.html", {
            "request": request, "headers": CSV_HEADERS_P,
            "preview": None, "errors": ["파일 인코딩을 인식할 수 없습니다."],
        })
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return templates.TemplateResponse("purchases/import_csv.html", {
            "request": request, "headers": CSV_HEADERS_P,
            "preview": None, "errors": ["빈 파일입니다."],
        })
    hdr = [c.strip() for c in rows[0]]
    if hdr[:len(CSV_HEADERS_P)] != CSV_HEADERS_P:
        return templates.TemplateResponse("purchases/import_csv.html", {
            "request": request, "headers": CSV_HEADERS_P,
            "preview": None, "errors": [f"헤더 불일치: 받은 {hdr}", f"기대 {CSV_HEADERS_P}"],
        })

    errs, preview = [], []
    for i, row in enumerate(rows[1:], start=2):
        if not any(c.strip() for c in row):
            continue
        if len(row) < len(CSV_HEADERS_P):
            row = row + [""] * (len(CSV_HEADERS_P) - len(row))
        dt = parse_d(row[0].strip())
        if not dt:
            errs.append(f"행 {i}: 일자 형식 오류 ({row[0]})"); continue
        if not row[1].strip():
            errs.append(f"행 {i}: 거래처명 누락"); continue
        try:
            supply = float(row[6] or 0); vat = float(row[7] or 0)
        except Exception:
            errs.append(f"행 {i}: 숫자 형식 오류"); continue
        preview.append({
            "row_no": i, "txn_date": dt, "party_name": row[1].strip(),
            "party_code": row[2].strip(), "product_code": row[3].strip(),
            "item_raw": row[4].strip(), "purchase_type": row[5].strip() or "기타",
            "supply": supply, "vat": vat, "total": supply + vat,
            "payment_method": row[8].strip(), "note": row[9].strip(),
        })
    encoded = json.dumps(preview, default=str, ensure_ascii=False)
    return templates.TemplateResponse("purchases/import_csv.html", {
        "request": request, "headers": CSV_HEADERS_P,
        "preview": preview, "errors": errs, "encoded": encoded,
    })


@router.post("/import-csv/commit")
def import_csv_commit(db: Session = Depends(get_db), payload: str = Form(...)):
    rows = json.loads(payload)
    n = 0
    ids = []
    for r in rows:
        dt = datetime.strptime(r["txn_date"], "%Y-%m-%d").date() if isinstance(r["txn_date"], str) else r["txn_date"]
        product_code = r.get("product_code") or "P999"
        prod = db.get(Product, product_code)
        product_name = prod.name if prod else "기타"
        q = (dt.month - 1) // 3 + 1
        p = Purchase(
            txn_date=dt, year=dt.year, month=dt.month,
            quarter=f"Q{q}", half="H1" if dt.month <= 6 else "H2",
            party_code=r.get("party_code") or None, party_name=r["party_name"],
            product_code=product_code, product_name=product_name,
            item_raw=r.get("item_raw") or None,
            purchase_type=r.get("purchase_type") or "기타",
            supply=r["supply"], vat=r["vat"], total=r["supply"] + r["vat"],
            payment_method=r.get("payment_method") or None,
            note=r.get("note") or None,
            source_file="web_app", source_sheet="csv_import",
        )
        db.add(p); db.flush()
        p.txn_id = f"P-CSV-{p.id:06d}"
        ids.append(p.id); n += 1
    db.commit()
    bid = _record_purchase_batch(db, "csv", ids, note=f"CSV 업로드 {n}건")
    return RedirectResponse(f"/purchases?imported={n}&batch={bid}", status_code=303)


@router.post("/import/undo/{batch_id}")
def import_undo_purchase(batch_id: int, db: Session = Depends(get_db)):
    """매입 업로드 적용을 행 단위로 정확히 되돌리기 — 적용 전 상태로 복원"""
    import json as _json
    from models import ImportBatch
    b = db.get(ImportBatch, batch_id)
    removed = 0
    if b and b.undone != "Y" and b.domain == "purchase":
        ids = _json.loads(b.row_ids or "[]")
        for pid in ids:
            row = db.get(Purchase, pid)
            # 안전: 웹 업로드 데이터만 삭제
            if row and row.source_file == "web_app":
                db.delete(row); removed += 1
        b.undone = "Y"
        db.commit()
    return RedirectResponse(f"/purchases?undone={removed}", status_code=303)


# ====================== 기존 CRUD ======================
@router.get("/new", response_class=HTMLResponse)
def new_form(request: Request, db: Session = Depends(get_db)):
    products = db.execute(select(Product).order_by(Product.code)).scalars().all()
    parties = db.execute(select(Party).where(Party.active == "Y").order_by(Party.name).limit(2000)).scalars().all()
    return templates.TemplateResponse("purchases/form.html", {
        "request": request, "row": None, "products": products, "parties": parties, "today": date.today(),
    })


@router.post("")
def create_purchase(
    db: Session = Depends(get_db),
    txn_date: str = Form(...), party_code: str = Form(""), party_name: str = Form(...),
    product_code: str = Form(""), item_raw: str = Form(""),
    supply: float = Form(0), vat: float = Form(0),
    purchase_type: str = Form("기타"), payment_method: str = Form(""), note: str = Form(""),
):
    dt = datetime.strptime(txn_date, "%Y-%m-%d").date()
    product_name = "기타"
    if product_code:
        prod = db.get(Product, product_code)
        product_name = prod.name if prod else "기타"
    p = Purchase(
        txn_date=dt, year=dt.year, month=dt.month,
        quarter=f"Q{(dt.month - 1) // 3 + 1}",
        half="H1" if dt.month <= 6 else "H2",
        party_code=party_code or None, party_name=party_name,
        product_code=product_code or "P999", product_name=product_name,
        item_raw=item_raw or None, purchase_type=purchase_type or "기타",
        supply=supply, vat=vat, total=supply + vat,
        payment_method=payment_method or None, note=note or None,
        source_file="web_app", source_sheet="manual",
    )
    db.add(p); db.commit(); db.refresh(p)
    p.txn_id = f"P-W-{p.id:06d}"; db.commit()
    return RedirectResponse(f"/purchases?year={dt.year}", status_code=303)


@router.get("/{pid}/edit", response_class=HTMLResponse)
def edit_form(pid: int, request: Request, db: Session = Depends(get_db), back: str = ""):
    row = db.get(Purchase, pid)
    if not row: raise HTTPException(404)
    products = db.execute(select(Product).order_by(Product.code)).scalars().all()
    parties = db.execute(select(Party).where(Party.active == "Y").order_by(Party.name).limit(2000)).scalars().all()
    safe_back = back if (back.startswith("/purchases") or back.startswith("/sales")) else ""
    return templates.TemplateResponse("purchases/form.html", {
        "request": request, "row": row, "products": products, "parties": parties,
        "today": date.today(), "back_url": safe_back,
    })


@router.post("/{pid}")
def update_purchase(
    pid: int, db: Session = Depends(get_db),
    txn_date: str = Form(...), party_code: str = Form(""), party_name: str = Form(...),
    product_code: str = Form(""), item_raw: str = Form(""),
    supply: float = Form(0), vat: float = Form(0),
    purchase_type: str = Form("기타"), payment_method: str = Form(""), note: str = Form(""),
    back: str = Form(""),
):
    row = db.get(Purchase, pid)
    if not row: raise HTTPException(404)
    dt = datetime.strptime(txn_date, "%Y-%m-%d").date()
    row.txn_date = dt; row.year = dt.year; row.month = dt.month
    row.quarter = f"Q{(dt.month - 1) // 3 + 1}"
    row.half = "H1" if dt.month <= 6 else "H2"
    row.party_code = party_code or None; row.party_name = party_name
    if product_code:
        row.product_code = product_code
        prod = db.get(Product, product_code)
        row.product_name = prod.name if prod else "기타"
    row.item_raw = item_raw or None; row.purchase_type = purchase_type or "기타"
    row.supply = supply; row.vat = vat; row.total = supply + vat
    row.payment_method = payment_method or None; row.note = note or None
    db.commit()
    if back and (back.startswith("/purchases") or back.startswith("/sales")):
        return RedirectResponse(back, status_code=303)
    return RedirectResponse(f"/purchases?year={dt.year}", status_code=303)


@router.post("/{pid}/delete")
def delete_purchase(pid: int, db: Session = Depends(get_db), back: str = Form("")):
    row = db.get(Purchase, pid)
    if row: db.delete(row); db.commit()
    if back and back.startswith("/purchases"):
        return RedirectResponse(back, status_code=303)
    return RedirectResponse("/purchases", status_code=303)
