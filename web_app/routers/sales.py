# -*- coding: utf-8 -*-
"""매출 라우터 — 기간 필터, 합계 카드, Excel/PDF 다운로드, CSV 업로드"""
import csv
import io
from datetime import date, datetime
from pathlib import Path
from typing import Optional
from fastapi import APIRouter, Request, Form, Depends, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from sqlalchemy import select, func, or_, and_
from sqlalchemy.orm import Session

from database import get_db
from helpers import templates
from models import Sale, Party, Product, ProductMapping
from export_util import build_xlsx, build_pdf

router = APIRouter()


def apply_product_mapping(item_name, db):
    if not item_name:
        return "P999", "기타"
    rules = db.execute(select(ProductMapping).order_by(ProductMapping.priority)).scalars().all()
    for r in rules:
        if r.pattern == "*":
            return r.product_code, r.product_name
        if r.pattern.lower() in item_name.lower():
            return r.product_code, r.product_name
    return "P999", "기타"


def parse_d(s):
    if not s: return None
    try: return datetime.strptime(s, "%Y-%m-%d").date()
    except: return None


def parse_i(s):
    """빈 문자열·None·잘못된 입력을 None으로 안전 변환"""
    if s is None or s == "":
        return None
    try:
        return int(s)
    except (ValueError, TypeError):
        return None


def build_sale_query(year, month, from_date, to_date, party, product, q, sale_type=""):
    """매출 조회용 공통 WHERE 절. SQLAlchemy 절들 반환."""
    conds = []
    if from_date: conds.append(Sale.txn_date >= from_date)
    if to_date: conds.append(Sale.txn_date <= to_date)
    if year: conds.append(Sale.year == year)
    if month: conds.append(Sale.month == month)
    if party: conds.append(Sale.party_name.contains(party))
    if product: conds.append(Sale.product_code == product)
    if sale_type: conds.append(Sale.sale_type == sale_type)
    if q:
        conds.append(or_(
            Sale.party_name.contains(q),
            Sale.item_raw.contains(q),
            Sale.note.contains(q),
        ))
    return conds


@router.get("", response_class=HTMLResponse)
def list_sales(
    request: Request,
    db: Session = Depends(get_db),
    year: str = "",
    month: str = "",
    from_date: str = "",
    to_date: str = "",
    party: str = "",
    product: str = "",
    q: str = "",
    sale_type: str = "",
    page: int = 1,
    per_page: int = 50,
    sort: str = "",
    dir: str = "desc",
):
    year_i = parse_i(year)
    month_i = parse_i(month)
    fd = parse_d(from_date)
    td = parse_d(to_date)
    # 설정: 검색 기본 기간 / 페이지당 건수 (쿼리 파라미터 없을 때만 적용)
    import settings_store as _ss
    if not request.query_params:
        _y, _fd, _td = _ss.default_search_filter()
        if _y: year_i = _y
        if _fd: fd = _fd
        if _td: td = _td
    if "per_page" not in request.query_params:
        per_page = _ss.get_int("search_per_page", per_page)
    conds = build_sale_query(year_i, month_i, fd, td, party, product, q, sale_type)
    base = select(Sale).where(*conds) if conds else select(Sale)

    # 합계
    total_count = db.scalar(select(func.count()).select_from(base.subquery())) or 0
    sum_q = select(
        func.coalesce(func.sum(Sale.supply), 0),
        func.coalesce(func.sum(Sale.vat), 0),
        func.coalesce(func.sum(Sale.total), 0),
    ).where(*conds) if conds else select(
        func.coalesce(func.sum(Sale.supply), 0),
        func.coalesce(func.sum(Sale.vat), 0),
        func.coalesce(func.sum(Sale.total), 0),
    )
    s_row = db.execute(sum_q).one()
    sum_supply, sum_vat, sum_total = float(s_row[0] or 0), float(s_row[1] or 0), float(s_row[2] or 0)

    # 거래처 TOP10 (필터 결과 기준)
    party_stmt = select(
        Sale.party_name, func.count().label("cnt"), func.sum(Sale.supply).label("supply")
    ).where(*conds, Sale.party_name.is_not(None)) if conds else select(
        Sale.party_name, func.count().label("cnt"), func.sum(Sale.supply).label("supply")
    ).where(Sale.party_name.is_not(None))
    party_top = db.execute(
        party_stmt.group_by(Sale.party_name).order_by(func.sum(Sale.supply).desc()).limit(10)
    ).all()

    # 제품별 (필터 결과 기준)
    prod_stmt = select(
        Sale.product_code, Sale.product_name, func.count().label("cnt"), func.sum(Sale.supply).label("supply")
    ).where(*conds) if conds else select(
        Sale.product_code, Sale.product_name, func.count().label("cnt"), func.sum(Sale.supply).label("supply")
    )
    prod_sum = db.execute(
        prod_stmt.group_by(Sale.product_code, Sale.product_name).order_by(func.sum(Sale.supply).desc())
    ).all()

    # 정렬 (사용자 지정 컬럼 + 방향)
    SORT_MAP = {
        "txn_date": Sale.txn_date, "party_name": Sale.party_name,
        "product_name": Sale.product_name, "item_raw": Sale.item_raw,
        "sale_type": Sale.sale_type, "supply": Sale.supply,
        "vat": Sale.vat, "total": Sale.total, "note": Sale.note,
    }
    col = SORT_MAP.get(sort)
    direction = "asc" if dir == "asc" else "desc"
    if col is not None:
        order_col = col.asc() if direction == "asc" else col.desc()
        order_stmt = base.order_by(order_col, Sale.id.desc())
    else:
        order_stmt = base.order_by(Sale.txn_date.desc(), Sale.id.desc())

    # 페이지 결과
    rows = db.execute(
        order_stmt.offset((page - 1) * per_page).limit(per_page)
    ).scalars().all()

    products = db.execute(select(Product).order_by(Product.code)).scalars().all()
    years = list(range(2021, datetime.now().year + 1))

    # 최근 업로드 적용 내역 (되돌리기용, 안 되돌린 것만)
    from models import ImportBatch
    recent_batches = db.execute(
        select(ImportBatch).where(ImportBatch.domain == "sale", ImportBatch.undone == "N")
        .order_by(ImportBatch.id.desc()).limit(5)
    ).scalars().all()

    # 현재 페이지 거래처 코드 → 사업자번호 매핑 (한 번에 조회)
    party_codes = list({r.party_code for r in rows if r.party_code})
    party_biz_map = {}
    if party_codes:
        for p in db.execute(select(Party).where(Party.code.in_(party_codes))).scalars():
            party_biz_map[p.code] = p.biz_no

    return templates.TemplateResponse("sales/list.html", {
        "request": request,
        "party_biz_map": party_biz_map,
        "rows": rows,
        "total_count": total_count,
        "sum_supply": sum_supply, "sum_vat": sum_vat, "sum_total": sum_total,
        "party_top": [(r[0], r[1], float(r[2] or 0)) for r in party_top],
        "prod_sum": [(r[0], r[1], r[2], float(r[3] or 0)) for r in prod_sum],
        "products": products,
        "years": years,
        "recent_batches": recent_batches,
        "filter": {"year": year_i, "month": month_i, "from_date": from_date, "to_date": to_date,
                   "party": party, "product": product, "q": q, "sale_type": sale_type,
                   "sort": sort, "dir": direction},
        "sale_types": [r[0] for r in db.execute(
            select(Sale.sale_type).where(Sale.sale_type.is_not(None))
            .group_by(Sale.sale_type).order_by(Sale.sale_type)
        ).all() if r[0]],
        "page": page, "per_page": per_page,
        "total_pages": (total_count + per_page - 1) // per_page,
        "qs": _qs(year_i, month_i, from_date, to_date, party, product, q, sale_type, sort, direction),
    })


def _qs(year, month, from_date, to_date, party, product, q, sale_type="", sort="", dir=""):
    """필터 쿼리스트링 — 페이지네이션/export 링크에 사용"""
    parts = []
    if year: parts.append(f"year={year}")
    if month: parts.append(f"month={month}")
    if from_date: parts.append(f"from_date={from_date}")
    if to_date: parts.append(f"to_date={to_date}")
    if party: parts.append(f"party={party}")
    if product: parts.append(f"product={product}")
    if sale_type: parts.append(f"sale_type={sale_type}")
    if q: parts.append(f"q={q}")
    if sort: parts.append(f"sort={sort}")
    if dir: parts.append(f"dir={dir}")
    return "&".join(parts)


# ====================== Export ======================

@router.get("/export.xlsx")
def export_xlsx(
    db: Session = Depends(get_db),
    year: str = "", month: str = "",
    from_date: str = "", to_date: str = "",
    party: str = "", product: str = "", q: str = "",
):
    year_i = parse_i(year); month_i = parse_i(month)
    fd = parse_d(from_date); td = parse_d(to_date)
    conds = build_sale_query(year_i, month_i, fd, td, party, product, q)
    base = select(Sale).where(*conds) if conds else select(Sale)
    rows = db.execute(base.order_by(Sale.txn_date, Sale.id)).scalars().all()

    headers = ["일자", "거래처", "제품코드", "제품명", "품명(원본)",
               "매출유형", "공급가액", "부가세", "합계", "비고"]
    data_rows = [[r.txn_date, r.party_name, r.product_code, r.product_name,
                  r.item_raw, r.sale_type,
                  float(r.supply or 0), float(r.vat or 0), float(r.total or 0),
                  r.note or ""] for r in rows]

    title = "인비즈 매출 명세"
    filter_desc = _filter_desc(year_i, month_i, fd, td, party, product, q)
    sums = {"공급가액": sum(r[6] for r in data_rows),
            "부가세": sum(r[7] for r in data_rows),
            "합계": sum(r[8] for r in data_rows)}
    xlsx_bytes = build_xlsx(title, filter_desc, headers, data_rows, sums,
                            money_cols=[6, 7, 8])
    fn = f"inviz_sales_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


@router.get("/export.pdf")
def export_pdf(
    db: Session = Depends(get_db),
    year: str = "", month: str = "",
    from_date: str = "", to_date: str = "",
    party: str = "", product: str = "", q: str = "",
):
    year_i = parse_i(year); month_i = parse_i(month)
    fd = parse_d(from_date); td = parse_d(to_date)
    conds = build_sale_query(year_i, month_i, fd, td, party, product, q)
    base = select(Sale).where(*conds) if conds else select(Sale)
    rows = db.execute(base.order_by(Sale.txn_date, Sale.id)).scalars().all()

    headers = ["일자", "거래처", "제품", "품명", "공급가액", "부가세", "합계"]
    data_rows = [[
        r.txn_date.strftime("%Y-%m-%d") if r.txn_date else "",
        (r.party_name or "")[:18],
        (r.product_name or "")[:12],
        (r.item_raw or "")[:20],
        float(r.supply or 0), float(r.vat or 0), float(r.total or 0),
    ] for r in rows]

    title = "매출 명세서"
    filter_desc = _filter_desc(year_i, month_i, fd, td, party, product, q)
    sums = {"공급가액": sum(r[4] for r in data_rows),
            "부가세": sum(r[5] for r in data_rows),
            "합계": sum(r[6] for r in data_rows)}
    pdf_bytes = build_pdf(title, filter_desc, headers, data_rows, sums,
                          money_cols=[4, 5, 6])
    fn = f"inviz_sales_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return Response(
        content=pdf_bytes, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


def _filter_desc(year, month, fd, td, party, product, q):
    parts = []
    if fd or td:
        parts.append(f"기간 {fd or '시작'} ~ {td or '종료'}")
    if year and not (fd or td):
        parts.append(f"{year}년" + (f" {month}월" if month else ""))
    if party: parts.append(f"거래처:{party}")
    if product: parts.append(f"제품:{product}")
    if q: parts.append(f"검색:{q}")
    return " · ".join(parts) if parts else "전체 기간"


# ====================== CSV 업로드 ======================

CSV_HEADERS = ["일자", "거래처명", "거래처코드", "제품코드", "품명",
               "매출유형", "공급가액", "부가세", "결제수단", "비고"]


# ====================== Excel 업로드 ======================

XLSX_HEADERS = ["일자", "거래처명", "거래처코드", "제품코드", "품명",
                "매출유형", "공급가액", "부가세", "결제수단", "비고"]


@router.get("/import-xlsx", response_class=HTMLResponse)
def import_xlsx_form(request: Request):
    return templates.TemplateResponse("sales/import_xlsx.html", {
        "request": request, "headers": XLSX_HEADERS,
        "preview": None, "errors": None, "sheets": None,
    })


@router.get("/import-xlsx/template")
def xlsx_template():
    """인비즈 브랜드 표준 Excel 양식 다운로드 (헤더 + 예시 행)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "매출"

    purple = "6B2C91"
    orange = "F47521"
    purple_light = "F5EDFA"
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # 안내 행
    ws["A1"] = "인비즈 매출 일괄 등록 양식"
    ws["A1"].font = Font(name="맑은 고딕", size=14, bold=True, color=purple)
    ws.merge_cells("A1:J1")
    ws["A2"] = "2행은 안내, 3행은 헤더, 4행부터 데이터 입력. 헤더 행은 수정 금지."
    ws["A2"].font = Font(name="맑은 고딕", size=9, italic=True, color="808080")
    ws.merge_cells("A2:J2")

    # 헤더
    fill = PatternFill("solid", fgColor=purple)
    for col_idx, h in enumerate(XLSX_HEADERS, 1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    # 예시 데이터 (노랑 배경)
    example_fill = PatternFill("solid", fgColor="FFF2CC")
    examples = [
        ["2026-06-01", "예시병원", "", "P001", "원격판독 6월분", "정기", 1500000, 150000, "이체", ""],
        ["2026-06-02", "예시메디칼", "", "P002", "Saintview PACS 유지보수", "정기", 500000, 50000, "이체", "예시"],
    ]
    for ri, row in enumerate(examples, start=4):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="맑은 고딕", size=10, italic=True, color="999999")
            c.fill = example_fill
            c.border = border
            if ci in (7, 8):  # 공급가액, 부가세
                c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="right")

    # 열 너비
    widths = [12, 18, 12, 10, 30, 10, 14, 12, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A4"

    # 데이터 유효성 검사 (드롭다운) — 매출유형
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"정기,신규,일회성,기타"', allow_blank=True)
    dv.add("F4:F1000")
    ws.add_data_validation(dv)

    import io
    buf = io.BytesIO()
    wb.save(buf)
    fn = f"inviz_sales_template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


@router.post("/import-xlsx/preview", response_class=HTMLResponse)
async def import_xlsx_preview(
    request: Request,
    file: UploadFile = File(...),
    sheet_name: str = Form(""),
    header_row: int = Form(3),
    db: Session = Depends(get_db),
):
    """Excel 업로드 → 시트 선택 → 미리보기"""
    from helpers import load_workbook_any
    raw = await file.read()
    try:
        wb = load_workbook_any(raw)
    except Exception as e:
        return templates.TemplateResponse("sales/import_xlsx.html", {
            "request": request, "headers": XLSX_HEADERS,
            "preview": None, "sheets": None,
            "errors": [f"Excel 파일을 읽을 수 없습니다: {e}"],
        })

    sheets = wb.sheetnames
    sel = sheet_name if sheet_name in sheets else sheets[0]
    ws = wb[sel]

    # 헤더 자동 검출 (사용자가 header_row 지정 또는 자동)
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = max(0, int(header_row) - 1) if header_row else 0
    # 자동 보정: header_row 위치에 일자/거래처명 같은 키워드 없으면 0~5행 탐색
    def is_header(r):
        if not r: return False
        cells = [str(c).strip() if c is not None else "" for c in r]
        return any(c in ("일자", "거래처명", "공급가액") for c in cells)

    if not is_header(rows[hdr_idx] if hdr_idx < len(rows) else None):
        for i in range(min(8, len(rows))):
            if is_header(rows[i]):
                hdr_idx = i; break

    if hdr_idx >= len(rows):
        return templates.TemplateResponse("sales/import_xlsx.html", {
            "request": request, "headers": XLSX_HEADERS,
            "preview": None, "sheets": sheets, "selected_sheet": sel,
            "errors": [f"헤더를 찾을 수 없습니다 (시트 '{sel}', 행 {hdr_idx + 1})"],
        })

    from helpers import normalize_header
    raw_hdr = rows[hdr_idx]
    hdr = normalize_header([str(c).strip() if c is not None else "" for c in raw_hdr])
    # 표준 헤더 / 국세청 홈택스 전자세금계산서 양식 자동 판별
    HOMETAX_REQUIRED = {"작성일자", "공급가액", "세액", "상호.1"}
    is_standard = hdr[:len(XLSX_HEADERS)] == XLSX_HEADERS
    is_hometax = HOMETAX_REQUIRED.issubset(set(hdr))
    if not (is_standard or is_hometax):
        return templates.TemplateResponse("sales/import_xlsx.html", {
            "request": request, "headers": XLSX_HEADERS,
            "preview": None, "sheets": sheets, "selected_sheet": sel,
            "errors": [f"헤더 불일치 (시트 '{sel}', 행 {hdr_idx + 1})",
                       f"받은 헤더: {hdr[:10]}",
                       f"기대 헤더(표준): {XLSX_HEADERS}",
                       f"또는 홈택스 양식 필수 컬럼: {sorted(HOMETAX_REQUIRED)}"],
        })

    errs, preview = [], []
    if is_hometax:
        ix = {h: i for i, h in enumerate(hdr)}
        def _v(row, key):
            i = ix.get(key)
            return row[i] if (i is not None and i < len(row)) else None
        for i, row in enumerate(rows[hdr_idx + 1:], start=hdr_idx + 2):
            if not row or not any(c not in (None, "") for c in row):
                continue
            d_raw = _v(row, "작성일자")
            if isinstance(d_raw, datetime):
                dt = d_raw.date()
            else:
                dt = parse_d(str(d_raw).strip() if d_raw else "")
            if not dt:
                errs.append(f"행 {i}: 작성일자 형식 오류 ({d_raw})"); continue
            party_name = str(_v(row, "상호.1") or "").strip()
            if not party_name:
                errs.append(f"행 {i}: 공급받는자 상호 누락"); continue
            try:
                supply = float(_v(row, "공급가액") or 0)
                vat = float(_v(row, "세액") or 0)
            except Exception:
                errs.append(f"행 {i}: 공급가액/세액 숫자 오류"); continue
            def _d(key):
                v = _v(row, key)
                if isinstance(v, datetime):
                    return v.date().isoformat()
                s = str(v or "").strip()
                return s[:10] if s else ""
            preview.append({
                "row_no": i, "txn_date": dt, "party_name": party_name,
                "party_code": str(_v(row, "공급받는자사업자등록번호") or "").strip(),
                "product_code": "",
                "item_raw": str(_v(row, "품목명") or "").strip(),
                "sale_type": "기타",
                "supply": supply, "vat": vat, "total": supply + vat,
                "payment_method": str(_v(row, "영수/청구 구분") or "").strip(),
                "note": str(_v(row, "비고") or "").strip(),
                # 홈택스 원본 메타 — commit 시 TaxInvoice(발급목록) 자동 생성용
                "_ht": 1,
                "_ht_approval": str(_v(row, "승인번호") or "").strip(),
                "_ht_issue_date": _d("발급일자"),
                "_ht_transmit_date": _d("전송일자"),
                "_ht_buyer_ceo": str(_v(row, "대표자명.1") or "").strip(),
                "_ht_buyer_email": str(_v(row, "이메일.1") or _v(row, "이메일") or "").strip(),
            })
    else:
        for i, row in enumerate(rows[hdr_idx + 1:], start=hdr_idx + 2):
            if not row or not any(c not in (None, "") for c in row):
                continue
            vals = list(row) + [None] * (len(XLSX_HEADERS) - len(row))
            d_raw = vals[0]
            if isinstance(d_raw, datetime):
                dt = d_raw.date()
            else:
                dt = parse_d(str(d_raw).strip() if d_raw else "")
            if not dt:
                errs.append(f"행 {i}: 일자 형식 오류 ({d_raw}) — YYYY-MM-DD"); continue
            party_name = str(vals[1]).strip() if vals[1] else ""
            if not party_name:
                errs.append(f"행 {i}: 거래처명 누락"); continue
            try:
                supply = float(vals[6]) if vals[6] not in (None, "") else 0
                vat = float(vals[7]) if vals[7] not in (None, "") else 0
            except Exception:
                errs.append(f"행 {i}: 공급가액·부가세 숫자 형식 오류 ({vals[6]}, {vals[7]})"); continue
            preview.append({
                "row_no": i, "txn_date": dt, "party_name": party_name,
                "party_code": str(vals[2]).strip() if vals[2] else "",
                "product_code": str(vals[3]).strip() if vals[3] else "",
                "item_raw": str(vals[4]).strip() if vals[4] else "",
                "sale_type": str(vals[5]).strip() if vals[5] else "기타",
                "supply": supply, "vat": vat, "total": supply + vat,
                "payment_method": str(vals[8]).strip() if vals[8] else "",
                "note": str(vals[9]).strip() if vals[9] else "",
            })

    # 🔍 중복 감지 — 기존 DB + 같은 배치 내 중복 표시
    from dedup import annotate_duplicates
    dup_stats = annotate_duplicates(db, "sale", preview)

    import json as _json
    encoded = _json.dumps(preview, default=str, ensure_ascii=False)
    return templates.TemplateResponse("sales/import_xlsx.html", {
        "request": request, "headers": XLSX_HEADERS,
        "preview": preview, "errors": errs, "encoded": encoded,
        "sheets": sheets, "selected_sheet": sel, "header_row": hdr_idx + 1,
        "dup_stats": dup_stats,
    })


def _create_hometax_invoices(db, rows) -> int:
    """홈택스 목록 Excel 업로드 → 발급목록(TaxInvoice) 자동 등록.
    승인번호가 이미 있으면 스킵 (재업로드 안전). 반환: 생성 건수."""
    from models import TaxInvoice
    created = 0
    for r in rows:
        appr = (r.get("_ht_approval") or "").strip()
        if not r.get("_ht") or not appr:
            continue
        exists = db.execute(select(TaxInvoice.id)
                            .where(TaxInvoice.invoice_no == appr).limit(1)).first()
        if exists:
            continue
        def _pd(s):
            try:
                return datetime.strptime(str(s)[:10], "%Y-%m-%d").date() if s else None
            except Exception:
                return None
        wd = _pd(r.get("txn_date"))
        supply = float(r.get("supply") or 0)
        vat = float(r.get("vat") or 0)
        inv = TaxInvoice(
            direction="sale",
            doc_kind="세금계산서" if vat else "계산서(면세)",
            invoice_no=appr,
            write_date=wd,
            issue_date=_pd(r.get("_ht_issue_date")),
            transmit_date=_pd(r.get("_ht_transmit_date")),
            buyer_corp_no=(r.get("party_code") or "").strip() or None,
            buyer_name=r.get("party_name"), party_name=r.get("party_name"),
            buyer_ceo=(r.get("_ht_buyer_ceo") or "").strip() or None,
            buyer_email=(r.get("_ht_buyer_email") or "").strip() or None,
            item_desc=(r.get("item_raw") or "").strip() or None,
            items_json=__import__("json").dumps(
                [{"월": f"{wd.month:02d}" if wd else "", "일": f"{wd.day:02d}" if wd else "",
                  "품목": (r.get("item_raw") or "").strip(),
                  "공급가액": supply, "세액": vat}], ensure_ascii=False),
            supply=supply, vat=vat, total=supply + vat,
            claim_kind=(r.get("payment_method") or "청구").strip() or "청구",
            note=(r.get("note") or "").strip() or None,
            status="sent", issue_method="hometax", source="hometax",
        )
        db.add(inv); created += 1
    if created:
        db.commit()
    return created


@router.post("/import-xlsx/commit")
def import_xlsx_commit(db: Session = Depends(get_db), payload: str = Form(...),
                       force_dup: str = Form("")):
    """Excel 미리보기 확정 적재 — 기본은 중복 행 자동 스킵."""
    import json as _json
    rows = _json.loads(payload)
    rows_all = list(rows)  # 홈택스 발급목록 생성용 (중복 스킵과 무관하게 전체)
    # commit 시점 기준으로 항상 중복 재검사 — 미리보기 이후 DB가 변했거나 payload 재전송(stale) 방지
    from dedup import annotate_duplicates
    annotate_duplicates(db, "sale", rows)
    from dedup import filter_for_commit
    keep, skipped = filter_for_commit(rows, allow_duplicates=bool(force_dup))
    rows = keep
    n = 0
    ids = []
    for r in rows:
        dt = datetime.strptime(r["txn_date"], "%Y-%m-%d").date() if isinstance(r["txn_date"], str) else r["txn_date"]
        product_code = r.get("product_code") or None
        product_name = "기타"
        if product_code:
            prod = db.get(Product, product_code)
            product_name = prod.name if prod else "기타"
        elif r.get("item_raw"):
            product_code, product_name = apply_product_mapping(r["item_raw"], db)
        q = (dt.month - 1) // 3 + 1
        s = Sale(
            txn_date=dt, year=dt.year, month=dt.month,
            quarter=f"Q{q}", half="H1" if dt.month <= 6 else "H2",
            party_code=r.get("party_code") or None,
            party_name=r["party_name"],
            product_code=product_code or "P999", product_name=product_name,
            item_raw=r.get("item_raw") or None,
            sale_type=r.get("sale_type") or "기타",
            supply=r["supply"], vat=r["vat"], total=r["supply"] + r["vat"],
            payment_method=r.get("payment_method") or None,
            note=r.get("note") or None,
            source_file="web_app", source_sheet="xlsx_import",
        )
        db.add(s); db.flush()
        s.txn_id = f"S-XLSX-{s.id:06d}"
        ids.append(s.id)
        n += 1
    db.commit()
    # 홈택스 목록 파일이면 발급목록(TaxInvoice)에도 자동 등록
    ht_created = 0
    try:
        ht_created = _create_hometax_invoices(db, rows_all)
    except Exception as e:
        print(f"[sales] 홈택스 발급목록 등록 실패: {e}")
    bid = _record_batch(db, "xlsx", ids,
                        f"Excel 업로드 {n}건 (중복 {skipped}건 스킵"
                        + (f", 계산서 {ht_created}건 발급목록 등록" if ht_created else "") + ")")
    return RedirectResponse(
        f"/sales?imported_xlsx={n}&batch={bid}&skipped={skipped}"
        + (f"&invoices={ht_created}" if ht_created else ""), status_code=303)


@router.get("/import-csv", response_class=HTMLResponse)
def import_csv_form(request: Request):
    return templates.TemplateResponse("sales/import_csv.html", {
        "request": request, "headers": CSV_HEADERS, "preview": None, "errors": None,
    })


@router.get("/import-csv/template")
def csv_template():
    """샘플 양식 CSV 다운로드"""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(CSV_HEADERS)
    w.writerow(["2026-05-27", "예시병원", "", "P001", "원격판독 5월분", "정기", "1500000", "150000", "이체", ""])
    w.writerow(["2026-05-28", "예시메디칼", "", "P002", "Saintview PACS 유지보수", "정기", "500000", "50000", "이체", ""])
    data = buf.getvalue().encode("utf-8-sig")  # BOM for Excel 한글
    return Response(
        content=data, media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="inviz_sales_template.csv"'},
    )


@router.post("/import-csv/preview", response_class=HTMLResponse)
async def import_csv_preview(request: Request, file: UploadFile = File(...),
                             db: Session = Depends(get_db)):
    raw = await file.read()
    # BOM 처리 + UTF-8/CP949 자동 감지
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            text = raw.decode(enc); break
        except Exception:
            continue
    if text is None:
        return templates.TemplateResponse("sales/import_csv.html", {
            "request": request, "headers": CSV_HEADERS, "preview": None,
            "errors": ["파일 인코딩을 인식할 수 없습니다. UTF-8 또는 CP949로 저장하세요."],
        })

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return templates.TemplateResponse("sales/import_csv.html", {
            "request": request, "headers": CSV_HEADERS, "preview": None,
            "errors": ["빈 파일입니다."],
        })

    # 헤더 검증
    hdr = [c.strip() for c in rows[0]]
    if hdr[:len(CSV_HEADERS)] != CSV_HEADERS:
        return templates.TemplateResponse("sales/import_csv.html", {
            "request": request, "headers": CSV_HEADERS, "preview": None,
            "errors": [f"헤더가 일치하지 않습니다. 받은 헤더: {hdr}", f"기대 헤더: {CSV_HEADERS}"],
        })

    # 행 검증
    errs = []
    preview = []
    for i, row in enumerate(rows[1:], start=2):
        if not any(c.strip() for c in row):
            continue
        if len(row) < len(CSV_HEADERS):
            row = row + [""] * (len(CSV_HEADERS) - len(row))
        dt = parse_d(row[0].strip())
        if not dt:
            errs.append(f"행 {i}: 일자 형식 오류 ({row[0]}) — YYYY-MM-DD")
            continue
        if not row[1].strip():
            errs.append(f"행 {i}: 거래처명 누락")
            continue
        try:
            supply = float(row[6] or 0); vat = float(row[7] or 0)
        except Exception:
            errs.append(f"행 {i}: 공급가액·부가세 숫자 형식 오류")
            continue
        preview.append({
            "row_no": i, "txn_date": dt, "party_name": row[1].strip(),
            "party_code": row[2].strip(), "product_code": row[3].strip(),
            "item_raw": row[4].strip(), "sale_type": row[5].strip() or "기타",
            "supply": supply, "vat": vat, "total": supply + vat,
            "payment_method": row[8].strip(), "note": row[9].strip(),
        })

    # 🔍 중복 감지
    from dedup import annotate_duplicates
    dup_stats = annotate_duplicates(db, "sale", preview)

    import json
    encoded = json.dumps(preview, default=str, ensure_ascii=False)
    return templates.TemplateResponse("sales/import_csv.html", {
        "request": request, "headers": CSV_HEADERS,
        "preview": preview, "errors": errs, "encoded": encoded,
        "dup_stats": dup_stats,
    })


@router.post("/import-csv/commit")
def import_csv_commit(
    db: Session = Depends(get_db),
    payload: str = Form(...),
    force_dup: str = Form(""),
):
    import json
    rows = json.loads(payload)
    # commit 시점 기준으로 항상 중복 재검사 (stale payload 방지)
    from dedup import annotate_duplicates
    annotate_duplicates(db, "sale", rows)
    from dedup import filter_for_commit
    keep, skipped = filter_for_commit(rows, allow_duplicates=bool(force_dup))
    rows = keep
    n = 0
    ids = []
    for r in rows:
        dt = datetime.strptime(r["txn_date"], "%Y-%m-%d").date() if isinstance(r["txn_date"], str) else r["txn_date"]
        product_code = r.get("product_code") or None
        product_name = "기타"
        if product_code:
            p = db.get(Product, product_code)
            product_name = p.name if p else "기타"
        elif r.get("item_raw"):
            product_code, product_name = apply_product_mapping(r["item_raw"], db)
        q = (dt.month - 1) // 3 + 1
        s = Sale(
            txn_date=dt, year=dt.year, month=dt.month,
            quarter=f"Q{q}", half="H1" if dt.month <= 6 else "H2",
            party_code=r.get("party_code") or None,
            party_name=r["party_name"],
            product_code=product_code or "P999", product_name=product_name,
            item_raw=r.get("item_raw") or None,
            sale_type=r.get("sale_type") or "기타",
            supply=r["supply"], vat=r["vat"], total=r["supply"] + r["vat"],
            payment_method=r.get("payment_method") or None,
            note=r.get("note") or None,
            source_file="web_app", source_sheet="csv_import",
        )
        db.add(s); db.flush()
        s.txn_id = f"S-CSV-{s.id:06d}"
        ids.append(s.id)
        n += 1
    db.commit()
    bid = _record_batch(db, "csv", ids, f"CSV 업로드 {n}건 (중복 {skipped}건 스킵)")
    return RedirectResponse(f"/sales?imported={n}&batch={bid}&skipped={skipped}", status_code=303)


def _record_batch(db, kind, ids, note=""):
    """업로드 적용 배치 기록 (되돌리기용). 배치 id 반환."""
    from models import ImportBatch
    if not ids:
        return ""
    b = ImportBatch(domain="sale", kind=kind, count=len(ids),
                    row_ids=__import__("json").dumps(ids), note=note)
    db.add(b); db.commit(); db.refresh(b)
    return b.id


@router.post("/import/undo/{batch_id}")
def import_undo(batch_id: int, db: Session = Depends(get_db)):
    """직전(또는 지정) 업로드 적용을 행 단위로 정확히 되돌리기 — 적용 전 상태로 복원"""
    import json as _json
    from models import ImportBatch
    b = db.get(ImportBatch, batch_id)
    removed = 0
    if b and b.undone != "Y" and b.domain == "sale":
        ids = _json.loads(b.row_ids or "[]")
        for sid in ids:
            row = db.get(Sale, sid)
            # 안전: 웹 업로드 데이터만 삭제
            if row and row.source_file == "web_app":
                db.delete(row); removed += 1
        b.undone = "Y"
        db.commit()
    return RedirectResponse(f"/sales?undone={removed}", status_code=303)


# ====================== 기존 CRUD ======================

@router.get("/new", response_class=HTMLResponse)
def new_form(request: Request, db: Session = Depends(get_db)):
    products = db.execute(select(Product).order_by(Product.code)).scalars().all()
    parties = db.execute(select(Party).where(Party.active == "Y").order_by(Party.name).limit(2000)).scalars().all()
    return templates.TemplateResponse("sales/form.html", {
        "request": request, "row": None, "products": products, "parties": parties, "today": date.today(),
    })


@router.post("")
def create_sale(
    db: Session = Depends(get_db),
    txn_date: str = Form(...), party_code: str = Form(""), party_name: str = Form(...),
    product_code: str = Form(""), item_raw: str = Form(""),
    supply: float = Form(0), vat: float = Form(0),
    sale_type: str = Form("기타"), payment_method: str = Form(""), note: str = Form(""),
):
    dt = datetime.strptime(txn_date, "%Y-%m-%d").date()
    if not product_code:
        product_code, product_name = apply_product_mapping(item_raw, db)
    else:
        prod = db.get(Product, product_code)
        product_name = prod.name if prod else "기타"
    s = Sale(
        txn_date=dt, year=dt.year, month=dt.month,
        quarter=f"Q{(dt.month - 1) // 3 + 1}",
        half="H1" if dt.month <= 6 else "H2",
        party_code=party_code or None, party_name=party_name,
        product_code=product_code, product_name=product_name,
        item_raw=item_raw or None,
        sale_type=sale_type or "기타",
        supply=supply, vat=vat, total=supply + vat,
        payment_method=payment_method or None, note=note or None,
        source_file="web_app", source_sheet="manual",
    )
    db.add(s); db.commit(); db.refresh(s)
    s.txn_id = f"S-W-{s.id:06d}"; db.commit()
    return RedirectResponse(f"/sales?year={dt.year}", status_code=303)


@router.get("/{sale_id}/edit", response_class=HTMLResponse)
def edit_form(sale_id: int, request: Request, db: Session = Depends(get_db), back: str = ""):
    row = db.get(Sale, sale_id)
    if not row: raise HTTPException(404)
    products = db.execute(select(Product).order_by(Product.code)).scalars().all()
    parties = db.execute(select(Party).where(Party.active == "Y").order_by(Party.name).limit(2000)).scalars().all()
    # 안전한 back URL: /sales 또는 /sales?... 만 허용 (external redirect 차단)
    safe_back = back if (back.startswith("/sales") or back.startswith("/purchases")) else ""
    return templates.TemplateResponse("sales/form.html", {
        "request": request, "row": row, "products": products, "parties": parties,
        "today": date.today(), "back_url": safe_back,
    })


@router.post("/{sale_id}")
def update_sale(
    sale_id: int, db: Session = Depends(get_db),
    txn_date: str = Form(...), party_code: str = Form(""), party_name: str = Form(...),
    product_code: str = Form(""), item_raw: str = Form(""),
    supply: float = Form(0), vat: float = Form(0),
    sale_type: str = Form("기타"), payment_method: str = Form(""), note: str = Form(""),
    back: str = Form(""),
):
    row = db.get(Sale, sale_id)
    if not row: raise HTTPException(404)
    dt = datetime.strptime(txn_date, "%Y-%m-%d").date()
    row.txn_date = dt; row.year = dt.year; row.month = dt.month
    row.quarter = f"Q{(dt.month - 1) // 3 + 1}"
    row.half = "H1" if dt.month <= 6 else "H2"
    row.party_code = party_code or None; row.party_name = party_name
    if product_code:
        row.product_code = product_code
        prod = db.get(Product, product_code)
        row.product_name = prod.name if prod else "기타"
    row.item_raw = item_raw or None
    row.sale_type = sale_type or "기타"
    row.supply = supply; row.vat = vat; row.total = supply + vat
    row.payment_method = payment_method or None
    row.note = note or None
    db.commit()
    # 필터 보존: back이 안전한 내부 경로면 그곳으로, 아니면 fallback
    if back and (back.startswith("/sales") or back.startswith("/purchases")):
        return RedirectResponse(back, status_code=303)
    return RedirectResponse(f"/sales?year={dt.year}", status_code=303)


@router.post("/{sale_id}/delete")
def delete_sale(sale_id: int, request: Request, db: Session = Depends(get_db), back: str = Form("")):
    row = db.get(Sale, sale_id)
    if row:
        # 🛡 User Intent Ledger — 삭제 의도 영구 기록 (재sync 시 재삽입 차단)
        try:
            from user_intent import record_deletion
            record_deletion(db, kind="sale", row=row,
                            reason="사용자 개별 삭제 (/sales UI)",
                            client_ip=request.client.host if request.client else "")
        except Exception as e:
            print(f"[sales] user_intent 기록 실패: {e}")
        db.delete(row); db.commit()
    if back and back.startswith("/sales"):
        return RedirectResponse(back, status_code=303)
    return RedirectResponse("/sales", status_code=303)
