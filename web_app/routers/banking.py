# -*- coding: utf-8 -*-
"""자금/계좌 관리 라우터 — 은행 계좌·거래내역 + 카드 한도·출금일.

- /banking                 — 계좌 현황(은행별 잔액, 카드 요약)
- /banking/transactions    — 은행 거래내역 (필터 + Excel/CSV 업로드)
- /banking/cards           — 카드 현황 (한도 사용률·출금예정일 + 이용내역 업로드)
- /banking/account/save|delete, /banking/card/save|delete — 등록/수정/삭제(설정에서 호출)
- POST /banking/upload, /banking/card-upload — 거래내역 파일 적재(중복 자동 제외)

거래내역은 엑셀/CSV 업로드로 적재. 오픈뱅킹 '기관' API 연동은 설정에 키가 있을 때만 활성(준비).
"""
import io
import csv
import hashlib
import re
from datetime import date, datetime, timedelta
from calendar import monthrange
from fastapi import APIRouter, Request, Form, Depends, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse
from sqlalchemy import select, func, or_
from sqlalchemy.orm import Session

from database import get_db
from helpers import templates
from models import BankAccount, BankTransaction, Card, CardTransaction

router = APIRouter()

BANKS = ["신한은행", "광주은행", "하나은행", "기업은행", "국민은행", "우리은행", "농협은행", "기타"]
CARD_ISSUERS = ["우리카드", "광주카드", "하나카드", "신한카드", "삼성카드", "현대카드", "기타"]
TX_CATEGORIES = ["매출", "매입", "급여", "세금", "카드대금", "임대료", "이자", "기타"]


# ───────────────────────── 파싱 유틸 ─────────────────────────
def _money(v):
    """문자열/숫자 → float (콤마·통화·괄호음수 처리)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("₩", "").replace("원", "").replace(" ", "")
    if not s or s in ("-", "."):
        return None
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True; s = s[1:-1]
    if s.startswith("-"):
        neg = True; s = s[1:]
    s = re.sub(r"[^0-9.]", "", s)
    if not s:
        return None
    try:
        f = float(s)
        return -f if neg else f
    except ValueError:
        return None


def _parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    s = s.split()[0].split("T")[0]  # 시간부 제거
    s = s.replace("/", "-").replace(".", "-")
    s = re.sub(r"-+$", "", s)
    for fmt in ("%Y-%m-%d", "%y-%m-%d", "%Y-%m-%d-", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    if re.fullmatch(r"\d{8}", s):
        try:
            return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
        except ValueError:
            return None
    return None


def _read_rows(upload: UploadFile):
    """업로드 파일을 행렬(list[list])로 읽기 (xlsx/xls/csv)."""
    name = (upload.filename or "").lower()
    upload.file.seek(0)
    data = upload.file.read()
    if name.endswith((".xlsx", ".xlsm", ".xls")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
        wb.close()
        return rows
    # CSV — 인코딩 추정
    text = None
    for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = data.decode("utf-8", "replace")
    return [row for row in csv.reader(io.StringIO(text))]


# 컬럼 키워드 (헤더 자동 인식)
_COL = {
    "date":   ["거래일시", "거래일자", "거래일", "거래날짜", "일자", "날짜", "거래 일자", "승인일", "이용일", "매출일"],
    "in":     ["입금", "맡기신금액", "입금액", "입금금액", "받으신금액"],
    "out":    ["출금", "찾으신금액", "출금액", "출금금액", "보내신금액", "이체금액"],
    "amount": ["거래금액", "금액", "거래 금액", "이용금액", "승인금액", "청구금액", "결제금액"],
    "balance": ["거래후잔액", "잔액", "잔고", "거래 후 잔액", "거래후 잔액"],
    "kind":   ["거래구분", "구분", "입출구분", "유형", "적요구분"],
    "party":  ["적요", "내용", "거래내용", "거래기록사항", "상대방", "기재내용", "거래점", "받는분", "보낸분", "의뢰인", "수취인", "가맹점", "가맹점명", "이용하신곳", "이용내역"],
    "install": ["할부", "할부개월", "할부기간"],
}


def _find_header(rows, need=("date",)):
    """헤더 행 인덱스 + {역할:열인덱스} 매핑 탐색 (상단 20행)."""
    best = None
    for ri, row in enumerate(rows[:20]):
        cells = [str(c).strip() if c is not None else "" for c in row]
        mapping = {}
        for ci, cell in enumerate(cells):
            cn = cell.replace(" ", "")
            for role, kws in _COL.items():
                if role in mapping:
                    continue
                for kw in kws:
                    if kw.replace(" ", "") in cn:
                        mapping[role] = ci
                        break
        score = len(mapping)
        if all(n in mapping for n in need) and score >= 2:
            if best is None or score > best[2]:
                best = (ri, mapping, score)
    return (best[0], best[1]) if best else (None, None)


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _hash(*parts):
    return hashlib.md5("|".join(str(p) for p in parts).encode("utf-8")).hexdigest()[:24]


# ───────────────────────── 계좌 현황(대시보드) ─────────────────────────
@router.get("", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db)):
    accounts = db.execute(select(BankAccount).order_by(
        BankAccount.active.desc(), BankAccount.sort_order, BankAccount.bank_name, BankAccount.id)).scalars().all()
    # 은행별 그룹
    by_bank = {}
    for a in accounts:
        by_bank.setdefault(a.bank_name, []).append(a)
    total_balance = sum(float(a.balance or 0) for a in accounts if a.active == "Y")

    cards = db.execute(select(Card).order_by(Card.active.desc(), Card.sort_order, Card.id)).scalars().all()
    today = date.today()
    month_start = today.replace(day=1)
    card_info = []
    for c in cards:
        used = db.scalar(select(func.coalesce(func.sum(CardTransaction.amount), 0)).where(
            CardTransaction.card_id == c.id, CardTransaction.tx_date >= month_start)) or 0
        used = float(used)
        limit = float(c.credit_limit or 0)
        pct = round(used / limit * 100, 1) if limit else 0
        nb = _next_billing(c.billing_day, today)
        card_info.append({"c": c, "used": used, "limit": limit, "pct": pct,
                          "next_billing": nb, "days_to": (nb - today).days if nb else None})

    recent = db.execute(select(BankTransaction).order_by(
        BankTransaction.tx_date.desc().nullslast(), BankTransaction.id.desc()).limit(12)).scalars().all()
    acc_name = {a.id: (a.account_alias or a.bank_name) for a in accounts}

    return templates.TemplateResponse("banking/dashboard.html", {
        "request": request, "by_bank": by_bank, "accounts": accounts,
        "total_balance": total_balance, "card_info": card_info,
        "recent": recent, "acc_name": acc_name,
        "tx_count": db.scalar(select(func.count()).select_from(BankTransaction)) or 0,
    })


def _next_billing(billing_day, today):
    if not billing_day:
        return None
    try:
        bd = int(billing_day)
    except (ValueError, TypeError):
        return None
    bd = max(1, min(bd, 31))

    def _safe(y, m, d):
        last = monthrange(y, m)[1]
        return date(y, m, min(d, last))
    cand = _safe(today.year, today.month, bd)
    if cand < today:
        ny, nm = (today.year + 1, 1) if today.month == 12 else (today.year, today.month + 1)
        cand = _safe(ny, nm, bd)
    return cand


# ───────────────────────── 은행 거래내역 ─────────────────────────
@router.get("/transactions", response_class=HTMLResponse)
def transactions(request: Request, db: Session = Depends(get_db),
                 account_id: str = "", direction: str = "", category: str = "",
                 q: str = "", date_from: str = "", date_to: str = "",
                 page: int = 1, per_page: int = 100):
    stmt = select(BankTransaction)
    aid = None
    if account_id:
        try:
            aid = int(account_id); stmt = stmt.where(BankTransaction.account_id == aid)
        except ValueError:
            pass
    if direction in ("in", "out"):
        stmt = stmt.where(BankTransaction.direction == direction)
    if category:
        stmt = stmt.where(BankTransaction.category == category)
    if q:
        stmt = stmt.where(or_(BankTransaction.counterparty.contains(q), BankTransaction.memo.contains(q)))
    df = _parse_date(date_from); dt = _parse_date(date_to)
    if df:
        stmt = stmt.where(BankTransaction.tx_date >= df)
    if dt:
        stmt = stmt.where(BankTransaction.tx_date <= dt)

    total_count = db.scalar(select(func.count()).select_from(stmt.subquery())) or 0
    sum_in = db.scalar(select(func.coalesce(func.sum(BankTransaction.amount), 0)).select_from(
        stmt.where(BankTransaction.direction == "in").subquery())) or 0
    sum_out = db.scalar(select(func.coalesce(func.sum(BankTransaction.amount), 0)).select_from(
        stmt.where(BankTransaction.direction == "out").subquery())) or 0

    rows = db.execute(stmt.order_by(
        BankTransaction.tx_date.desc().nullslast(), BankTransaction.id.desc())
        .offset((page - 1) * per_page).limit(per_page)).scalars().all()

    accounts = db.execute(select(BankAccount).order_by(BankAccount.sort_order, BankAccount.bank_name)).scalars().all()
    acc_name = {a.id: (a.account_alias and f"{a.bank_name} · {a.account_alias}" or a.bank_name) for a in accounts}

    return templates.TemplateResponse("banking/transactions.html", {
        "request": request, "rows": rows, "accounts": accounts, "acc_name": acc_name,
        "categories": TX_CATEGORIES,
        "total_count": total_count, "sum_in": float(sum_in), "sum_out": float(sum_out),
        "filter": {"account_id": account_id, "direction": direction, "category": category,
                   "q": q, "date_from": date_from, "date_to": date_to},
        "page": page, "per_page": per_page, "total_pages": (total_count + per_page - 1) // per_page,
    })


@router.post("/upload")
def upload_bank(db: Session = Depends(get_db), account_id: int = Form(...),
                upload: UploadFile = File(...)):
    acc = db.get(BankAccount, account_id)
    if not acc:
        return RedirectResponse("/banking/transactions?err=계좌를+먼저+선택하세요", status_code=303)
    try:
        rows = _read_rows(upload)
    except Exception as e:
        return RedirectResponse(f"/banking/transactions?err=파일+읽기+실패:{e}", status_code=303)
    hi, cmap = _find_header(rows, need=("date",))
    if hi is None or not ("in" in cmap or "out" in cmap or "amount" in cmap):
        return RedirectResponse(
            "/banking/transactions?err=거래일·금액+열을+찾지+못했습니다.+은행+양식+파일인지+확인하세요", status_code=303)

    existing = {r[0] for r in db.execute(
        select(BankTransaction.raw_ref).where(BankTransaction.account_id == account_id)).all()}
    new_n = dup_n = skip_n = 0
    last_date = None; last_balance = None
    for row in rows[hi + 1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        d = _parse_date(_cell(row, cmap.get("date")))
        if not d:
            skip_n += 1; continue
        v_in = _money(_cell(row, cmap.get("in")))
        v_out = _money(_cell(row, cmap.get("out")))
        amt = None; direction = None
        if v_in:
            amt = abs(v_in); direction = "in"
        elif v_out:
            amt = abs(v_out); direction = "out"
        else:
            a = _money(_cell(row, cmap.get("amount")))
            if a is None or a == 0:
                skip_n += 1; continue
            kind = str(_cell(row, cmap.get("kind")) or "").strip()
            if a < 0 or any(k in kind for k in ("출금", "이체", "지급", "결제")):
                direction = "out"; amt = abs(a)
            else:
                direction = "in"; amt = abs(a)
        bal = _money(_cell(row, cmap.get("balance")))
        party = str(_cell(row, cmap.get("party")) or "").strip()[:200] or None
        rref = _hash(account_id, d, direction, amt, bal, (party or "")[:20])
        if rref in existing:
            dup_n += 1; continue
        existing.add(rref)
        db.add(BankTransaction(
            account_id=account_id, tx_date=d, direction=direction,
            type_text=str(_cell(row, cmap.get("kind")) or "").strip()[:20] or None,
            amount=amt, balance_after=bal, counterparty=party,
            source="upload", raw_ref=rref))
        new_n += 1
        if last_date is None or d >= last_date:
            last_date = d
            if bal is not None:
                last_balance = bal
    # 최신 잔액으로 계좌 갱신
    if last_balance is not None and last_date is not None:
        if acc.balance_date is None or last_date >= acc.balance_date:
            acc.balance = last_balance; acc.balance_date = last_date
    db.commit()
    try:
        from activity import log_event
        log_event("자금", f"{acc.bank_name} 거래내역 업로드 — 신규 {new_n}·중복 {dup_n}·제외 {skip_n}")
    except Exception:
        pass
    return RedirectResponse(
        f"/banking/transactions?account_id={account_id}&up_new={new_n}&up_dup={dup_n}&up_skip={skip_n}",
        status_code=303)


# ───────────────────────── 카드 ─────────────────────────
@router.get("/cards", response_class=HTMLResponse)
def cards_page(request: Request, db: Session = Depends(get_db), card_id: str = ""):
    cards = db.execute(select(Card).order_by(Card.active.desc(), Card.sort_order, Card.id)).scalars().all()
    accounts = db.execute(select(BankAccount).order_by(BankAccount.bank_name)).scalars().all()
    acc_name = {a.id: f"{a.bank_name}{(' · ' + a.account_alias) if a.account_alias else ''}" for a in accounts}
    today = date.today()
    month_start = today.replace(day=1)
    info = []
    for c in cards:
        used = float(db.scalar(select(func.coalesce(func.sum(CardTransaction.amount), 0)).where(
            CardTransaction.card_id == c.id, CardTransaction.tx_date >= month_start)) or 0)
        limit = float(c.credit_limit or 0)
        nb = _next_billing(c.billing_day, today)
        info.append({"c": c, "used": used, "limit": limit,
                     "pct": round(used / limit * 100, 1) if limit else 0,
                     "next_billing": nb, "days_to": (nb - today).days if nb else None,
                     "pay_acc": acc_name.get(c.payment_account_id, "")})

    sel = None
    recent = []
    if card_id:
        try:
            sel = int(card_id)
            recent = db.execute(select(CardTransaction).where(CardTransaction.card_id == sel)
                                .order_by(CardTransaction.tx_date.desc().nullslast(), CardTransaction.id.desc())
                                .limit(100)).scalars().all()
        except ValueError:
            sel = None

    return templates.TemplateResponse("banking/cards.html", {
        "request": request, "info": info, "cards": cards, "accounts": accounts,
        "acc_name": acc_name, "issuers": CARD_ISSUERS, "sel_card": sel, "recent": recent,
        "card_name_map": {c.id: c.card_name for c in cards},
    })


@router.post("/card-upload")
def upload_card(db: Session = Depends(get_db), card_id: int = Form(...), upload: UploadFile = File(...)):
    card = db.get(Card, card_id)
    if not card:
        return RedirectResponse("/banking/cards?err=카드를+먼저+선택하세요", status_code=303)
    try:
        rows = _read_rows(upload)
    except Exception as e:
        return RedirectResponse(f"/banking/cards?err=파일+읽기+실패:{e}", status_code=303)
    hi, cmap = _find_header(rows, need=("date",))
    if hi is None or "amount" not in cmap:
        return RedirectResponse("/banking/cards?err=이용일·이용금액+열을+찾지+못했습니다", status_code=303)
    existing = {r[0] for r in db.execute(
        select(CardTransaction.raw_ref).where(CardTransaction.card_id == card_id)).all()}
    new_n = dup_n = skip_n = 0
    for row in rows[hi + 1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        d = _parse_date(_cell(row, cmap.get("date")))
        amt = _money(_cell(row, cmap.get("amount")))
        if not d or amt is None or amt == 0:
            skip_n += 1; continue
        merchant = str(_cell(row, cmap.get("party")) or "").strip()[:200] or None
        install = str(_cell(row, cmap.get("install")) or "").strip()[:20] or None
        rref = _hash(card_id, d, abs(amt), (merchant or "")[:24])
        if rref in existing:
            dup_n += 1; continue
        existing.add(rref)
        db.add(CardTransaction(card_id=card_id, tx_date=d, amount=abs(amt),
                               merchant=merchant, installment=install,
                               source="upload", raw_ref=rref))
        new_n += 1
    db.commit()
    try:
        from activity import log_event
        log_event("자금", f"{card.card_name} 카드 이용내역 업로드 — 신규 {new_n}·중복 {dup_n}")
    except Exception:
        pass
    return RedirectResponse(f"/banking/cards?card_id={card_id}&up_new={new_n}&up_dup={dup_n}&up_skip={skip_n}",
                            status_code=303)


# ───────────────────────── 계좌/카드 등록(설정에서 호출) ─────────────────────────
def _parse_d(s):
    return _parse_date(s)


@router.post("/account/save")
def account_save(db: Session = Depends(get_db),
                 id: str = Form(""), bank_name: str = Form(...), account_no: str = Form(""),
                 account_alias: str = Form(""), account_type: str = Form("입출금"),
                 holder: str = Form(""), balance: str = Form(""), balance_date: str = Form(""),
                 active: str = Form("Y"), note: str = Form(""), back: str = Form("/settings/banking")):
    bal = _money(balance) or 0
    if id.strip():
        a = db.get(BankAccount, int(id))
        if not a:
            raise HTTPException(404)
    else:
        a = BankAccount()
        db.add(a)
    a.bank_name = bank_name.strip()
    a.account_no = account_no.strip() or None
    a.account_alias = account_alias.strip() or None
    a.account_type = account_type.strip() or "입출금"
    a.holder = holder.strip() or None
    a.balance = bal
    bd = _parse_date(balance_date)
    if bd:
        a.balance_date = bd
    a.active = "Y" if active == "Y" else "N"
    a.note = note.strip() or None
    db.commit()
    sep = "&" if "?" in back else "?"
    return RedirectResponse(f"{back}{sep}acc_saved=1", status_code=303)


@router.post("/account/{aid}/delete")
def account_delete(aid: int, db: Session = Depends(get_db), back: str = Form("/settings/banking")):
    a = db.get(BankAccount, aid)
    if a:
        # 거래내역도 함께 삭제
        db.query(BankTransaction).filter(BankTransaction.account_id == aid).delete()
        db.delete(a); db.commit()
    sep = "&" if "?" in back else "?"
    return RedirectResponse(f"{back}{sep}acc_deleted=1", status_code=303)


@router.post("/card/save")
def card_save(db: Session = Depends(get_db),
              id: str = Form(""), card_name: str = Form(...), issuer: str = Form(""),
              card_no_last4: str = Form(""), card_type: str = Form("신용"),
              credit_limit: str = Form(""), billing_day: str = Form(""),
              payment_account_id: str = Form(""), active: str = Form("Y"),
              note: str = Form(""), back: str = Form("/settings/banking")):
    if id.strip():
        c = db.get(Card, int(id))
        if not c:
            raise HTTPException(404)
    else:
        c = Card()
        db.add(c)
    c.card_name = card_name.strip()
    c.issuer = issuer.strip() or None
    c.card_no_last4 = card_no_last4.strip() or None
    c.card_type = card_type.strip() or "신용"
    c.credit_limit = _money(credit_limit) or 0
    try:
        c.billing_day = int(billing_day) if billing_day.strip() else None
    except ValueError:
        c.billing_day = None
    try:
        c.payment_account_id = int(payment_account_id) if payment_account_id.strip() else None
    except ValueError:
        c.payment_account_id = None
    c.active = "Y" if active == "Y" else "N"
    c.note = note.strip() or None
    db.commit()
    sep = "&" if "?" in back else "?"
    return RedirectResponse(f"{back}{sep}card_saved=1", status_code=303)


@router.post("/card/{cid}/delete")
def card_delete(cid: int, db: Session = Depends(get_db), back: str = Form("/settings/banking")):
    c = db.get(Card, cid)
    if c:
        db.query(CardTransaction).filter(CardTransaction.card_id == cid).delete()
        db.delete(c); db.commit()
    sep = "&" if "?" in back else "?"
    return RedirectResponse(f"{back}{sep}card_deleted=1", status_code=303)
