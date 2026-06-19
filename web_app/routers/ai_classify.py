# -*- coding: utf-8 -*-
"""AI 자동 분류 라우터 — 파일 업로드 → LLM이 도메인·컬럼 추정 → 미리보기 → 적용

플로우:
  1) POST /ai-classify/analyze — 파일 받아 헤더+샘플 5행을 LLM에 보냄
     → 도메인(매출/매입/급여/비용) + 컬럼 매핑 + 신뢰도 JSON 반환
  2) 화면에서 사용자가 도메인·매핑 확인/수정 가능
  3) POST /ai-classify/apply — 확정된 도메인·매핑대로 DB 일괄 등록
"""
import io
import json
import re
from datetime import date, datetime
from fastapi import APIRouter, Request, Form, Depends, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from sqlalchemy.orm import Session

from database import get_db
from helpers import templates
from models import Sale, Purchase, Payroll, Expense, Product, ProductMapping
from chat_engine import ollama_chat, ollama_available, DEFAULT_MODEL

router = APIRouter()


def _ai_ready() -> bool:
    """설정된 AI 공급자(클라우드 또는 Ollama)가 사용 가능한지"""
    try:
        import llm_provider
        ok, _ = llm_provider.provider_ready()
        return ok
    except Exception:
        return ollama_available()


# ============== 도메인 스키마 ==============
DOMAIN_SCHEMAS = {
    "sale": {
        "label": "매출",
        "icon": "💰",
        "color": "purple",
        "page": "/sales",
        "fields": {
            "txn_date":      {"label": "일자",     "required": True,  "type": "date"},
            "party_name":    {"label": "거래처",   "required": True,  "type": "text"},
            "product_code":  {"label": "제품코드", "required": False, "type": "text"},
            "item_raw":      {"label": "품명",     "required": False, "type": "text"},
            "sale_type":     {"label": "매출유형", "required": False, "type": "text"},
            "supply":        {"label": "공급가액", "required": True,  "type": "number"},
            "vat":           {"label": "부가세",   "required": False, "type": "number"},
            "payment_method":{"label": "결제수단", "required": False, "type": "text"},
            "note":          {"label": "비고",     "required": False, "type": "text"},
        },
    },
    "purchase": {
        "label": "매입",
        "icon": "📦",
        "color": "orange",
        "page": "/purchases",
        "fields": {
            "txn_date":      {"label": "일자",     "required": True,  "type": "date"},
            "party_name":    {"label": "거래처",   "required": True,  "type": "text"},
            "product_code":  {"label": "제품코드", "required": False, "type": "text"},
            "item_raw":      {"label": "품명",     "required": False, "type": "text"},
            "purchase_type": {"label": "매입유형", "required": False, "type": "text"},
            "supply":        {"label": "공급가액", "required": True,  "type": "number"},
            "vat":           {"label": "부가세",   "required": False, "type": "number"},
            "payment_method":{"label": "결제수단", "required": False, "type": "text"},
            "note":          {"label": "비고",     "required": False, "type": "text"},
        },
    },
    "payroll": {
        "label": "급여",
        "icon": "👥",
        "color": "blue",
        "page": "/payroll",
        "fields": {
            "period":         {"label": "지급월 (YYYY-MM)", "required": True,  "type": "text"},
            "employee_name":  {"label": "직원명",   "required": True,  "type": "text"},
            "department":     {"label": "부서",     "required": False, "type": "text"},
            "basic":          {"label": "기본급",   "required": False, "type": "number"},
            "gross_pay":      {"label": "지급합계", "required": True,  "type": "number"},
            "total_deduction":{"label": "공제합계", "required": False, "type": "number"},
            "net_pay":        {"label": "실수령액", "required": False, "type": "number"},
            "note":           {"label": "비고",     "required": False, "type": "text"},
        },
    },
    "expense": {
        "label": "비용",
        "icon": "💳",
        "color": "red",
        "page": "/dashboard",
        "fields": {
            "use_date":       {"label": "사용일",   "required": True,  "type": "date"},
            "employee_name":  {"label": "사용자",   "required": False, "type": "text"},
            "department":     {"label": "부서",     "required": False, "type": "text"},
            "party_or_place": {"label": "거래처·장소","required": False, "type": "text"},
            "amount":         {"label": "금액",     "required": True,  "type": "number"},
            "category_main":  {"label": "대분류",   "required": False, "type": "text"},
            "category_sub":   {"label": "소분류",   "required": False, "type": "text"},
            "payment_method": {"label": "결제수단", "required": False, "type": "text"},
            "note":           {"label": "비고",     "required": False, "type": "text"},
        },
    },
}


# ============== 파일 파싱 ==============
def _parse_uploaded_file(content: bytes, filename: str):
    """업로드 파일 파싱 → (sheets_meta, parsed_rows)
    sheets_meta = [{"name": ..., "headers": [...], "row_count": N}]
    parsed_rows = {시트명: [{컬럼명: 값}, ...]}
    """
    name = filename.lower()
    if name.endswith((".xlsx", ".xlsm", ".xls")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
        result_meta, result_rows = [], {}
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            # 헤더 행 자동 감지 (처음 5행 중 비어있지 않은 셀이 가장 많은 행)
            hdr_idx = 0
            max_filled = 0
            for i, r in enumerate(rows[:5]):
                filled = sum(1 for c in (r or []) if c not in (None, ""))
                if filled > max_filled:
                    max_filled = filled
                    hdr_idx = i
            hdr = [str(c).strip() if c is not None else "" for c in (rows[hdr_idx] or [])]
            data_rows = []
            for r in rows[hdr_idx + 1:]:
                if not r or not any(c not in (None, "") for c in r):
                    continue
                d = {}
                for i, col_name in enumerate(hdr):
                    if not col_name:
                        continue
                    d[col_name] = r[i] if i < len(r) else None
                data_rows.append(d)
            result_meta.append({"name": sn, "headers": hdr, "row_count": len(data_rows), "header_row": hdr_idx + 1})
            result_rows[sn] = data_rows
        return result_meta, result_rows
    elif name.endswith(".csv"):
        import csv
        text = content.decode("utf-8-sig", errors="ignore")
        if not text.strip():
            try:
                text = content.decode("cp949", errors="ignore")
            except Exception:
                pass
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return [], {}
        hdr = [c.strip() for c in rows[0]]
        data_rows = []
        for r in rows[1:]:
            if not any(c.strip() for c in r if c):
                continue
            d = {}
            for i, col_name in enumerate(hdr):
                if not col_name:
                    continue
                d[col_name] = r[i] if i < len(r) else None
            data_rows.append(d)
        return [{"name": "CSV", "headers": hdr, "row_count": len(data_rows), "header_row": 1}], {"CSV": data_rows}
    elif name.endswith(".pdf"):
        return _parse_pdf(content)
    else:
        raise ValueError(f"지원하지 않는 형식: {filename} (xlsx/xls/csv/pdf만)")


def _parse_pdf(content: bytes):
    """PDF 파싱 → (sheets_meta, parsed_rows)
    1순위: pdfplumber로 표(table) 추출 (페이지·표별로 분리)
    2순위: pypdf 텍스트 추출 후 줄/공백 기반 행 분해
    """
    # ----- 1순위: pdfplumber 표 추출 -----
    try:
        import pdfplumber
        result_meta, result_rows = [], {}
        # 표 검출 전략: 선(line) 기반 → 실패 시 텍스트(text) 기반
        STRATEGIES = [
            None,  # 기본 (선 기반)
            {"vertical_strategy": "text", "horizontal_strategy": "text",
             "snap_y_tolerance": 5, "intersection_x_tolerance": 15},
        ]
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            tbl_no = 0
            for page_idx, page in enumerate(pdf.pages, 1):
                tables = []
                for strat in STRATEGIES:
                    try:
                        tables = page.extract_tables(table_settings=strat) if strat else page.extract_tables()
                    except Exception:
                        tables = []
                    # 의미 있는 표(2행+ 2열+)가 나오면 채택
                    if any(t and len(t) >= 2 and max(len(r or []) for r in t) >= 2 for t in tables):
                        break
                for t_idx, table in enumerate(tables, 1):
                    # 빈 표 스킵
                    if not table or len(table) < 2:
                        continue
                    # 헤더 행 자동 감지 (처음 3행 중 비어있지 않은 셀 최다)
                    hdr_idx = 0
                    max_filled = 0
                    for i, r in enumerate(table[:3]):
                        filled = sum(1 for c in (r or []) if c not in (None, ""))
                        if filled > max_filled:
                            max_filled = filled
                            hdr_idx = i
                    hdr_raw = table[hdr_idx] or []
                    hdr = []
                    for j, c in enumerate(hdr_raw):
                        nm = str(c).strip().replace("\n", " ") if c is not None else ""
                        hdr.append(nm or f"열{j+1}")
                    data_rows = []
                    for r in table[hdr_idx + 1:]:
                        if not r or not any(c not in (None, "") for c in r):
                            continue
                        d = {}
                        for j, col_name in enumerate(hdr):
                            d[col_name] = (str(r[j]).strip().replace("\n", " ")
                                           if j < len(r) and r[j] is not None else None)
                        data_rows.append(d)
                    if not data_rows:
                        continue
                    tbl_no += 1
                    sn = f"P{page_idx}-표{t_idx}"
                    result_meta.append({"name": sn, "headers": hdr,
                                        "row_count": len(data_rows), "header_row": hdr_idx + 1})
                    result_rows[sn] = data_rows
        if result_meta:
            return result_meta, result_rows
    except ImportError:
        pass
    except Exception as e:
        print(f"[ai_classify] pdfplumber 실패, 텍스트 폴백: {e}")

    # ----- 2순위: pypdf 텍스트 추출 + 줄 파싱 -----
    text = ""
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(content))
        text = "\n".join((p.extract_text() or "") for p in reader.pages)
    except Exception:
        try:
            from PyPDF2 import PdfReader as _R2
            reader = _R2(io.BytesIO(content))
            text = "\n".join((p.extract_text() or "") for p in reader.pages)
        except Exception as e:
            raise ValueError(f"PDF 텍스트 추출 실패: {e}")

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not lines:
        raise ValueError("PDF에서 표나 텍스트를 찾지 못했습니다. (스캔본 PDF는 OCR 필요)")

    # 공백 2칸+ 또는 탭으로 컬럼 분리 시도
    split_rows = []
    for ln in lines:
        cols = re.split(r"\s{2,}|\t", ln)
        cols = [c.strip() for c in cols if c.strip()]
        if len(cols) >= 2:
            split_rows.append(cols)
    if not split_rows:
        # 단일 컬럼이라도 텍스트 행으로 제공
        hdr = ["내용"]
        data_rows = [{"내용": ln} for ln in lines]
        return ([{"name": "PDF-텍스트", "headers": hdr, "row_count": len(data_rows), "header_row": 1}],
                {"PDF-텍스트": data_rows})

    # 컬럼 수가 가장 많은 행을 헤더로 추정
    max_cols = max(len(r) for r in split_rows)
    hdr_idx = next((i for i, r in enumerate(split_rows) if len(r) == max_cols), 0)
    hdr_raw = split_rows[hdr_idx]
    hdr = [c if c else f"열{j+1}" for j, c in enumerate(hdr_raw)]
    data_rows = []
    for r in split_rows[hdr_idx + 1:]:
        d = {}
        for j, col_name in enumerate(hdr):
            d[col_name] = r[j] if j < len(r) else None
        data_rows.append(d)
    return ([{"name": "PDF-텍스트", "headers": hdr, "row_count": len(data_rows), "header_row": hdr_idx + 1}],
            {"PDF-텍스트": data_rows})


# ============== LLM 도메인 추정 ==============
def _llm_classify(headers, sample_rows, model=DEFAULT_MODEL):
    """LLM에게 도메인 추정 + 컬럼 매핑 요청. JSON 반환."""
    domain_descriptions = "\n".join([
        f"- {k}: {v['label']} — 필드: {', '.join(v['fields'].keys())}"
        for k, v in DOMAIN_SCHEMAS.items()
    ])
    # 샘플 행 텍스트 (앞 5개)
    sample_text = ""
    for i, row in enumerate(sample_rows[:5], 1):
        sample_text += f"\n  행 {i}: " + ", ".join([f"{k}={v}" for k, v in row.items() if v not in (None, "")][:8])

    system = "당신은 한국 의료 IT 회사 인비즈(Inviz)의 경영관리 데이터 분류 비서입니다. JSON으로만 답하세요."
    user = f"""아래 파일의 헤더와 샘플 행을 보고 어떤 도메인인지 판단하고, 각 도메인 필드가 어느 헤더에 매핑되는지 알려주세요.

[헤더]
{headers}

[샘플]
{sample_text}

[가능한 도메인]
{domain_descriptions}

[출력 JSON 형식]
{{
  "domain": "sale|purchase|payroll|expense|unknown",
  "confidence": 0.0~1.0 사이 숫자,
  "reason": "한국어로 판단 이유 한 문장",
  "column_mapping": {{
    "도메인_필드명": "원본_헤더명 (없으면 빈 문자열)"
  }}
}}

JSON만 출력. 마크다운 코드블록 사용 금지."""

    try:
        import llm_provider
        resp = llm_provider.chat_complete(
            [{"role": "system", "content": system},
             {"role": "user", "content": user}],
            temperature=0.1, json_mode=True, max_tokens=600,
        )
        # JSON 추출 (```json ... ``` 또는 그냥 JSON)
        m = re.search(r"\{[\s\S]*\}", resp)
        if not m:
            return {"domain": "unknown", "confidence": 0, "reason": "LLM 응답에 JSON 없음", "column_mapping": {}}
        return json.loads(m.group(0))
    except Exception as e:
        return {"domain": "unknown", "confidence": 0,
                "reason": f"LLM 호출 실패: {e}", "column_mapping": {}}


# ============== 라우트 ==============
@router.get("", response_class=HTMLResponse)
@router.get("/", response_class=HTMLResponse)
def ai_classify_form(request: Request):
    return templates.TemplateResponse("ai_classify/upload.html", {
        "request": request, "schemas": DOMAIN_SCHEMAS,
        "analysis": None, "preview": None, "errors": None,
        "ollama_ok": _ai_ready(),
    })


@router.post("/analyze")
async def ai_classify_analyze(
    request: Request,
    db: Session = Depends(get_db),
    file: UploadFile = File(...),
    sheet_name: str = Form(""),
    model: str = Form(DEFAULT_MODEL),
):
    """파일 업로드 → LLM 분류 → 미리보기 화면"""
    content = await file.read()
    try:
        sheets_meta, all_rows = _parse_uploaded_file(content, file.filename)
    except Exception as e:
        return templates.TemplateResponse("ai_classify/upload.html", {
            "request": request, "schemas": DOMAIN_SCHEMAS,
            "errors": [f"파일 파싱 실패: {e}"],
            "ollama_ok": _ai_ready(),
        })

    if not sheets_meta:
        return templates.TemplateResponse("ai_classify/upload.html", {
            "request": request, "schemas": DOMAIN_SCHEMAS,
            "errors": ["빈 파일입니다."], "ollama_ok": ollama_available(),
        })

    # 시트 선택 (지정 없으면 첫 시트)
    sel_sheet = sheet_name or sheets_meta[0]["name"]
    if sel_sheet not in all_rows:
        sel_sheet = sheets_meta[0]["name"]
    headers = next((s["headers"] for s in sheets_meta if s["name"] == sel_sheet), [])
    rows = all_rows.get(sel_sheet, [])

    if not headers or not rows:
        return templates.TemplateResponse("ai_classify/upload.html", {
            "request": request, "schemas": DOMAIN_SCHEMAS,
            "errors": [f"시트 '{sel_sheet}'에 데이터가 없습니다."],
            "ollama_ok": _ai_ready(),
        })

    # LLM 호출 (설정된 AI 공급자 사용)  # _ai_ready 헬퍼는 아래 정의
    import llm_provider
    _ready, _msg = llm_provider.provider_ready()
    if not _ready:
        analysis = {
            "domain": "unknown", "confidence": 0,
            "reason": f"AI 공급자 미준비({_msg}) — 수동으로 도메인과 매핑을 선택하세요.",
            "column_mapping": {},
        }
    else:
        analysis = _llm_classify(headers, rows[:5], model=model)

    # 정규화
    domain = analysis.get("domain", "unknown")
    if domain not in DOMAIN_SCHEMAS:
        domain = "unknown"
    analysis["domain"] = domain
    mapping = analysis.get("column_mapping") or {}

    # 미리보기 데이터 (최대 50행)
    preview_rows = []
    for i, r in enumerate(rows[:50], 1):
        preview_rows.append({"_row_no": i, **{k: (v if v is not None else "") for k, v in r.items()}})

    # 전체 데이터 인코딩 (apply 단계로 전달)
    encoded = json.dumps({
        "domain": domain, "mapping": mapping, "headers": headers,
        "rows": [{k: (str(v) if isinstance(v, (date, datetime)) else v)
                  for k, v in r.items()} for r in rows],
    }, default=str, ensure_ascii=False)

    return templates.TemplateResponse("ai_classify/upload.html", {
        "request": request, "schemas": DOMAIN_SCHEMAS,
        "analysis": analysis, "preview": preview_rows,
        "headers": headers, "selected_sheet": sel_sheet,
        "sheets": sheets_meta, "total_rows": len(rows),
        "encoded": encoded, "errors": None,
        "ollama_ok": _ai_ready(),
    })


@router.post("/suggest-domain")
async def ai_classify_suggest_domain(request: Request):
    """현재 미리보기의 headers+sample을 LLM에 보내 도메인·매핑 자동 추정 → JSON 반환."""
    form = dict(await request.form())
    try:
        payload = json.loads(form.get("payload", "{}"))
        headers = payload.get("headers") or []
        rows = payload.get("rows") or []
        model = form.get("model") or DEFAULT_MODEL
        if not headers:
            return JSONResponse({"ok": False, "error": "헤더 정보가 없습니다."}, status_code=400)
        # _llm_classify는 row를 dict로 기대 → list 행이면 headers와 zip해서 변환
        sample_dicts = []
        for r in rows[:6]:
            if isinstance(r, dict):
                sample_dicts.append(r)
            elif isinstance(r, list):
                sample_dicts.append({h: v for h, v in zip(headers, r)})
        analysis = _llm_classify(headers, sample_dicts, model=model)
        domain = analysis.get("domain", "unknown")
        if domain not in DOMAIN_SCHEMAS:
            domain = "unknown"
        mapping = {}
        if domain != "unknown":
            llm_map = analysis.get("column_mapping", {}) or {}
            for field in DOMAIN_SCHEMAS[domain]["fields"]:
                v = (llm_map.get(field) or "").strip()
                mapping[field] = v if v in headers else ""
        return JSONResponse({
            "ok": True,
            "domain": domain,
            "label": DOMAIN_SCHEMAS.get(domain, {}).get("label", "미분류"),
            "icon": DOMAIN_SCHEMAS.get(domain, {}).get("icon", "❓"),
            "confidence": analysis.get("confidence", 0),
            "reason": analysis.get("reason", ""),
            "mapping": mapping,
            "model": model,
        })
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)[:300]}, status_code=500)


# ===== 멀티시트 자동 분석 + 적용 (통합 마스터처럼 시트별 도메인이 다른 파일용) =====
_MULTI_JOBS: dict = {}


def _run_multi_sheet_job(run_id: str, content: bytes, filename: str, model: str, auto_threshold: float):
    """백그라운드: 모든 시트 LLM 분류 → 신뢰도 ≥ 임계면 자동 적재."""
    import threading as _t
    import time as _time
    from database import SessionLocal
    job = _MULTI_JOBS[run_id]
    try:
        sheets_meta, all_rows = _parse_uploaded_file(content, filename)
        if not sheets_meta:
            job.update(stage="❌ 빈 파일", ok=False, finished_at=_time.time(), progress=100)
            return
        job["sheets_total"] = len(sheets_meta)
        results = []
        applied_total = {"sale": 0, "purchase": 0, "payroll": 0, "expense": 0}
        applied_errors = []
        for i, meta in enumerate(sheets_meta):
            sn = meta["name"]
            rows = all_rows.get(sn, [])
            job.update(stage=f"🤖 시트 분석: {sn}", progress=int(5 + i * 85 / len(sheets_meta)))
            if not rows or not meta["headers"]:
                results.append({"sheet": sn, "row_count": meta["row_count"], "domain": "skip",
                                "reason": "헤더 또는 데이터 없음", "applied": 0})
                continue
            analysis = _llm_classify(meta["headers"], rows[:5], model=model)
            domain = analysis.get("domain", "unknown")
            if domain not in DOMAIN_SCHEMAS:
                domain = "unknown"
            conf = analysis.get("confidence", 0)
            mapping = analysis.get("column_mapping") or {}
            entry = {
                "sheet": sn, "row_count": meta["row_count"],
                "domain": domain,
                "label": DOMAIN_SCHEMAS.get(domain, {}).get("label", "미분류"),
                "icon": DOMAIN_SCHEMAS.get(domain, {}).get("icon", "❓"),
                "confidence": conf, "reason": analysis.get("reason", ""),
                "mapping": mapping, "applied": 0,
            }
            # 자동 적재 조건: 신뢰도 ≥ 임계 + 핸들러 있는 도메인 + 필수 필드 매핑됨
            if conf >= auto_threshold and domain in DOMAIN_SCHEMAS and domain != "unknown":
                req_fields = [k for k, v in DOMAIN_SCHEMAS[domain]["fields"].items() if v.get("required")]
                missing = [f for f in req_fields if not (mapping or {}).get(f)]
                if missing:
                    entry["skipped"] = f"필수 매핑 누락: {', '.join(missing)}"
                else:
                    db = SessionLocal()
                    try:
                        fn = {"sale": _apply_sale, "purchase": _apply_purchase,
                              "payroll": _apply_payroll, "expense": _apply_expense}.get(domain)
                        if fn:
                            n, errs = fn(db, mapping, rows)
                            db.commit()
                            entry["applied"] = n
                            applied_total[domain] += n
                            if errs:
                                applied_errors.append({"sheet": sn, "errors": errs[:5]})
                    except Exception as e:
                        entry["skipped"] = f"적재 오류: {type(e).__name__}: {str(e)[:120]}"
                        try: db.rollback()
                        except Exception: pass
                    finally:
                        db.close()
            else:
                entry["skipped"] = f"신뢰도 {conf:.0%} < 임계 {auto_threshold:.0%}" if conf < auto_threshold else "도메인 미분류"
            results.append(entry)
            job["sheets_done"] = i + 1
        job.update(
            stage="✅ 완료", progress=100, ok=True,
            results=results, applied_total=applied_total, applied_errors=applied_errors,
            finished_at=_time.time(),
        )
    except Exception as e:
        import traceback
        job.update(stage="❌ 오류", ok=False, error=str(e)[:300],
                   traceback=traceback.format_exc()[:1500], finished_at=_time.time())


@router.post("/multi-sheet/start")
async def ai_classify_multi_start(
    file: UploadFile = File(...),
    model: str = Form(DEFAULT_MODEL),
    auto_threshold: float = Form(0.7),
):
    """업로드 파일의 모든 시트를 LLM이 분석 → 신뢰도 ≥ 임계면 자동 적재. run_id 반환."""
    import threading as _t
    import time as _time
    import uuid as _uuid
    content = await file.read()
    run_id = _uuid.uuid4().hex[:12]
    _MULTI_JOBS[run_id] = {
        "run_id": run_id, "started_at": _time.time(),
        "filename": file.filename, "model": model, "auto_threshold": auto_threshold,
        "stage": "🚀 파일 파싱 중", "progress": 0,
        "sheets_total": 0, "sheets_done": 0, "ok": None, "finished_at": None,
    }
    t = _t.Thread(target=_run_multi_sheet_job,
                  args=(run_id, content, file.filename, model, auto_threshold), daemon=True)
    t.start()
    return JSONResponse({"run_id": run_id})


@router.get("/multi-sheet/status")
def ai_classify_multi_status(run_id: str):
    job = _MULTI_JOBS.get(run_id)
    if not job:
        return JSONResponse({"error": "unknown_run_id"}, status_code=404)
    return JSONResponse({k: v for k, v in job.items() if k != "traceback"})


@router.post("/apply")
def ai_classify_apply(db: Session = Depends(get_db), payload: str = Form(...)):
    """확정된 도메인·매핑대로 DB 적재"""
    try:
        data = json.loads(payload)
    except Exception as e:
        return JSONResponse({"error": f"payload 파싱 실패: {e}"}, status_code=400)

    domain = data.get("domain")
    if domain not in DOMAIN_SCHEMAS:
        return JSONResponse({"error": f"지원하지 않는 도메인: {domain}"}, status_code=400)
    mapping = data.get("mapping") or {}
    rows = data.get("rows") or []

    n = 0
    errs = []
    sale_ids = []
    if domain == "sale":
        n, errs = _apply_sale(db, mapping, rows, sale_ids)
    elif domain == "purchase":
        n, errs = _apply_purchase(db, mapping, rows)
    elif domain == "payroll":
        n, errs = _apply_payroll(db, mapping, rows)
    elif domain == "expense":
        n, errs = _apply_expense(db, mapping, rows)

    db.commit()
    # 매출 AI 적용은 되돌리기 배치로 기록 (매출 페이지에서 취소 가능)
    if domain == "sale" and sale_ids:
        try:
            from models import ImportBatch
            import json as _j
            b = ImportBatch(domain="sale", kind="ai", count=len(sale_ids),
                            row_ids=_j.dumps(sale_ids), note=f"AI 분류 적용 {n}건")
            db.add(b); db.commit()
        except Exception as e:
            print(f"[ai_classify] 배치 기록 실패: {e}")
    schema = DOMAIN_SCHEMAS[domain]
    page = schema["page"]
    sep = "&" if "?" in page else "?"
    return RedirectResponse(f"{page}{sep}ai_classified={n}", status_code=303)


# ============== 도메인별 적재 ==============
def _get(row, mapping, field, default=None):
    """매핑된 헤더명으로 row에서 값 추출"""
    src = mapping.get(field, "")
    if not src:
        return default
    return row.get(src, default)


def _to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, date):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    return None


def _to_num(v):
    if v is None or v == "":
        return 0.0
    try:
        # 쉼표, 원 등 제거
        s = str(v).replace(",", "").replace("원", "").strip()
        return float(s) if s else 0.0
    except Exception:
        return 0.0


def _apply_sale(db, mapping, rows, ids_out=None):
    from routers.sales import apply_product_mapping
    n, errs = 0, []
    for i, r in enumerate(rows, 1):
        dt = _to_date(_get(r, mapping, "txn_date"))
        party_name = (_get(r, mapping, "party_name") or "").strip() if _get(r, mapping, "party_name") else ""
        supply = _to_num(_get(r, mapping, "supply"))
        if not dt or not party_name or supply == 0:
            errs.append(f"행 {i}: 일자/거래처/공급가액 누락"); continue
        vat = _to_num(_get(r, mapping, "vat"))
        product_code = (_get(r, mapping, "product_code") or "").strip() if _get(r, mapping, "product_code") else ""
        item_raw = _get(r, mapping, "item_raw") or ""
        product_name = "기타"
        if product_code:
            prod = db.get(Product, product_code)
            product_name = prod.name if prod else "기타"
        elif item_raw:
            product_code, product_name = apply_product_mapping(item_raw, db)
        q = (dt.month - 1) // 3 + 1
        s = Sale(
            txn_date=dt, year=dt.year, month=dt.month,
            quarter=f"Q{q}", half="H1" if dt.month <= 6 else "H2",
            party_name=party_name,
            product_code=product_code or "P999", product_name=product_name,
            item_raw=str(item_raw) if item_raw else None,
            sale_type=str(_get(r, mapping, "sale_type") or "기타"),
            supply=supply, vat=vat, total=supply + vat,
            payment_method=str(_get(r, mapping, "payment_method") or "") or None,
            note=str(_get(r, mapping, "note") or "") or None,
            source_file="web_app", source_sheet="ai_classify",
        )
        db.add(s); db.flush()
        s.txn_id = f"S-AI-{s.id:06d}"
        if ids_out is not None:
            ids_out.append(s.id)
        n += 1
    return n, errs


def _apply_purchase(db, mapping, rows):
    from routers.purchases import apply_product_mapping
    n, errs = 0, []
    for i, r in enumerate(rows, 1):
        dt = _to_date(_get(r, mapping, "txn_date"))
        party_name = (_get(r, mapping, "party_name") or "").strip() if _get(r, mapping, "party_name") else ""
        supply = _to_num(_get(r, mapping, "supply"))
        if not dt or not party_name or supply == 0:
            errs.append(f"행 {i}: 일자/거래처/공급가액 누락"); continue
        vat = _to_num(_get(r, mapping, "vat"))
        product_code = (_get(r, mapping, "product_code") or "").strip() if _get(r, mapping, "product_code") else ""
        item_raw = _get(r, mapping, "item_raw") or ""
        product_name = "기타"
        if product_code:
            prod = db.get(Product, product_code)
            product_name = prod.name if prod else "기타"
        elif item_raw:
            product_code, product_name = apply_product_mapping(item_raw, db)
        q = (dt.month - 1) // 3 + 1
        p = Purchase(
            txn_date=dt, year=dt.year, month=dt.month,
            quarter=f"Q{q}", half="H1" if dt.month <= 6 else "H2",
            party_name=party_name,
            product_code=product_code or "P999", product_name=product_name,
            item_raw=str(item_raw) if item_raw else None,
            purchase_type=str(_get(r, mapping, "purchase_type") or "기타"),
            supply=supply, vat=vat, total=supply + vat,
            payment_method=str(_get(r, mapping, "payment_method") or "") or None,
            note=str(_get(r, mapping, "note") or "") or None,
            source_file="web_app", source_sheet="ai_classify",
        )
        db.add(p); db.flush()
        p.txn_id = f"P-AI-{p.id:06d}"
        n += 1
    return n, errs


def _apply_payroll(db, mapping, rows):
    n, errs = 0, []
    for i, r in enumerate(rows, 1):
        period_raw = _get(r, mapping, "period")
        emp_name = (_get(r, mapping, "employee_name") or "").strip() if _get(r, mapping, "employee_name") else ""
        if not emp_name:
            errs.append(f"행 {i}: 직원명 누락"); continue
        # YYYY-MM 추출
        period = ""
        year, month = 0, 0
        if period_raw:
            s = str(period_raw).strip()
            m = re.search(r"(\d{4})[-./년\s]+(\d{1,2})", s)
            if m:
                year, month = int(m.group(1)), int(m.group(2))
                period = f"{year:04d}-{month:02d}"
        if not period:
            errs.append(f"행 {i}: 지급월 형식 오류 ({period_raw})"); continue
        gross = _to_num(_get(r, mapping, "gross_pay"))
        deduct = _to_num(_get(r, mapping, "total_deduction"))
        p = Payroll(
            period=period, year=year, month=month,
            employee_name=emp_name,
            department=str(_get(r, mapping, "department") or "") or None,
            basic=_to_num(_get(r, mapping, "basic")),
            gross_pay=gross,
            total_deduction=deduct,
            net_pay=_to_num(_get(r, mapping, "net_pay")) or (gross - deduct),
            note=str(_get(r, mapping, "note") or "") or None,
        )
        db.add(p)
        n += 1
    return n, errs


def _apply_expense(db, mapping, rows):
    n, errs = 0, []
    for i, r in enumerate(rows, 1):
        dt = _to_date(_get(r, mapping, "use_date"))
        amount = _to_num(_get(r, mapping, "amount"))
        if not dt or amount == 0:
            errs.append(f"행 {i}: 사용일/금액 누락"); continue
        q = (dt.month - 1) // 3 + 1
        e = Expense(
            use_date=dt, year=dt.year, month=dt.month, quarter=f"Q{q}",
            employee_name=str(_get(r, mapping, "employee_name") or "") or None,
            department=str(_get(r, mapping, "department") or "") or None,
            party_or_place=str(_get(r, mapping, "party_or_place") or "") or None,
            amount=amount,
            category_main=str(_get(r, mapping, "category_main") or "") or None,
            category_sub=str(_get(r, mapping, "category_sub") or "") or None,
            payment_method=str(_get(r, mapping, "payment_method") or "") or None,
            note=str(_get(r, mapping, "note") or "") or None,
        )
        db.add(e)
        n += 1
    return n, errs
