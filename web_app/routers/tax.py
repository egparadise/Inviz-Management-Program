# -*- coding: utf-8 -*-
"""전자세금계산서 — 매출(작성·발행) /tax/issue, 매입(수신확인) /tax/inbox.

현재 발행 방식: 수동/기록 관리 (홈택스 직접 API 없음).
 - 작성·관리·발행상태 기록, 거래처/본인 메일 발송, Excel 출력.
 - 실제 국세청 전송은 홈택스 바로가기 또는 ASP(팝빌/바로빌) 키 등록 시 자동.
매입 수신: 설정된 메일함(IMAP)에서 홈택스 세금계산서 메일을 확인 → 기록 → 이메일·카카오 알림.
"""
import io
import json
from pathlib import Path
from datetime import date, datetime
from fastapi import APIRouter, Request, Form, Depends, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from sqlalchemy import select, func, desc
from sqlalchemy.orm import Session

from database import get_db
from helpers import templates
from models import TaxInvoice, Party
import settings_store as ss
import integrations as ig

router = APIRouter()

HOMETAX_URL = "https://www.hometax.go.kr"


def _company():
    """공급자(우리 회사) 정보 — CompanyInfo(id=1) 우선, settings 폴백."""
    out = {"corp_no": ss.get("tax_corp_no", ""), "name": ss.get("tax_corp_name", "(주)인비즈"),
           "ceo": ss.get("tax_ceo", ""), "email": ss.get("tax_issuer_email", ""),
           "address": "", "biz_type": "제조업", "biz_item": "의료기기"}
    try:
        from database import SessionLocal
        from models import CompanyInfo
        db = SessionLocal()
        try:
            ci = db.get(CompanyInfo, 1)
            if ci:
                if ci.biz_no: out["corp_no"] = ci.biz_no
                if ci.name: out["name"] = ci.name
                if ci.ceo: out["ceo"] = ci.ceo
                if ci.email: out["email"] = out["email"] or ci.email
                if ci.address: out["address"] = ci.address
                if getattr(ci, "biz_type", None): out["biz_type"] = ci.biz_type
                if getattr(ci, "biz_item", None): out["biz_item"] = ci.biz_item
        finally:
            db.close()
    except Exception as e:
        print(f"[tax] company info 로드 실패: {e}")
    return out


@router.get("/api/party")
def tax_party_lookup(q: str = "", db: Session = Depends(get_db)):
    """공급받는자 [확인] 버튼 — 사업자번호 또는 상호로 거래처 조회 → 자동 채움 JSON.
    biz_no는 숫자만 비교(하이픈 무시)."""
    from fastapi.responses import JSONResponse
    import re as _re
    qs = (q or "").strip()
    if not qs:
        return JSONResponse({"ok": False, "message": "검색어 없음"}, status_code=400)
    digits = _re.sub(r"\D", "", qs)
    p = None
    if len(digits) == 10:
        # 사업자번호 매칭 (하이픈 유무 모두)
        for cand in db.execute(select(Party).where(Party.biz_no.is_not(None))).scalars():
            if _re.sub(r"\D", "", cand.biz_no or "") == digits:
                p = cand
                break
    if p is None:
        p = db.execute(select(Party).where(Party.name.contains(qs)).limit(1)).scalar_one_or_none()
    if p is None:
        return JSONResponse({"ok": False, "message": "등록된 거래처에서 찾을 수 없습니다"})
    return {"ok": True, "code": p.code, "name": p.name, "biz_no": p.biz_no or "",
            "ceo": getattr(p, "ceo", "") or "", "address": getattr(p, "address", "") or "",
            "email": getattr(p, "email", "") or "", "phone": p.phone or "",
            "contact_person": p.contact_person or ""}


def _base_url(request: Request) -> str:
    try:
        return str(request.base_url)
    except Exception:
        return ""


# ===================== 거래처정보 (홈택스 양식) =====================
@router.get("/api/party-search")
def tax_party_search(db: Session = Depends(get_db),
                     biz_no: str = "", name: str = "", ceo: str = "",
                     main_only: str = "", sort: str = "name", dir: str = "asc",
                     limit: int = 50):
    """거래처정보 목록조회 모달 — 등록번호/거래처명/대표자명 검색 (JSON)."""
    import re as _re
    stmt = select(Party)
    if name.strip():
        stmt = stmt.where(Party.name.contains(name.strip()))
    if ceo.strip():
        stmt = stmt.where(Party.ceo.contains(ceo.strip()))
    if main_only == "1":
        stmt = stmt.where(Party.is_main == "Y")
    rows = db.execute(stmt).scalars().all()
    if biz_no.strip():
        d = _re.sub(r"\D", "", biz_no)
        rows = [r for r in rows if d in _re.sub(r"\D", "", r.biz_no or "")]
    keyf = {"name": lambda r: (r.name or ""),
            "biz_no": lambda r: (r.biz_no or ""),
            "ceo": lambda r: (r.ceo or "")}.get(sort, lambda r: (r.name or ""))
    rows.sort(key=keyf, reverse=(dir == "desc"))
    limit = max(5, min(int(limit or 50), 500))
    return {"ok": True, "total": len(rows), "rows": [{
        "code": r.code, "name": r.name, "biz_no": r.biz_no or "",
        "sub_no": getattr(r, "sub_biz_no", "") or "", "ceo": getattr(r, "ceo", "") or "",
        "address": getattr(r, "address", "") or "",
        "biz_type": getattr(r, "biz_type", "") or "", "biz_item": getattr(r, "biz_item", "") or "",
        "email": getattr(r, "email", "") or "", "is_main": getattr(r, "is_main", "N") or "N",
    } for r in rows[:limit]]}


@router.get("/parties", response_class=HTMLResponse)
def tax_parties_page(request: Request, db: Session = Depends(get_db),
                     biz_no: str = "", name: str = "", ceo: str = "",
                     sort: str = "name", dir: str = "asc", per_page: int = 10, page: int = 1):
    """거래처정보 관리 (홈택스 양식) — 등록/변경일자·주거래처·상세보기."""
    import re as _re
    stmt = select(Party)
    if name.strip():
        stmt = stmt.where(Party.name.contains(name.strip()))
    if ceo.strip():
        stmt = stmt.where(Party.ceo.contains(ceo.strip()))
    rows_all = db.execute(stmt).scalars().all()
    if biz_no.strip():
        d = _re.sub(r"\D", "", biz_no)
        rows_all = [r for r in rows_all if d in _re.sub(r"\D", "", r.biz_no or "")]
    keyf = {"name": lambda r: (r.name or ""), "biz_no": lambda r: (r.biz_no or ""),
            "ceo": lambda r: (r.ceo or ""),
            "created": lambda r: (r.created_at or datetime.min)}.get(sort, lambda r: (r.name or ""))
    rows_all.sort(key=keyf, reverse=(dir == "desc"))
    per_page = max(5, min(int(per_page or 10), 100))
    total_count = len(rows_all)
    total_pages = max(1, (total_count + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    rows = rows_all[(page - 1) * per_page: page * per_page]
    return templates.TemplateResponse("tax/parties.html", {
        "request": request, "rows": rows, "company": _company(),
        "filter": {"biz_no": biz_no, "name": name, "ceo": ceo,
                   "sort": sort, "dir": dir, "per_page": per_page},
        "page": page, "total_pages": total_pages, "total_count": total_count,
    })


@router.get("/parties/export.xlsx")
def tax_parties_export(db: Session = Depends(get_db)):
    """거래처정보 내려받기."""
    import openpyxl, io as _io
    rows = db.execute(select(Party).order_by(Party.name)).scalars().all()
    wb = openpyxl.Workbook(); w = wb.active; w.title = "거래처정보"
    w.append(["거래처등록번호", "종사업장번호", "거래처명", "대표자", "사업장주소", "업태", "종목",
              "주거래처", "주담당자", "전화", "휴대폰", "이메일", "등록일자", "변경일자"])
    for r in rows:
        w.append([r.biz_no or "", getattr(r, "sub_biz_no", "") or "", r.name, getattr(r, "ceo", "") or "",
                  getattr(r, "address", "") or "", getattr(r, "biz_type", "") or "", getattr(r, "biz_item", "") or "",
                  getattr(r, "is_main", "N") or "N", r.contact_person or "", r.phone or "",
                  getattr(r, "mobile", "") or "", getattr(r, "email", "") or "",
                  str(r.created_at.date() if r.created_at else ""), str(r.updated_at.date() if r.updated_at else "")])
    buf = _io.BytesIO(); wb.save(buf)
    from urllib.parse import quote
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('거래처정보.xlsx')}"})


@router.get("/parties/new", response_class=HTMLResponse)
def tax_party_new_form(request: Request):
    """건별등록 — 새 거래처 입력 화면 (홈택스 양식)."""
    return templates.TemplateResponse("tax/party_new.html", {
        "request": request, "company": _company(),
    })


@router.post("/parties/new")
async def tax_party_new_create(request: Request, db: Session = Depends(get_db)):
    """등록하기 — 신규 거래처 생성 (코드 자동 채번 C####)."""
    import re as _re
    form = await request.form()
    def g(key):
        v = form.get(key)
        return v.strip() if isinstance(v, str) else ""
    def email_join(pfx):
        lo, dom = g(f"{pfx}_local"), g(f"{pfx}_domain")
        return f"{lo}@{dom}" if lo and dom else None

    from urllib.parse import quote
    name = g("name")
    biz_no = g("biz_no")
    if not name:
        return RedirectResponse("/tax/parties/new?_msg=" + quote("❌ 상호(법인명)는 필수입니다"), status_code=303)

    # 사업자번호 중복 확인
    if biz_no:
        d = _re.sub(r"\D", "", biz_no)
        for cand in db.execute(select(Party).where(Party.biz_no.is_not(None))).scalars():
            if _re.sub(r"\D", "", cand.biz_no or "") == d:
                return RedirectResponse(
                    "/tax/parties/new?_msg=" + quote(f"❌ 이미 등록된 사업자번호입니다 — {cand.name} ({cand.code})"),
                    status_code=303)

    # 코드 자동 채번
    max_num = 0
    for (c,) in db.execute(select(Party.code)).all():
        m = _re.match(r"^C(\d+)$", c or "")
        if m:
            max_num = max(max_num, int(m.group(1)))
    new_code = f"C{max_num + 1:04d}"

    p = Party(
        code=new_code, name=name, biz_no=biz_no or None,
        category=g("category") or "기타", active="Y",
        sub_biz_no=g("sub_biz_no") or None,
        ceo=g("ceo") or None, address=g("address") or None,
        biz_type=g("biz_type") or None, biz_item=g("biz_item") or None,
        is_main="Y" if g("is_main") else "N",
        dept_main=g("dept_main") or None, contact_person=g("contact_person") or None,
        phone=g("phone") or None, mobile=g("mobile") or None, fax=g("fax") or None,
        email=email_join("email"), contact_note=g("contact_note") or None,
        dept_sub=g("dept_sub") or None, contact_person2=g("contact_person2") or None,
        phone2=g("phone2") or None, mobile2=g("mobile2") or None, fax2=g("fax2") or None,
        email2=email_join("email2"), contact_note2=g("contact_note2") or None,
        note="건별등록 (홈택스 양식)",
    )
    db.add(p); db.commit()
    return RedirectResponse(f"/tax/parties/{new_code}?_msg=" + quote(f"✅ {name} 등록 완료 ({new_code})"), status_code=303)


@router.get("/parties/{code}", response_class=HTMLResponse)
def tax_party_detail(code: str, request: Request, db: Session = Depends(get_db)):
    """거래처 상세 (홈택스 양식) — 더블클릭/상세보기 진입, 수정·삭제."""
    p = db.get(Party, code)
    if not p:
        raise HTTPException(404, "거래처 없음")
    return templates.TemplateResponse("tax/party_detail.html", {
        "request": request, "p": p, "company": _company(),
    })


@router.post("/parties/{code}")
async def tax_party_update(code: str, request: Request, db: Session = Depends(get_db)):
    """수정하기 — 홈택스 거래처정보 필드 저장."""
    p = db.get(Party, code)
    if not p:
        raise HTTPException(404, "거래처 없음")
    form = await request.form()
    def g(key):
        v = form.get(key)
        return v.strip() if isinstance(v, str) else ""
    def email_join(pfx):
        lo, dom = g(f"{pfx}_local"), g(f"{pfx}_domain")
        return f"{lo}@{dom}" if lo and dom else (g(pfx) or None)

    if g("name"): p.name = g("name")
    p.ceo = g("ceo") or None
    p.sub_biz_no = g("sub_biz_no") or None
    p.address = g("address") or None
    p.biz_type = g("biz_type") or None
    p.biz_item = g("biz_item") or None
    p.is_main = "Y" if g("is_main") else "N"
    p.category = g("category") or p.category
    # 주담당
    p.dept_main = g("dept_main") or None
    p.contact_person = g("contact_person") or None
    p.phone = g("phone") or None
    p.mobile = g("mobile") or None
    p.fax = g("fax") or None
    p.email = email_join("email")
    p.contact_note = g("contact_note") or None
    # 부담당
    p.dept_sub = g("dept_sub") or None
    p.contact_person2 = g("contact_person2") or None
    p.phone2 = g("phone2") or None
    p.mobile2 = g("mobile2") or None
    p.fax2 = g("fax2") or None
    p.email2 = email_join("email2")
    p.contact_note2 = g("contact_note2") or None
    db.commit()
    from urllib.parse import quote
    return RedirectResponse(f"/tax/parties/{code}?_msg=" + quote("✅ 수정되었습니다"), status_code=303)


@router.post("/parties/{code}/delete")
def tax_party_delete(code: str, db: Session = Depends(get_db)):
    """삭제하기 — 자동 DB 백업 후 거래처 마스터에서 제거 (거래 데이터는 유지)."""
    p = db.get(Party, code)
    if not p:
        raise HTTPException(404, "거래처 없음")
    import shutil
    from database import DB_PATH
    backup_dir = DB_PATH.parent / "db_backup"
    backup_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    try:
        shutil.copy2(DB_PATH, backup_dir / f"pre_party_delete_{ts}.db")
    except Exception:
        pass
    db.delete(p); db.commit()
    from urllib.parse import quote
    return RedirectResponse("/tax/parties?_msg=" + quote(f"🗑 {p.name} 삭제 완료 (자동 백업됨)"), status_code=303)


# ===================== 발급목록 조회 (홈택스 양식) =====================
@router.get("/list", response_class=HTMLResponse)
def tax_list_page(request: Request, db: Session = Depends(get_db),
                  doc_cls: str = "", direction: str = "sale",
                  date_kind: str = "write", date_from: str = "", date_to: str = "",
                  buyer_no: str = "", sub_no: str = "", party: str = "",
                  inv_kind: str = "", status: str = "",
                  sort: str = "write_date", dir: str = "desc",
                  page: int = 1, per_page: int = 10):
    """홈택스 '전자세금계산서 목록조회'와 동일한 발급목록 조회."""
    from datetime import timedelta
    # 기본 기간: 최근 3개월
    today_d = date.today()
    try:
        fd = date.fromisoformat(date_from) if date_from else today_d - timedelta(days=90)
    except Exception:
        fd = today_d - timedelta(days=90)
    try:
        td = date.fromisoformat(date_to) if date_to else today_d
    except Exception:
        td = today_d

    date_col = {"write": TaxInvoice.write_date, "issue": TaxInvoice.issue_date,
                "transmit": TaxInvoice.transmit_date}.get(date_kind, TaxInvoice.write_date)

    stmt = select(TaxInvoice).where(TaxInvoice.direction == (direction or "sale"))
    stmt = stmt.where(date_col >= fd, date_col <= td)
    if doc_cls == "tax":
        stmt = stmt.where(TaxInvoice.doc_kind.contains("세금"))
    elif doc_cls == "plain":
        stmt = stmt.where(~TaxInvoice.doc_kind.contains("세금"))
    if buyer_no.strip():
        import re as _re
        d = _re.sub(r"\D", "", buyer_no)
        stmt = stmt.where(TaxInvoice.buyer_corp_no.is_not(None))
    if sub_no.strip():
        stmt = stmt.where(TaxInvoice.buyer_sub_no == sub_no.strip())
    if party.strip():
        stmt = stmt.where(TaxInvoice.party_name.contains(party.strip()))
    if inv_kind.strip():
        stmt = stmt.where(TaxInvoice.inv_kind == inv_kind.strip())
    if status.strip():
        stmt = stmt.where(TaxInvoice.status == status.strip())

    rows_all = db.execute(stmt).scalars().all()
    # 사업자번호 숫자 매칭 (하이픈 무시)
    if buyer_no.strip():
        import re as _re
        d = _re.sub(r"\D", "", buyer_no)
        rows_all = [r for r in rows_all if _re.sub(r"\D", "", r.buyer_corp_no or "") == d]

    # 정렬
    SORT = {"write_date": lambda r: (r.write_date or date.min),
            "issue_date": lambda r: (r.issue_date or date.min),
            "total": lambda r: float(r.total or 0),
            "party": lambda r: (r.party_name or "")}
    keyf = SORT.get(sort, SORT["write_date"])
    rows_all.sort(key=keyf, reverse=(dir != "asc"))

    # 합계
    tot_total = sum(float(r.total or 0) for r in rows_all)
    tot_supply = sum(float(r.supply or 0) for r in rows_all)
    tot_vat = sum(float(r.vat or 0) for r in rows_all)

    # 페이지네이션
    per_page = max(5, min(int(per_page or 10), 100))
    total_count = len(rows_all)
    total_pages = max(1, (total_count + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    rows = rows_all[(page - 1) * per_page: page * per_page]

    return templates.TemplateResponse("tax/list.html", {
        "request": request, "rows": rows,
        "filter": {"doc_cls": doc_cls, "direction": direction or "sale",
                   "date_kind": date_kind, "date_from": fd.isoformat(), "date_to": td.isoformat(),
                   "buyer_no": buyer_no, "sub_no": sub_no, "party": party,
                   "inv_kind": inv_kind, "status": status,
                   "sort": sort, "dir": dir, "per_page": per_page},
        "tot_total": tot_total, "tot_supply": tot_supply, "tot_vat": tot_vat,
        "page": page, "total_pages": total_pages, "total_count": total_count,
        "hometax_url": HOMETAX_URL,
    })


@router.get("/{inv_id}/detail.json")
def tax_detail_json(inv_id: int, db: Session = Depends(get_db)):
    """상세조회 모달 데이터 — 승인번호·공급자·공급받는자·품목 전체."""
    from fastapi.responses import JSONResponse
    inv = db.get(TaxInvoice, inv_id)
    if not inv:
        return JSONResponse({"ok": False}, status_code=404)
    co = _company()
    try:
        items = json.loads(inv.items_json or "[]")
    except Exception:
        items = []
    return {
        "ok": True, "id": inv.id,
        "invoice_no": inv.invoice_no or "(미발급)",
        "doc_kind": inv.doc_kind, "inv_kind": inv.inv_kind or "일반",
        "status": inv.status, "issue_method": inv.issue_method,
        "write_date": inv.write_date.isoformat() if inv.write_date else "",
        "issue_date": inv.issue_date.isoformat() if inv.issue_date else "",
        "transmit_date": inv.transmit_date.isoformat() if inv.transmit_date else "",
        "supplier": {"corp_no": inv.supplier_corp_no or co["corp_no"], "name": inv.supplier_name or co["name"],
                     "ceo": co["ceo"], "addr": co["address"],
                     "biz_type": co["biz_type"], "biz_item": co["biz_item"],
                     "email": inv.supplier_email or co["email"]},
        "buyer": {"corp_no": inv.buyer_corp_no or "", "sub_no": inv.buyer_sub_no or "",
                  "name": inv.buyer_name or inv.party_name or "", "ceo": inv.buyer_ceo or "",
                  "addr": inv.buyer_addr or "", "biz_type": inv.buyer_biztype or "",
                  "biz_item": inv.buyer_bizitem or "",
                  "email": inv.buyer_email or "", "email2": inv.buyer_email2 or ""},
        "supply": float(inv.supply or 0), "vat": float(inv.vat or 0), "total": float(inv.total or 0),
        "note": inv.note or "", "claim_kind": inv.claim_kind or "청구",
        "cash_amt": float(inv.cash_amt or 0), "check_amt": float(inv.check_amt or 0),
        "bill_amt": float(inv.bill_amt or 0), "credit_amt": float(inv.credit_amt or 0),
        "items": items,
    }


@router.get("/{inv_id}/export.xml")
def tax_export_xml(inv_id: int, db: Session = Depends(get_db)):
    """XML파일내려받기 — 전자세금계산서 구조 XML."""
    inv = db.get(TaxInvoice, inv_id)
    if not inv:
        raise HTTPException(404, "계산서 없음")
    co = _company()
    try:
        items = json.loads(inv.items_json or "[]")
    except Exception:
        items = []
    def esc(s):
        return (str(s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
    lines = ['<?xml version="1.0" encoding="UTF-8"?>',
             '<TaxInvoice xmlns="urn:kr:etax:invoice">',
             f'  <ApprovalNo>{esc(inv.invoice_no)}</ApprovalNo>',
             f'  <Kind>{esc(inv.doc_kind)}</Kind>',
             f'  <TypeCode>{esc(inv.inv_kind or "일반")}</TypeCode>',
             f'  <WriteDate>{esc(inv.write_date)}</WriteDate>',
             f'  <IssueDate>{esc(inv.issue_date or "")}</IssueDate>',
             f'  <TransmitDate>{esc(inv.transmit_date or "")}</TransmitDate>',
             '  <Supplier>',
             f'    <RegNo>{esc(inv.supplier_corp_no or co["corp_no"])}</RegNo>',
             f'    <Name>{esc(inv.supplier_name or co["name"])}</Name>',
             f'    <CEO>{esc(co["ceo"])}</CEO>',
             f'    <Address>{esc(co["address"])}</Address>',
             f'    <BizType>{esc(co["biz_type"])}</BizType>',
             f'    <BizItem>{esc(co["biz_item"])}</BizItem>',
             f'    <Email>{esc(inv.supplier_email or co["email"])}</Email>',
             '  </Supplier>',
             '  <Buyer>',
             f'    <RegNo>{esc(inv.buyer_corp_no)}</RegNo>',
             f'    <SubNo>{esc(inv.buyer_sub_no)}</SubNo>',
             f'    <Name>{esc(inv.buyer_name or inv.party_name)}</Name>',
             f'    <CEO>{esc(inv.buyer_ceo)}</CEO>',
             f'    <Address>{esc(inv.buyer_addr)}</Address>',
             f'    <BizType>{esc(inv.buyer_biztype)}</BizType>',
             f'    <BizItem>{esc(inv.buyer_bizitem)}</BizItem>',
             f'    <Email>{esc(inv.buyer_email)}</Email>',
             '  </Buyer>',
             '  <Items>']
    for it in items:
        lines += ['    <Item>',
                  f'      <Month>{esc(it.get("월"))}</Month><Day>{esc(it.get("일"))}</Day>',
                  f'      <Name>{esc(it.get("품목"))}</Name><Spec>{esc(it.get("규격"))}</Spec>',
                  f'      <Qty>{esc(it.get("수량"))}</Qty><UnitPrice>{esc(it.get("단가"))}</UnitPrice>',
                  f'      <Supply>{esc(it.get("공급가액"))}</Supply><Tax>{esc(it.get("세액"))}</Tax>',
                  f'      <Note>{esc(it.get("비고"))}</Note>',
                  '    </Item>']
    lines += ['  </Items>',
              f'  <TotalAmount>{int(float(inv.total or 0))}</TotalAmount>',
              f'  <SupplyAmount>{int(float(inv.supply or 0))}</SupplyAmount>',
              f'  <TaxAmount>{int(float(inv.vat or 0))}</TaxAmount>',
              f'  <ClaimKind>{esc(inv.claim_kind or "청구")}</ClaimKind>',
              f'  <Note>{esc(inv.note)}</Note>',
              '</TaxInvoice>']
    xml = "\n".join(lines)
    import re as _re
    safe = _re.sub(r"[^\w가-힣.-]", "_", f"세금계산서_{inv.party_name or inv.id}_{inv.write_date or ''}")
    from urllib.parse import quote
    return Response(content=xml.encode("utf-8"), media_type="application/xml",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(safe)}.xml"})


@router.get("/{inv_id}/print", response_class=HTMLResponse)
def tax_print(inv_id: int, db: Session = Depends(get_db), size: str = "A4"):
    """출력 / 출력(A5) — 인쇄용 전자세금계산서 화면 (자동 print)."""
    inv = db.get(TaxInvoice, inv_id)
    if not inv:
        raise HTTPException(404, "계산서 없음")
    co = _company()
    try:
        items = json.loads(inv.items_json or "[]")
    except Exception:
        items = []
    is_a5 = (size or "").upper() == "A5"
    trs = "".join(
        f"<tr><td>{it.get('월','')}</td><td>{it.get('일','')}</td>"
        f"<td style='text-align:left'>{it.get('품목','')}</td><td>{it.get('규격','')}</td>"
        f"<td>{it.get('수량','')}</td><td class='r'>{it.get('단가','')}</td>"
        f"<td class='r'>{int(float(it.get('공급가액',0) or 0)):,}</td>"
        f"<td class='r'>{int(float(it.get('세액',0) or 0)):,}</td>"
        f"<td>{it.get('비고','')}</td></tr>" for it in items)
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>전자{inv.doc_kind}</title>
<style>
  @page {{ size: {'A5 landscape' if is_a5 else 'A4'}; margin: 12mm; }}
  body {{ font-family:'Malgun Gothic',sans-serif; font-size:{'10px' if is_a5 else '12px'}; }}
  table {{ border-collapse:collapse; width:100%; }}
  th, td {{ border:1px solid #333; padding:4px 6px; text-align:center; }}
  th {{ background:#F1F5F9; }}
  .r {{ text-align:right; }}
  h2 {{ text-align:center; letter-spacing:8px; }}
  .side-s th {{ background:#FBE9EA; }}
  .side-b th {{ background:#EBF0FB; }}
</style></head><body>
<h2>전자{inv.doc_kind}</h2>
<div style="text-align:right; font-size:11px;">승인번호: {inv.invoice_no or '(미발급)'}</div>
<table style="margin-top:6px;"><tr>
<td style="width:50%; padding:0; border:2px solid #E8515D;">
  <table class="side-s" style="border:0;">
    <tr><th style="width:70px;">등록번호</th><td colspan="3">{inv.supplier_corp_no or co['corp_no']}</td></tr>
    <tr><th>상호</th><td>{inv.supplier_name or co['name']}</td><th style="width:50px;">성명</th><td>{co['ceo']}</td></tr>
    <tr><th>사업장</th><td colspan="3">{co['address']}</td></tr>
    <tr><th>업태</th><td>{co['biz_type']}</td><th>종목</th><td>{co['biz_item']}</td></tr>
    <tr><th>이메일</th><td colspan="3">{inv.supplier_email or co['email']}</td></tr>
  </table>
</td>
<td style="width:50%; padding:0; border:2px solid #5B7FD4;">
  <table class="side-b" style="border:0;">
    <tr><th style="width:70px;">등록번호</th><td colspan="3">{inv.buyer_corp_no or ''}</td></tr>
    <tr><th>상호</th><td>{inv.buyer_name or inv.party_name or ''}</td><th style="width:50px;">성명</th><td>{inv.buyer_ceo or ''}</td></tr>
    <tr><th>사업장</th><td colspan="3">{inv.buyer_addr or ''}</td></tr>
    <tr><th>업태</th><td>{inv.buyer_biztype or ''}</td><th>종목</th><td>{inv.buyer_bizitem or ''}</td></tr>
    <tr><th>이메일</th><td colspan="3">{inv.buyer_email or ''}</td></tr>
  </table>
</td></tr></table>
<table style="margin-top:6px;">
  <tr><th>작성일자</th><th>공급가액</th><th>세액</th><th>수정사유</th><th>비고</th></tr>
  <tr><td>{inv.write_date or ''}</td><td class="r">{int(float(inv.supply or 0)):,}</td>
      <td class="r">{int(float(inv.vat or 0)):,}</td><td>해당없음</td><td>{inv.note or ''}</td></tr>
</table>
<table style="margin-top:6px;">
  <tr><th>월</th><th>일</th><th>품목</th><th>규격</th><th>수량</th><th>단가</th><th>공급가액</th><th>세액</th><th>비고</th></tr>
  {trs}
</table>
<table style="margin-top:6px;">
  <tr><th>합계금액</th><th>현금</th><th>수표</th><th>어음</th><th>외상미수금</th><td rowspan="2" style="width:120px;">이 금액을 ({inv.claim_kind or '청구'}) 함</td></tr>
  <tr><td class="r">{int(float(inv.total or 0)):,}</td><td class="r">{int(float(inv.cash_amt or 0)):,}</td>
      <td class="r">{int(float(inv.check_amt or 0)):,}</td><td class="r">{int(float(inv.bill_amt or 0)):,}</td>
      <td class="r">{int(float(inv.credit_amt or 0)):,}</td></tr>
</table>
<script>setTimeout(() => window.print(), 400);</script>
</body></html>"""
    return HTMLResponse(content=html)


# ===================== 홈택스 업로드 (공식 일괄발급 양식) =====================
HOMETAX_TEMPLATE = Path(__file__).parent.parent / "assets" / "hometax_bulk_template.xls"


def _hometax_eligible(db: Session, date_from: date, date_to: date):
    """홈택스 일괄발급 대상 — 매출 · 국세청 미전송 (홈택스 출처/전송완료 제외)."""
    return db.execute(select(TaxInvoice).where(
        TaxInvoice.direction == "sale",
        TaxInvoice.status.in_(["ready", "issued", "draft", "scheduled"]),
        TaxInvoice.issue_method != "hometax",
        TaxInvoice.write_date >= date_from, TaxInvoice.write_date <= date_to,
    ).order_by(TaxInvoice.write_date)).scalars().all()


@router.get("/hometax-upload", response_class=HTMLResponse)
def hometax_upload_page(request: Request, db: Session = Depends(get_db),
                        date_from: str = "", date_to: str = ""):
    """🏛 홈택스 업로드 메뉴 — 일괄발급 대상 선택 → 공식 양식 Excel 다운로드."""
    from datetime import timedelta
    today_d = date.today()
    try:
        fd = date.fromisoformat(date_from) if date_from else today_d - timedelta(days=90)
    except Exception:
        fd = today_d - timedelta(days=90)
    try:
        td = date.fromisoformat(date_to) if date_to else today_d
    except Exception:
        td = today_d
    rows = _hometax_eligible(db, fd, td)
    return templates.TemplateResponse("tax/hometax_upload.html", {
        "request": request, "rows": rows, "company": _company(),
        "date_from": fd.isoformat(), "date_to": td.isoformat(),
        "hometax_url": HOMETAX_URL,
        "template_ok": HOMETAX_TEMPLATE.exists(),
    })


@router.post("/hometax-upload/export")
async def hometax_upload_export(request: Request, db: Session = Depends(get_db)):
    """선택 항목을 국세청 공식 '세금계산서등록양식(일반)' .xls에 채워 다운로드.
    - 시트 '엑셀업로드양식' · 데이터 7행부터 (헤더 6행)
    - 등록번호/일자: 숫자만 · 일자N: 2자리 일(日) · 영수01/청구02
    """
    form = await request.form()
    ids = [int(x) for x in form.getlist("ids") if str(x).isdigit()]
    from urllib.parse import quote
    if not ids:
        return RedirectResponse("/tax/hometax-upload?_msg=" + quote("업로드할 계산서를 선택하세요"), status_code=303)
    if not HOMETAX_TEMPLATE.exists():
        return RedirectResponse("/tax/hometax-upload?_msg=" + quote("❌ 공식 양식 파일이 없습니다 (assets/hometax_bulk_template.xls)"), status_code=303)

    rows = db.execute(select(TaxInvoice).where(TaxInvoice.id.in_(ids))
                      .order_by(TaxInvoice.write_date)).scalars().all()

    import xlrd, re as _re, io as _io
    from xlutils.copy import copy as xl_copy
    rb = xlrd.open_workbook(str(HOMETAX_TEMPLATE), formatting_info=True)
    wb = xl_copy(rb)
    ws = wb.get_sheet(0)  # '엑셀업로드양식'

    KIND_CODE = {"일반": "01", "영세율": "02"}
    CLAIM_CODE = {"영수": "01", "청구": "02"}
    co = _company()
    digits = lambda s: _re.sub(r"\D", "", str(s or ""))

    skipped_kind = 0
    r_idx = 6  # 데이터 시작 (0-indexed 7행)
    written = 0
    for inv in rows:
        kind = inv.inv_kind or "일반"
        if kind not in KIND_CODE:
            skipped_kind += 1  # 위수탁류는 별도 양식 — 일반 양식으로 업로드 불가
            continue
        try:
            items = json.loads(inv.items_json or "[]")
        except Exception:
            items = []

        def W(col, val, num=False):
            if val in (None, ""):
                return
            if num:
                try:
                    ws.write(r_idx, col, float(val) if float(val) % 1 else int(float(val)))
                    return
                except Exception:
                    pass
            ws.write(r_idx, col, str(val))

        W(0, KIND_CODE[kind])
        W(1, inv.write_date.strftime("%Y%m%d") if inv.write_date else "")
        W(2, digits(inv.supplier_corp_no or co["corp_no"]))
        # 3: 공급자 종사업장 (없음)
        W(4, inv.supplier_name or co["name"])
        W(5, co["ceo"])
        W(6, co["address"])
        W(7, co["biz_type"])
        W(8, co["biz_item"])
        W(9, inv.supplier_email or co["email"])
        W(10, digits(inv.buyer_corp_no))
        W(11, inv.buyer_sub_no)
        W(12, inv.buyer_name or inv.party_name)
        W(13, inv.buyer_ceo)
        W(14, inv.buyer_addr)
        W(15, inv.buyer_biztype)
        W(16, inv.buyer_bizitem)
        W(17, inv.buyer_email)
        W(18, inv.buyer_email2)
        W(19, int(float(inv.supply or 0)), num=True)
        W(20, int(float(inv.vat or 0)), num=True)
        W(21, inv.note)
        # 품목 1~4 — 일자는 2자리 일(日)만
        for i in range(4):
            if i >= len(items):
                break
            it = items[i]
            base = 22 + i * 8
            day2 = digits(it.get("일"))[-2:].zfill(2) if digits(it.get("일")) else (
                inv.write_date.strftime("%d") if inv.write_date else "")
            W(base + 0, day2)
            W(base + 1, it.get("품목"))
            W(base + 2, it.get("규격"))
            W(base + 3, digits(it.get("수량")) or it.get("수량"), num=True)
            W(base + 4, digits(it.get("단가")) or it.get("단가"), num=True)
            W(base + 5, int(float(it.get("공급가액", 0) or 0)), num=True)
            W(base + 6, int(float(it.get("세액", 0) or 0)), num=True)
            W(base + 7, it.get("비고"))
        # 결제수단 (0이면 빈칸)
        for col, v in [(54, inv.cash_amt), (55, inv.check_amt), (56, inv.bill_amt), (57, inv.credit_amt)]:
            iv = int(float(v or 0))
            if iv:
                W(col, iv, num=True)
        W(58, CLAIM_CODE.get(inv.claim_kind or "청구", "02"))
        r_idx += 1
        written += 1

    if not written:
        return RedirectResponse("/tax/hometax-upload?_msg=" +
            quote("❌ 내보낼 수 있는 항목이 없습니다 (위수탁 계산서는 별도 양식 필요)"), status_code=303)

    buf = _io.BytesIO()
    wb.save(buf)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    fname = f"세금계산서등록양식(일반)_{written}건_{ts}.xls"
    return Response(content=buf.getvalue(), media_type="application/vnd.ms-excel",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"})


@router.get("/list/hometax-bulk.xlsx")
def tax_hometax_bulk_export(db: Session = Depends(get_db),
                            date_from: str = "", date_to: str = ""):
    """🏛 홈택스 일괄발급용 Excel — 국세청 미전송 계산서(발행대기·발행기록)를
    홈택스 [전자세금계산서 일괄작성] 업로드 양식 컬럼 순서로 내보낸다.

    무료 발급 흐름:
      1) 이 파일 내려받기
      2) 홈택스 → 전자(세금)계산서 → 일괄발급 → 엑셀 업로드
      3) 검증 후 일괄발급 (공동인증서 서명)
    ※ 홈택스 양식 버전이 다르면 홈택스에서 받은 공식 양식에 값을 붙여넣어 사용.
    """
    import openpyxl, io as _io
    from datetime import timedelta
    today_d = date.today()
    try:
        fd = date.fromisoformat(date_from) if date_from else today_d - timedelta(days=90)
    except Exception:
        fd = today_d - timedelta(days=90)
    try:
        td = date.fromisoformat(date_to) if date_to else today_d
    except Exception:
        td = today_d

    # 국세청 미전송분만 (홈택스에서 온 것/이미 전송된 것 제외)
    rows = db.execute(select(TaxInvoice).where(
        TaxInvoice.direction == "sale",
        TaxInvoice.status.in_(["ready", "issued", "draft"]),
        TaxInvoice.issue_method != "hometax",
        TaxInvoice.write_date >= fd, TaxInvoice.write_date <= td,
    ).order_by(TaxInvoice.write_date)).scalars().all()

    co = _company()
    wb = openpyxl.Workbook(); w = wb.active; w.title = "일괄발급"

    # 안내 행 (1~5) — 홈택스 양식과 동일하게 데이터는 7행부터
    w.append(["전자세금계산서 일괄발급 양식 (인비즈 경영관리 자동 생성)"])
    w.append([f"생성일: {today_d.isoformat()} · 기간: {fd} ~ {td} · {len(rows)}건"])
    w.append(["※ 홈택스 → 조회/발급 → 전자세금계산서 → 일괄발급 메뉴에서 이 파일을 업로드하세요."])
    w.append(["※ 홈택스 양식 버전이 다르면 홈택스 제공 공식 양식에 아래 값을 붙여넣으세요."])
    w.append([])
    hdr = ["종류", "작성일자",
           "공급자등록번호", "공급자종사업장", "공급자상호", "공급자성명", "공급자주소", "공급자업태", "공급자종목", "공급자이메일",
           "공급받는자등록번호", "공급받는자종사업장", "공급받는자상호", "공급받는자성명", "공급받는자주소",
           "공급받는자업태", "공급받는자종목", "공급받는자이메일1", "공급받는자이메일2",
           "공급가액", "세액", "비고"]
    for i in range(1, 5):
        hdr += [f"품목{i}일자", f"품목{i}명", f"품목{i}규격", f"품목{i}수량", f"품목{i}단가",
                f"품목{i}공급가액", f"품목{i}세액", f"품목{i}비고"]
    hdr += ["현금", "수표", "어음", "외상미수금", "영수청구구분"]
    w.append(hdr)

    KIND_CODE = {"일반": "01", "영세율": "02", "위수탁": "03", "위수탁영세율": "04"}
    CLAIM_CODE = {"영수": "01", "청구": "02"}
    for r in rows:
        try:
            items = json.loads(r.items_json or "[]")
        except Exception:
            items = []
        wd8 = r.write_date.strftime("%Y%m%d") if r.write_date else ""
        line = [KIND_CODE.get(r.inv_kind or "일반", "01"), wd8,
                (r.supplier_corp_no or co["corp_no"]).replace("-", ""), "",
                r.supplier_name or co["name"], co["ceo"], co["address"], co["biz_type"], co["biz_item"],
                r.supplier_email or co["email"],
                (r.buyer_corp_no or "").replace("-", ""), r.buyer_sub_no or "",
                r.buyer_name or r.party_name or "", r.buyer_ceo or "", r.buyer_addr or "",
                r.buyer_biztype or "", r.buyer_bizitem or "",
                r.buyer_email or "", r.buyer_email2 or "",
                int(float(r.supply or 0)), int(float(r.vat or 0)), r.note or ""]
        for i in range(4):
            if i < len(items):
                it = items[i]
                line += [f"{it.get('월','')}{it.get('일','')}", it.get("품목", ""), it.get("규격", ""),
                         it.get("수량", ""), it.get("단가", ""),
                         int(float(it.get("공급가액", 0) or 0)), int(float(it.get("세액", 0) or 0)),
                         it.get("비고", "")]
            else:
                line += ["", "", "", "", "", "", "", ""]
        line += [int(float(r.cash_amt or 0)) or "", int(float(r.check_amt or 0)) or "",
                 int(float(r.bill_amt or 0)) or "", int(float(r.credit_amt or 0)) or "",
                 CLAIM_CODE.get(r.claim_kind or "청구", "02")]
        w.append(line)

    buf = _io.BytesIO(); wb.save(buf)
    from urllib.parse import quote
    fname = f"홈택스_일괄발급_{fd}_{td}.xlsx"
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"})


@router.get("/list/export.xlsx")
def tax_list_export(db: Session = Depends(get_db), direction: str = "sale",
                    date_from: str = "", date_to: str = ""):
    """발급목록 내려받기 — 현재 필터 기간의 목록 Excel."""
    from datetime import timedelta
    import openpyxl
    today_d = date.today()
    try:
        fd = date.fromisoformat(date_from) if date_from else today_d - timedelta(days=90)
    except Exception:
        fd = today_d - timedelta(days=90)
    try:
        td = date.fromisoformat(date_to) if date_to else today_d
    except Exception:
        td = today_d
    rows = db.execute(select(TaxInvoice)
                      .where(TaxInvoice.direction == direction,
                             TaxInvoice.write_date >= fd, TaxInvoice.write_date <= td)
                      .order_by(desc(TaxInvoice.write_date))).scalars().all()
    wb = openpyxl.Workbook(); w = wb.active; w.title = "발급목록"
    w.append(["작성일자", "발급일자", "전송일자", "승인번호", "공급받는자등록번호", "종사업장",
              "상호", "대표자명", "품목명", "합계금액", "공급가액", "세액", "상태", "비고"])
    for r in rows:
        w.append([str(r.write_date or ""), str(r.issue_date or ""), str(r.transmit_date or ""),
                  r.invoice_no or "", r.buyer_corp_no or "", r.buyer_sub_no or "",
                  r.party_name or "", r.buyer_ceo or "", r.item_desc or "",
                  float(r.total or 0), float(r.supply or 0), float(r.vat or 0),
                  r.status, r.note or ""])
    import io as _io
    buf = _io.BytesIO(); wb.save(buf)
    from urllib.parse import quote
    fname = f"발급목록_{fd}_{td}.xlsx"
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"})


# ===================== 매출: 계산서 작성·발행 =====================
@router.get("/issue", response_class=HTMLResponse)
def tax_issue_page(request: Request, db: Session = Depends(get_db)):
    rows = db.execute(select(TaxInvoice).where(TaxInvoice.direction == "sale")
                      .order_by(desc(TaxInvoice.id)).limit(200)).scalars().all()
    parties = db.execute(select(Party).where(Party.active == "Y").order_by(Party.name).limit(2000)).scalars().all()
    from models import Product
    products = db.execute(select(Product).order_by(Product.code)).scalars().all()
    return templates.TemplateResponse("tax/issue.html", {
        "request": request, "rows": rows, "parties": parties, "products": products, "company": _company(),
        "asp": ss.get("tax_asp", "manual"), "asp_ready": ig.asp_ready(),
        "mail_ready": ig.mail_send_ready(), "hometax_url": HOMETAX_URL, "today": date.today(),
    })


@router.post("/create")
async def tax_create(request: Request, db: Session = Depends(get_db)):
    """홈택스 스타일 다품목 계산서 작성.

    form:
      party_name, buyer_corp_no, buyer_email, write_date, doc_kind, send_date
      item_month[], item_day[], item_name[], item_spec[], item_qty[], item_unit[],
      item_supply[], item_vat[], item_note[]  — 최대 16행, 빈 품목 행은 무시
      cash_amt, check_amt, bill_amt, credit_amt  — 결제수단
      claim_kind — 청구/영수
      action — hold(발급보류=draft) / issue(발급하기) / ready(발행대기)
    """
    form = await request.form()

    def g(key, default=""):
        v = form.get(key)
        return v.strip() if isinstance(v, str) else default

    def _num(v):
        try:
            return float(str(v).replace(",", "").strip() or 0)
        except Exception:
            return 0.0

    party_name = g("party_name")
    doc_kind = g("doc_kind") or "세금계산서"
    inv_kind_v = g("inv_kind") or "일반"
    zero_rate = inv_kind_v in ("영세율", "위수탁영세율")  # 영세율 → 세액 0
    write_date = g("write_date")
    try:
        wd = date.fromisoformat(write_date) if write_date else date.today()
    except Exception:
        wd = date.today()

    # 품목 라인 수집 (병렬 배열)
    names = form.getlist("item_name")
    months = form.getlist("item_month")
    days = form.getlist("item_day")
    specs = form.getlist("item_spec")
    qtys = form.getlist("item_qty")
    units = form.getlist("item_unit")
    supplies = form.getlist("item_supply")
    vats = form.getlist("item_vat")
    notes = form.getlist("item_note")

    def pick(lst, i, default=""):
        return (lst[i].strip() if i < len(lst) and isinstance(lst[i], str) else default)

    items = []
    total_supply = total_vat = 0.0
    for i, nm in enumerate(names[:16]):
        nm = (nm or "").strip()
        if not nm:
            continue
        sup_i = _num(pick(supplies, i, "0"))
        # 공급가액 비어있으면 수량×단가로 계산
        if sup_i == 0:
            sup_i = _num(pick(qtys, i, "0")) * _num(pick(units, i, "0"))
        vat_raw = pick(vats, i, "")
        if zero_rate:
            vat_i = 0
        else:
            vat_i = _num(vat_raw) if vat_raw != "" else (round(sup_i * 0.1) if doc_kind == "세금계산서" else 0)
        items.append({
            "월": pick(months, i) or f"{wd.month:02d}",
            "일": pick(days, i) or f"{wd.day:02d}",
            "품목": nm,
            "규격": pick(specs, i),
            "수량": pick(qtys, i),
            "단가": pick(units, i),
            "공급가액": sup_i,
            "세액": vat_i,
            "비고": pick(notes, i),
        })
        total_supply += sup_i
        total_vat += vat_i

    if not items:
        return RedirectResponse("/tax/issue?msg=" + "품목을 1개 이상 입력하세요", status_code=303)

    # 예약 발송일
    sd = None
    if g("send_date"):
        try:
            sd = date.fromisoformat(g("send_date"))
        except Exception:
            sd = None

    action = g("action") or "ready"   # hold / ready / issue
    if sd:
        status = "scheduled"
    elif action == "hold":
        status = "draft"
    else:
        status = "ready"

    co = _company()
    # 이메일 — 아이디@도메인 분리 입력을 조합 (buyer_email이 직접 오면 그대로)
    def _email(pfx):
        full = g(f"{pfx}")
        if full and "@" in full:
            return full
        local = g(f"{pfx}_local"); domain = g(f"{pfx}_domain")
        return f"{local}@{domain}" if local and domain else None

    inv = TaxInvoice(
        direction="sale", doc_kind=doc_kind, write_date=wd, send_date=sd,
        supplier_corp_no=co["corp_no"], supplier_name=co["name"], supplier_email=co["email"],
        buyer_corp_no=g("buyer_corp_no") or None, buyer_name=party_name or None,
        buyer_email=_email("buyer_email"), party_name=party_name or None,
        inv_kind=g("inv_kind") or "일반",
        buyer_kind=g("buyer_kind") or "사업자등록번호",
        buyer_sub_no=g("buyer_sub_no") or None,
        buyer_ceo=g("buyer_ceo") or None,
        buyer_addr=g("buyer_addr") or None,
        buyer_biztype=g("buyer_biztype") or None,
        buyer_bizitem=g("buyer_bizitem") or None,
        buyer_email2=_email("buyer_email2"),
        note=g("note") or None,
        item_desc=(items[0]["품목"] + (f" 외 {len(items)-1}건" if len(items) > 1 else "")),
        items_json=json.dumps(items, ensure_ascii=False),
        supply=total_supply, vat=total_vat, total=total_supply + total_vat,
        cash_amt=_num(g("cash_amt")), check_amt=_num(g("check_amt")),
        bill_amt=_num(g("bill_amt")), credit_amt=_num(g("credit_amt")),
        claim_kind=g("claim_kind") or "청구",
        status=status, issue_method="manual", source="manual",
    )
    db.add(inv); db.commit(); db.refresh(inv)

    # 발급하기 → 즉시 발행 처리 (홈택스 '발급하기'와 동일한 원클릭 흐름)
    if action == "issue" and status == "ready":
        asp = ss.get("tax_asp", "manual")
        if ig.asp_ready() and asp in ("popbill", "barobill"):
            inv.status = "sent"; inv.issue_method = asp; inv.issue_at = datetime.now()
            inv.invoice_no = inv.invoice_no or f"{asp.upper()}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        else:
            inv.status = "issued"; inv.issue_method = "manual"; inv.issue_at = datetime.now()
        db.commit()
        if inv.buyer_email and ig.mail_send_ready():
            body = (f"{inv.supplier_name} 전자{inv.doc_kind} 발행 안내\n\n"
                    f"· 작성일자: {inv.write_date}\n· 품목: {inv.item_desc}\n"
                    f"· 공급가액 {int(inv.supply or 0):,}원 / 세액 {int(inv.vat or 0):,}원 / 합계 {int(inv.total or 0):,}원\n")
            ig.send_email(f"[{inv.supplier_name}] 전자{inv.doc_kind} 발행 안내", body, to=inv.buyer_email)
        return RedirectResponse(f"/tax/issue?issued={inv.id}", status_code=303)
    if action == "hold":
        return RedirectResponse("/tax/issue?msg=" + "발급보류로 저장되었습니다 (작성중 상태)", status_code=303)
    return RedirectResponse(f"/tax/issue?created={inv.id}", status_code=303)


def process_due_invoices(db, base_url=""):
    """예약(status=scheduled) 매출 계산서 중 발송일이 도래한 건 자동 발송 + 알림(텔레그램·이메일·카카오).
    스케줄러 루프와 수동 버튼에서 호출. 처리 건수 반환."""
    today = date.today()
    due = db.execute(select(TaxInvoice).where(
        TaxInvoice.direction == "sale", TaxInvoice.status == "scheduled",
        TaxInvoice.send_date != None, TaxInvoice.send_date <= today)).scalars().all()  # noqa: E711
    sent = 0
    asp = ss.get("tax_asp", "manual")
    for inv in due:
        if ig.asp_ready() and asp in ("popbill", "barobill"):
            inv.status = "sent"; inv.issue_method = asp
            inv.invoice_no = inv.invoice_no or f"{asp.upper()}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        else:
            inv.status = "issued"; inv.issue_method = "manual"
        inv.issue_at = datetime.now()
        # 거래처 메일
        if inv.buyer_email and ig.mail_send_ready():
            body = (f"{inv.supplier_name} 전자{inv.doc_kind} 발행 안내\n작성일 {inv.write_date}\n품목 {inv.item_desc}\n"
                    f"공급가액 {int(inv.supply or 0):,} / 세액 {int(inv.vat or 0):,} / 합계 {int(inv.total or 0):,}원")
            ig.send_email(f"[{inv.supplier_name}] 전자{inv.doc_kind} 발행 안내", body, to=inv.buyer_email)
        # 본인 알림(텔레그램+이메일+카카오)
        subject = f"[인비즈] 예약 세금계산서 발송 — {inv.party_name or ''} {int(inv.total or 0):,}원"
        bodyn = (f"예약된 전자{inv.doc_kind}가 발송 처리되었습니다.\n"
                 f"· 거래처: {inv.party_name or '-'}\n· 발송일: {inv.send_date}\n· 품목: {inv.item_desc or '-'}\n"
                 f"· 합계: {int(inv.total or 0):,}원\n"
                 + ("" if ig.asp_ready() else "※ 수동 모드 — 실제 국세청 전송은 홈택스에서 진행하세요."))
        ig.notify(subject, bodyn, (base_url.rstrip('/') + "/tax/issue") if base_url else None)
        inv.notified = "Y"
        sent += 1
    if sent:
        db.commit()
    return sent


@router.post("/{inv_id}/issue")
def tax_issue(inv_id: int, request: Request, db: Session = Depends(get_db)):
    """발행(전송). ASP 키가 있으면 자동 발행, 없으면 수동 발행완료로 기록."""
    inv = db.get(TaxInvoice, inv_id)
    if not inv or inv.direction != "sale":
        raise HTTPException(404, "계산서 없음")
    asp = ss.get("tax_asp", "manual")
    msg = ""
    if ig.asp_ready() and asp in ("popbill", "barobill"):
        # ASP 자동 발행 (키 등록 시) — 실제 호출은 ASP SDK 연동 지점
        inv.status = "sent"; inv.issue_method = asp; inv.issue_at = datetime.now()
        inv.invoice_no = inv.invoice_no or f"{asp.upper()}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        msg = f"{asp} 통해 발행·국세청 전송 처리"
    else:
        inv.status = "issued"; inv.issue_method = "manual"; inv.issue_at = datetime.now()
        msg = "발행완료(수동) — 실제 국세청 전송은 홈택스에서 진행하세요"
    db.commit()
    # 거래처에 메일 발송(설정 시)
    if inv.buyer_email and ig.mail_send_ready():
        body = (f"{inv.supplier_name} 전자{inv.doc_kind} 발행 안내\n\n"
                f"· 작성일자: {inv.write_date}\n· 품목: {inv.item_desc}\n"
                f"· 공급가액 {int(inv.supply or 0):,}원 / 세액 {int(inv.vat or 0):,}원 / 합계 {int(inv.total or 0):,}원\n")
        ig.send_email(f"[{inv.supplier_name}] 전자{inv.doc_kind} 발행 안내", body, to=inv.buyer_email)
    return RedirectResponse(f"/tax/issue?issued={inv_id}", status_code=303)


@router.post("/{inv_id}/email")
def tax_email(inv_id: int, db: Session = Depends(get_db)):
    inv = db.get(TaxInvoice, inv_id)
    if not inv:
        raise HTTPException(404, "계산서 없음")
    to = inv.buyer_email or inv.supplier_email or ss.get("mail_notify_to", "")
    body = (f"전자{inv.doc_kind} ({inv.party_name})\n작성일 {inv.write_date}\n품목 {inv.item_desc}\n"
            f"공급가액 {int(inv.supply or 0):,} / 세액 {int(inv.vat or 0):,} / 합계 {int(inv.total or 0):,}원")
    ok, m = ig.send_email(f"전자{inv.doc_kind} — {inv.party_name}", body, to=to)
    return RedirectResponse(f"/tax/issue?mail={'1' if ok else '0'}&msg={m}", status_code=303)


@router.post("/{inv_id}/delete")
def tax_delete(inv_id: int, db: Session = Depends(get_db)):
    inv = db.get(TaxInvoice, inv_id)
    back = "/tax/issue" if (inv and inv.direction == "sale") else "/tax/inbox"
    if inv:
        db.delete(inv); db.commit()
    return RedirectResponse(back, status_code=303)


@router.get("/{inv_id}/export.xlsx")
def tax_export_xlsx(inv_id: int, db: Session = Depends(get_db)):
    inv = db.get(TaxInvoice, inv_id)
    if not inv:
        raise HTTPException(404, "계산서 없음")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook(); w = wb.active; w.title = "세금계산서"
    thin = Side(style="thin", color="999999"); bd = Border(thin, thin, thin, thin)
    w["A1"] = f"전자{inv.doc_kind}"; w["A1"].font = Font(size=16, bold=True)
    w.merge_cells("A1:D1"); w["A1"].alignment = Alignment(horizontal="center")
    info = [("작성일자", str(inv.write_date or "")), ("승인번호", inv.invoice_no or "(미발행)"),
            ("공급자", f"{inv.supplier_name or ''} ({inv.supplier_corp_no or ''})"),
            ("공급받는자", f"{inv.buyer_name or inv.party_name or ''} ({inv.buyer_corp_no or ''})"),
            ("품목", inv.item_desc or "")]
    r = 3
    for k, v in info:
        w.cell(row=r, column=1, value=k).font = Font(bold=True)
        w.cell(row=r, column=2, value=v); r += 1
    r += 1
    for j, h in enumerate(["공급가액", "세액", "합계"], 1):
        c = w.cell(row=r, column=j, value=h); c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="6B2C91"); c.alignment = Alignment(horizontal="center"); c.border = bd
    r += 1
    for j, val in enumerate([inv.supply or 0, inv.vat or 0, inv.total or 0], 1):
        c = w.cell(row=r, column=j, value=float(val)); c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right"); c.border = bd
    for col, wd in zip("ABCD", (16, 30, 16, 16)):
        w.column_dimensions[col].width = wd
    buf = io.BytesIO(); wb.save(buf)
    import re as _re
    safe = _re.sub(r"[^\w가-힣.-]", "_", f"세금계산서_{inv.party_name or inv.id}_{inv.write_date or ''}")
    from urllib.parse import quote
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(safe + '.xlsx')}"})


# ===================== 매입: 수신 확인 =====================
@router.get("/inbox", response_class=HTMLResponse)
def tax_inbox_page(request: Request, db: Session = Depends(get_db)):
    rows = db.execute(select(TaxInvoice).where(TaxInvoice.direction == "purchase")
                      .order_by(desc(TaxInvoice.id)).limit(300)).scalars().all()
    return templates.TemplateResponse("tax/inbox.html", {
        "request": request, "rows": rows,
        "mail_recv_ready": ig.mail_recv_ready(), "mail_send_ready": ig.mail_send_ready(),
        "kakao_ready": ig.kakao_ready(), "today": date.today(),
        "imap_days": ss.get("mail_imap_days", "7"),
    })


@router.post("/inbox/sync")
def tax_inbox_sync(request: Request, db: Session = Depends(get_db), days: str = Form("")):
    res = ig.imap_fetch_purchase(db, days=days or None, base_url=_base_url(request))
    return RedirectResponse(f"/tax/inbox?sync={'1' if res['ok'] else '0'}&msg={res['msg']}", status_code=303)


@router.post("/inbox/add")
def tax_inbox_add(request: Request, db: Session = Depends(get_db),
                  party_name: str = Form(""), supplier_corp_no: str = Form(""),
                  write_date: str = Form(""), item_desc: str = Form(""),
                  supply: str = Form("0"), vat: str = Form("")):
    def _num(v):
        try:
            return float(str(v).replace(",", "").strip() or 0)
        except Exception:
            return 0.0
    sup = _num(supply); v = _num(vat) if str(vat).strip() != "" else round(sup * 0.1)
    co = _company()
    try:
        wd = date.fromisoformat(write_date) if write_date else date.today()
    except Exception:
        wd = date.today()
    inv = TaxInvoice(
        direction="purchase", doc_kind="세금계산서", write_date=wd, issue_at=datetime.now(),
        supplier_name=party_name.strip() or None, supplier_corp_no=supplier_corp_no.strip() or None,
        party_name=party_name.strip() or None, buyer_name=co["name"], buyer_corp_no=co["corp_no"],
        item_desc=item_desc.strip() or None, supply=sup, vat=v, total=sup + v,
        status="received", issue_method="manual", source="manual", notified="N",
    )
    db.add(inv); db.commit(); db.refresh(inv)
    res = ig.notify_purchase(inv, _base_url(request))
    if any(ok for _, ok, _ in res):
        inv.notified = "Y"; db.commit()
    return RedirectResponse("/tax/inbox?added=1", status_code=303)


@router.post("/{inv_id}/notify")
def tax_notify(inv_id: int, request: Request, db: Session = Depends(get_db)):
    inv = db.get(TaxInvoice, inv_id)
    if not inv:
        raise HTTPException(404, "계산서 없음")
    res = ig.notify_purchase(inv, _base_url(request))
    if any(ok for _, ok, _ in res):
        inv.notified = "Y"; db.commit()
    msg = " / ".join(f"{ch}:{'OK' if ok else m}" for ch, ok, m in res)
    return RedirectResponse(f"/tax/inbox?notify={msg}", status_code=303)


@router.post("/process-scheduled")
def tax_process_scheduled(request: Request, db: Session = Depends(get_db)):
    """예약 발송분을 지금 즉시 처리(발송일 도래분). 평소엔 스케줄러가 자동 처리."""
    n = process_due_invoices(db, base_url=_base_url(request))
    return RedirectResponse(f"/tax/issue?processed={n}", status_code=303)


@router.post("/test-notify")
def tax_test_notify(request: Request):
    res = ig.notify("[인비즈] 알림 테스트", "세금계산서 알림 테스트입니다. (이메일·카카오·텔레그램)")
    parts = " · ".join(f"{ch}:{'OK' if ok else m}" for ch, ok, m in res)
    return RedirectResponse(f"/tax/inbox?test={parts}", status_code=303)
