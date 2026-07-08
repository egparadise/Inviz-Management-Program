# -*- coding: utf-8 -*-
"""템플릿 헬퍼·필터 — 단일 Jinja2Templates 인스턴스 공유"""
from pathlib import Path
from datetime import date, datetime
from fastapi.templating import Jinja2Templates


def load_workbook_any(raw: bytes):
    """업로드된 Excel을 openpyxl Workbook으로 반환.

    .xlsx/.xlsm → openpyxl 직접 로드
    .xls (BIFF binary) → pandas+xlrd로 읽어 메모리 .xlsx로 변환 후 로드
    HTML 위장 .xls (국세청 홈택스 등) → pandas.read_html로 파싱 후 변환

    모두 실패하면 마지막 예외를 raise.
    """
    import openpyxl
    import io as _io
    bio = _io.BytesIO(raw)
    last_err = None
    try:
        return openpyxl.load_workbook(bio, data_only=True, read_only=True)
    except Exception as e:
        last_err = e
    try:
        import pandas as pd
        dfs = pd.read_excel(_io.BytesIO(raw), sheet_name=None, header=None, engine="xlrd")
        buf = _io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            for name, df in dfs.items():
                df.to_excel(w, sheet_name=str(name)[:31] or "Sheet1", header=False, index=False)
        buf.seek(0)
        return openpyxl.load_workbook(buf, data_only=True, read_only=True)
    except Exception as e:
        last_err = e
    try:
        import pandas as pd
        dfs = pd.read_html(_io.BytesIO(raw))
        if not dfs:
            raise ValueError("HTML 표를 찾지 못했습니다")
        buf = _io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            for i, df in enumerate(dfs):
                df.to_excel(w, sheet_name=f"Sheet{i+1}", header=False, index=False)
        buf.seek(0)
        return openpyxl.load_workbook(buf, data_only=True, read_only=True)
    except Exception as e:
        last_err = e
    raise last_err


def normalize_header(hdr: list) -> list:
    """중복 헤더(예: '상호' 두 번) 발견 시 두 번째부터 .1, .2 suffix를 붙인다.
    pandas DataFrame과 동일 규칙. 홈택스 양식의 공급자/공급받는자 구분에 필요."""
    seen = {}
    out = []
    for h in hdr:
        key = (h or "").strip()
        if not key:
            out.append("")
            continue
        if key in seen:
            seen[key] += 1
            out.append(f"{key}.{seen[key]}")
        else:
            seen[key] = 0
            out.append(key)
    return out


def fmt_money(v):
    if v is None:
        return ""
    try:
        return f"{float(v):,.0f}"
    except Exception:
        return str(v)


def fmt_pct(v, digits=1):
    if v is None:
        return ""
    try:
        return f"{float(v):.{digits}f}%"
    except Exception:
        return str(v)


def fmt_date(v):
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.strftime("%Y-%m-%d")
    return str(v)


def fmt_period(v):
    return "" if v is None else str(v)


# 단일 공유 인스턴스 — Jinja2 캐시 비활성화 (Python 3.14 LRUCache 호환성 이슈 회피)
TEMPLATES_DIR = Path(__file__).parent / "templates"
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
templates.env.cache = None
templates.env.auto_reload = True

# 신구 시그니처 호환 어댑터:
#   구: TemplateResponse(name, {"request": req, ...})
#   신: TemplateResponse(request, name, {...})
_orig_tr = templates.TemplateResponse


def _compat_TemplateResponse(*args, **kwargs):
    if args and isinstance(args[0], str):
        name = args[0]
        ctx = args[1] if len(args) > 1 else kwargs.get("context", {})
        req = ctx.get("request") if isinstance(ctx, dict) else None
        return _orig_tr(req, name, ctx, **{k: v for k, v in kwargs.items() if k != "context"})
    return _orig_tr(*args, **kwargs)


templates.TemplateResponse = _compat_TemplateResponse
templates.env.filters["money"] = fmt_money
templates.env.filters["pct"] = fmt_pct
templates.env.filters["date"] = fmt_date
templates.env.filters["period"] = fmt_period


# 전역 설정 접근자 — 모든 템플릿에서 setting('키','기본값') 사용 가능
def _setting(key, default=None):
    try:
        import settings_store
        return settings_store.get(key, default)
    except Exception:
        return default


templates.env.globals["setting"] = _setting


# ===== 네비게이션 메뉴 정의 (설정에서 순서 조정 가능) =====
# 기본 순서: 대시보드 바로 아래 회사정보
NAV_ITEMS = [
    {"key": "dashboard",   "path": "/",            "icon": "🏠", "label": "대시보드", "exact": True},
    {"key": "calendar",    "path": "/calendar",    "icon": "📅", "label": "캘린더"},
    {"key": "company",     "path": "/company",     "icon": "🏢", "label": "회사정보",
     "children": [
        {"key": "employees", "path": "/employees", "icon": "👤", "label": "직원"},
        {"key": "payroll",   "path": "/payroll",   "icon": "💵", "label": "급여"},
     ]},
    {"key": "sales",       "path": "/sales",       "icon": "💰", "label": "매출",
     "children": [
        {"key": "product-sales", "path": "/product-sales", "icon": "📊", "label": "제품별 매출"},
        {"key": "products",      "path": "/products",      "icon": "📦", "label": "제품"},
        {"key": "tax-issue",     "path": "/tax/issue",     "icon": "🧾", "label": "계산서 작성·발행"},
        {"key": "tax-list",      "path": "/tax/list",      "icon": "🗂", "label": "발급목록 조회"},
        {"key": "tax-hometax",   "path": "/tax/hometax-upload", "icon": "🏛", "label": "홈택스 업로드"},
     ]},
    {"key": "purchases",   "path": "/purchases",   "icon": "🛒", "label": "매입",
     "children": [
        {"key": "tax-inbox", "path": "/tax/inbox", "icon": "📥", "label": "매입계산서 수신확인"},
        {"key": "expense",   "path": "/expense",   "icon": "💸", "label": "지출"},
     ]},
    {"key": "contracts",   "path": "/contracts",   "icon": "📋", "label": "계약"},
    {"key": "loans",       "path": "/loans",       "icon": "🏦", "label": "차입금"},
    {"key": "banking",     "path": "/banking",     "icon": "💳", "label": "자금/계좌",
     "children": [
        {"key": "bank-accounts", "path": "/banking",              "icon": "🏦", "label": "계좌 현황"},
        {"key": "bank-tx",       "path": "/banking/transactions", "icon": "💸", "label": "거래내역"},
        {"key": "bank-cards",    "path": "/banking/cards",        "icon": "💳", "label": "카드"},
     ]},
    {"key": "parties",     "path": "/parties",     "icon": "🤝", "label": "거래처"},
    {"key": "documents",   "path": "/documents",   "icon": "📁", "label": "서류·인증",
     "children": [
        {"key": "doc-company",   "path": "/documents/cat/company",       "icon": "🏢", "label": "회사 서류"},
        {"key": "doc-cert",      "path": "/documents/cat/certification", "icon": "📜", "label": "인증 서류"},
        {"key": "doc-product",   "path": "/documents/cat/product",       "icon": "📦", "label": "제품 서류"},
        {"key": "doc-mgmt",      "path": "/documents/cat/mgmt_contract", "icon": "📋", "label": "경영 계약서"},
        {"key": "doc-customer",  "path": "/documents/contracts",         "icon": "🤝", "label": "고객 계약서"},
     ]},
    {"key": "reports",     "path": "/reports",     "icon": "📑", "label": "보고서",
     "children": [
        {"key": "reports-closing",   "path": "/reports/closing",   "icon": "🧾", "label": "결산 보고"},
        {"key": "reports-financial", "path": "/reports/financial", "icon": "📊", "label": "재무제표"},
        {"key": "reports-investor",  "path": "/reports/investor",  "icon": "📈", "label": "투자사 보고"},
        {"key": "reports-tax",       "path": "/reports/tax",       "icon": "🧾", "label": "세금"},
     ]},
    {"key": "ai-analysis", "path": "/ai-classify", "icon": "🤖", "label": "AI 분석",
     "children": [
        {"key": "ai-classify", "path": "/ai-classify", "icon": "🗂", "label": "AI 분류"},
        {"key": "knowledge",   "path": "/knowledge",   "icon": "📚", "label": "AI 학습"},
     ]},
    {"key": "self-dev",    "path": "/self-dev",    "icon": "🛡", "label": "자가발전"},
    {"key": "sync",        "path": "/sync",        "icon": "🔄", "label": "동기화"},
]
NAV_DEFAULT_ORDER = [it["key"] for it in NAV_ITEMS]


def _nav_items():
    """설정(nav_order)에 따라 정렬된 메뉴 목록 반환. 누락/신규 항목은 기본 순서로 뒤에 추가."""
    import json
    by_key = {it["key"]: it for it in NAV_ITEMS}
    order = []
    try:
        import settings_store
        raw = settings_store.get("nav_order")
        if raw:
            order = json.loads(raw)
    except Exception:
        order = []
    keys = [k for k in order if k in by_key]
    seen = set(keys)
    # 저장된 순서에 없는(신규) 항목은 기본 순서상 위치에 삽입 (끝에 몰아넣지 않음)
    for idx, it in enumerate(NAV_ITEMS):
        if it["key"] in seen:
            continue
        pos = len(keys)
        for j in range(idx - 1, -1, -1):
            prev = NAV_ITEMS[j]["key"]
            if prev in keys:
                pos = keys.index(prev) + 1
                break
        else:
            pos = 0
        keys.insert(pos, it["key"]); seen.add(it["key"])
    return [by_key[k] for k in keys]


templates.env.globals["nav_items"] = _nav_items


# ===== 현재 활성 AI 모델 (모든 AI 사용 페이지에 표시) =====
def _ai_label(default="AI"):
    try:
        import llm_provider
        return llm_provider.active_label()
    except Exception:
        return default


def _ai_ready():
    try:
        import llm_provider
        ok, _ = llm_provider.provider_ready()
        return ok
    except Exception:
        return False


templates.env.globals["ai_label"] = _ai_label
templates.env.globals["ai_ready"] = _ai_ready


def _now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M")


templates.env.globals["now_str"] = _now_str


# 하위 호환을 위한 함수 (기존 라우터 코드 영향 최소화)
def register_filters(t):
    pass
